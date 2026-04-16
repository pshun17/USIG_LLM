"""
backtest_real.py  v2
Bloomberg US IG Corporate Bond — 완전한 Integrated_Score 기반 월별 리밸런싱 백테스팅

[Integrated_Score 구성]  Sentiment 제외 (역사적 데이터 불가), 나머지 4개 동일비중
  Bond_TR_Score  x 0.25  (Carry + Compression + DP_Rating_Score, class내 percentile)
  Eq_Mom_Score   x 0.25  (1M/3M return, 30D vol, 52W high — yfinance 역사적)
  Eq_Fund_Score  x 0.25  (D/E, Margin, Growth, CR, EV/EBITDA — yfinance 현재, 근사치)
  AI_Macro_Score x 0.25  (Sector×0.40 + Maturity×0.35 + RatingBuf×0.25)

[백테스트 방법론]
  - T 시점 스냅샷 스코어링 → T+1 실제 1mo Total Return 측정
  - 생존편향 없음: 각 월 실제 유니버스 사용
  - Look-ahead 없음: T 시점 데이터만 사용
"""

import os, re, sys, time, ssl, warnings, json
import numpy as np
import pandas as pd
from datetime import date, timedelta, datetime
from itertools import groupby
from collections import defaultdict

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import matplotlib.gridspec as gridspec

warnings.filterwarnings('ignore')

# SSL 우회 (사내망)
ssl._create_default_https_context = ssl._create_unverified_context
os.environ['CURL_CA_BUNDLE'] = ''
os.environ['REQUESTS_CA_BUNDLE'] = ''

# ── 설정 ──────────────────────────────────────────────────────────────────────
PATH_OLD  = r"Z:\USIGAIQ\USIG Market data\LUACSTAT"
PATH_NEW  = r"Z:\USIGAIQ\USIG Market data\LUACSTAT_PREP"
OUT       = r"C:\Users\sh.park\Documents\USIG_LLM\backtest_real.png"
OUT_XLSX  = r"C:\Users\sh.park\Documents\USIG_LLM\backtest_real_detail.xlsx"
EQ_CACHE  = r"C:\Users\sh.park\Documents\USIG_LLM\eq_price_cache.pkl"
FUND_CACHE= r"C:\Users\sh.park\Documents\USIG_LLM\eq_fund_cache.pkl"

BACKTEST_START = date(2021, 1, 1)
TOP_N          = 3
MIN_CLASS_SIZE = 5

# ── 등급 매핑 ────────────────────────────────────────────────────────────────
RTG_TEXT_TO_NUM = {
    'AAA':1,'AA+':2,'AA':3,'AA-':4,
    'A+':5,'A':6,'A-':7,
    'BBB+':8,'BBB':9,'BBB-':10,
    'BB+':11,'BB':12,'BB-':13,
}
RATING_CUT = 7   # A- 이상 = Comp <= 7

DP_RATING_MAP = {
    'AAA':1,'AA1':2,'AA2':3,'AA3':4,
    'A1':5,'A2':6,'A3':7,
    'BAA1':8,'BAA2':9,'BAA3':10,
    'BA1':11,'BA2':12,'BA3':13,
    'B1':14,'B2':15,'B3':16,
    'CAA1':17,'CAA2':18,'CAA3':19,
    'CA':20,'C':21,
}

# ── AI Sector Score 맵 (SUBGROUP_SCORE_MAP — ai_macro_score.py 동일) ──────────
SUBGROUP_SCORE_MAP = {
    'Electric-Integrated':0.85,'Electric-Distribution':0.85,'Electric-Transmission':0.80,
    'Electric-Generation':0.75,'Gas-Distribution':0.80,'Water':0.85,
    'Non-hazardous Waste Disp':0.70,'Medical-Drugs':0.70,'Medical-Hospitals':0.75,
    'Medical-HMO':0.65,'Medical-Biomedical/Gene':0.60,'Medical Products':0.60,
    'Medical Instruments':0.55,'Medical Labs&Testing Srv':0.65,'Pharmacy Services':0.60,
    'Medical-Whsle Drug Dist':0.55,'Medical Imaging Systems':0.55,'Diagnostic Equipment':0.55,
    'Drug Delivery Systems':0.55,'Medical-Generic Drugs':0.50,'Phys Practice Mgmnt':0.60,
    'Medical-Outptnt/Home Med':0.60,'Diversified Banking Inst':0.55,
    'Super-Regional Banks-US':0.60,'Commer Banks-Eastern US':0.55,
    'Commer Banks-Southern US':0.55,'Commer Banks-Central US':0.50,
    'Commer Banks-Western US':0.50,'Commer Banks Non-US':0.30,'Money Center Banks':0.45,
    'Fiduciary Banks':0.65,'Life/Health Insurance':0.50,'Property/Casualty Ins':0.55,
    'Reinsurance':0.50,'Multi-line Insurance':0.45,'Insurance Brokers':0.55,
    'Financial Guarantee Ins':0.20,'Finance-Invest Bnkr/Brkr':0.35,
    'Invest Mgmnt/Advis Serv':0.40,'Private Equity':0.20,'Finance-Credit Card':0.30,
    'Finance-Leasing Compan':0.25,'Finance-Auto Loans':0.10,'Finance-Other Services':0.30,
    'Finance-Mtge Loan/Banker':0.20,'Finance-Consumer Loans':0.10,'Venture Capital':0.00,
    'Investment Companies':0.15,'Pipelines':0.30,'Oil Comp-Integrated':0.10,
    'Oil Comp-Explor&Prodtn':-0.10,'Oil Refining&Marketing':-0.05,'Oil-Field Services':-0.20,
    'Oil&Gas Drilling':-0.25,'Agricultural Chemicals':0.20,
    'Auto-Cars/Light Trucks':-0.90,'Auto-Med&Heavy Duty Trks':-0.70,
    'Auto/Trk Prts&Equip-Orig':-0.80,'Retail-Automobile':-0.60,
    'Retail-Discount':-0.20,'Retail-Building Products':-0.30,'Retail-Major Dept Store':-0.50,
    'Retail-Apparel/Shoe':-0.60,'Retail-Auto Parts':-0.40,'Retail-Sporting Goods':-0.40,
    'Retail-Consumer Electron':-0.50,'Retail-Restaurants':-0.10,'Retail-Gardening Prod':-0.30,
    'Hotels&Motels':-0.15,'Casino Hotels':-0.20,'Cruise Lines':-0.25,
    'E-Commerce/Products':-0.20,'E-Commerce/Services':-0.10,'Internet Content-Entmnt':0.00,
    'Apparel Manufacturers':-0.55,'Athletic Footwear':-0.55,'Recreational Vehicles':-0.40,
    'Beverages-Non-alcoholic':0.55,'Beverages-Wine/Spirits':0.45,'Brewery':0.45,
    'Food-Misc/Diversified':0.50,'Food-Retail':0.40,'Food-Confectionery':0.50,
    'Food-Meat Products':0.45,'Food-Baking':0.45,'Food-Wholesale/Distrib':0.40,
    'Poultry':0.40,'Coffee':0.45,'Cosmetics&Toiletries':0.55,'Soap&Cleaning Prepar':0.50,
    'Consumer Products-Misc':0.40,'Tobacco':0.60,'Agricultural Operations':0.35,
    'Electronic Compo-Semicon':-0.30,'Semicon Compo-Intg Circu':-0.25,
    'Semiconductor Equipment':-0.35,'Enterprise Software/Serv':0.10,
    'Applications Software':0.05,'Computer Services':0.15,'Data Processing/Mgmt':0.20,
    'Computer Aided Design':0.00,'Software Tools':0.00,'E-Marketing/Info':0.00,
    'Decision Support Softwar':0.05,'Computers':-0.15,'Computers-Memory Devices':-0.20,
    'Computers-Other':-0.10,'Electronic Connectors':-0.20,'Electronic Compo-Misc':-0.15,
    'Electronic Parts Distrib':-0.10,'Electronic Measur Instr':-0.10,
    'Electronic Secur Devices':-0.05,'Electronic Forms':0.00,'Networking Products':-0.10,
    'Wireless Equipment':-0.15,'Telecom Eq Fiber Optics':-0.10,
    'Telecommunication Equip':-0.10,'Industrial Automat/Robot':-0.10,
    'Instruments-Controls':-0.05,'Telephone-Integrated':0.10,'Cellular Telecom':0.05,
    'Cable/Satellite TV':0.00,'Multimedia':0.00,'Broadcast Serv/Program':0.00,
    'Advertising Agencies':-0.15,'Advertising Services':-0.15,'Telecom Services':0.10,
    'Web Portals/ISP':0.00,'Aerospace/Defense':0.40,'Aerospace/Defense-Equip':0.35,
    'Machinery-Farm':0.20,'Machinery-General Indust':-0.10,'Machinery-Constr&Mining':-0.10,
    'Machinery-Pumps':-0.05,'Machinery-Electric Util':0.10,'Tools-Hand Held':-0.10,
    'Diversified Manufact Op':-0.05,'Industrial Gases':0.30,'Chemicals-Diversified':-0.10,
    'Chemicals-Specialty':-0.05,'Coatings/Paint':-0.10,'Containers-Paper/Plastic':-0.20,
    'Paper&Related Products':-0.15,'Commercial Serv-Finance':0.10,'Commercial Services':0.05,
    'Consulting Services':0.10,'Non-Profit Charity':0.00,'Distribution/Wholesale':-0.10,
    'Office Automation&Equip':-0.15,'Office Supplies&Forms':-0.15,
    'Bldg Prod-Cement/Aggreg':-0.10,'Bldg Prod-Air&Heating':-0.10,'Bldg Prod-Wood':-0.15,
    'Bldg&Construct Prod-Misc':-0.10,'Bldg-Residential/Commer':-0.20,
    'Building-Maint&Service':-0.05,'Shipbuilding':0.20,'Power Conv/Supply Equip':0.10,
    'Transport-Rail':0.30,'Transport-Services':0.00,'Transport-Equip&Leasng':-0.10,
    'Transport-Truck':-0.10,'Transport-Marine':-0.10,'Airlines':-0.40,
    'Steel-Producers':0.20,'Steel Pipe&Tube':0.15,'Metal-Iron':0.10,'Metal-Aluminum':0.10,
    'Metal-Copper':0.00,'Metal-Diversified':0.00,'Diversified Minerals':0.00,
    'Gold Mining':0.30,'Metal Processors&Fabrica':0.05,'Schools':0.40,'Toys':-0.40,
    'Entertainment Software':0.00,'Engineering/R&D Services':0.15,'Vitamins&Nutrition Prod':0.30,
    'Rental Auto/Equipment':-0.10,'Mach Tools&Rel Products':-0.10,
    'Motorcycle/Motor Scooter':-0.50,'Chemicals-Plastics':-0.15,
    'Miscellaneous Manufactur':-0.10,'REITS-Industrial':0.10,'REITS-Warehouse/Industr':0.20,
    'REITS-Diversified':-0.10,'REITS-Apartments':-0.05,'REITS-Shopping Centers':-0.30,
    'REITS-Regional Malls':-0.40,'REITS-Office Property':-0.35,'REITS-Health Care':0.30,
    'REITS-Storage':0.10,'REITS-Single Tenant':0.00,'REITS-Hotels':-0.20,
    'REITS-Manufactured Homes':0.00,'REITS-Mortgage':-0.20,'Real Estate Mgmnt/Servic':-0.10,
}

# BCLASS3 fallback sector scores (AI Sector Score 매핑 없을 때)
BCLASS3_SCORE_MAP = {
    'Utility':0.75, 'Healthcare':0.60, 'Banking':0.50, 'Financial Other':0.30,
    'Insurance':0.50, 'Energy':0.05, 'Consumer Cyclical':0.20,
    'Consumer Non-Cyclical':0.45, 'Technology':0.00, 'Communications':0.05,
    'Industrial':0.00, 'Transportation':0.00, 'Basic Industry':-0.05,
    'Real Estate':-0.10, 'Government Related':0.30, 'Sovereign':0.20,
    'Quasi-Government':0.25, 'Supranational':0.40,
}

RATING_BUFFER_MAP = {
    'AAA':0.50,'AA1':0.55,'AA2':0.55,'AA3':0.50,
    'A1':0.70,'A2':0.70,'A3':0.65,
    'BAA1':0.25,'BAA2':0.05,'BAA3':-0.65,
    'BA1':-0.85,'BA2':-1.00,'BA3':-1.00,
    'B1':-1.00,'B2':-1.00,'B3':-1.00,
}

def maturity_bucket(yrs):
    if pd.isna(yrs): return None
    if yrs <= 3:     return '~3Y'
    elif yrs <= 7:   return '3~7Y'
    elif yrs <= 15:  return '7~15Y'
    else:            return '15Y+'

def maturity_score(oad):
    """OAD 기반 만기 포지셔닝 점수 (커브 스티프닝 환경)."""
    if pd.isna(oad): return 0.0
    oad = float(oad)
    if oad < 2:    return  0.00
    elif oad < 4:  return  0.30
    elif oad < 7:  return  1.00
    elif oad < 10: return  0.70
    elif oad < 13: return  0.00
    elif oad < 16: return -0.50
    else:          return -1.00

def percentile_norm(series):
    """Rank-normalize to [-1, +1]. 높을수록 좋음."""
    n = series.notna().sum()
    if n < 2: return pd.Series(np.nan, index=series.index)
    ranked = series.rank(method='average', na_option='keep')
    return (ranked - 1) / (n - 1) * 2 - 1

# ═══════════════════════════════════════════════════════════════════════════════
# 1. 월별 파일 목록
# ═══════════════════════════════════════════════════════════════════════════════
pattern = re.compile(r'^LUACSTAT_(\d{4})_(\d{2})_(\d{2})\.(xls|xlsx)$')
dated_dict = {}
for p in [PATH_OLD, PATH_NEW]:
    for f in os.listdir(p):
        m = pattern.match(f)
        if m:
            y,mo,d,ext = int(m.group(1)),int(m.group(2)),int(m.group(3)),m.group(4)
            try:
                dt = date(y,mo,d)
                if dt not in dated_dict or ext=='xlsx':
                    dated_dict[dt] = (os.path.join(p,f), ext)
            except: pass

dated_sorted = sorted(dated_dict.items())

monthly = []
for (yr,mo), grp in groupby(dated_sorted, key=lambda x:(x[0].year, x[0].month)):
    grp_list = list(grp)
    monthly.append(grp_list[-1])

monthly_bt = [(d,fp) for d,(fp,ext) in monthly if d.year >= BACKTEST_START.year-1]

print("="*65)
print(f"Backtest files: {len(monthly_bt)}")
print(f"Range: {monthly_bt[0][0]} ~ {monthly_bt[-1][0]}")

# ═══════════════════════════════════════════════════════════════════════════════
# 2. 파일 로딩 함수
# ═══════════════════════════════════════════════════════════════════════════════
def load_file(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    df_raw = pd.read_excel(filepath, header=None, nrows=15)
    header_row = None
    for i, row in df_raw.iterrows():
        vals = [str(v) for v in row.values]
        if 'ISIN' in vals and ('OAS' in vals or 'Yield to Worst' in vals):
            header_row = i
            break
    if header_row is None:
        return None, None

    df = pd.read_excel(filepath, header=header_row)

    # Rating 표준화
    if 'Composite Rating Num' not in df.columns and 'BB Comp' in df.columns:
        df['Composite Rating Num'] = df['BB Comp'].map(RTG_TEXT_TO_NUM)
    elif 'Composite Rating Num' not in df.columns and 'Index Rtg' in df.columns:
        df['Composite Rating Num'] = df['Index Rtg'].map(RTG_TEXT_TO_NUM)

    # Maturity 표준화
    if 'Mty (Yrs)' not in df.columns and 'Eff Mty (Yrs)' in df.columns:
        df.rename(columns={'Eff Mty (Yrs)': 'Mty (Yrs)'}, inplace=True)

    # OASD 표준화
    if 'OASD' not in df.columns and 'Spd Dur' in df.columns:
        df.rename(columns={'Spd Dur': 'OASD'}, inplace=True)

    # 집계행 분리
    isin_col = df['ISIN'] if 'ISIN' in df.columns else pd.Series(dtype=str)
    bond_mask = isin_col.astype(str).str.match(r'^[A-Z]{2}[A-Z0-9]{10}$', na=False)
    df_bonds = df[bond_mask].copy()

    # 숫자형 변환
    num_cols = ['OAS','OASD','OAD','Yield to Worst','Mty (Yrs)',
                'Composite Rating Num','1Y Dflt','Total Return - 1mo']
    for col in num_cols:
        if col in df_bonds.columns:
            df_bonds[col] = pd.to_numeric(df_bonds[col], errors='coerce')

    df_bonds['ISIN'] = df_bonds['ISIN'].astype(str).str.strip()
    df_bonds = df_bonds[df_bonds['ISIN'].str.match(r'^[A-Z]{2}[A-Z0-9]{10}$', na=False)]

    return df_bonds, None

# ═══════════════════════════════════════════════════════════════════════════════
# 3. Equity 데이터 사전 수집
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[Phase 1] Scanning monthly files for equity tickers...")

def get_eq_ticker(ticker_str):
    if pd.isna(ticker_str): return None
    t = str(ticker_str).strip()
    if t in ('', 'N/A', 'nan', 'None'): return None
    parts = t.split()
    t = parts[0] if parts else t
    # Bloomberg 형식 제거 (e.g. "AAPL US Equity" -> "AAPL")
    return t if t else None

# 모든 월 파일 스캔해서 unique tickers 수집
all_tickers = set()
ticker_col_candidates = ['Eqty Ticker', 'Equity Ticker', 'Ticker']

for dt_scan, fp_scan in monthly_bt:
    try:
        df_s, _ = load_file(fp_scan)
        if df_s is None: continue
        for col in ticker_col_candidates:
            if col in df_s.columns:
                ts = df_s[col].dropna().apply(get_eq_ticker).dropna().unique()
                all_tickers.update(ts)
                break
    except Exception as e:
        pass

all_tickers = sorted([t for t in all_tickers if t and len(t) <= 10])
print(f"  Found {len(all_tickers)} unique equity tickers across all monthly files")

# ── 주가 히스토리 수집 (캐시 활용) ──────────────────────────────────────────
PRICE_START = '2019-01-01'
PRICE_END   = date.today().strftime('%Y-%m-%d')

if os.path.exists(EQ_CACHE):
    print(f"  Loading price cache from {EQ_CACHE}...")
    price_cache = pd.read_pickle(EQ_CACHE)
    print(f"  Cache loaded: {len(price_cache)} tickers")
else:
    print(f"  Fetching price history for {len(all_tickers)} tickers (this may take a while)...")
    price_cache = {}
    BATCH = 50

    try:
        import yfinance as yf
        # SSL fix for yfinance
        try:
            import curl_cffi.requests as cffi_req
            yf.utils.requests = cffi_req
        except ImportError:
            pass

        for i in range(0, len(all_tickers), BATCH):
            batch = all_tickers[i:i+BATCH]
            pct = (i+len(batch))/len(all_tickers)*100
            print(f"  Batch {i//BATCH+1}/{(len(all_tickers)-1)//BATCH+1} ({pct:.0f}%)  ...", end='', flush=True)
            try:
                raw = yf.download(
                    batch, start=PRICE_START, end=PRICE_END,
                    progress=False, auto_adjust=True, threads=True
                )
                if isinstance(raw.columns, pd.MultiIndex):
                    close = raw['Close'] if 'Close' in raw.columns.get_level_values(0) else raw.iloc[:,0]
                else:
                    close = raw[['Close']] if 'Close' in raw.columns else raw
                for t in batch:
                    if t in close.columns:
                        s = close[t].dropna()
                        if len(s) > 0:
                            price_cache[t] = s
                print(f" OK ({sum(1 for t in batch if t in price_cache)} valid)")
            except Exception as e:
                print(f" WARN: {e}")
                # 개별 fetch fallback
                for t in batch:
                    if t in price_cache: continue
                    try:
                        s = yf.download(t, start=PRICE_START, end=PRICE_END,
                                        progress=False, auto_adjust=True)
                        if 'Close' in s.columns and len(s) > 5:
                            price_cache[t] = s['Close'].dropna()
                    except: pass
            time.sleep(0.3)

    except ImportError:
        print("  yfinance not available — Eq_Mom_Score will be 0")

    pd.to_pickle(price_cache, EQ_CACHE)
    print(f"  Price cache saved: {len(price_cache)} tickers")

# ── 펀더멘탈 수집 (현재 시점 1회, 백테스트 전 기간 근사치로 사용) ──────────
import requests as req_lib
import urllib3
urllib3.disable_warnings()

SESS = req_lib.Session()
SESS.verify = False
SESS.headers.update({
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'application/json',
})

def fetch_fundamentals_yf(ticker):
    url = (f'https://query2.finance.yahoo.com/v10/finance/quoteSummary/{ticker}'
           f'?modules=financialData,defaultKeyStatistics')
    try:
        r = SESS.get(url, timeout=10)
        if r.status_code != 200: return {}
        data = r.json()
        result = data.get('quoteSummary',{}).get('result')
        if not result: return {}
        res = result[0]
        fd = res.get('financialData',{})
        ks = res.get('defaultKeyStatistics',{})
        def g(d, k):
            v = d.get(k)
            if isinstance(v, dict): v = v.get('raw')
            if v is not None:
                try:
                    f = float(v)
                    return None if np.isnan(f) else f
                except: pass
            return None
        return {
            'Debt_to_Equity':   g(fd,'debtToEquity'),
            'Profit_Margin':    g(fd,'profitMargins'),
            'Revenue_Growth':   g(fd,'revenueGrowth'),
            'Current_Ratio':    g(fd,'currentRatio'),
            'EV_EBITDA':        g(ks,'enterpriseToEbitda'),
        }
    except: return {}

if os.path.exists(FUND_CACHE):
    print(f"\n  Loading fundamental cache from {FUND_CACHE}...")
    fund_cache = pd.read_pickle(FUND_CACHE)
    print(f"  Cache loaded: {len(fund_cache)} tickers")
else:
    print(f"\n[Phase 2] Fetching fundamentals for {len(all_tickers)} tickers...")
    fund_cache = {}
    for i, t in enumerate(all_tickers):
        if i % 100 == 0:
            print(f"  {i}/{len(all_tickers)}...", end='\r', flush=True)
        fund_cache[t] = fetch_fundamentals_yf(t)
        time.sleep(0.05)
    pd.to_pickle(fund_cache, FUND_CACHE)
    print(f"\n  Fundamental cache saved: {len(fund_cache)} tickers")

# ── 전체 펀더멘탈 DataFrame 구성 & 스코어 ────────────────────────────────────
def rank_norm_01(series, ascending=True):
    s = pd.to_numeric(series, errors='coerce')
    valid = s.notna()
    result = pd.Series(np.nan, index=s.index)
    if valid.sum() < 2: return result
    ranked = s[valid].rank(ascending=ascending, method='average')
    result[valid] = ranked / valid.sum()
    return result

if fund_cache:
    fund_rows = [{'ticker': t, **v} for t, v in fund_cache.items() if v]
    fund_df = pd.DataFrame(fund_rows).set_index('ticker') if fund_rows else pd.DataFrame()
    if len(fund_df) > 0:
        fd_ = rank_norm_01(fund_df.get('Debt_to_Equity', pd.Series(dtype=float)), ascending=False)
        fp_ = rank_norm_01(fund_df.get('Profit_Margin', pd.Series(dtype=float)), ascending=True)
        fg_ = rank_norm_01(fund_df.get('Revenue_Growth', pd.Series(dtype=float)), ascending=True)
        fc_ = rank_norm_01(fund_df.get('Current_Ratio', pd.Series(dtype=float)), ascending=True)
        ev_ = fund_df.get('EV_EBITDA', pd.Series(dtype=float)).copy()
        ev_[ev_ <= 0] = np.nan
        fe_ = rank_norm_01(ev_, ascending=False)
        comps = pd.DataFrame({'fd':fd_,'fp':fp_,'fg':fg_,'fc':fc_,'fe':fe_})
        fund_df['Eq_Fund_Score'] = comps.mean(axis=1) * 2 - 1
        print(f"  Eq_Fund_Score computed for {fund_df['Eq_Fund_Score'].notna().sum()} tickers")
    else:
        fund_df = pd.DataFrame(columns=['Eq_Fund_Score'])
else:
    fund_df = pd.DataFrame(columns=['Eq_Fund_Score'])

# ═══════════════════════════════════════════════════════════════════════════════
# 4. 월별 주식 모멘텀 스코어 함수
# ═══════════════════════════════════════════════════════════════════════════════
def compute_eq_mom_scores(tickers, as_of_date):
    """
    tickers: list of equity tickers
    as_of_date: datetime.date — 이 날짜 기준으로 지표 계산
    Returns: dict {ticker: score}
    """
    as_of_ts = pd.Timestamp(as_of_date)
    one_mo  = as_of_ts - pd.DateOffset(months=1)
    three_mo= as_of_ts - pd.DateOffset(months=3)
    yr52    = as_of_ts - pd.DateOffset(weeks=52)

    metrics_list = []
    for t in tickers:
        prices = price_cache.get(t)
        if prices is None or len(prices) < 10:
            metrics_list.append({'ticker': t})
            continue
        # as_of_date 이전 데이터만
        p = prices[prices.index <= as_of_ts]
        if len(p) < 5:
            metrics_list.append({'ticker': t})
            continue

        cur = float(p.iloc[-1])
        m = {'ticker': t}

        # 1M return
        p1 = p[p.index <= one_mo]
        if len(p1) > 0:
            m['ret_1m'] = cur / float(p1.iloc[-1]) - 1

        # 3M return
        p3 = p[p.index <= three_mo]
        if len(p3) > 0:
            m['ret_3m'] = cur / float(p3.iloc[-1]) - 1

        # 30D vol
        log_r = np.log(p / p.shift(1)).dropna()
        if len(log_r) >= 20:
            m['vol_30d'] = float(log_r.iloc[-30:].std() * np.sqrt(252))

        # vs 52W high
        p52 = p[p.index >= yr52]
        if len(p52) > 0:
            hi = float(p52.max())
            if hi > 0:
                m['vs_52w'] = cur / hi - 1

        metrics_list.append(m)

    mom_df = pd.DataFrame(metrics_list).set_index('ticker')

    def rn(series, asc=True):
        s = pd.to_numeric(series, errors='coerce')
        valid = s.notna()
        res = pd.Series(np.nan, index=s.index)
        if valid.sum() < 2: return res
        ranked = s[valid].rank(ascending=asc, method='average')
        n = valid.sum()
        res[valid] = (ranked - 1) / (n - 1) * 2 - 1
        return res

    n1 = rn(mom_df.get('ret_1m',  pd.Series(dtype=float)), asc=True)
    n3 = rn(mom_df.get('ret_3m',  pd.Series(dtype=float)), asc=True)
    nv = rn(mom_df.get('vol_30d', pd.Series(dtype=float)), asc=False)
    nh = rn(mom_df.get('vs_52w',  pd.Series(dtype=float)), asc=True)
    comps = pd.DataFrame({'n1':n1,'n3':n3,'nv':nv,'nh':nh})
    mom_df['Eq_Mom_Score'] = comps.mean(axis=1)

    result = {}
    for t in tickers:
        if t in mom_df.index and not pd.isna(mom_df.loc[t,'Eq_Mom_Score']):
            result[t] = float(mom_df.loc[t,'Eq_Mom_Score'])
    return result

# ═══════════════════════════════════════════════════════════════════════════════
# 5. 스코어링 함수 — 전체 Integrated_Score
# ═══════════════════════════════════════════════════════════════════════════════
def score_and_pick(df, as_of_date, top_n=TOP_N, min_size=MIN_CLASS_SIZE):
    """
    월말 스냅샷 + 날짜 → 클래스별 Top-N 종목 선정
    Integrated_Score = Bond_TR×0.25 + Eq_Mom×0.25 + Eq_Fund×0.25 + AI_Macro×0.25
    """
    df = df.copy()

    # ── 활성 채권 필터 ─────────────────────────────────────────────────────────
    mask = (
        df['OAS'].notna() & df['OAS'].between(-50, 600) &
        df['OASD'].notna() & df['OASD'].gt(0) &
        df['Yield to Worst'].notna() & df['Yield to Worst'].gt(0) &
        df['Mty (Yrs)'].notna() &
        df['BCLASS3'].notna() &
        df['Composite Rating Num'].notna()
    )
    df = df[mask].copy()

    # ── 클래스 정의 ────────────────────────────────────────────────────────────
    df['_mty_bkt'] = df['Mty (Yrs)'].apply(maturity_bucket)
    df['_rtg_bkt'] = df['Composite Rating Num'].apply(
        lambda x: 'A-이상' if x <= RATING_CUT else 'BBB+이하'
    )
    df['_bclass3'] = df['BCLASS3'].astype(str).str.strip()
    df['_class'] = df['_bclass3'] + '|' + df['_mty_bkt'] + '|' + df['_rtg_bkt']

    # ── A. Bond_TR_Score ────────────────────────────────────────────────────────
    # Carry
    df['_carry'] = df['Yield to Worst'] / 12 * 2.5

    # Compression
    df['_dflt'] = df['1Y Dflt'].fillna(0)
    df['_comp_raw'] = (df['OAS'] - df['_dflt'] * 60) * df['OASD']

    # DP Rating Score (class 전체 기준 percentile)
    if 'DPFundamentalRating' in df.columns and 'DPSpreadRating' in df.columns:
        df['_dp_fund_n'] = df['DPFundamentalRating'].map(DP_RATING_MAP)
        df['_dp_spd_n']  = df['DPSpreadRating'].map(DP_RATING_MAP)
        df['_dp_gap'] = df['_dp_spd_n'] - df['_dp_fund_n']
        df['_dp_score'] = percentile_norm(df['_dp_gap'])
    else:
        df['_dp_score'] = 0.0

    # Bond_TR_Est per bond, then class-normalized
    df['_bond_tr_est'] = (
        df['_carry'].fillna(0) +
        df['_comp_raw'].fillna(0) +
        df['_dp_score'].fillna(0) * 0.05
    )

    # ── B. AI_Macro_Score ──────────────────────────────────────────────────────
    # Sector score
    subgroup_col = None
    for c in ['Industry Subgroup','BCLASS4','BCLASS3']:
        if c in df.columns:
            subgroup_col = c
            break
    if subgroup_col:
        df['_ai_sector'] = df[subgroup_col].astype(str).map(SUBGROUP_SCORE_MAP)
        # BCLASS3 fallback
        fallback_mask = df['_ai_sector'].isna()
        if fallback_mask.any():
            df.loc[fallback_mask,'_ai_sector'] = df.loc[fallback_mask,'_bclass3'].map(BCLASS3_SCORE_MAP)
        df['_ai_sector'] = df['_ai_sector'].fillna(0.0)
    else:
        df['_ai_sector'] = 0.0

    # Maturity score (OAD 기반)
    if 'OAD' in df.columns:
        df['_ai_mty'] = df['OAD'].apply(maturity_score)
    else:
        df['_ai_mty'] = df['Mty (Yrs)'].apply(
            lambda y: maturity_score(y * 0.9) if not pd.isna(y) else 0.0  # OAD ~ Mty * 0.9 근사
        )

    # RatingBuf score (DPFundamentalRating 기반)
    rtg_buf_col = None
    for c in ['DPFundamentalRating','Issuer Rtg','BB Comp']:
        if c in df.columns:
            rtg_buf_col = c
            break
    if rtg_buf_col:
        df['_ai_rtg'] = df[rtg_buf_col].map(RATING_BUFFER_MAP)
        df['_ai_rtg'] = df['_ai_rtg'].fillna(0.0)
    else:
        df['_ai_rtg'] = 0.0

    df['_ai_macro'] = (
        df['_ai_sector'] * 0.40 +
        df['_ai_mty']    * 0.35 +
        df['_ai_rtg']    * 0.25
    ).clip(-1.0, 1.0)

    # ── C. Equity 스코어 매핑 ──────────────────────────────────────────────────
    eq_ticker_col = None
    for c in ticker_col_candidates:
        if c in df.columns:
            eq_ticker_col = c
            break

    if eq_ticker_col:
        df['_eq_ticker'] = df[eq_ticker_col].apply(get_eq_ticker)
        uniq_tickers = [t for t in df['_eq_ticker'].dropna().unique() if t]

        # Eq_Mom_Score (as_of_date 기준 역사적)
        mom_scores = compute_eq_mom_scores(uniq_tickers, as_of_date)
        df['_eq_mom'] = df['_eq_ticker'].map(mom_scores)

        # Eq_Fund_Score (static)
        if 'Eq_Fund_Score' in fund_df.columns:
            df['_eq_fund'] = df['_eq_ticker'].map(fund_df['Eq_Fund_Score'])
        else:
            df['_eq_fund'] = np.nan
    else:
        df['_eq_mom']  = np.nan
        df['_eq_fund'] = np.nan

    # ── D. Integrated_Score per class ─────────────────────────────────────────
    picks = {}
    for cls, grp in df.groupby('_class'):
        if len(grp) < min_size:
            continue

        # Bond_TR_Score: class내 percentile-rank
        bond_tr_score = percentile_norm(grp['_bond_tr_est'])

        # Eq scores: class 내 re-rank (uniform 분포 유지)
        eq_mom_score  = percentile_norm(grp['_eq_mom'])   if grp['_eq_mom'].notna().sum() >= 3  else pd.Series(0.0, index=grp.index)
        eq_fund_score = percentile_norm(grp['_eq_fund'])  if grp['_eq_fund'].notna().sum() >= 3 else pd.Series(0.0, index=grp.index)

        ai_macro = grp['_ai_macro']  # 이미 [-1,+1]

        # 0.25 × 4 equal weight
        integrated = (
            bond_tr_score.fillna(0) * 0.25 +
            eq_mom_score.fillna(0)  * 0.25 +
            eq_fund_score.fillna(0) * 0.25 +
            ai_macro.fillna(0)      * 0.25
        )

        top_idx = integrated.nlargest(top_n).index
        picks[cls] = grp.loc[top_idx, 'ISIN'].tolist()

    return picks

# ═══════════════════════════════════════════════════════════════════════════════
# 6. 백테스팅 실행
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[Phase 3] Running backtest...")
print(f"  TOP_N per class = {TOP_N},  MIN_CLASS_SIZE = {MIN_CLASS_SIZE}")
print(f"  Score = Bond_TR×0.25 + Eq_Mom×0.25 + Eq_Fund×0.25 + AI_Macro×0.25")

port_records  = []
all_picks_log = []

for i in range(len(monthly_bt) - 1):
    dt_t,  fp_t  = monthly_bt[i]
    dt_t1, fp_t1 = monthly_bt[i + 1]

    if dt_t < BACKTEST_START:
        continue

    # T 시점 로드
    df_t, _ = load_file(fp_t)
    if df_t is None or len(df_t) < 100:
        print(f"  SKIP {dt_t}: load failed or too few bonds")
        continue

    # 스코어링 (T 시점 데이터 + T 날짜 기준 주식 데이터)
    picks = score_and_pick(df_t, as_of_date=dt_t)
    n_classes  = len(picks)
    n_selected = sum(len(v) for v in picks.values())

    # T+1 로드 → 실제 수익률
    df_t1, _ = load_file(fp_t1)
    if df_t1 is None:
        print(f"  SKIP {dt_t1}: load T+1 failed")
        continue

    df_t1['_tr'] = pd.to_numeric(df_t1['Total Return - 1mo'], errors='coerce')
    isin_to_tr = dict(zip(df_t1['ISIN'], df_t1['_tr']))

    # 벤치마크: T+1 유니버스 시가총액 가중평균 TR
    if 'Mkt Val' in df_t1.columns:
        df_t1['_mv'] = pd.to_numeric(df_t1['Mkt Val'], errors='coerce')
        valid_bch = df_t1.dropna(subset=['_mv','_tr'])
        valid_bch = valid_bch[valid_bch['_mv'] > 0]
        if len(valid_bch) > 100:
            bench_tr = float(np.average(valid_bch['_tr'], weights=valid_bch['_mv']))
        else:
            bench_tr = float(df_t1['_tr'].mean()) if df_t1['_tr'].notna().sum() > 0 else np.nan
    else:
        tr_num = df_t1['_tr'].dropna()
        bench_tr = float(tr_num.mean()) if len(tr_num) > 0 else np.nan

    # 포트폴리오 수익률
    all_isins = [isin for isins in picks.values() for isin in isins]
    tr_valid  = [float(isin_to_tr[i]) for i in all_isins
                 if i in isin_to_tr and isin_to_tr[i] is not None
                 and not np.isnan(float(isin_to_tr[i]))]

    if len(tr_valid) < 5:
        print(f"  SKIP {dt_t1}: too few return matches ({len(tr_valid)})")
        continue

    port_tr = np.mean(tr_valid)

    port_records.append({
        'date':       dt_t1,
        'port_tr':    port_tr / 100,
        'bench_tr':   bench_tr / 100 if (bench_tr is not None and not np.isnan(bench_tr)) else np.nan,
        'n_classes':  n_classes,
        'n_selected': n_selected,
        'n_matched':  len(tr_valid),
    })

    all_picks_log.append({
        'score_date': dt_t,
        'return_date': dt_t1,
        'picks': picks,
    })

    print(f"  {dt_t} -> {dt_t1}: classes={n_classes}, sel={n_selected}, "
          f"matched={len(tr_valid)}, port={port_tr:.3f}%, bench={bench_tr:.3f}%")

if not port_records:
    print("ERROR: No valid backtest records!")
    exit(1)

res = pd.DataFrame(port_records).set_index('date')
res.index = pd.to_datetime(res.index)

print(f"\n  Total months: {len(res)}")
print(f"  Period: {res.index[0].strftime('%Y-%m')} ~ {res.index[-1].strftime('%Y-%m')}")

# ═══════════════════════════════════════════════════════════════════════════════
# 7. 성과 통계
# ═══════════════════════════════════════════════════════════════════════════════
def cum_wealth(ret, start=100):
    return start * (1 + ret).cumprod()

def perf_stats(ret, label, bench=None):
    ann    = (1 + ret.mean()) ** 12 - 1
    vol    = ret.std() * np.sqrt(12)
    sharpe = ann / vol if vol > 0 else 0
    cw     = cum_wealth(ret)
    max_dd = (cw / cw.cummax() - 1).min()
    total  = (1 + ret).prod() - 1
    win    = (ret > 0).mean()
    d = {
        'Portfolio': label,
        'Total': f'{total:.1%}',
        'Ann.Ret': f'{ann:.2%}',
        'Vol': f'{vol:.2%}',
        'Sharpe': f'{sharpe:.2f}',
        'MaxDD': f'{max_dd:.2%}',
        'WinRate': f'{win:.0%}',
    }
    if bench is not None:
        excess_m = ret - bench
        info_r = excess_m.mean() / excess_m.std() * np.sqrt(12) if excess_m.std() > 0 else 0
        d['InfoRatio'] = f'{info_r:.2f}'
        d['CumAlpha']  = f'{(1+ret).prod() - (1+bench).prod():.1%}'
    return d

bench   = res['bench_tr'].dropna()
port    = res['port_tr']
aligned = port.loc[bench.index]

print("\n" + "="*70)
print("PERFORMANCE SUMMARY  (Full Integrated_Score: Bond_TR + EqMom + EqFund + AI_Macro)")
stats = pd.DataFrame([
    perf_stats(bench, 'LUACTRUU (벤치마크)'),
    perf_stats(aligned, f'Model Top{TOP_N}/Class', bench),
]).set_index('Portfolio')
print(stats.to_string())
print("="*70)

print("\n연도별 성과:")
print(f"{'Year':<6} {'Bench':>8} {'Model':>8} {'Alpha':>8}")
bench_valid = res['bench_tr'].dropna()
for yr in sorted(res.index.year.unique()):
    yr_mask_p = res.index.year == yr
    yr_mask_b = bench_valid.index.year == yr
    b = (1 + bench_valid.loc[yr_mask_b]).prod() - 1 if yr_mask_b.any() else np.nan
    p = (1 + port.loc[yr_mask_p]).prod() - 1
    b_str = f'{b:>7.2%}' if not np.isnan(b) else '    N/A'
    a_str = f'{p-b:>+7.2%}' if not np.isnan(b) else '    N/A'
    print(f"  {yr}  {b_str}  {p:>7.2%}  {a_str}")

# ═══════════════════════════════════════════════════════════════════════════════
# 8. 클래스 기여도 분석
# ═══════════════════════════════════════════════════════════════════════════════
class_perf = defaultdict(list)
fp_map = {d: fp for d, fp in monthly_bt}
for log in all_picks_log:
    rd = pd.Timestamp(log['return_date'])
    if rd not in res.index: continue
    fp_t1 = fp_map.get(log['return_date'])
    if fp_t1 is None: continue
    df_t1, _ = load_file(fp_t1)
    if df_t1 is None: continue
    isin_tr = dict(zip(df_t1['ISIN'], pd.to_numeric(df_t1['Total Return - 1mo'], errors='coerce')))
    for cls, isins in log['picks'].items():
        trs = [isin_tr.get(isin) for isin in isins
               if isin in isin_tr and not pd.isna(isin_tr.get(isin, np.nan))]
        if trs:
            bclass = cls.split('|')[0]
            class_perf[bclass].append(np.mean(trs) / 100)

class_ann = {}
for cls, rets in class_perf.items():
    if len(rets) >= 6:
        ann = (1 + np.mean(rets)) ** 12 - 1
        class_ann[cls] = (ann, len(rets))

# ═══════════════════════════════════════════════════════════════════════════════
# 9. 차트
# ═══════════════════════════════════════════════════════════════════════════════
EVENTS = [
    ('2022-01', 'Fed 긴축\n시작'),
    ('2022-06', '금리\n피크'),
    ('2023-03', 'SVB\n사태'),
    ('2023-10', '금리\n재급등'),
    ('2024-09', 'Fed\n피벗'),
]

fig = plt.figure(figsize=(22, 14))
gs  = gridspec.GridSpec(3, 3, figure=fig, hspace=0.45, wspace=0.32, top=0.91, bottom=0.06)
COLORS = {'bench': '#1F3864', 'model': '#C00000', 'excess': '#70AD47'}

# ─── Plot 1: 누적수익 ─────────────────────────────────────────────────────────
ax1 = fig.add_subplot(gs[0, :2])
cw_bench = cum_wealth(bench)
cw_model = cum_wealth(aligned)
ax1.fill_between(cw_model.index, cw_model.values, cw_bench.values,
                 where=(cw_model.values >= cw_bench.values),
                 alpha=0.15, color=COLORS['model'])
ax1.fill_between(cw_model.index, cw_model.values, cw_bench.values,
                 where=(cw_model.values < cw_bench.values),
                 alpha=0.15, color='gray')
ax1.plot(cw_bench.index, cw_bench.values, color=COLORS['bench'], lw=2.5,
         label=f"LUACTRUU: {(1+bench).prod()-1:.1%}", zorder=5)
ax1.plot(cw_model.index, cw_model.values, color=COLORS['model'], lw=2.3,
         label=f"Model Top{TOP_N}/Class: {(1+aligned).prod()-1:.1%}", zorder=7)
for ev_dt, ev_lbl in EVENTS:
    xv = pd.Timestamp(ev_dt)
    if res.index[0] <= xv <= res.index[-1]:
        ax1.axvline(xv, color='#AAAAAA', lw=0.9, ls='--', alpha=0.7)
        y0 = ax1.get_ylim()[0]
        ax1.text(xv, y0*1.01 if y0 > 0 else y0*0.99,
                 ev_lbl, fontsize=7.5, color='#555555', ha='center')
ax1.axhline(100, color='gray', lw=0.5, ls=':', alpha=0.6)
ax1.set_title(
    f'누적 수익률 ($100 시작) — Full Integrated Score / Top{TOP_N}/Class / 월별 리밸\n'
    f'Bond_TR×0.25 + Eq_Mom×0.25 + Eq_Fund×0.25 + AI_Macro×0.25  (Sentiment 제외)',
    fontsize=10.5, fontweight='bold', pad=6
)
ax1.set_ylabel('포트폴리오 가치 ($)', fontsize=10)
ax1.yaxis.set_major_formatter(mtick.FormatStrFormatter('$%.0f'))
ax1.legend(fontsize=10, loc='upper left', framealpha=0.9)
ax1.grid(True, alpha=0.2)
ax1.set_xlim(res.index[0], res.index[-1])

# ─── Plot 2: 성과 테이블 ──────────────────────────────────────────────────────
ax2 = fig.add_subplot(gs[0, 2])
ax2.axis('off')
b_ann = (1 + bench.mean()) ** 12 - 1
m_ann = (1 + aligned.mean()) ** 12 - 1
b_vol = bench.std() * np.sqrt(12)
m_vol = aligned.std() * np.sqrt(12)
excess_m = aligned - bench
rows = [
    ['', 'LUACTRUU', f'Model Top{TOP_N}'],
    ['총수익', f"{(1+bench).prod()-1:.1%}", f"{(1+aligned).prod()-1:.1%}"],
    ['연환산수익', f"{b_ann:.2%}", f"{m_ann:.2%}"],
    ['연환산 변동성', f"{b_vol:.2%}", f"{m_vol:.2%}"],
    ['Sharpe', f"{b_ann/b_vol:.2f}" if b_vol>0 else 'N/A', f"{m_ann/m_vol:.2f}" if m_vol>0 else 'N/A'],
    ['최대낙폭',
     f"{(cum_wealth(bench)/cum_wealth(bench).cummax()-1).min():.2%}",
     f"{(cum_wealth(aligned)/cum_wealth(aligned).cummax()-1).min():.2%}"],
    ['월별승률', f"{(bench>0).mean():.0%}", f"{(aligned>0).mean():.0%}"],
    ['누적 초과수익', '', f"{(1+aligned).prod()-(1+bench).prod():.1%}"],
    ['연환산 Alpha', '', f"{excess_m.mean()*12:.2%}"],
    ['Info Ratio', '', f"{excess_m.mean()/excess_m.std()*np.sqrt(12):.2f}" if excess_m.std()>0 else 'N/A'],
]
tbl = ax2.table(cellText=rows[1:], colLabels=rows[0],
                cellLoc='center', loc='center', bbox=[0,0,1,1])
tbl.auto_set_font_size(False)
tbl.set_fontsize(9.5)
for (r,c), cell in tbl.get_celld().items():
    cell.set_edgecolor('#DDDDDD')
    cell.set_height(0.09)
    if r == 0:
        cell.set_facecolor('#1F3864')
        cell.set_text_props(color='white', fontweight='bold')
    elif c == 2 and r > 0:
        cell.set_facecolor('#FFF0F0')
        cell.set_text_props(fontweight='bold', color='#C00000', fontsize=10)
    elif r % 2 == 0:
        cell.set_facecolor('#F5F5F5')
ax2.set_title('성과 요약\n(Full Integrated Score)', fontsize=10, fontweight='bold', pad=4)

# ─── Plot 3: 연도별 바 차트 ───────────────────────────────────────────────────
ax3 = fig.add_subplot(gs[1, :2])
years_list = sorted(res.index.year.unique())
bench_yr = [(1+bench[bench.index.year==y]).prod()-1
            if (bench.index.year==y).any() else np.nan for y in years_list]
model_yr = [(1+aligned[aligned.index.year==y]).prod()-1 for y in years_list]
alpha_yr = [m-b if not np.isnan(b) else np.nan for m,b in zip(model_yr, bench_yr)]

x = np.arange(len(years_list))
w = 0.35
ax3.bar(x-w/2, [v*100 for v in bench_yr], w, label='LUACTRUU', color=COLORS['bench'], alpha=0.85)
ax3.bar(x+w/2, [v*100 for v in model_yr], w, label=f'Model Top{TOP_N}/Class', color=COLORS['model'], alpha=0.85)
ax3.axhline(0, color='black', lw=0.8)
ax3.set_xticks(x)
ax3.set_xticklabels([str(y) for y in years_list], fontsize=10)
ax3.set_title('연도별 수익률 비교 (%)', fontsize=12, fontweight='bold', pad=6)
ax3.set_ylabel('연간 수익률 (%)', fontsize=10)
ax3.yaxis.set_major_formatter(mtick.PercentFormatter())
ax3.legend(fontsize=10)
ax3.grid(True, alpha=0.2, axis='y')
for i, (bv, mv, av) in enumerate(zip(bench_yr, model_yr, alpha_yr)):
    if bv is not None and not np.isnan(bv):
        ax3.text(x[i]-w/2, bv*100+(0.2 if bv>=0 else -0.8),
                 f'{bv*100:.1f}%', ha='center', va='bottom' if bv>=0 else 'top',
                 fontsize=8, color=COLORS['bench'])
    if not np.isnan(mv):
        color = '#C00000' if mv > (bv or 0) else '#888888'
        ax3.text(x[i]+w/2, mv*100+(0.2 if mv>=0 else -0.8),
                 f'{mv*100:.1f}%', ha='center', va='bottom' if mv>=0 else 'top',
                 fontsize=8, color=color, fontweight='bold')
        if av is not None and not np.isnan(av):
            ax3.text(x[i], max(mv*100,(bv or 0)*100)+0.8, f'a{av*100:+.1f}%',
                     ha='center', va='bottom', fontsize=7.5,
                     color=COLORS['excess'], fontweight='bold')

# ─── Plot 4: 롤링 Alpha (12M) ─────────────────────────────────────────────────
ax4 = fig.add_subplot(gs[1, 2])
excess_cum = (1 + excess_m).cumprod() - 1
ax4.plot(excess_cum.index, excess_cum.values * 100, color=COLORS['excess'], lw=2)
ax4.fill_between(excess_cum.index, excess_cum.values * 100, 0,
                 where=(excess_cum.values >= 0), alpha=0.3, color=COLORS['excess'])
ax4.fill_between(excess_cum.index, excess_cum.values * 100, 0,
                 where=(excess_cum.values < 0), alpha=0.3, color='#FF6B6B')
ax4.axhline(0, color='black', lw=0.8)
ax4.set_title('누적 초과수익 (vs 벤치마크, %)', fontsize=10, fontweight='bold', pad=4)
ax4.yaxis.set_major_formatter(mtick.PercentFormatter())
ax4.grid(True, alpha=0.2)
ax4.set_xlim(res.index[0], res.index[-1])

# ─── Plot 5: 월별 Alpha 분포 ─────────────────────────────────────────────────
ax5 = fig.add_subplot(gs[2, 0])
excess_pct = excess_m * 100
ax5.hist(excess_pct.values, bins=25, color=COLORS['model'], alpha=0.7, edgecolor='white')
ax5.axvline(0, color='black', lw=1)
ax5.axvline(excess_pct.mean(), color='orange', lw=1.5, ls='--',
            label=f'Mean: {excess_pct.mean():.2f}%')
ax5.set_title('월별 초과수익 분포 (%)', fontsize=10, fontweight='bold', pad=4)
ax5.set_xlabel('초과수익 (%)', fontsize=9)
ax5.legend(fontsize=9)
ax5.grid(True, alpha=0.2)

# ─── Plot 6: 섹터별 기여도 ───────────────────────────────────────────────────
ax6 = fig.add_subplot(gs[2, 1:])
if class_ann:
    sorted_cls = sorted(class_ann.items(), key=lambda x: x[1][0], reverse=True)[:15]
    cls_labels = [c[:20] for c, _ in sorted_cls]
    cls_vals   = [ann * 100 for _, (ann, _n) in sorted_cls]
    colors_bar = ['#C00000' if v >= 0 else '#1F3864' for v in cls_vals]
    ybars = np.arange(len(cls_labels))
    ax6.barh(ybars, cls_vals, color=colors_bar, alpha=0.85)
    ax6.set_yticks(ybars)
    ax6.set_yticklabels(cls_labels, fontsize=8)
    ax6.axvline(0, color='black', lw=0.8)
    ax6.set_title('BCLASS3별 연환산 수익률 (상위 15)', fontsize=10, fontweight='bold', pad=4)
    ax6.set_xlabel('연환산 수익률 (%)', fontsize=9)
    ax6.xaxis.set_major_formatter(mtick.PercentFormatter())
    ax6.grid(True, alpha=0.2, axis='x')
else:
    ax6.text(0.5, 0.5, 'No class data', ha='center', va='center', transform=ax6.transAxes)

fig.suptitle(
    f'Bloomberg US IG Corp — Full Integrated Score Backtest  '
    f'({res.index[0].strftime("%Y-%m")} ~ {res.index[-1].strftime("%Y-%m")})\n'
    f'Bond_TR×0.25 + Eq_Mom×0.25 + Eq_Fund×0.25 + AI_Macro×0.25  |  Top{TOP_N}/Class Equal-Weighted',
    fontsize=13, fontweight='bold', y=0.97
)

plt.savefig(OUT, dpi=150, bbox_inches='tight', facecolor='white')
print(f"\n[DONE] Chart saved: {OUT}")

# ═══════════════════════════════════════════════════════════════════════════════
# 10. Excel 상세 결과
# ═══════════════════════════════════════════════════════════════════════════════
try:
    from openpyxl import Workbook as OWB
    from openpyxl.styles import Font as OFont, PatternFill as OFill, Alignment as OAlign

    wb = OWB()
    ws = wb.active
    ws.title = 'Monthly_Performance'

    headers = ['Date','Port_TR(%)','Bench_TR(%)','Excess(%)','n_Classes','n_Selected','n_Matched']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = OFont(bold=True, color='FFFFFF')
        c.fill = OFill('solid', fgColor='1F3864')
        c.alignment = OAlign(horizontal='center')

    for ri, (idx, row) in enumerate(res.iterrows(), 2):
        bench_v = row['bench_tr'] * 100 if not pd.isna(row['bench_tr']) else None
        port_v  = row['port_tr'] * 100
        exc_v   = (port_v - bench_v) if bench_v is not None else None
        vals = [
            idx.strftime('%Y-%m'), round(port_v,3),
            round(bench_v,3) if bench_v else None,
            round(exc_v,3) if exc_v else None,
            row['n_classes'], row['n_selected'], row['n_matched'],
        ]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=ri, column=ci, value=v)

    wb.save(OUT_XLSX)
    print(f"[DONE] Excel saved: {OUT_XLSX}")
except Exception as e:
    print(f"[WARN] Excel save failed: {e}")
