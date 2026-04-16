"""update_methodology.py — Methodology 시트 현행 버전으로 전면 재작성"""
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

FILE = 'C:/Users/sh.park/Documents/USIG_LLM/LUACSTAT_2026_03_31_SCORED.xlsx'
wb = load_workbook(FILE)
ws = wb['Methodology']

# 전체 초기화
for row in ws.iter_rows():
    for cell in row:
        cell.value = None
        cell.font = Font(name='Arial', size=10)
        cell.fill = PatternFill(fill_type=None)
        cell.alignment = Alignment()

ws.column_dimensions['A'].width = 100

FILL_H1   = PatternFill('solid', fgColor='1F3864')  # dark blue  — 대제목
FILL_H2   = PatternFill('solid', fgColor='2E75B6')  # mid blue   — 섹션
FILL_H3   = PatternFill('solid', fgColor='BDD7EE')  # light blue — 소제목
FILL_AI   = PatternFill('solid', fgColor='4A148C')  # purple     — AI 섹션
FILL_AI_L = PatternFill('solid', fgColor='E1BEE7')  # light purple — AI 서브
FILL_NEW  = PatternFill('solid', fgColor='E2EFDA')  # light green — 변경사항
FILL_WARN = PatternFill('solid', fgColor='FFF2CC')  # yellow     — 주의

r = [1]  # mutable row counter

def w(text, bold=False, size=10, color='000000', fill=None, indent=0):
    val = (' ' * indent * 2) + text if indent else text
    c = ws.cell(row=r[0], column=1, value=val)
    c.font = Font(name='Arial', bold=bold, size=size, color=color)
    c.alignment = Alignment(vertical='center', wrap_text=True)
    if fill:
        c.fill = fill
    ws.row_dimensions[r[0]].height = 16
    r[0] += 1

def blank():
    ws.row_dimensions[r[0]].height = 8
    r[0] += 1

# ── 제목 ─────────────────────────────────────────────────────────────────
w('BLOOMBERG US IG CORPORATE BOND  —  SCORING METHODOLOGY  (v3, 2026-04-11)',
  bold=True, size=13, color='FFFFFF', fill=FILL_H1)
blank()

# ── DATA SOURCE ──────────────────────────────────────────────────────────
w('DATA SOURCE', bold=True, color='FFFFFF', fill=FILL_H2)
w('Input  : Bloomberg LUAC Index constituents as of 2026-03-31  |  Universe: 8,704 US IG Corporate Bonds', indent=1)
w('Equity : Yahoo Finance quoteSummary & chart API (fetched at runtime)', indent=1)
w('News   : Yahoo Finance search API + Google News RSS, up to 20 headlines per ticker (deduplicated)', indent=1)
w('Trends : Google Trends via pytrends, 3-month window, US geography', indent=1)
w('AI     : Claude API (claude-opus-4-5) — market analysis refreshed dynamically per run via update_ai_scores.py', indent=1)
blank()

# ── STEP 1: Bond TR Estimate ─────────────────────────────────────────────
w('STEP 1  —  BOND TOTAL RETURN ESTIMATE  (Bond_TR_Est_pct)', bold=True, color='FFFFFF', fill=FILL_H2)

w('A.  Carry_2.5M_pct', bold=True, fill=FILL_H3)
w('Formula : YTM (Yield to Worst, %) / 12 x 2.5', indent=2)
w('Estimated income accrual (% of face value) over a 2.5-month holding horizon, assuming yield held constant.', indent=2)

w('B.  Compression_Score_pct', bold=True, fill=FILL_H3)
w('Formula : (OAS - Spread_floor_bp) x OASD / 100', indent=2)
w('Spread_floor_bp = 1Y_Dflt(%) x 60   [Recovery Rate 40% assumed -> (1 - 0.40) x 100 = 60]', indent=2)
w('Compression_gap = OAS minus default-probability-implied fair spread.', indent=2)
w('  OAS > Spread_floor -> undervalued -> positive score  (spread can compress toward fair value)', indent=2)
w('  OAS < Spread_floor -> overvalued  -> negative score  (no max(0) clamp: overvalued bonds are penalized)', indent=2)

w('C.  DP_Rating_Score', bold=True, fill=FILL_H3)
w('Source  : Bloomberg DPFundamentalRating + DPSpreadRating  (Moody-style letter scale: AAA=1 ... C=21)', indent=2)
w('Gap     : Rating_Gap = DPSpreadRating_num - DPFundamentalRating_num', indent=2)
w('          Positive gap -> market prices bond WORSE than fundamentals justify -> undervalued -> high score', indent=2)
w('          Negative gap -> market prices bond BETTER than fundamentals justify -> overvalued -> low score', indent=2)
w('Formula : DP_Rating_Score = percentile_rank(Rating_Gap, ascending=True)  ->  [-1, +1]', indent=2)
w('Example : DPSpreadRating=BAA3, DPFundamentalRating=A1 -> gap = +5 -> high positive score (cheap vs fundamentals)', indent=2)

w('D.  Bond_TR_Est_pct  (absolute return estimate)', bold=True, fill=FILL_H3)
w('Formula : Carry_2.5M_pct + Compression_Score_pct + DP_Rating_Score x 0.05', indent=2)
blank()

# ── STEP 2: Bond TR Score ────────────────────────────────────────────────
w('STEP 2  —  BOND_TR_SCORE  :  CLASS-LEVEL NORMALIZATION', bold=True, color='FFFFFF', fill=FILL_H2)
w('Bond_TR_Est_pct is percentile-ranked WITHIN EACH CLASS separately  ->  [-1, +1]', indent=1)
w('Formula : (rank - 1) / (N_class - 1) x 2 - 1', indent=2)
w('Best bond in class = +1.0  |  Worst bond in class = -1.0', indent=2)
w('Cross-class comparison intentionally excluded — each class is evaluated on its own fundamentals.', indent=2)
w('Class sizes: S6_T4_A = 58 bonds (smallest) ... S7_T1_A = 680 bonds (largest).', indent=2)
blank()

# ── STEP 3: Equity Scores ────────────────────────────────────────────────
w('STEP 3  —  EQUITY MOMENTUM SCORE  (Eq_Mom_Score)', bold=True, color='FFFFFF', fill=FILL_H2)
w('Universe : ~702 unique equity tickers  (cross-class — same issuer bonds share the same score)', indent=1)
w('Formula per component : (rank / N) x 2 - 1', indent=1)
w('Components:', indent=1)
w('Eq_Ret_1M      — 1-month equity return          (higher = better)', indent=3)
w('Eq_Ret_3M      — 3-month equity return          (higher = better)', indent=3)
w('Eq_Vol_30D     — 30-day realized vol (ann.)     (lower  = better)', indent=3)
w('Eq_vs_52w_High — current price / 52-week high   (higher = better)', indent=3)
w('Eq_Mom_Score = simple average of 4 normalized components', indent=1)
w('[NOTE] Practical range is approx [-0.998, +0.840]. Averaging prevents any single ticker from reaching ±1 unless it tops all 4 simultaneously.',
  indent=1, fill=FILL_WARN)
blank()

w('STEP 3  —  EQUITY FUNDAMENTAL SCORE  (Eq_Fund_Score)', bold=True, color='FFFFFF', fill=FILL_H2)
w('Universe : ~702 unique equity tickers  (cross-class)', indent=1)
w('Formula  : (rank / N) -> [0,1] per component, then average, then x 2 - 1 -> [-1, +1]', indent=1)
w('Components:', indent=1)
w('Debt_to_Equity  lower  = better  |  Current_Ratio   higher = better  (capped at 3.0)', indent=3)
w('Profit_Margin   higher = better  |  EV_EBITDA       lower  = better  (negatives excluded)', indent=3)
w('Revenue_Growth  higher = better  |  PE_Ratio        lower  = better  (0-100, negatives excluded)', indent=3)
w('Eq_Fund_Score = average of available components (min 2 required), rescaled to [-1, +1]', indent=1)
w('[NOTE] Practical range is approx [-0.810, +0.879] due to 2-stage normalization and component averaging.',
  indent=1, fill=FILL_WARN)
blank()

# ── STEP 4: Sentiment ────────────────────────────────────────────────────
w('STEP 4  —  SENTIMENT SCORE  (Sentiment_Score)', bold=True, color='FFFFFF', fill=FILL_H2)

w('Component A — News Sentiment  (primary signal)', bold=True, fill=FILL_H3)
w('Source   : Yahoo Finance search API + Google News RSS (merged, deduplicated by title)', indent=2)
w('NLP      : VADER SentimentIntensityAnalyzer -> compound score [-1, +1] per headline', indent=2)
w('Weighting: exp(-days_old / 14.0)  — recent articles up-weighted, half-life ~14 days', indent=2)
w('Raw score: weighted mean of compound scores  |  News_Norm = rank-normalized to [-1, +1]', indent=2)
w('Top_Headline : headline with highest |compound x recency_weight| stored for reference', indent=2)

w('Component B — Google Trends  (amplitude modifier, NOT direction)', bold=True, fill=FILL_H3)
w('Trends_Momentum : (mean search interest last 4 weeks - prior 4 weeks) / (prior 4 weeks + 1)', indent=2)
w('Trends_Norm     : rank-normalized Trends_Momentum  ->  [-1, +1]', indent=2)
w('Trends_Factor   : 1.0 + Trends_Norm x 0.3   ->  range [0.70, 1.30]', indent=2)
w('Interpretation  : rising interest -> factor > 1.0 (amplify news signal) | falling -> factor < 1.0 (dampen)', indent=2)
w('[IMPORTANT] Trends_Factor only modifies the MAGNITUDE of the news signal; NEWS determines direction.', indent=2, fill=FILL_WARN)

w('Generic News Filtering', bold=True, fill=FILL_H3)
w('Yahoo Finance returns generic market news when a ticker cannot be found.', indent=2)
w('Detection: same News_Sentiment_Raw value (rounded to 5 d.p.) appears for > 5 different tickers', indent=2)
w('Action   : flagged (News_Generic_Flag=1) -> set to NaN -> contributes 0 to Integrated_Score', indent=2)

w('Composite Formula', bold=True, fill=FILL_H3)
w('Sentiment_Score = News_Norm x Trends_Factor         (both available — news x amplitude)', indent=2)
w('                = News_Norm                          (trends unavailable — news direction only)', indent=2)
w('                = Trends_Norm x 0.50                (news unavailable — trends at half weight)', indent=2)
w('                = NaN -> 0 in Integrated_Score       (neither available, or generic flag)', indent=2)
blank()

# ── STEP 5: AI Macro Score ───────────────────────────────────────────────
w('STEP 5  —  AI MACRO SCORE  (AI_Macro_Score)  [DYNAMIC — refreshed via update_ai_scores.py]',
  bold=True, color='FFFFFF', fill=FILL_AI)
w('Purpose  : Captures qualitative macro regime insights that cannot be derived from price/fundamental data alone.', indent=1)
w('Source   : Claude API (claude-opus-4-5) called with current date + sector OAS context each refresh.', indent=1)
w('Frequency: Manually triggered (run update_ai_scores.py). ai_macro_score.py is overwritten with fresh scores.', indent=1)
blank()

w('Sub-score A — AI_Sector_Score  (weight 40%)', bold=True, fill=FILL_AI_L)
w('Granularity : Industry Subgroup level  (~150 categories, e.g. Electric-Integrated, Auto-Cars/Light Trucks)', indent=2)
w('Range       : [-1.0, +1.0] per subgroup  |  0.0 = neutral', indent=2)
w('Example     : Electric-Integrated = +0.85  (regulated utility, defensive in risk-off)', indent=2)
w('Example     : Auto-Cars/Light Trucks = -0.90  (direct tariff exposure, earnings risk)', indent=2)
w('Unmapped subgroups receive 0.0 (neutral fallback).', indent=2)

w('Sub-score B — AI_Maturity_Score  (weight 35%)', bold=True, fill=FILL_AI_L)
w('Based on OAD (Option-Adjusted Duration) — reflects preferred curve positioning in current macro regime:', indent=2)
w('  OAD < 2   (ultra short)  |  2-4   (short)  |  4-7   (medium, typical sweet spot)', indent=2)
w('  7-10 (medium-long)  |  10-13 (long)  |  13-16 (very long)  |  16+ (ultra long)', indent=2)
w('Scores vary per run: e.g. in rate-cut expectations, medium (4-7) = +1.0, ultra-long = -1.0.', indent=2)

w('Sub-score C — AI_RatingBuf_Score  (weight 25%)', bold=True, fill=FILL_AI_L)
w('Per-notch rating safety score  (AAA through BA2+)  reflecting fallen-angel risk in current credit cycle.', indent=2)
w('Primary lookup: DPFundamentalRating  |  Fallback: Issuer Rtg  |  Unmapped -> 0.0', indent=2)
w('Example: BAA3 = -0.65 (high fallen-angel risk) | A1/A2 = +0.70 (solid IG, low downgrade risk)', indent=2)

w('Combined Formula', bold=True, fill=FILL_AI_L)
w('AI_Macro_Score = AI_Sector_Score x 0.40 + AI_Maturity_Score x 0.35 + AI_RatingBuf_Score x 0.25', indent=2)
w('Clipped to [-1.0, +1.0]', indent=2)
blank()

# ── STEP 6: Integrated ───────────────────────────────────────────────────
w('STEP 6  —  INTEGRATED SCORE & RANKING  (Equal-weight 5-factor)', bold=True, color='FFFFFF', fill=FILL_H2)
w('All 5 components normalized to [-1, +1] before combining. NaN -> 0 (neutral, no penalty).', indent=1)
blank()
w('Integrated_Score  =  Bond_TR_Score    x 0.20', indent=3, bold=True, fill=FILL_NEW)
w('                  +  Eq_Mom_Score     x 0.20   (NaN -> 0)', indent=3, fill=FILL_NEW)
w('                  +  Eq_Fund_Score    x 0.20   (NaN -> 0)', indent=3, fill=FILL_NEW)
w('                  +  Sentiment_Score  x 0.20   (NaN -> 0, generic news -> 0)', indent=3, fill=FILL_NEW)
w('                  +  AI_Macro_Score   x 0.20   (refreshed via Claude API each run)', indent=3, fill=FILL_NEW)
blank()
w('Integrated_Rank_in_Class : descending rank of Integrated_Score WITHIN each class (1 = best)', indent=1)
w('Bond_TR_Rank_in_Class    : descending rank of Bond_TR_Est_pct within class (reference only)', indent=1)
w('Top_Pick_Flag            : ★★★ TOP3 (rank <= 3)  |  ★★ TOP10 (rank <= 10)  |  ★ TOP25 (rank <= 25)', indent=1)
blank()

# ── OUTPUT SHEETS ────────────────────────────────────────────────────────
w('OUTPUT SHEETS', bold=True, color='FFFFFF', fill=FILL_H2)
w('Score_BondTR      : Carry / Compression / DPRating -> Bond_TR_Est_pct -> Bond_TR_Score (class-normalized)', indent=1)
w('Score_EqMom       : Eq_Ret_1M / 3M / Vol / 52wHigh -> Eq_Mom_Score', indent=1)
w('Score_EqFund      : D/E / Margin / Growth / CR / EV/EBITDA / P/E -> Eq_Fund_Score', indent=1)
w('Score_Sentiment   : News raw / article count / top headline / trends / sentiment score', indent=1)
w('Score_AI          : Industry Subgroup / sector score / maturity score / rating buffer -> AI_Macro_Score', indent=1)
w('Score_Integrated  : All 5 scores + LQA + Integrated_Score + Rank + Flag  (sorted by class then rank)', indent=1)
w('Detail_Scored     : Full universe with all computed columns (raw data + all scores)', indent=1)
blank()

# ── REFRESH WORKFLOW ─────────────────────────────────────────────────────
w('REFRESH WORKFLOW', bold=True, color='FFFFFF', fill=FILL_H2)
w('Standard refresh (sentiment + scores):  python sentiment_update.py', indent=1)
w('  -> Fetches Yahoo + Google News, Google Trends, recomputes Sentiment_Score', indent=2)
w('  -> Calls build_score_sheets.py internally (rebuilds all Score_* sheets)', indent=2)
w('AI score refresh (dynamic macro view):  python update_ai_scores.py', indent=1)
w('  -> Calls Claude API with current sector OAS context + date', indent=2)
w('  -> Rewrites ai_macro_score.py with fresh subgroup/maturity/rating scores', indent=2)
w('  -> Calls build_score_sheets.py internally (rebuilds all Score_* sheets)', indent=2)
w('  Requires: ANTHROPIC_API_KEY environment variable', indent=2)
blank()

# ── DISCLAIMER ───────────────────────────────────────────────────────────
w('DISCLAIMER', bold=True, color='FFFFFF', fill=FILL_H2)
w('For informational purposes only. Scores are quantitative and AI-assisted estimates.', indent=1)
w('Combine with fundamental analysis, credit research, and portfolio construction guidelines.', indent=1)
w('AI macro scores reflect Claude\'s analysis as of the last refresh date stored in ai_macro_score.py.', indent=1)
w('Past performance and model outputs do not guarantee future results.', indent=1)

wb.save(FILE)
print(f'Done. Methodology rows: {r[0]-1}  |  File: {os.path.getsize(FILE)/1024/1024:.2f} MB')
