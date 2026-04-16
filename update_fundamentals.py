import sys
import io
import re
import time
import warnings
import json
import requests
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
warnings.filterwarnings('ignore')

EXCEL_PATH = 'C:/Users/sh.park/Documents/USIG_LLM/LUACSTAT_2026_03_31_SCORED.xlsx'

# ─────────────────────────────────────────────
# STEP 1: Read file and extract tickers
# ─────────────────────────────────────────────
print("=" * 60)
print("STEP 1: Reading Excel file and extracting tickers")
print("=" * 60)

df = pd.read_excel(EXCEL_PATH, sheet_name='Detail_Scored', header=1)
print(f"Detail_Scored shape: {df.shape}")

tickers_raw = df['Eqty Ticker'].dropna().unique().tolist()
tickers_raw = [str(t).strip() for t in tickers_raw if str(t).strip() not in ('', 'nan', 'N/A')]

def clean_ticker(t):
    """Strip exchange suffixes to get Yahoo Finance compatible tickers."""
    for suffix in [' US', ' LN', ' FP', ' GR', ' JP', ' HK', ' CN', ' AU', ' SM', ' IM', ' NA', ' SW']:
        if t.endswith(suffix):
            return t[:-len(suffix)].strip()
    return t.strip()

# Build mapping: raw ticker -> clean ticker
ticker_map = {t: clean_ticker(t) for t in tickers_raw}
# Only keep tickers that look like valid Yahoo Finance symbols
def is_valid_yahoo_ticker(t):
    return bool(re.match(r'^[A-Z][A-Z0-9\.\-]{0,9}$', t))

valid_clean_tickers = list({v for v in ticker_map.values() if is_valid_yahoo_ticker(v)})
print(f"Total raw unique tickers: {len(tickers_raw)}")
print(f"Valid Yahoo Finance tickers: {len(valid_clean_tickers)}")

# ─────────────────────────────────────────────
# STEP 2: Fetch fundamentals via Yahoo Finance API
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 2: Fetching fundamental data from Yahoo Finance")
print("=" * 60)

api_headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'application/json',
    'Accept-Language': 'en-US,en;q=0.9',
}

def init_session():
    """Initialize session and get Yahoo Finance crumb."""
    session = requests.Session()
    session.verify = False
    browse_headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
    }
    session.get('https://finance.yahoo.com', headers=browse_headers, timeout=15)
    r = session.get('https://query1.finance.yahoo.com/v1/test/getcrumb',
                    headers={'User-Agent': browse_headers['User-Agent']}, timeout=10)
    crumb = r.text.strip()
    print(f"  Got crumb: {crumb}")
    return session, crumb

def fetch_ticker_fundamentals(session, crumb, ticker):
    """Fetch fundamental data for a single ticker."""
    modules = 'defaultKeyStatistics,financialData,summaryDetail'

    for base in ['https://query1.finance.yahoo.com', 'https://query2.finance.yahoo.com']:
        url = f'{base}/v10/finance/quoteSummary/{ticker}?modules={modules}&crumb={crumb}'
        try:
            r = session.get(url, headers=api_headers, timeout=10)
            if r.status_code == 401:
                return None  # Need new crumb
            data = r.json()
            result = data.get('quoteSummary', {}).get('result', [])
            if result:
                fd = result[0].get('financialData', {})
                ks = result[0].get('defaultKeyStatistics', {})
                sd = result[0].get('summaryDetail', {})

                def get_raw(d, key):
                    v = d.get(key, {})
                    if isinstance(v, dict):
                        return v.get('raw', None)
                    return None

                return {
                    'Debt_to_Equity': get_raw(fd, 'debtToEquity'),
                    'Profit_Margin': get_raw(fd, 'profitMargins'),
                    'Revenue_Growth': get_raw(fd, 'revenueGrowth'),
                    'Current_Ratio': get_raw(fd, 'currentRatio'),
                    'EV_EBITDA': get_raw(ks, 'enterpriseToEbitda'),
                    'PE_Ratio': get_raw(sd, 'trailingPE'),
                }
        except Exception:
            continue
    return None

# Initialize session
session, crumb = init_session()

fundamentals = {}
success_count = 0
fail_count = 0
total = len(valid_clean_tickers)
refresh_interval = 200  # Refresh crumb every 200 requests

print(f"Fetching data for {total} tickers...")
for i, ticker in enumerate(valid_clean_tickers):
    # Refresh crumb periodically
    if i > 0 and i % refresh_interval == 0:
        print(f"  Refreshing session at ticker {i}...")
        try:
            session, crumb = init_session()
        except Exception as e:
            print(f"  Warning: Could not refresh crumb: {e}")

    result = fetch_ticker_fundamentals(session, crumb, ticker)

    # If unauthorized, refresh immediately
    if result is None and i < total - 1:
        try:
            session, crumb = init_session()
            result = fetch_ticker_fundamentals(session, crumb, ticker)
        except Exception:
            pass

    if result is not None:
        has_data = any(v is not None for v in result.values())
        if has_data:
            fundamentals[ticker] = result
            success_count += 1
        else:
            fail_count += 1
    else:
        fail_count += 1

    if (i + 1) % 50 == 0:
        print(f"  Progress: {i+1}/{total} tickers processed, {success_count} successful")

    # Rate limiting
    if i > 0 and i % 30 == 0:
        time.sleep(0.5)
    else:
        time.sleep(0.05)

print(f"\nFetch complete: {success_count}/{total} tickers with data")
print(f"Success rate: {success_count/total*100:.1f}%")

# If success rate is very low, try v8 API as alternative
if success_count < total * 0.10:
    print("\nLow success rate - trying v8 API as fallback...")
    for ticker in valid_clean_tickers[:20]:
        try:
            url = f'https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1d&range=1d'
            r = session.get(url, headers=api_headers, timeout=10)
            print(f"  v8 test {ticker}: {r.status_code}")
            break
        except Exception as e:
            print(f"  v8 test error: {e}")

# ─────────────────────────────────────────────
# STEP 3: Compute Eq_Fund_Score
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 3: Computing Eq_Fund_Score")
print("=" * 60)

fund_cols = ['Debt_to_Equity', 'Profit_Margin', 'Revenue_Growth', 'Current_Ratio', 'EV_EBITDA', 'PE_Ratio']

# Build DataFrame from fetched fundamentals
fund_df = pd.DataFrame.from_dict(fundamentals, orient='index')
fund_df.index.name = 'ticker'
fund_df = fund_df.reset_index()

print(f"Tickers with fundamental data: {len(fund_df)}")
for col in fund_cols:
    if col in fund_df.columns:
        print(f"  {col}: {fund_df[col].notna().sum()} non-null")

# Apply caps/adjustments before normalization
fund_df_norm = fund_df.copy()

# Cap Current_Ratio at 3
if 'Current_Ratio' in fund_df_norm.columns:
    fund_df_norm['Current_Ratio'] = fund_df_norm['Current_Ratio'].clip(upper=3.0)

# Cap PE_Ratio at 100, exclude negatives
if 'PE_Ratio' in fund_df_norm.columns:
    fund_df_norm['PE_Ratio'] = fund_df_norm['PE_Ratio'].where(fund_df_norm['PE_Ratio'] > 0)
    fund_df_norm['PE_Ratio'] = fund_df_norm['PE_Ratio'].clip(upper=100.0)

# EV_EBITDA: exclude negatives
if 'EV_EBITDA' in fund_df_norm.columns:
    fund_df_norm['EV_EBITDA'] = fund_df_norm['EV_EBITDA'].where(fund_df_norm['EV_EBITDA'] > 0)

# Normalize each metric: rank percentile 0=worst, 1=best
def percentile_rank(series, higher_is_better=True):
    """Returns percentile rank (0 to 1)."""
    if higher_is_better:
        ranked = series.rank(method='average', na_option='keep')
    else:
        ranked = (-series).rank(method='average', na_option='keep')
    n = series.notna().sum()
    if n == 0:
        return pd.Series(np.nan, index=series.index)
    return (ranked - 1) / (n - 1) if n > 1 else ranked / n

norm_cols = {}
for col, higher_is_better in [
    ('Debt_to_Equity', False),
    ('Profit_Margin', True),
    ('Revenue_Growth', True),
    ('Current_Ratio', True),
    ('EV_EBITDA', False),
    ('PE_Ratio', False),
]:
    if col in fund_df_norm.columns:
        norm_cols[col] = percentile_rank(fund_df_norm[col], higher_is_better)

norm_df = pd.DataFrame(norm_cols, index=fund_df_norm.index)
norm_df['ticker'] = fund_df_norm['ticker']

# Compute raw score = mean of available components (need >= 2)
score_cols = [c for c in fund_cols if c in norm_df.columns]
norm_df['available_count'] = norm_df[score_cols].notna().sum(axis=1)
norm_df['Eq_Fund_Score_raw'] = norm_df[score_cols].mean(axis=1)
# Set to NaN if fewer than 2 metrics available
norm_df.loc[norm_df['available_count'] < 2, 'Eq_Fund_Score_raw'] = np.nan
# Rescale to [-1, +1]
norm_df['Eq_Fund_Score'] = norm_df['Eq_Fund_Score_raw'] * 2 - 1
norm_df.loc[norm_df['available_count'] < 2, 'Eq_Fund_Score'] = np.nan

# Merge back to fund_df
fund_df = fund_df.merge(norm_df[['ticker', 'Eq_Fund_Score']], on='ticker', how='left')

print(f"\nEq_Fund_Score computed:")
print(f"  Non-null scores: {fund_df['Eq_Fund_Score'].notna().sum()}")
print(f"  Score range: [{fund_df['Eq_Fund_Score'].min():.4f}, {fund_df['Eq_Fund_Score'].max():.4f}]")

# ─────────────────────────────────────────────
# STEP 4: Map back to df and recompute Integrated_Score
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 4: Recomputing Integrated_Score and rankings")
print("=" * 60)

# Create ticker lookup: clean_ticker -> fundamentals
# We need to map clean tickers back to the Eqty Ticker (with ' US' suffix etc.)
# ticker_map = {raw_ticker: clean_ticker}
# Reverse: clean_ticker -> list of raw tickers
reverse_map = {}
for raw, clean in ticker_map.items():
    if clean not in reverse_map:
        reverse_map[clean] = []
    reverse_map[clean].append(raw)

# Build fundamental lookup by raw Eqty Ticker
fund_by_raw = {}
for _, row in fund_df.iterrows():
    clean_t = row['ticker']
    raw_tickers = reverse_map.get(clean_t, [])
    for raw_t in raw_tickers:
        fund_by_raw[raw_t] = row

# Apply fundamentals to df
for col in fund_cols + ['Eq_Fund_Score']:
    df[col] = np.nan

for idx, row in df.iterrows():
    raw_t = str(row.get('Eqty Ticker', '')).strip()
    if raw_t in fund_by_raw:
        fund_row = fund_by_raw[raw_t]
        for col in fund_cols + ['Eq_Fund_Score']:
            if col in fund_row.index:
                df.at[idx, col] = fund_row[col]

print(f"Fundamental columns updated in df:")
for col in fund_cols + ['Eq_Fund_Score']:
    print(f"  {col}: {df[col].notna().sum()} non-null")

# Recompute Integrated_Score
eq_mom = df['Eq_Mom_Score'].fillna(0)
eq_fund = df['Eq_Fund_Score'].fillna(0)
bond_tr = df['Bond_TR_Est_pct']

# Normalize Bond_TR_Est_pct → [-1, +1] within each class
bond_tr_score = pd.Series(np.nan, index=bond_tr.index)
active_mask = bond_tr.notna() & df['class'].notna() & (df['class'].astype(str).str.lower() != 'off')
for cls in df.loc[active_mask, 'class'].unique():
    cls_mask = active_mask & (df['class'] == cls)
    tr_cls = bond_tr[cls_mask]
    n_tr = tr_cls.notna().sum()
    if n_tr > 1:
        tr_ranked = tr_cls.rank(method='average', na_option='keep')
        bond_tr_score[cls_mask] = (tr_ranked - 1) / (n_tr - 1) * 2 - 1
df['Bond_TR_Score'] = bond_tr_score

df['Integrated_Score'] = bond_tr_score.fillna(0) * 0.25 + eq_mom * 0.25 + eq_fund * 0.25

# Recompute Integrated_Rank_in_Class (excluding 'off' class)
df['Integrated_Rank_in_Class'] = np.nan

for cls, grp in df.groupby('class'):
    if str(cls).lower() == 'off':
        continue
    valid_mask = df['class'] == cls
    ranks = df.loc[valid_mask, 'Integrated_Score'].rank(method='min', ascending=False)
    df.loc[valid_mask, 'Integrated_Rank_in_Class'] = ranks

# Recompute Top_Pick_Flag
def assign_flag(rank):
    if pd.isna(rank):
        return ''
    r = int(rank)
    if r <= 3:
        return '★★★ TOP3'
    elif r <= 10:
        return '★★ TOP10'
    elif r <= 25:
        return '★ TOP25'
    return ''

df['Top_Pick_Flag'] = df['Integrated_Rank_in_Class'].apply(assign_flag)

print(f"\nTop_Pick_Flag distribution:")
print(df['Top_Pick_Flag'].value_counts().to_string())

# ─────────────────────────────────────────────
# STEP 5: Update Excel file with openpyxl
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 5: Updating Excel file")
print("=" * 60)

wb = load_workbook(EXCEL_PATH)

# ── Detail_Scored sheet ──
ws_detail = wb['Detail_Scored']

# Find header row (row 2) and build column index map
header_row = 2
col_map = {}
for cell in ws_detail[header_row]:
    if cell.value is not None:
        col_map[str(cell.value).strip()] = cell.column

print(f"Column map keys (sample): {list(col_map.keys())[:10]}")

update_cols = [
    'Debt_to_Equity', 'Profit_Margin', 'Revenue_Growth', 'Current_Ratio',
    'EV_EBITDA', 'PE_Ratio', 'Eq_Fund_Score',
    'Integrated_Score', 'Integrated_Rank_in_Class', 'Top_Pick_Flag'
]

# Verify all columns found
for col in update_cols:
    if col in col_map:
        print(f"  Found column '{col}' at col {col_map[col]}")
    else:
        print(f"  WARNING: Column '{col}' NOT found in header!")

# Flag formatting
flag_styles = {
    '★★★ TOP3': {'font': Font(bold=True, color='FF0000'), 'fill': PatternFill('solid', start_color='FFFF00', end_color='FFFF00')},
    '★★ TOP10': {'font': Font(bold=True, color='C55A11'), 'fill': PatternFill('solid', start_color='FCE4D6', end_color='FCE4D6')},
    '★ TOP25':  {'font': Font(bold=True, color='1F3864'), 'fill': PatternFill('solid', start_color='DDEBF7', end_color='DDEBF7')},
    '':         {'font': Font(bold=False, color='000000'), 'fill': PatternFill(fill_type=None)},
}

# Data rows start at row 3
df_reset = df.reset_index(drop=True)
for df_idx, row in df_reset.iterrows():
    excel_row = df_idx + 3  # row 3 onwards (header=row2, data starts row3)

    for col_name in update_cols:
        if col_name not in col_map:
            continue
        col_idx = col_map[col_name]
        val = row.get(col_name, np.nan)

        # Convert NaN/nan to None for Excel
        if isinstance(val, float) and np.isnan(val):
            val = None
        elif isinstance(val, (np.floating, np.integer)):
            val = val.item()

        cell = ws_detail.cell(row=excel_row, column=col_idx)
        cell.value = val

        if col_name == 'Top_Pick_Flag':
            flag_val = str(val) if val is not None else ''
            style = flag_styles.get(flag_val, flag_styles[''])
            cell.font = style['font']
            cell.fill = style['fill']

print(f"  Detail_Scored updated ({len(df_reset)} data rows)")

# ── Equity_Data sheet ──
ws_equity = wb['Equity_Data']

# Header is row 1 in Equity_Data
eq_col_map = {}
for cell in ws_equity[1]:
    if cell.value is not None:
        eq_col_map[str(cell.value).strip()] = cell.column

eq_fund_update_cols = [
    'Debt_to_Equity', 'Profit_Margin', 'Revenue_Growth', 'Current_Ratio',
    'EV_EBITDA', 'PE_Ratio', 'Eq_Fund_Score'
]

# Build lookup by clean ticker (Equity_Data uses clean tickers)
fund_by_clean = {}
for _, frow in fund_df.iterrows():
    fund_by_clean[frow['ticker']] = frow

# Also build Eq_Fund_Score lookup by clean ticker from df
eq_fund_score_by_clean = {}
for _, drow in df.iterrows():
    raw_t = str(drow.get('Eqty Ticker', '')).strip()
    if raw_t in ticker_map:
        clean_t = ticker_map[raw_t]
        if pd.notna(drow.get('Eq_Fund_Score')):
            eq_fund_score_by_clean[clean_t] = drow['Eq_Fund_Score']

eq_ticker_col = eq_col_map.get('Ticker')
for eq_row_idx in range(2, ws_equity.max_row + 1):
    if eq_ticker_col is None:
        break
    ticker_cell = ws_equity.cell(row=eq_row_idx, column=eq_ticker_col)
    clean_t = str(ticker_cell.value).strip() if ticker_cell.value else ''
    if not clean_t or clean_t == 'nan':
        continue

    fund_row = fund_by_clean.get(clean_t)
    for col_name in eq_fund_update_cols:
        if col_name not in eq_col_map:
            continue
        col_idx = eq_col_map[col_name]
        if fund_row is not None and col_name in fund_row.index:
            val = fund_row[col_name]
            if isinstance(val, float) and np.isnan(val):
                val = None
            elif isinstance(val, (np.floating, np.integer)):
                val = val.item()
        else:
            val = None
        ws_equity.cell(row=eq_row_idx, column=col_idx).value = val

print(f"  Equity_Data sheet updated")

# ── Top_Picks_by_Class sheet ──
# Drop and recreate
if 'Top_Picks_by_Class' in wb.sheetnames:
    del wb['Top_Picks_by_Class']

ws_top = wb.create_sheet('Top_Picks_by_Class')

# Title row
top_cols = ['class', 'Des', 'ISIN', 'Ticker', 'Cpn', 'OAS', 'OASD',
            'Carry_2.5M_pct', 'Compression_Score_pct', 'Bond_TR_Est_pct',
            'Eq_Ret_1M', 'Eq_Ret_3M', 'Eq_Mom_Score', 'Eq_Fund_Score',
            'Integrated_Score', 'Integrated_Rank_in_Class', 'Top_Pick_Flag',
            'Issuer Rtg', 'S&P Outlook', "Moody's Outlook", 'BCLASS3', 'Industry Sector']

title_cell = ws_top.cell(row=1, column=1)
title_cell.value = 'Top 5 Integrated Score Candidates by Class (2-3M Horizon) | As of 2026-03-31'
title_cell.font = Font(bold=True)

# Header row
header_font = Font(bold=True)
header_fill = PatternFill('solid', start_color='D9E1F2', end_color='D9E1F2')
for c_idx, col_name in enumerate(top_cols, start=1):
    cell = ws_top.cell(row=2, column=c_idx)
    cell.value = col_name
    cell.font = header_font
    cell.fill = header_fill

# Get top 5 per class
top_df = (df[df['class'].notna() & (df['class'].str.lower() != 'off')]
          .sort_values(['class', 'Integrated_Rank_in_Class'])
          .groupby('class')
          .head(5)
          .reset_index(drop=True))

data_row = 3
for _, row in top_df.iterrows():
    for c_idx, col_name in enumerate(top_cols, start=1):
        val = row.get(col_name, None)
        if isinstance(val, float) and np.isnan(val):
            val = None
        elif isinstance(val, (np.floating, np.integer)):
            val = val.item()
        cell = ws_top.cell(row=data_row, column=c_idx)
        cell.value = val

        if col_name == 'Top_Pick_Flag':
            flag_val = str(val) if val is not None else ''
            style = flag_styles.get(flag_val, flag_styles[''])
            cell.font = style['font']
            cell.fill = style['fill']
    data_row += 1

print(f"  Top_Picks_by_Class rebuilt with {len(top_df)} rows")

# Save
wb.save(EXCEL_PATH)
print(f"\nFile saved: {EXCEL_PATH}")

# ─────────────────────────────────────────────
# STEP 6: Verify
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 6: Verification")
print("=" * 60)

import os
file_size = os.path.getsize(EXCEL_PATH)
print(f"File size: {file_size / 1024 / 1024:.2f} MB")

# Reload and verify
df_verify = pd.read_excel(EXCEL_PATH, sheet_name='Detail_Scored', header=1)
print(f"\nFundamental data in updated file:")
for col in fund_cols + ['Eq_Fund_Score']:
    nn = df_verify[col].notna().sum()
    print(f"  {col}: {nn} non-null")

print(f"\nTickers with fundamentals successfully fetched: {success_count}")

print(f"\nSample 5 rows with non-null Eq_Fund_Score:")
sample = df_verify[df_verify['Eq_Fund_Score'].notna()][
    ['Eqty Ticker', 'Debt_to_Equity', 'Profit_Margin', 'Revenue_Growth',
     'Current_Ratio', 'EV_EBITDA', 'PE_Ratio', 'Eq_Fund_Score', 'Integrated_Score', 'Top_Pick_Flag']
].head(5)
print(sample.to_string())

print("\nFUNDAMENTALS UPDATE COMPLETE")
