# -*- coding: utf-8 -*-
"""
backtest_combo.py
4-component weight combination comparison backtest
  B=Bond_TR_Score  M=Eq_Mom_Score  F=Eq_Fund_Score  A=AI_Macro_Score
Uses cached price/fund data for speed.
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import os, re, time, ssl, warnings
import numpy as np
import pandas as pd
from datetime import date
from itertools import groupby
from collections import defaultdict

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import matplotlib.gridspec as gridspec

warnings.filterwarnings('ignore')
ssl._create_default_https_context = ssl._create_unverified_context
os.environ['CURL_CA_BUNDLE'] = ''

# ── 설정 ──────────────────────────────────────────────────────────────────────
PATH_OLD   = r"Z:\USIGAIQ\USIG Market data\LUACSTAT"
PATH_NEW   = r"Z:\USIGAIQ\USIG Market data\LUACSTAT_PREP"
EQ_CACHE   = r"C:\Users\sh.park\Documents\USIG_LLM\eq_price_cache.pkl"
FUND_CACHE = r"C:\Users\sh.park\Documents\USIG_LLM\eq_fund_cache.pkl"
OUT_PNG    = r"C:\Users\sh.park\Documents\USIG_LLM\backtest_combo.png"
OUT_XLSX   = r"C:\Users\sh.park\Documents\USIG_LLM\backtest_combo.xlsx"

BACKTEST_START = date(2021, 1, 1)
TOP_N          = 3
MIN_CLASS_SIZE = 5

# ── 비교할 가중치 조합 ─────────────────────────────────────────────────────────
# (name, w_bond_tr, w_eq_mom, w_eq_fund, w_ai_macro)
COMBOS = [
    # ── 단일 컴포넌트 ─────────────────────────────────────────────────────────
    ("Bond_TR Only",       1.00, 0.00, 0.00, 0.00),
    ("Eq_Mom Only",        0.00, 1.00, 0.00, 0.00),
    ("Eq_Fund Only",       0.00, 0.00, 1.00, 0.00),
    ("AI_Macro Only",      0.00, 0.00, 0.00, 1.00),
    # ── 2-way ─────────────────────────────────────────────────────────────────
    ("BondTR+AI (50/50)",  0.50, 0.00, 0.00, 0.50),
    ("BondTR+Mom (50/50)", 0.50, 0.50, 0.00, 0.00),
    ("BondTR+Fund(50/50)", 0.50, 0.00, 0.50, 0.00),
    ("Mom+AI (50/50)",     0.00, 0.50, 0.00, 0.50),
    # ── 3-way (Eq_Fund 제외) ──────────────────────────────────────────────────
    ("B+M+AI (33/33/33)",  0.33, 0.33, 0.00, 0.34),
    ("B+F+AI (33/33/33)",  0.33, 0.00, 0.33, 0.34),
    ("B+M+F  (33/33/33)",  0.33, 0.33, 0.34, 0.00),
    # ── Equal 4-way ───────────────────────────────────────────────────────────
    ("Equal 4way (25×4)",  0.25, 0.25, 0.25, 0.25),
    # ── BondTR 중심 ───────────────────────────────────────────────────────────
    ("B40+M20+F20+A20",    0.40, 0.20, 0.20, 0.20),
    ("B50+M17+F17+A16",    0.50, 0.17, 0.17, 0.16),
    ("B40+M30+AI30",       0.40, 0.30, 0.00, 0.30),
    ("B60+M20+AI20",       0.60, 0.20, 0.00, 0.20),
    # ── AI_Macro 중심 ─────────────────────────────────────────────────────────
    ("B20+M20+F20+A40",    0.20, 0.20, 0.20, 0.40),
    ("B25+M25+AI50",       0.25, 0.25, 0.00, 0.50),
    # ── Eq_Mom 중심 ───────────────────────────────────────────────────────────
    ("B20+M40+F20+A20",    0.20, 0.40, 0.20, 0.20),
    ("B30+M40+AI30",       0.30, 0.40, 0.00, 0.30),
]

# ── 매핑 테이블 ───────────────────────────────────────────────────────────────
RTG_TEXT_TO_NUM = {
    'AAA':1,'AA+':2,'AA':3,'AA-':4,'A+':5,'A':6,'A-':7,
    'BBB+':8,'BBB':9,'BBB-':10,'BB+':11,'BB':12,'BB-':13,
}
RATING_CUT = 7

DP_RATING_MAP = {
    'AAA':1,'AA1':2,'AA2':3,'AA3':4,'A1':5,'A2':6,'A3':7,
    'BAA1':8,'BAA2':9,'BAA3':10,'BA1':11,'BA2':12,'BA3':13,
    'B1':14,'B2':15,'B3':16,'CAA1':17,'CAA2':18,'CAA3':19,'CA':20,'C':21,
}

SUBGROUP_SCORE_MAP = {
    'Electric-Integrated':0.85,'Electric-Distribution':0.85,'Electric-Transmission':0.80,
    'Electric-Generation':0.75,'Gas-Distribution':0.80,'Water':0.85,
    'Non-hazardous Waste Disp':0.70,'Medical-Drugs':0.70,'Medical-Hospitals':0.75,
    'Medical-HMO':0.65,'Medical-Biomedical/Gene':0.60,'Medical Products':0.60,
    'Medical Instruments':0.55,'Medical Labs&Testing Srv':0.65,'Pharmacy Services':0.60,
    'Medical-Whsle Drug Dist':0.55,'Medical Imaging Systems':0.55,'Diagnostic Equipment':0.55,
    'Drug Delivery Systems':0.55,'Medical-Generic Drugs':0.50,'Phys Practice Mgmnt':0.60,
    'Medical-Outptnt/Home Med':0.60,'Diversified Banking Inst':0.55,
    'Super-Regional Banks-US':0.60,'Commer Banks-Eastern US':0.55,
    'Commer Banks-Southern US':0.55,'Commer Banks-Central US':0.50,
    'Commer Banks-Western US':0.50,'Commer Banks Non-US':0.30,'Money Center Banks':0.45,
    'Fiduciary Banks':0.65,'Life/Health Insurance':0.50,'Property/Casualty Ins':0.55,
    'Reinsurance':0.50,'Multi-line Insurance':0.45,'Insurance Brokers':0.55,
    'Financial Guarantee Ins':0.20,'Finance-Invest Bnkr/Brkr':0.35,
    'Invest Mgmnt/Advis Serv':0.40,'Private Equity':0.20,'Finance-Credit Card':0.30,
    'Finance-Leasing Compan':0.25,'Finance-Auto Loans':0.10,'Finance-Other Services':0.30,
    'Finance-Mtge Loan/Banker':0.20,'Finance-Consumer Loans':0.10,'Venture Capital':0.00,
    'Investment Companies':0.15,'Pipelines':0.30,'Oil Comp-Integrated':0.10,
    'Oil Comp-Explor&Prodtn':-0.10,'Oil Refining&Marketing':-0.05,'Oil-Field Services':-0.20,
    'Oil&Gas Drilling':-0.25,'Agricultural Chemicals':0.20,
    'Auto-Cars/Light Trucks':-0.90,'Auto-Med&Heavy Duty Trks':-0.70,
    'Auto/Trk Prts&Equip-Orig':-0.80,'Retail-Automobile':-0.60,
    'Retail-Discount':-0.20,'Retail-Building Products':-0.30,'Retail-Major Dept Store':-0.50,
    'Retail-Apparel/Shoe':-0.60,'Retail-Auto Parts':-0.40,'Retail-Sporting Goods':-0.40,
    'Retail-Consumer Electron':-0.50,'Retail-Restaurants':-0.10,'Retail-Gardening Prod':-0.30,
    'Hotels&Motels':-0.15,'Casino Hotels':-0.20,'Cruise Lines':-0.25,
    'E-Commerce/Products':-0.20,'E-Commerce/Services':-0.10,'Internet Content-Entmnt':0.00,
    'Apparel Manufacturers':-0.55,'Athletic Footwear':-0.55,'Recreational Vehicles':-0.40,
    'Beverages-Non-alcoholic':0.55,'Beverages-Wine/Spirits':0.45,'Brewery':0.45,
    'Food-Misc/Diversified':0.50,'Food-Retail':0.40,'Food-Confectionery':0.50,
    'Food-Meat Products':0.45,'Food-Baking':0.45,'Food-Wholesale/Distrib':0.40,
    'Poultry':0.40,'Coffee':0.45,'Cosmetics&Toiletries':0.55,'Soap&Cleaning Prepar':0.50,
    'Consumer Products-Misc':0.40,'Tobacco':0.60,'Agricultural Operations':0.35,
    'Electronic Compo-Semicon':-0.30,'Semicon Compo-Intg Circu':-0.25,
    'Semiconductor Equipment':-0.35,'Enterprise Software/Serv':0.10,
    'Applications Software':0.05,'Computer Services':0.15,'Data Processing/Mgmt':0.20,
    'Computer Aided Design':0.00,'Software Tools':0.00,'E-Marketing/Info':0.00,
    'Decision Support Softwar':0.05,'Computers':-0.15,'Computers-Memory Devices':-0.20,
    'Computers-Other':-0.10,'Electronic Connectors':-0.20,'Electronic Compo-Misc':-0.15,
    'Electronic Parts Distrib':-0.10,'Electronic Measur Instr':-0.10,
    'Electronic Secur Devices':-0.05,'Electronic Forms':0.00,'Networking Products':-0.10,
    'Wireless Equipment':-0.15,'Telecom Eq Fiber Optics':-0.10,
    'Telecommunication Equip':-0.10,'Industrial Automat/Robot':-0.10,
    'Instruments-Controls':-0.05,'Telephone-Integrated':0.10,'Cellular Telecom':0.05,
    'Cable/Satellite TV':0.00,'Multimedia':0.00,'Broadcast Serv/Program':0.00,
    'Advertising Agencies':-0.15,'Advertising Services':-0.15,'Telecom Services':0.10,
    'Web Portals/ISP':0.00,'Aerospace/Defense':0.40,'Aerospace/Defense-Equip':0.35,
    'Machinery-Farm':0.20,'Machinery-General Indust':-0.10,'Machinery-Constr&Mining':-0.10,
    'Machinery-Pumps':-0.05,'Machinery-Electric Util':0.10,'Tools-Hand Held':-0.10,
    'Diversified Manufact Op':-0.05,'Industrial Gases':0.30,'Chemicals-Diversified':-0.10,
    'Chemicals-Specialty':-0.05,'Coatings/Paint':-0.10,'Containers-Paper/Plastic':-0.20,
    'Paper&Related Products':-0.15,'Commercial Serv-Finance':0.10,'Commercial Services':0.05,
    'Consulting Services':0.10,'Non-Profit Charity':0.00,'Distribution/Wholesale':-0.10,
    'Office Automation&Equip':-0.15,'Office Supplies&Forms':-0.15,
    'Bldg Prod-Cement/Aggreg':-0.10,'Bldg Prod-Air&Heating':-0.10,'Bldg Prod-Wood':-0.15,
    'Bldg&Construct Prod-Misc':-0.10,'Bldg-Residential/Commer':-0.20,
    'Building-Maint&Service':-0.05,'Shipbuilding':0.20,'Power Conv/Supply Equip':0.10,
    'Transport-Rail':0.30,'Transport-Services':0.00,'Transport-Equip&Leasng':-0.10,
    'Transport-Truck':-0.10,'Transport-Marine':-0.10,'Airlines':-0.40,
    'Steel-Producers':0.20,'Steel Pipe&Tube':0.15,'Metal-Iron':0.10,'Metal-Aluminum':0.10,
    'Metal-Copper':0.00,'Metal-Diversified':0.00,'Diversified Minerals':0.00,
    'Gold Mining':0.30,'Metal Processors&Fabrica':0.05,'Schools':0.40,'Toys':-0.40,
    'Entertainment Software':0.00,'Engineering/R&D Services':0.15,'Vitamins&Nutrition Prod':0.30,
    'Rental Auto/Equipment':-0.10,'Mach Tools&Rel Products':-0.10,
    'Motorcycle/Motor Scooter':-0.50,'Chemicals-Plastics':-0.15,
    'Miscellaneous Manufactur':-0.10,'REITS-Industrial':0.10,'REITS-Warehouse/Industr':0.20,
    'REITS-Diversified':-0.10,'REITS-Apartments':-0.05,'REITS-Shopping Centers':-0.30,
    'REITS-Regional Malls':-0.40,'REITS-Office Property':-0.35,'REITS-Health Care':0.30,
    'REITS-Storage':0.10,'REITS-Single Tenant':0.00,'REITS-Hotels':-0.20,
    'REITS-Manufactured Homes':0.00,'REITS-Mortgage':-0.20,'Real Estate Mgmnt/Servic':-0.10,
}
BCLASS3_SCORE_MAP = {
    'Utility':0.75,'Healthcare':0.60,'Banking':0.50,'Financial Other':0.30,
    'Insurance':0.50,'Energy':0.05,'Consumer Cyclical':0.20,
    'Consumer Non-Cyclical':0.45,'Technology':0.00,'Communications':0.05,
    'Industrial':0.00,'Transportation':0.00,'Basic Industry':-0.05,
    'Real Estate':-0.10,'Government Related':0.30,'Sovereign':0.20,
    'Quasi-Government':0.25,'Supranational':0.40,
}
RATING_BUFFER_MAP = {
    'AAA':0.50,'AA1':0.55,'AA2':0.55,'AA3':0.50,
    'A1':0.70,'A2':0.70,'A3':0.65,
    'BAA1':0.25,'BAA2':0.05,'BAA3':-0.65,
    'BA1':-0.85,'BA2':-1.00,'BA3':-1.00,
    'B1':-1.00,'B2':-1.00,'B3':-1.00,
}

def maturity_bucket(yrs):
    if pd.isna(yrs): return None
    if yrs <= 3:   return '~3Y'
    elif yrs <= 7: return '3~7Y'
    elif yrs <= 15:return '7~15Y'
    else:          return '15Y+'

def maturity_score(oad):
    if pd.isna(oad): return 0.0
    oad = float(oad)
    if oad < 2:    return  0.00
    elif oad < 4:  return  0.30
    elif oad < 7:  return  1.00
    elif oad < 10: return  0.70
    elif oad < 13: return  0.00
    elif oad < 16: return -0.50
    else:          return -1.00

def percentile_norm(series):
    n = series.notna().sum()
    if n < 2: return pd.Series(np.nan, index=series.index)
    ranked = series.rank(method='average', na_option='keep')
    return (ranked - 1) / (n - 1) * 2 - 1

# ═══════════════════════════════════════════════════════════════════════════════
# 1. 월별 파일 목록
# ═══════════════════════════════════════════════════════════════════════════════
pattern = re.compile(r'^LUACSTAT_(\d{4})_(\d{2})_(\d{2})\.(xls|xlsx)$')
dated_dict = {}
for p in [PATH_OLD, PATH_NEW]:
    for f in os.listdir(p):
        m = pattern.match(f)
        if m:
            y,mo,d,ext = int(m.group(1)),int(m.group(2)),int(m.group(3)),m.group(4)
            try:
                dt = date(y,mo,d)
                if dt not in dated_dict or ext=='xlsx':
                    dated_dict[dt] = (os.path.join(p,f), ext)
            except: pass

dated_sorted = sorted(dated_dict.items())
monthly = []
for (yr,mo), grp in groupby(dated_sorted, key=lambda x:(x[0].year,x[0].month)):
    monthly.append(list(grp)[-1])
monthly_bt = [(d,fp) for d,(fp,ext) in monthly if d.year >= BACKTEST_START.year-1]

print(f"Files: {len(monthly_bt)}, Range: {monthly_bt[0][0]} ~ {monthly_bt[-1][0]}")

# ═══════════════════════════════════════════════════════════════════════════════
# 2. 파일 로딩
# ═══════════════════════════════════════════════════════════════════════════════
def load_file(filepath):
    df_raw = pd.read_excel(filepath, header=None, nrows=15)
    header_row = None
    for i, row in df_raw.iterrows():
        vals = [str(v) for v in row.values]
        if 'ISIN' in vals and ('OAS' in vals or 'Yield to Worst' in vals):
            header_row = i; break
    if header_row is None: return None

    df = pd.read_excel(filepath, header=header_row)
    if 'Composite Rating Num' not in df.columns and 'BB Comp' in df.columns:
        df['Composite Rating Num'] = df['BB Comp'].map(RTG_TEXT_TO_NUM)
    elif 'Composite Rating Num' not in df.columns and 'Index Rtg' in df.columns:
        df['Composite Rating Num'] = df['Index Rtg'].map(RTG_TEXT_TO_NUM)
    if 'Mty (Yrs)' not in df.columns and 'Eff Mty (Yrs)' in df.columns:
        df.rename(columns={'Eff Mty (Yrs)': 'Mty (Yrs)'}, inplace=True)
    if 'OASD' not in df.columns and 'Spd Dur' in df.columns:
        df.rename(columns={'Spd Dur': 'OASD'}, inplace=True)

    isin_mask = df['ISIN'].astype(str).str.match(r'^[A-Z]{2}[A-Z0-9]{10}$', na=False) if 'ISIN' in df.columns else pd.Series(False, index=df.index)
    df_bonds = df[isin_mask].copy()
    for col in ['OAS','OASD','OAD','Yield to Worst','Mty (Yrs)',
                'Composite Rating Num','1Y Dflt','Total Return - 1mo']:
        if col in df_bonds.columns:
            df_bonds[col] = pd.to_numeric(df_bonds[col], errors='coerce')
    df_bonds['ISIN'] = df_bonds['ISIN'].astype(str).str.strip()
    df_bonds = df_bonds[df_bonds['ISIN'].str.match(r'^[A-Z]{2}[A-Z0-9]{10}$', na=False)]
    return df_bonds

# ═══════════════════════════════════════════════════════════════════════════════
# 3. 캐시 로드 및 사전 계산
# ═══════════════════════════════════════════════════════════════════════════════
print("\nLoading caches...")
price_cache = pd.read_pickle(EQ_CACHE)   if os.path.exists(EQ_CACHE)   else {}
fund_raw    = pd.read_pickle(FUND_CACHE) if os.path.exists(FUND_CACHE) else {}
print(f"  Price cache: {len(price_cache)} tickers")
print(f"  Fund cache:  {len(fund_raw)} tickers")

def rank_norm_01(series, ascending=True):
    s = pd.to_numeric(series, errors='coerce')
    valid = s.notna()
    result = pd.Series(np.nan, index=s.index)
    if valid.sum() < 2: return result
    ranked = s[valid].rank(ascending=ascending, method='average')
    result[valid] = ranked / valid.sum()
    return result

# Eq_Fund_Score 사전 계산
if fund_raw:
    fund_rows = [{'ticker': t, **v} for t, v in fund_raw.items() if v]
    fund_df = pd.DataFrame(fund_rows).set_index('ticker') if fund_rows else pd.DataFrame()
    if len(fund_df) > 0:
        ev_ = fund_df.get('EV_EBITDA', pd.Series(dtype=float)).copy()
        ev_[ev_ <= 0] = np.nan
        comps = pd.DataFrame({
            'fd': rank_norm_01(fund_df.get('Debt_to_Equity', pd.Series(dtype=float)), ascending=False),
            'fp': rank_norm_01(fund_df.get('Profit_Margin',  pd.Series(dtype=float)), ascending=True),
            'fg': rank_norm_01(fund_df.get('Revenue_Growth', pd.Series(dtype=float)), ascending=True),
            'fc': rank_norm_01(fund_df.get('Current_Ratio',  pd.Series(dtype=float)), ascending=True),
            'fe': rank_norm_01(ev_, ascending=False),
        })
        fund_df['Eq_Fund_Score'] = comps.mean(axis=1) * 2 - 1
    else:
        fund_df = pd.DataFrame(columns=['Eq_Fund_Score'])
else:
    fund_df = pd.DataFrame(columns=['Eq_Fund_Score'])

print(f"  Eq_Fund_Score: {fund_df['Eq_Fund_Score'].notna().sum() if 'Eq_Fund_Score' in fund_df.columns else 0} tickers computed")

def get_eq_ticker(t):
    if pd.isna(t): return None
    t = str(t).strip()
    if t in ('','N/A','nan','None'): return None
    return t.split()[0]

ticker_col_candidates = ['Eqty Ticker','Equity Ticker','Ticker']

# ── Eq_Mom_Score 함수 ──────────────────────────────────────────────────────────
def compute_eq_mom_scores(tickers, as_of_date):
    as_of_ts = pd.Timestamp(as_of_date)
    one_mo   = as_of_ts - pd.DateOffset(months=1)
    three_mo = as_of_ts - pd.DateOffset(months=3)
    yr52     = as_of_ts - pd.DateOffset(weeks=52)
    metrics_list = []
    for t in tickers:
        prices = price_cache.get(t)
        if prices is None or len(prices) < 10:
            metrics_list.append({'ticker': t}); continue
        p = prices[prices.index <= as_of_ts]
        if len(p) < 5:
            metrics_list.append({'ticker': t}); continue
        cur = float(p.iloc[-1])
        m = {'ticker': t}
        p1 = p[p.index <= one_mo]
        if len(p1) > 0: m['r1'] = cur / float(p1.iloc[-1]) - 1
        p3 = p[p.index <= three_mo]
        if len(p3) > 0: m['r3'] = cur / float(p3.iloc[-1]) - 1
        lr = np.log(p / p.shift(1)).dropna()
        if len(lr) >= 20: m['vol'] = float(lr.iloc[-30:].std() * np.sqrt(252))
        p52 = p[p.index >= yr52]
        if len(p52) > 0:
            hi = float(p52.max())
            if hi > 0: m['h52'] = cur / hi - 1
        metrics_list.append(m)
    mdf = pd.DataFrame(metrics_list).set_index('ticker')
    def rn(s, asc=True):
        s = pd.to_numeric(s, errors='coerce'); valid = s.notna()
        res = pd.Series(np.nan, index=s.index)
        if valid.sum() < 2: return res
        ranked = s[valid].rank(ascending=asc, method='average')
        res[valid] = (ranked - 1) / (valid.sum() - 1) * 2 - 1
        return res
    c = pd.DataFrame({
        'n1': rn(mdf.get('r1',  pd.Series(dtype=float)), True),
        'n3': rn(mdf.get('r3',  pd.Series(dtype=float)), True),
        'nv': rn(mdf.get('vol', pd.Series(dtype=float)), False),
        'nh': rn(mdf.get('h52', pd.Series(dtype=float)), True),
    })
    mdf['score'] = c.mean(axis=1)
    return {t: float(mdf.loc[t,'score']) for t in tickers
            if t in mdf.index and not pd.isna(mdf.loc[t,'score'])}

# ═══════════════════════════════════════════════════════════════════════════════
# 4. 월별 공통 요소 사전 계산 (모든 콤보가 공유)
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[Phase 1] Pre-computing monthly bond features...")

monthly_data = []   # list of dicts: date, isin_to_tr, bench_tr, df_with_scores

for i in range(len(monthly_bt) - 1):
    dt_t,  fp_t  = monthly_bt[i]
    dt_t1, fp_t1 = monthly_bt[i + 1]
    if dt_t < BACKTEST_START: continue

    df_t = load_file(fp_t)
    if df_t is None or len(df_t) < 100: continue

    df_t1 = load_file(fp_t1)
    if df_t1 is None: continue

    # 벤치마크 TR
    df_t1['_tr'] = pd.to_numeric(df_t1['Total Return - 1mo'], errors='coerce')
    if 'Mkt Val' in df_t1.columns:
        df_t1['_mv'] = pd.to_numeric(df_t1['Mkt Val'], errors='coerce')
        vb = df_t1.dropna(subset=['_mv','_tr'])
        vb = vb[vb['_mv'] > 0]
        bench_tr = float(np.average(vb['_tr'], weights=vb['_mv'])) if len(vb) > 100 else float(df_t1['_tr'].mean())
    else:
        bench_tr = float(df_t1['_tr'].mean())
    isin_to_tr = dict(zip(df_t1['ISIN'], df_t1['_tr']))

    # ── T 시점 필터 ─────────────────────────────────────────────────────────
    mask = (
        df_t['OAS'].notna() & df_t['OAS'].between(-50,600) &
        df_t['OASD'].notna() & df_t['OASD'].gt(0) &
        df_t['Yield to Worst'].notna() & df_t['Yield to Worst'].gt(0) &
        df_t['Mty (Yrs)'].notna() &
        df_t['BCLASS3'].notna() &
        df_t['Composite Rating Num'].notna()
    )
    df = df_t[mask].copy()
    df['_mty_bkt'] = df['Mty (Yrs)'].apply(maturity_bucket)
    df['_rtg_bkt'] = df['Composite Rating Num'].apply(
        lambda x: 'A-이상' if x <= RATING_CUT else 'BBB+이하')
    df['_bclass3'] = df['BCLASS3'].astype(str).str.strip()
    df['_class'] = df['_bclass3'] + '|' + df['_mty_bkt'] + '|' + df['_rtg_bkt']

    # Bond_TR 요소
    df['_carry']    = df['Yield to Worst'] / 12 * 2.5
    df['_comp_raw'] = (df['OAS'] - df['1Y Dflt'].fillna(0) * 60) * df['OASD']
    if 'DPFundamentalRating' in df.columns and 'DPSpreadRating' in df.columns:
        df['_dp_gap'] = df['DPSpreadRating'].map(DP_RATING_MAP) - df['DPFundamentalRating'].map(DP_RATING_MAP)
        df['_dp_score'] = percentile_norm(df['_dp_gap'])
    else:
        df['_dp_score'] = 0.0
    df['_bond_tr_est'] = df['_carry'].fillna(0) + df['_comp_raw'].fillna(0) + df['_dp_score'].fillna(0)*0.05

    # AI_Macro 요소
    sg_col = next((c for c in ['Industry Subgroup','BCLASS4','BCLASS3'] if c in df.columns), None)
    if sg_col:
        df['_ai_sector'] = df[sg_col].astype(str).map(SUBGROUP_SCORE_MAP)
        fallback = df['_ai_sector'].isna()
        if fallback.any():
            df.loc[fallback,'_ai_sector'] = df.loc[fallback,'_bclass3'].map(BCLASS3_SCORE_MAP)
        df['_ai_sector'] = df['_ai_sector'].fillna(0.0)
    else:
        df['_ai_sector'] = 0.0
    oad_col = 'OAD' if 'OAD' in df.columns else None
    if oad_col:
        df['_ai_mty'] = df['OAD'].apply(maturity_score)
    else:
        df['_ai_mty'] = df['Mty (Yrs)'].apply(lambda y: maturity_score(y*0.9) if not pd.isna(y) else 0.0)
    rtg_col = next((c for c in ['DPFundamentalRating','Issuer Rtg'] if c in df.columns), None)
    df['_ai_rtg'] = df[rtg_col].map(RATING_BUFFER_MAP).fillna(0.0) if rtg_col else 0.0
    df['_ai_macro'] = (df['_ai_sector']*0.40 + df['_ai_mty']*0.35 + df['_ai_rtg']*0.25).clip(-1,1)

    # Equity 스코어
    eq_col = next((c for c in ticker_col_candidates if c in df.columns), None)
    if eq_col:
        df['_eq_tk'] = df[eq_col].apply(get_eq_ticker)
        tickers_this = [t for t in df['_eq_tk'].dropna().unique() if t]
        mom_scores = compute_eq_mom_scores(tickers_this, dt_t)
        df['_eq_mom']  = df['_eq_tk'].map(mom_scores)
        df['_eq_fund'] = df['_eq_tk'].map(fund_df['Eq_Fund_Score']) if 'Eq_Fund_Score' in fund_df.columns else np.nan
    else:
        df['_eq_mom']  = np.nan
        df['_eq_fund'] = np.nan

    monthly_data.append({
        'dt_t':      dt_t,
        'dt_t1':     dt_t1,
        'df':        df,
        'isin_to_tr': isin_to_tr,
        'bench_tr':  bench_tr,
    })
    print(f"  {dt_t} prepared ({len(df)} bonds)", end='\r', flush=True)

print(f"\n  {len(monthly_data)} months ready")

# ═══════════════════════════════════════════════════════════════════════════════
# 5. 각 조합별 백테스트
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[Phase 2] Running all combos...")

def run_combo(w_b, w_m, w_f, w_a):
    records = []
    for md in monthly_data:
        df   = md['df'].copy()
        picks_all = []
        for cls, grp in df.groupby('_class'):
            if len(grp) < MIN_CLASS_SIZE: continue
            # per-class percentile rank for each component
            score_b = percentile_norm(grp['_bond_tr_est']) if w_b > 0 else pd.Series(0.0, index=grp.index)
            score_m = percentile_norm(grp['_eq_mom'])  if (w_m > 0 and grp['_eq_mom'].notna().sum() >= 3) else pd.Series(0.0, index=grp.index)
            score_f = percentile_norm(grp['_eq_fund']) if (w_f > 0 and grp['_eq_fund'].notna().sum() >= 3) else pd.Series(0.0, index=grp.index)
            score_a = grp['_ai_macro'] if w_a > 0 else pd.Series(0.0, index=grp.index)
            integrated = (
                score_b.fillna(0) * w_b +
                score_m.fillna(0) * w_m +
                score_f.fillna(0) * w_f +
                score_a.fillna(0) * w_a
            )
            top_idx = integrated.nlargest(TOP_N).index
            picks_all.extend(grp.loc[top_idx, 'ISIN'].tolist())

        tr_valid = [float(md['isin_to_tr'][i]) for i in picks_all
                    if i in md['isin_to_tr'] and
                    not pd.isna(md['isin_to_tr'].get(i, np.nan))]
        if len(tr_valid) < 5: continue
        records.append({
            'date':     md['dt_t1'],
            'port_tr':  np.mean(tr_valid) / 100,
            'bench_tr': md['bench_tr'] / 100,
        })
    if not records: return pd.Series(dtype=float), pd.Series(dtype=float)
    r = pd.DataFrame(records).set_index('date')
    r.index = pd.to_datetime(r.index)
    return r['port_tr'], r['bench_tr']

results = {}   # combo_name -> (port_ret, bench_ret)
for name, wb_, wm_, wf_, wa_ in COMBOS:
    print(f"  Running: {name:<35}", end='', flush=True)
    port, bench = run_combo(wb_, wm_, wf_, wa_)
    results[name] = (port, bench)
    if len(port) > 0 and len(bench) > 0:
        al = port.loc[bench.index]
        exc = al - bench
        ir = exc.mean()/exc.std()*np.sqrt(12) if exc.std()>0 else 0
        total = (1+al).prod()-1
        print(f" total={total:.1%}  IR={ir:.2f}")
    else:
        print(" no data")

# ═══════════════════════════════════════════════════════════════════════════════
# 6. 성과 요약 DataFrame
# ═══════════════════════════════════════════════════════════════════════════════
def perf(port, bench_s):
    if len(port) == 0: return {}
    al = port.loc[bench_s.index] if len(bench_s) > 0 else port
    ann  = (1+al.mean())**12 - 1
    vol  = al.std()*np.sqrt(12)
    total= (1+al).prod()-1
    exc  = al - bench_s if len(bench_s) > 0 else pd.Series(0.0, index=al.index)
    ir   = exc.mean()/exc.std()*np.sqrt(12) if exc.std()>0 else 0
    mdd  = (lambda cw: (cw/cw.cummax()-1).min())(100*(1+al).cumprod())
    calmar = ann/abs(mdd) if mdd<0 else 0
    win  = (al>0).mean()
    alpha_ann = exc.mean()*12
    return {
        'Total': total, 'Ann': ann, 'Vol': vol,
        'Sharpe': ann/vol if vol>0 else 0,
        'MaxDD': mdd, 'Calmar': calmar,
        'WinRate': win, 'IR': ir, 'Alpha_Ann': alpha_ann,
    }

bench_ref = results[list(results.keys())[0]][1]  # 공통 벤치마크
summary_rows = []
for name, (wb_, wm_, wf_, wa_) in zip([c[0] for c in COMBOS], [(c[1],c[2],c[3],c[4]) for c in COMBOS]):
    port, bench_s = results[name]
    p = perf(port, bench_s)
    if not p: continue
    summary_rows.append({
        'Name': name, 'w_B': wb_, 'w_M': wm_, 'w_F': wf_, 'w_A': wa_,
        **p
    })

summary = pd.DataFrame(summary_rows).set_index('Name')
summary_sorted = summary.sort_values('IR', ascending=False)

print("\n" + "="*90)
print("COMBO COMPARISON -- sorted by Info Ratio")
print("="*90)
disp = summary_sorted[['w_B','w_M','w_F','w_A','Total','Ann','Vol','Sharpe','MaxDD','IR','Alpha_Ann']].copy()
disp['Total']     = disp['Total'].map('{:.1%}'.format)
disp['Ann']       = disp['Ann'].map('{:.2%}'.format)
disp['Vol']       = disp['Vol'].map('{:.2%}'.format)
disp['Sharpe']    = disp['Sharpe'].map('{:.2f}'.format)
disp['MaxDD']     = disp['MaxDD'].map('{:.2%}'.format)
disp['IR']        = disp['IR'].map('{:.2f}'.format)
disp['Alpha_Ann'] = disp['Alpha_Ann'].map('{:.2%}'.format)
print(disp.to_string())
print("="*90)

# ═══════════════════════════════════════════════════════════════════════════════
# 7. 차트
# ═══════════════════════════════════════════════════════════════════════════════
# 색상 팔레트
N = len(COMBOS)
cmap = plt.cm.get_cmap('tab20', N)
combo_colors = {COMBOS[i][0]: cmap(i) for i in range(N)}

# 하이라이트: IR Top5
top5 = summary_sorted.index[:5].tolist()
bench_label = 'LUACTRUU'

fig = plt.figure(figsize=(26, 18))
gs = gridspec.GridSpec(3, 3, figure=fig, hspace=0.45, wspace=0.32, top=0.91, bottom=0.06)

# ─── Plot 1: 누적 수익 (Top5 + Bench) ──────────────────────────────────────
ax1 = fig.add_subplot(gs[0, :2])
bench_port, bench_s = results[COMBOS[0][0]]
cw_bench = 100 * (1 + bench_s).cumprod()
ax1.plot(cw_bench.index, cw_bench.values, color='#1F3864', lw=2.5,
         label=f"{bench_label}: {(1+bench_s).prod()-1:.1%}", zorder=10, ls='--')
for name in top5:
    port, bs = results[name]
    al = port.loc[bs.index]
    cw = 100 * (1 + al).cumprod()
    ax1.plot(cw.index, cw.values, lw=2.0, color=combo_colors[name],
             label=f"{name}: {(1+al).prod()-1:.1%}")
ax1.axhline(100, color='gray', lw=0.5, ls=':', alpha=0.6)
ax1.set_title('누적 수익률 ($100) -- IR 상위 5개 조합', fontsize=12, fontweight='bold', pad=6)
ax1.set_ylabel('포트폴리오 가치 ($)', fontsize=10)
ax1.yaxis.set_major_formatter(mtick.FormatStrFormatter('$%.0f'))
ax1.legend(fontsize=8.5, loc='upper left', framealpha=0.9)
ax1.grid(True, alpha=0.2)

# ─── Plot 2: IR vs Alpha scatter ─────────────────────────────────────────────
ax2 = fig.add_subplot(gs[0, 2])
for name in summary_sorted.index:
    row = summary_sorted.loc[name]
    is_top5 = name in top5
    ax2.scatter(float(row['IR'].replace('%','') if isinstance(row['IR'],str) else row['IR']),
                summary.loc[name,'Alpha_Ann']*100,
                color=combo_colors[name],
                s=120 if is_top5 else 60,
                zorder=5 if is_top5 else 3,
                alpha=1.0 if is_top5 else 0.6)
    if is_top5:
        ax2.annotate(name[:18], (summary_sorted.loc[name,'IR'] if isinstance(summary_sorted.loc[name,'IR'],float)
                                  else float(summary_sorted.loc[name,'IR']),
                                  summary.loc[name,'Alpha_Ann']*100),
                     fontsize=6.5, ha='left', va='bottom')

# replot with numeric IR
ax2.cla()
for name in summary.index:
    ir_v = summary.loc[name,'IR']
    alp_v= summary.loc[name,'Alpha_Ann']*100
    is_top5 = name in top5
    ax2.scatter(ir_v, alp_v, color=combo_colors[name],
                s=130 if is_top5 else 55,
                zorder=5 if is_top5 else 3,
                alpha=1.0 if is_top5 else 0.55,
                edgecolors='black' if is_top5 else 'none', linewidths=0.8)
    if is_top5:
        ax2.annotate(name[:18], (ir_v, alp_v), fontsize=6.5, ha='left', va='bottom',
                     xytext=(3,3), textcoords='offset points')
ax2.axhline(0, color='gray', lw=0.8, ls='--')
ax2.axvline(0, color='gray', lw=0.8, ls='--')
ax2.set_xlabel('Info Ratio', fontsize=10)
ax2.set_ylabel('연환산 Alpha (%)', fontsize=10)
ax2.set_title('IR vs Alpha (점=각 조합)', fontsize=11, fontweight='bold', pad=4)
ax2.yaxis.set_major_formatter(mtick.PercentFormatter())
ax2.grid(True, alpha=0.2)

# ─── Plot 3: IR 순위 바 차트 ────────────────────────────────────────────────
ax3 = fig.add_subplot(gs[1, :])
names_sorted = summary_sorted.index.tolist()
ir_vals  = [summary_sorted.loc[n,'IR'] for n in names_sorted]
colors_b = [combo_colors[n] for n in names_sorted]
edge_c   = ['black' if n in top5 else 'none' for n in names_sorted]
y = np.arange(len(names_sorted))
bars = ax3.barh(y, ir_vals, color=colors_b, edgecolor=edge_c, linewidth=1.2, alpha=0.85)
for i, (n, v) in enumerate(zip(names_sorted, ir_vals)):
    lbl = f"B={COMBOS[[c[0] for c in COMBOS].index(n)][1]:.2f} M={COMBOS[[c[0] for c in COMBOS].index(n)][2]:.2f} F={COMBOS[[c[0] for c in COMBOS].index(n)][3]:.2f} A={COMBOS[[c[0] for c in COMBOS].index(n)][4]:.2f}"
    ax3.text(v + 0.01, i, f'{v:.2f}  [{lbl}]', va='center', fontsize=8,
             fontweight='bold' if n in top5 else 'normal')
ax3.set_yticks(y)
ax3.set_yticklabels(names_sorted, fontsize=9)
ax3.axvline(0, color='black', lw=0.8)
ax3.set_title('조합별 Information Ratio 순위 (높을수록 우수)', fontsize=12, fontweight='bold', pad=6)
ax3.set_xlabel('Information Ratio', fontsize=10)
ax3.grid(True, alpha=0.2, axis='x')

# ─── Plot 4: 연도별 Best combo Alpha ────────────────────────────────────────
ax4 = fig.add_subplot(gs[2, :2])
best_name = top5[0]
port_best, bench_best = results[best_name]
al_best = port_best.loc[bench_best.index]
years_list = sorted(al_best.index.year.unique())
bench_yr = [(1+bench_best[bench_best.index.year==y]).prod()-1 if (bench_best.index.year==y).any() else np.nan for y in years_list]
model_yr = [(1+al_best[al_best.index.year==y]).prod()-1 for y in years_list]
alpha_yr = [m-b if b is not None and not np.isnan(b) else np.nan for m,b in zip(model_yr, bench_yr)]
x = np.arange(len(years_list))
w_ = 0.35
ax4.bar(x-w_/2, [v*100 if v and not np.isnan(v) else 0 for v in bench_yr], w_,
        label='LUACTRUU', color='#1F3864', alpha=0.85)
ax4.bar(x+w_/2, [v*100 for v in model_yr], w_,
        label=f'Best: {best_name}', color=combo_colors[best_name], alpha=0.85)
ax4.axhline(0, color='black', lw=0.8)
for i,(bv,mv,av) in enumerate(zip(bench_yr,model_yr,alpha_yr)):
    if av is not None and not np.isnan(av):
        ax4.text(x[i], max(mv*100, (bv or 0)*100)+0.4, f'a{av*100:+.1f}%',
                 ha='center', va='bottom', fontsize=8, color='#70AD47', fontweight='bold')
ax4.set_xticks(x)
ax4.set_xticklabels([str(y) for y in years_list], fontsize=10)
ax4.set_title(f'연도별 수익률 -- Best Combo: {best_name}', fontsize=11, fontweight='bold', pad=4)
ax4.yaxis.set_major_formatter(mtick.PercentFormatter())
ax4.legend(fontsize=9)
ax4.grid(True, alpha=0.2, axis='y')

# ─── Plot 5: 성과 요약 테이블 (Top 10) ───────────────────────────────────────
ax5 = fig.add_subplot(gs[2, 2])
ax5.axis('off')
top10 = summary_sorted.head(10)
tbl_data = []
for nm, row in top10.iterrows():
    tbl_data.append([
        nm[:22],
        f"{row['Total']:.1%}",
        f"{row['Ann']:.2%}",
        f"{row['Sharpe']:.2f}",
        f"{row['IR']:.2f}",
        f"{row['Alpha_Ann']:.2%}",
    ])
tbl = ax5.table(
    cellText=tbl_data,
    colLabels=['조합', '총수익', '연환산', 'Sharpe', 'IR', 'Alpha/yr'],
    cellLoc='center', loc='center', bbox=[0,0,1,1]
)
tbl.auto_set_font_size(False)
tbl.set_fontsize(7.5)
for (r,c), cell in tbl.get_celld().items():
    cell.set_edgecolor('#DDDDDD')
    cell.set_height(0.085)
    if r == 0:
        cell.set_facecolor('#1F3864')
        cell.set_text_props(color='white', fontweight='bold', fontsize=7)
    elif r <= 3:
        cell.set_facecolor('#FFF0F0')
        cell.set_text_props(color='#C00000', fontweight='bold')
    elif r % 2 == 0:
        cell.set_facecolor('#F5F5F5')
ax5.set_title('Top 10 조합 요약', fontsize=10, fontweight='bold', pad=4)

fig.suptitle(
    f'Bloomberg US IG Corp -- 스코어 가중치 조합 비교  '
    f'({monthly_data[0]["dt_t"]} ~ {monthly_data[-1]["dt_t1"]})\n'
    f'B=Bond_TR  M=Eq_Mom  F=Eq_Fund  A=AI_Macro  |  {len(COMBOS)} 조합  |  Top{TOP_N}/Class',
    fontsize=13, fontweight='bold', y=0.97
)

plt.savefig(OUT_PNG, dpi=150, bbox_inches='tight', facecolor='white')
print(f"\n[DONE] Chart saved: {OUT_PNG}")

# ═══════════════════════════════════════════════════════════════════════════════
# 8. Excel 저장
# ═══════════════════════════════════════════════════════════════════════════════
try:
    from openpyxl import Workbook as OWB
    from openpyxl.styles import Font as OF, PatternFill as OFill, Alignment as OA, PatternFill
    from openpyxl.utils import get_column_letter

    wb = OWB()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Summary'
    hdrs = ['Rank','Name','w_Bond_TR','w_Eq_Mom','w_Eq_Fund','w_AI_Macro',
            'Total(%)','Ann(%)','Vol(%)','Sharpe','MaxDD(%)','Calmar','IR','Alpha_Ann(%)','WinRate(%)']
    for ci, h in enumerate(hdrs, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font = OF(bold=True, color='FFFFFF')
        c.fill = OFill('solid', fgColor='1F3864')
        c.alignment = OA(horizontal='center')

    for rank, (nm, row) in enumerate(summary_sorted.iterrows(), 1):
        ri = rank + 1
        vals = [rank, nm,
                row['w_B'], row['w_M'], row['w_F'], row['w_A'],
                round(row['Total']*100,2), round(row['Ann']*100,2),
                round(row['Vol']*100,2), round(row['Sharpe'],3),
                round(row['MaxDD']*100,2), round(row['Calmar'],3),
                round(row['IR'],3), round(row['Alpha_Ann']*100,2),
                round(row['WinRate']*100,1)]
        for ci, v in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=ci, value=v)
            if rank <= 3:
                c.fill = OFill('solid', fgColor='FFF0F0')
                if ci >= 7:
                    c.font = OF(bold=True, color='C00000')

    ws1.column_dimensions['B'].width = 28
    for i in range(3, 16):
        ws1.column_dimensions[get_column_letter(i)].width = 12

    # ── Sheet 2: Monthly Returns per Combo ──────────────────────────────────
    ws2 = wb.create_sheet('Monthly_Returns')
    all_dates = sorted(set(pd.Timestamp(md['dt_t1']) for md in monthly_data))
    combo_names = [c[0] for c in COMBOS]

    ws2.cell(row=1, column=1, value='Date')
    ws2.cell(row=1, column=2, value='Benchmark(%)')
    for ci, n in enumerate(combo_names, 3):
        c = ws2.cell(row=1, column=ci, value=n[:25])
        c.font = OF(bold=True)

    bench_series = results[combo_names[0]][1]
    for ri, dt in enumerate(all_dates, 2):
        ws2.cell(row=ri, column=1, value=dt.strftime('%Y-%m'))
        bv = bench_series.get(dt, np.nan)
        ws2.cell(row=ri, column=2, value=round(bv*100,3) if not pd.isna(bv) else None)
        for ci, n in enumerate(combo_names, 3):
            port_s, _ = results[n]
            pv = port_s.get(dt, np.nan)
            ws2.cell(row=ri, column=ci, value=round(pv*100,3) if not pd.isna(pv) else None)

    wb.save(OUT_XLSX)
    print(f"[DONE] Excel saved: {OUT_XLSX}")
except Exception as e:
    print(f"[WARN] Excel save failed: {e}")
