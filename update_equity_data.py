"""
Equity Data Update Script
Updates momentum, fundamentals, and sentiment data in LUACSTAT_2026_03_31_SCORED.xlsx
"""

import pandas as pd
import numpy as np
import requests
import time
import math
import re
import warnings
from datetime import datetime, date
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.filterwarnings('ignore')

TODAY = date(2026, 4, 3)
FILE_PATH = 'C:/Users/sh.park/Documents/USIG_LLM/LUACSTAT_2026_03_31_SCORED.xlsx'

# ─────────────────────────────────────────────
# STEP 1: Read data and build ticker list
# ─────────────────────────────────────────────
print("=" * 60)
print("STEP 1: Reading data and building ticker list")
print("=" * 60)

df = pd.read_excel(FILE_PATH, sheet_name='Detail_Scored', header=1)
print(f"Loaded {len(df)} rows, {len(df.columns)} columns")

# Build sets from each column
ticker_from_ticker_col = set()
for v in df['Ticker'].dropna():
    s = str(v).strip()
    if s and s.lower() not in ('nan', 'n/a', ''):
        ticker_from_ticker_col.add(s)

ticker_from_eqty_col = set()
for v in df['Eqty Ticker'].dropna():
    s = str(v).strip()
    # Strip exchange suffix (everything after first space)
    s = s.split(' ')[0].strip()
    if s and s.lower() not in ('nan', 'n/a', ''):
        ticker_from_eqty_col.add(s)

# Combine and filter
all_raw = ticker_from_ticker_col | ticker_from_eqty_col

def is_valid_ticker(t):
    if not t or len(t) < 1:
        return False
    if t.lower() in ('nan', 'n/a'):
        return False
    if re.search(r'\d', t):  # has digits → Bloomberg internal ID
        return False
    return True

all_tickers = sorted([t for t in all_raw if is_valid_ticker(t)])
print(f"Total unique tickers to try: {len(all_tickers)}")

# Build per-row ticker mapping: prefer Ticker col, fallback to stripped Eqty Ticker
def get_row_ticker(row):
    t = str(row.get('Ticker', '')).strip()
    if t and t.lower() not in ('nan', 'n/a', '') and is_valid_ticker(t):
        return t, 'Ticker'
    eq = str(row.get('Eqty Ticker', '')).strip().split(' ')[0]
    if eq and eq.lower() not in ('nan', 'n/a', '') and is_valid_ticker(eq):
        return eq, 'Eqty Ticker'
    return None, None

row_tickers = []
row_sources = []
for _, row in df.iterrows():
    t, src = get_row_ticker(row)
    row_tickers.append(t)
    row_sources.append(src)

df['_ticker'] = row_tickers
df['_source'] = row_sources

# ─────────────────────────────────────────────
# Yahoo Finance session with crumb
# ─────────────────────────────────────────────
print("\nSetting up Yahoo Finance session...")

session = requests.Session()
session.verify = False
session.headers.update({
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'application/json',
})

crumb = None
for attempt in range(5):
    try:
        r = session.get('https://query1.finance.yahoo.com/v1/test/getcrumb', timeout=10)
        if r.status_code == 200 and r.text.strip():
            crumb = r.text.strip()
            print(f"Got crumb: {crumb}")
            break
    except Exception as e:
        print(f"  Crumb attempt {attempt+1} failed: {e}")
    # Try alternative crumb endpoint
    try:
        r2 = session.get('https://finance.yahoo.com/', timeout=10)
        r3 = session.get('https://query1.finance.yahoo.com/v1/test/getcrumb', timeout=10)
        if r3.status_code == 200 and r3.text.strip():
            crumb = r3.text.strip()
            print(f"Got crumb via cookie refresh: {crumb}")
            break
    except Exception as e2:
        print(f"  Cookie refresh attempt {attempt+1} failed: {e2}")
    time.sleep(1)

if not crumb:
    print("WARNING: Could not get crumb, will try without")

# ─────────────────────────────────────────────
# STEP 2: Fetch price/momentum data
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 2: Fetching equity price data (momentum)")
print("=" * 60)

price_data = {}  # ticker -> dict of metrics

BATCH = 50
SLEEP_BETWEEN = 0.5

def fetch_price(ticker):
    url = f'https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1d&range=6mo'
    try:
        r = session.get(url, timeout=10)
        if r.status_code != 200:
            return None
        j = r.json()
        result = j.get('chart', {}).get('result', [])
        if not result:
            return None
        res = result[0]
        closes = res.get('indicators', {}).get('quote', [{}])[0].get('close', [])
        closes = [c for c in closes if c is not None]
        if len(closes) < 22:
            return None

        prices = np.array(closes, dtype=float)
        n = len(prices)

        ret_1m = (prices[-1] / prices[max(0, n-22)] - 1) if n >= 22 else np.nan
        ret_3m = (prices[-1] / prices[max(0, n-63)] - 1) if n >= 63 else np.nan

        # Vol 30D: annualized from last 30 log returns
        if n >= 31:
            log_rets = np.diff(np.log(prices[-31:]))
            vol_30d = np.std(log_rets, ddof=1) * math.sqrt(252)
        else:
            log_rets = np.diff(np.log(prices))
            vol_30d = np.std(log_rets, ddof=1) * math.sqrt(252) if len(log_rets) > 1 else np.nan

        # vs 52w high
        high_252 = np.max(prices[-252:]) if n >= 252 else np.max(prices)
        vs_high = prices[-1] / high_252 - 1

        return {
            'Eq_Ret_1M': ret_1m,
            'Eq_Ret_3M': ret_3m,
            'Eq_Vol_30D': vol_30d,
            'Eq_vs_52w_High': vs_high,
        }
    except Exception:
        return None

price_success = 0
for i, ticker in enumerate(all_tickers):
    if i % 100 == 0:
        print(f"  Price: {i}/{len(all_tickers)} done, {price_success} succeeded")
    if i > 0 and i % BATCH == 0:
        time.sleep(SLEEP_BETWEEN)
    result = fetch_price(ticker)
    if result:
        price_data[ticker] = result
        price_success += 1

print(f"  Price data: {price_success}/{len(all_tickers)} tickers succeeded")

# ─────────────────────────────────────────────
# STEP 3: Fetch fundamentals
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 3: Fetching fundamentals")
print("=" * 60)

fund_data = {}

def fetch_fundamentals(ticker):
    modules = 'financialData,defaultKeyStatistics,summaryDetail'
    url = f'https://query1.finance.yahoo.com/v10/finance/quoteSummary/{ticker}?modules={modules}'
    if crumb:
        url += f'&crumb={crumb}'
    try:
        r = session.get(url, timeout=10)
        if r.status_code != 200:
            return None
        j = r.json()
        res = j.get('quoteSummary', {}).get('result', [])
        if not res:
            return None
        data = res[0]
        fd = data.get('financialData', {})
        ks = data.get('defaultKeyStatistics', {})
        sd = data.get('summaryDetail', {})

        def safe_raw(d, key):
            v = d.get(key, {})
            if isinstance(v, dict):
                return v.get('raw', np.nan)
            return np.nan

        return {
            'Debt_to_Equity': safe_raw(fd, 'debtToEquity'),
            'Profit_Margin': safe_raw(fd, 'profitMargins'),
            'Revenue_Growth': safe_raw(fd, 'revenueGrowth'),
            'Current_Ratio': safe_raw(fd, 'currentRatio'),
            'EV_EBITDA': safe_raw(ks, 'enterpriseToEbitda'),
            'PE_Ratio': safe_raw(sd, 'trailingPE'),
        }
    except Exception:
        return None

fund_success = 0
for i, ticker in enumerate(all_tickers):
    if i % 100 == 0:
        print(f"  Fundamentals: {i}/{len(all_tickers)} done, {fund_success} succeeded")
    if i > 0 and i % BATCH == 0:
        time.sleep(SLEEP_BETWEEN)
    result = fetch_fundamentals(ticker)
    if result:
        # Check at least one field is not nan
        vals = [v for v in result.values() if not (isinstance(v, float) and math.isnan(v))]
        if vals:
            fund_data[ticker] = result
            fund_success += 1

print(f"  Fundamentals: {fund_success}/{len(all_tickers)} tickers succeeded")

# ─────────────────────────────────────────────
# STEP 4: Fetch news sentiment + Google Trends
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 4: Fetching news sentiment and Google Trends")
print("=" * 60)

analyzer = SentimentIntensityAnalyzer()
news_data = {}

def fetch_news_sentiment(ticker):
    url = f'https://query1.finance.yahoo.com/v1/finance/search?q={ticker}&newsCount=20&quotesCount=0'
    if crumb:
        url += f'&crumb={crumb}'
    try:
        r = session.get(url, timeout=10)
        if r.status_code != 200:
            return None
        j = r.json()
        news = j.get('news', [])
        if not news:
            return None

        today_ts = datetime.combine(TODAY, datetime.min.time()).timestamp()
        weighted_scores = []
        weights = []
        for item in news:
            headline = item.get('title', '')
            pub_time = item.get('providerPublishTime', today_ts)
            days_ago = max(0, (today_ts - pub_time) / 86400)
            weight = math.exp(-days_ago / 14)
            score = analyzer.polarity_scores(headline)['compound']
            weighted_scores.append(score * weight)
            weights.append(weight)

        if not weights or sum(weights) == 0:
            return None
        sentiment = sum(weighted_scores) / sum(weights)
        return {'News_Sentiment_Raw': sentiment, 'News_Article_Count': len(news)}
    except Exception:
        return None

news_success = 0
for i, ticker in enumerate(all_tickers):
    if i % 100 == 0:
        print(f"  News: {i}/{len(all_tickers)} done, {news_success} succeeded")
    if i > 0 and i % BATCH == 0:
        time.sleep(SLEEP_BETWEEN)
    result = fetch_news_sentiment(ticker)
    if result:
        news_data[ticker] = result
        news_success += 1

print(f"  News sentiment: {news_success}/{len(all_tickers)} tickers succeeded")

# Google Trends
print("  Fetching Google Trends...")
trends_data = {}
trends_success = 0

try:
    from pytrends.request import TrendReq
    pytrends = TrendReq(hl='en-US', tz=360, timeout=(10, 25), retries=2, backoff_factor=0.5)

    # Process in groups of 5
    ticker_groups = [all_tickers[i:i+5] for i in range(0, len(all_tickers), 5)]

    for gi, group in enumerate(ticker_groups):
        if gi % 20 == 0:
            print(f"  Trends group {gi}/{len(ticker_groups)}")
        try:
            pytrends.build_payload(group, timeframe='today 3-m', geo='US')
            df_trends = pytrends.interest_over_time()
            if df_trends.empty:
                time.sleep(1)
                continue
            # For each ticker in group
            for ticker in group:
                if ticker in df_trends.columns:
                    series = df_trends[ticker].dropna()
                    if len(series) >= 8:
                        last4w = series.iloc[-4:].mean()
                        prev4w = series.iloc[-8:-4].mean()
                        momentum = (last4w - prev4w) / (prev4w + 1)
                        trends_data[ticker] = momentum
                        trends_success += 1
            time.sleep(1)
        except Exception as e:
            time.sleep(2)
            continue
    print(f"  Trends: {trends_success}/{len(all_tickers)} tickers succeeded")
except Exception as e:
    print(f"  Google Trends unavailable: {e}")
    print("  Continuing without Trends data")

# ─────────────────────────────────────────────
# STEP 5: Compute all scores
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 5: Computing scores")
print("=" * 60)

def rank_pct_to_score(series, invert=False):
    """Rank percentile → [-1, +1]. NaN stays NaN."""
    valid = series.dropna()
    if len(valid) == 0:
        return series * np.nan
    ranks = series.rank(method='average', na_option='keep')
    n = series.notna().sum()
    pct = (ranks - 1) / (n - 1) if n > 1 else ranks * 0
    if invert:
        pct = 1 - pct
    return pct * 2 - 1  # → [-1, +1]

# Map data to df rows
def map_metric(col_name, data_dict, metric_key):
    return df['_ticker'].map(lambda t: data_dict.get(t, {}).get(metric_key, np.nan) if t else np.nan)

# Price metrics
df['Eq_Ret_1M'] = map_metric('Eq_Ret_1M', price_data, 'Eq_Ret_1M')
df['Eq_Ret_3M'] = map_metric('Eq_Ret_3M', price_data, 'Eq_Ret_3M')
df['Eq_Vol_30D'] = map_metric('Eq_Vol_30D', price_data, 'Eq_Vol_30D')
df['Eq_vs_52w_High'] = map_metric('Eq_vs_52w_High', price_data, 'Eq_vs_52w_High')

# Fundamental metrics
df['Debt_to_Equity'] = map_metric('Debt_to_Equity', fund_data, 'Debt_to_Equity')
df['Profit_Margin'] = map_metric('Profit_Margin', fund_data, 'Profit_Margin')
df['Revenue_Growth'] = map_metric('Revenue_Growth', fund_data, 'Revenue_Growth')
df['Current_Ratio'] = map_metric('Current_Ratio', fund_data, 'Current_Ratio')
df['EV_EBITDA'] = map_metric('EV_EBITDA', fund_data, 'EV_EBITDA')
df['PE_Ratio'] = map_metric('PE_Ratio', fund_data, 'PE_Ratio')

# News sentiment
df['News_Sentiment_Raw'] = df['_ticker'].map(lambda t: news_data.get(t, {}).get('News_Sentiment_Raw', np.nan) if t else np.nan)
df['News_Article_Count'] = df['_ticker'].map(lambda t: news_data.get(t, {}).get('News_Article_Count', np.nan) if t else np.nan)

# Trends momentum
df['Trends_Momentum'] = df['_ticker'].map(lambda t: trends_data.get(t, np.nan) if t else np.nan)

# ---- Momentum Score ----
mom_ret1m = rank_pct_to_score(df['Eq_Ret_1M'], invert=False)
mom_ret3m = rank_pct_to_score(df['Eq_Ret_3M'], invert=False)
mom_vol = rank_pct_to_score(df['Eq_Vol_30D'], invert=True)   # lower = better
mom_52w = rank_pct_to_score(df['Eq_vs_52w_High'], invert=False)

mom_components = pd.DataFrame({'r1m': mom_ret1m, 'r3m': mom_ret3m, 'vol': mom_vol, 'h52': mom_52w})
df['Eq_Mom_Score'] = mom_components.mean(axis=1, skipna=False)  # NaN if any component missing
# Actually: mean of available (need ≥1), use skipna for partial
df['Eq_Mom_Score'] = mom_components.mean(axis=1)  # averages available ones
# But only assign if price data exists
has_price = df['_ticker'].map(lambda t: t in price_data if t else False)
df.loc[~has_price, 'Eq_Mom_Score'] = np.nan

# ---- Fundamental Score ----
# Cap Current_Ratio at 3
cr_capped = df['Current_Ratio'].clip(upper=3)

# Exclude negatives for EV_EBITDA and PE_Ratio
ev_filtered = df['EV_EBITDA'].where(df['EV_EBITDA'] > 0)
pe_filtered = df['PE_Ratio'].where((df['PE_Ratio'] > 0) & (df['PE_Ratio'] <= 100))

fund_d2e = rank_pct_to_score(df['Debt_to_Equity'], invert=True)
fund_pm = rank_pct_to_score(df['Profit_Margin'], invert=False)
fund_rg = rank_pct_to_score(df['Revenue_Growth'], invert=False)
fund_cr = rank_pct_to_score(cr_capped, invert=False)
fund_ev = rank_pct_to_score(ev_filtered, invert=True)
fund_pe = rank_pct_to_score(pe_filtered, invert=True)

fund_components = pd.DataFrame({
    'd2e': fund_d2e, 'pm': fund_pm, 'rg': fund_rg,
    'cr': fund_cr, 'ev': fund_ev, 'pe': fund_pe
})
fund_count = fund_components.notna().sum(axis=1)
fund_raw = fund_components.mean(axis=1)  # avg of available
fund_raw[fund_count < 2] = np.nan  # need ≥2 components

df['Eq_Fund_Score'] = fund_raw * 2 - 1  # → [-1, +1]

# ---- Sentiment Score ----
news_norm = rank_pct_to_score(df['News_Sentiment_Raw'], invert=False)
trends_norm = rank_pct_to_score(df['Trends_Momentum'], invert=False)

def calc_sentiment(row_news, row_trends):
    n_valid = not (isinstance(row_news, float) and math.isnan(row_news))
    t_valid = not (isinstance(row_trends, float) and math.isnan(row_trends))
    if n_valid and t_valid:
        return row_news * 0.7 + row_trends * 0.3
    elif n_valid:
        return row_news
    elif t_valid:
        return row_trends * 0.5
    return np.nan

df['Sentiment_Score'] = [calc_sentiment(n, t) for n, t in zip(news_norm, trends_norm)]

# ---- Integrated Score ----
bond_tr = df['Bond_TR_Est_pct'].fillna(0)
mom_s = df['Eq_Mom_Score'].fillna(0)
fund_s = df['Eq_Fund_Score'].fillna(0)
sent_s = df['Sentiment_Score'].fillna(0)

df['Integrated_Score'] = (
    bond_tr * 0.6
    + mom_s * 0.025
    + fund_s * 0.015
    + sent_s * 0.010
)

# ---- Recompute rankings per class ----
df['Integrated_Rank_in_Class'] = np.nan
df['Top_Pick_Flag'] = ''

classes = df['class'].dropna().unique()
classes = [c for c in classes if str(c).lower() != 'off']

for cls in classes:
    mask = (df['class'] == cls)
    sub = df[mask].copy()
    if len(sub) == 0:
        continue
    ranks = sub['Integrated_Score'].rank(method='first', ascending=False, na_option='bottom')
    df.loc[mask, 'Integrated_Rank_in_Class'] = ranks.values

def assign_flag(rank):
    if pd.isna(rank):
        return ''
    r = int(rank)
    if r <= 3:
        return '★★★ TOP3'
    elif r <= 10:
        return '★★ TOP10'
    elif r <= 25:
        return '★ TOP25'
    return ''

df['Top_Pick_Flag'] = df['Integrated_Rank_in_Class'].map(assign_flag)
df.loc[df['class'].str.lower().fillna('') == 'off', 'Top_Pick_Flag'] = ''
df.loc[df['class'].str.lower().fillna('') == 'off', 'Integrated_Rank_in_Class'] = np.nan

print(f"  Eq_Mom_Score non-null: {df['Eq_Mom_Score'].notna().sum()}")
print(f"  Eq_Fund_Score non-null: {df['Eq_Fund_Score'].notna().sum()}")
print(f"  Sentiment_Score non-null: {df['Sentiment_Score'].notna().sum()}")
print(f"  Integrated_Score non-null: {df['Integrated_Score'].notna().sum()}")

# ─────────────────────────────────────────────
# STEP 6: Write to Excel
# ─────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 6: Writing to Excel")
print("=" * 60)

wb = load_workbook(FILE_PATH)
ws = wb['Detail_Scored']

# Find header row (row 2, 1-indexed)
# Our df used header=1 (0-indexed) so row 2 in Excel = header
# Data starts at row 3

# Find column positions by scanning header row (row 2)
header_row_idx = 2  # 1-indexed Excel row
col_map = {}
for col in ws.iter_cols(min_row=header_row_idx, max_row=header_row_idx, values_only=False):
    for cell in col:
        if cell.value is not None:
            col_map[str(cell.value).strip()] = cell.column

TARGET_COLS = [
    'Eq_Ret_1M', 'Eq_Ret_3M', 'Eq_Vol_30D', 'Eq_vs_52w_High', 'Eq_Mom_Score',
    'Debt_to_Equity', 'Profit_Margin', 'Revenue_Growth', 'Current_Ratio',
    'EV_EBITDA', 'PE_Ratio', 'Eq_Fund_Score',
    'News_Sentiment_Raw', 'News_Article_Count', 'Trends_Momentum', 'Sentiment_Score',
    'Integrated_Score', 'Integrated_Rank_in_Class', 'Top_Pick_Flag'
]

print(f"Found columns in Excel: {[c for c in TARGET_COLS if c in col_map]}")
missing_cols = [c for c in TARGET_COLS if c not in col_map]
if missing_cols:
    print(f"WARNING: Missing columns: {missing_cols}")

# Formatting for Top_Pick_Flag
flag_styles = {
    '★★★ TOP3': (Font(bold=True, color='FF0000'), PatternFill('solid', fgColor='FFFF00')),
    '★★ TOP10': (Font(bold=True, color='C55A11'), PatternFill('solid', fgColor='FCE4D6')),
    '★ TOP25':  (Font(bold=True, color='1F3864'), PatternFill('solid', fgColor='DDEBF7')),
    '':         (Font(bold=False, color='000000'), PatternFill('none')),
}

flag_col_idx = col_map.get('Top_Pick_Flag')

# Write data row by row (df row 0 = Excel row 3)
data_start_row = 3
for df_idx, row in df.iterrows():
    excel_row = data_start_row + df_idx

    for col_name in TARGET_COLS:
        if col_name not in col_map:
            continue
        col_idx = col_map[col_name]
        val = row.get(col_name, np.nan)

        if col_name == 'Top_Pick_Flag':
            str_val = str(val) if val and str(val) != 'nan' else ''
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.value = str_val
            fnt, fill = flag_styles.get(str_val, flag_styles[''])
            cell.font = fnt
            cell.fill = fill
        elif col_name in ('News_Article_Count', 'Integrated_Rank_in_Class'):
            if pd.isna(val) or (isinstance(val, float) and math.isnan(val)):
                ws.cell(row=excel_row, column=col_idx).value = None
            else:
                ws.cell(row=excel_row, column=col_idx).value = int(val)
        else:
            if isinstance(val, float) and math.isnan(val):
                ws.cell(row=excel_row, column=col_idx).value = None
            else:
                ws.cell(row=excel_row, column=col_idx).value = val

print(f"  Updated Detail_Scored sheet ({len(df)} rows)")

# ─────────────────────────────────────────────
# Regenerate Equity_Data sheet
# ─────────────────────────────────────────────
print("  Regenerating Equity_Data sheet...")

DARK_BLUE = '1F3864'
BLUE_HDR  = '4472C4'
WHITE     = 'FFFFFF'

def make_title_row(ws, title, ncols):
    ws.append([title] + [''] * (ncols - 1))
    row = ws.max_row
    cell = ws.cell(row=row, column=1)
    cell.font = Font(name='Arial', bold=True, size=13, color=WHITE)
    cell.fill = PatternFill('solid', fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 22
    for c in range(2, ncols + 1):
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=DARK_BLUE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)

def make_header_row(ws, headers):
    ws.append(headers)
    row = ws.max_row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name='Arial', bold=True, color=WHITE)
        cell.fill = PatternFill('solid', fgColor=BLUE_HDR)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 18

# Remove and recreate Equity_Data sheet
if 'Equity_Data' in wb.sheetnames:
    del wb['Equity_Data']
eq_ws = wb.create_sheet('Equity_Data')

eq_headers = [
    'Ticker', 'Company_Name',
    'Eq_Ret_1M', 'Eq_Ret_3M', 'Eq_Vol_30D', 'Eq_vs_52w_High', 'Eq_Mom_Score',
    'Debt_to_Equity', 'Profit_Margin', 'Revenue_Growth', 'Current_Ratio',
    'EV_EBITDA', 'PE_Ratio', 'Eq_Fund_Score',
    'News_Sentiment_Raw', 'News_Article_Count', 'Trends_Momentum', 'Sentiment_Score',
    'Data_Source'
]
ncols = len(eq_headers)

make_title_row(eq_ws, f'Equity Data — Price Momentum & Fundamentals | As of {TODAY}', ncols)
make_header_row(eq_ws, eq_headers)

# Build unique ticker rows (one row per ticker, best data)
ticker_rows = {}
for _, row in df.iterrows():
    t = row['_ticker']
    if not t:
        continue
    src = row['_source'] or ''
    if t not in ticker_rows:
        ticker_rows[t] = {
            'Ticker': t,
            'Company_Name': row.get('Company Name', ''),
            'Eq_Ret_1M': row.get('Eq_Ret_1M', np.nan),
            'Eq_Ret_3M': row.get('Eq_Ret_3M', np.nan),
            'Eq_Vol_30D': row.get('Eq_Vol_30D', np.nan),
            'Eq_vs_52w_High': row.get('Eq_vs_52w_High', np.nan),
            'Eq_Mom_Score': row.get('Eq_Mom_Score', np.nan),
            'Debt_to_Equity': row.get('Debt_to_Equity', np.nan),
            'Profit_Margin': row.get('Profit_Margin', np.nan),
            'Revenue_Growth': row.get('Revenue_Growth', np.nan),
            'Current_Ratio': row.get('Current_Ratio', np.nan),
            'EV_EBITDA': row.get('EV_EBITDA', np.nan),
            'PE_Ratio': row.get('PE_Ratio', np.nan),
            'Eq_Fund_Score': row.get('Eq_Fund_Score', np.nan),
            'News_Sentiment_Raw': row.get('News_Sentiment_Raw', np.nan),
            'News_Article_Count': row.get('News_Article_Count', np.nan),
            'Trends_Momentum': row.get('Trends_Momentum', np.nan),
            'Sentiment_Score': row.get('Sentiment_Score', np.nan),
            'Data_Source': src,
        }

# Sort by Eq_Mom_Score desc (nulls last)
eq_rows = sorted(ticker_rows.values(),
                 key=lambda x: (math.isnan(x['Eq_Mom_Score']) if isinstance(x['Eq_Mom_Score'], float) else False,
                                -(x['Eq_Mom_Score'] if not (isinstance(x['Eq_Mom_Score'], float) and math.isnan(x['Eq_Mom_Score'])) else 0)))

for r in eq_rows:
    row_vals = []
    for h in eq_headers:
        v = r.get(h, '')
        if isinstance(v, float) and math.isnan(v):
            v = None
        row_vals.append(v)
    eq_ws.append(row_vals)

# Column widths
col_widths = [12, 30, 10, 10, 10, 12, 12, 14, 13, 14, 13, 10, 10, 13, 18, 16, 15, 14, 12]
for i, w in enumerate(col_widths, 1):
    eq_ws.column_dimensions[get_column_letter(i)].width = w

# Freeze row 2
eq_ws.freeze_panes = 'A3'
eq_ws.auto_filter.ref = f'A2:{get_column_letter(ncols)}2'

print(f"  Equity_Data: {len(eq_rows)} unique tickers")

# ─────────────────────────────────────────────
# Regenerate Sentiment_Analysis sheet
# ─────────────────────────────────────────────
print("  Regenerating Sentiment_Analysis sheet...")

if 'Sentiment_Analysis' in wb.sheetnames:
    del wb['Sentiment_Analysis']
sent_ws = wb.create_sheet('Sentiment_Analysis')

sent_headers = ['Ticker', 'Company_Name', 'News_Sentiment_Raw', 'News_Article_Count',
                'Trends_Momentum', 'Sentiment_Score', 'Data_Source']
nsent = len(sent_headers)

make_title_row(sent_ws, f'Sentiment Analysis | As of {TODAY}', nsent)
make_header_row(sent_ws, sent_headers)

sent_rows = sorted(ticker_rows.values(),
                   key=lambda x: (x['Sentiment_Score'] is None or (isinstance(x['Sentiment_Score'], float) and math.isnan(x['Sentiment_Score'])),
                                  -(x['Sentiment_Score'] or 0)))

for r in sent_rows:
    row_vals = []
    for h in sent_headers:
        v = r.get(h, '')
        if isinstance(v, float) and math.isnan(v):
            v = None
        row_vals.append(v)
    sent_ws.append(row_vals)

sent_widths = [12, 30, 18, 16, 15, 14, 12]
for i, w in enumerate(sent_widths, 1):
    sent_ws.column_dimensions[get_column_letter(i)].width = w
sent_ws.freeze_panes = 'A3'
sent_ws.auto_filter.ref = f'A2:{get_column_letter(nsent)}2'

# ─────────────────────────────────────────────
# Regenerate Top_Picks_by_Class sheet
# ─────────────────────────────────────────────
print("  Regenerating Top_Picks_by_Class sheet...")

if 'Top_Picks_by_Class' in wb.sheetnames:
    del wb['Top_Picks_by_Class']
top_ws = wb.create_sheet('Top_Picks_by_Class')

top_headers = ['Class', 'Rank', 'Ticker', 'Company_Name', 'Bond_TR_Est_pct',
               'Eq_Mom_Score', 'Eq_Fund_Score', 'Sentiment_Score', 'Integrated_Score',
               'Integrated_Rank_in_Class', 'Top_Pick_Flag']
ntop = len(top_headers)

make_title_row(top_ws, f'Top Picks by Class (Top 5 per Class) | As of {TODAY}', ntop)
make_header_row(top_ws, top_headers)

for cls in sorted(classes):
    sub = df[df['class'] == cls].copy()
    sub = sub[sub['Integrated_Rank_in_Class'].notna()]
    sub = sub.sort_values('Integrated_Rank_in_Class').head(5)
    for rank_i, (_, row) in enumerate(sub.iterrows(), 1):
        flag_val = row.get('Top_Pick_Flag', '')
        row_vals = [
            cls, rank_i,
            row.get('_ticker', ''), row.get('Company Name', ''),
            row.get('Bond_TR_Est_pct', None),
            None if pd.isna(row.get('Eq_Mom_Score', np.nan)) else row.get('Eq_Mom_Score'),
            None if pd.isna(row.get('Eq_Fund_Score', np.nan)) else row.get('Eq_Fund_Score'),
            None if pd.isna(row.get('Sentiment_Score', np.nan)) else row.get('Sentiment_Score'),
            None if pd.isna(row.get('Integrated_Score', np.nan)) else row.get('Integrated_Score'),
            int(row.get('Integrated_Rank_in_Class', 0)) if not pd.isna(row.get('Integrated_Rank_in_Class', np.nan)) else None,
            str(flag_val) if flag_val and str(flag_val) != 'nan' else '',
        ]
        top_ws.append(row_vals)
        # Style top pick flag cell
        flag_cell = top_ws.cell(row=top_ws.max_row, column=ntop)
        fnt, fill = flag_styles.get(str(flag_val) if flag_val and str(flag_val) != 'nan' else '', flag_styles[''])
        flag_cell.font = fnt
        flag_cell.fill = fill

top_widths = [15, 6, 12, 30, 14, 12, 12, 14, 14, 18, 14]
for i, w in enumerate(top_widths, 1):
    top_ws.column_dimensions[get_column_letter(i)].width = w
top_ws.freeze_panes = 'A3'

# Save workbook
print("  Saving workbook...")
wb.save(FILE_PATH)
print("  Saved!")

# ─────────────────────────────────────────────
# STEP 7: Summary
# ─────────────────────────────────────────────
import os

print("\n" + "=" * 60)
print("STEP 7: SUMMARY")
print("=" * 60)
file_size_mb = os.path.getsize(FILE_PATH) / (1024 * 1024)

print(f"Total tickers attempted: {len(all_tickers)}")
print(f"Price data success: {price_success} ({100*price_success/len(all_tickers):.1f}%)")
print(f"Fundamentals success: {fund_success} ({100*fund_success/len(all_tickers):.1f}%)")
print(f"News sentiment success: {news_success} ({100*news_success/len(all_tickers):.1f}%)")
print(f"Trends success: {trends_success} ({100*trends_success/len(all_tickers):.1f}%)")
print(f"Bonds with Eq_Mom_Score: {df['Eq_Mom_Score'].notna().sum()}")
print(f"Bonds with Eq_Fund_Score: {df['Eq_Fund_Score'].notna().sum()}")
print(f"Bonds with Sentiment_Score: {df['Sentiment_Score'].notna().sum()}")
print(f"File size: {file_size_mb:.2f} MB")
print("FULL RERUN COMPLETE")
