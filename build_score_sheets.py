"""
build_score_sheets.py
─────────────────────
현재 LUACSTAT_2026_03_31_SCORED.xlsx 에서 데이터를 읽어
  1. DP_Rating_Score / Bond_TR_Score 재산출
  2. Integrated_Score 동일비중(0.25×4) 재산출
  3. 시트 5개 생성:
       Score_BondTR / Score_EqMom / Score_EqFund / Score_Sentiment / Score_Integrated
  4. Detail_Scored 의 Score·Rank·Flag 열 업데이트
"""

import math
import sys
import os
import warnings
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# AI Macro Score 모듈 임포트
sys.path.insert(0, os.path.dirname(os.path.abspath('C:/Users/sh.park/Documents/USIG_LLM/build_score_sheets.py')))
from ai_macro_score import compute_ai_macro_score

FILE = 'C:/Users/sh.park/Documents/USIG_LLM/LUACSTAT_2026_03_31_SCORED.xlsx'

# ─── 1. Load ───────────────────────────────────────────────────────────────────
print("Loading Detail_Scored...")
df = pd.read_excel(FILE, sheet_name='Detail_Scored', header=1)
print(f"  {len(df)} rows x {len(df.columns)} cols")

for col in ['OAS', 'OASD', 'Yield to Worst', '1Y Dflt', 'LQA']:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')

mask = (df['class'].notna()) & (df['class'].astype(str).str.lower() != 'off') \
       & df['OAS'].notna() & df['OASD'].notna()
print(f"  Active bonds: {mask.sum()}")

# ─── Carry_2.5M_pct 재산출 (YTM-based) ────────────────────────────────────────
carry_mask = mask & df['Yield to Worst'].notna()
df.loc[carry_mask, 'Carry_2.5M_pct'] = df.loc[carry_mask, 'Yield to Worst'] / 12 * 2.5
print(f"  Carry_2.5M_pct (YTM-based): non-null={df['Carry_2.5M_pct'].notna().sum()}, "
      f"range=[{df['Carry_2.5M_pct'].min():.3f}, {df['Carry_2.5M_pct'].max():.3f}]")

# ─── Compression_Score_pct 재산출 (1Y Dflt implied spread vs OAS) ─────────────
# Spread_floor(bp) = 1Y_Dflt(%) × (1 - Recovery_40%) × 100 = 1Y_Dflt × 60
comp_mask = mask & df['1Y Dflt'].notna()
df.loc[comp_mask, 'Spread_floor_bp']   = df.loc[comp_mask, '1Y Dflt'] * 60
df.loc[comp_mask, 'Compression_gap_bp'] = (
    df.loc[comp_mask, 'OAS'] - df.loc[comp_mask, 'Spread_floor_bp']
)
df.loc[comp_mask, 'Compression_Score_pct'] = (
    df.loc[comp_mask, 'Compression_gap_bp'] * df.loc[comp_mask, 'OASD'] / 100
)
print(f"  Compression_Score_pct (1Y Dflt-based): non-null={df['Compression_Score_pct'].notna().sum()}, "
      f"range=[{df['Compression_Score_pct'].min():.3f}, {df['Compression_Score_pct'].max():.3f}]")

# ─── 2. DP Rating Score ────────────────────────────────────────────────────────
DP_RATING_MAP = {
    'AAA': 1, 'AA1': 2, 'AA2': 3, 'AA3': 4,
    'A1': 5, 'A2': 6, 'A3': 7,
    'BAA1': 8, 'BAA2': 9, 'BAA3': 10,
    'BA1': 11, 'BA2': 12, 'BA3': 13,
    'B1': 14, 'B2': 15, 'B3': 16,
    'CAA1': 17, 'CAA2': 18, 'CAA3': 19,
    'CA': 20, 'C': 21,
}

def percentile_norm(series, ascending=True):
    """Rank-normalize to [-1, +1]. ascending=True → higher value = better."""
    n = series.notna().sum()
    if n < 2:
        return pd.Series(np.nan, index=series.index)
    if ascending:
        ranked = series.rank(method='average', na_option='keep')
    else:
        ranked = (-series).rank(method='average', na_option='keep')
    return (ranked - 1) / (n - 1) * 2 - 1

df['_dp_fund_num'] = df['DPFundamentalRating'].map(DP_RATING_MAP)
df['_dp_spd_num']  = df['DPSpreadRating'].map(DP_RATING_MAP)

# 갭 = DPSpreadRating - DPFundamentalRating
# 양수 → 시장이 펀더멘털보다 나쁘게 pricing → 저평가, 스프레드 압축 여지 → 좋은 신호
# 음수 → 시장이 펀더멘털보다 좋게 pricing → 고평가 → 나쁜 신호
df['_dp_gap'] = df['_dp_spd_num'] - df['_dp_fund_num']

# 갭이 클수록 좋은 신호 → ascending=True
df['DP_Rating_Score'] = np.nan
df.loc[mask, 'DP_Rating_Score'] = percentile_norm(df.loc[mask, '_dp_gap'], ascending=True).values
df.drop(columns=['_dp_fund_num', '_dp_spd_num', '_dp_gap'], inplace=True)

print(f"  DP_Rating_Score: non-null={df['DP_Rating_Score'].notna().sum()}, "
      f"range=[{df['DP_Rating_Score'].min():.3f}, {df['DP_Rating_Score'].max():.3f}]")

# ─── 3. Bond_TR_Est_pct 재산출 ─────────────────────────────────────────────────
# (Carry + Compression 이미 있음 → DP_Rating_Score 반영해 재계산)
df.loc[mask, 'Bond_TR_Est_pct'] = (
    df.loc[mask, 'Carry_2.5M_pct'].fillna(0) +
    df.loc[mask, 'Compression_Score_pct'].fillna(0) +
    df.loc[mask, 'DP_Rating_Score'].fillna(0) * 0.05
)

# ─── 4. Bond_TR_Score: class 내 percentile-rank → [-1, +1] ──────────────────
# 같은 class 안에서만 상대 비교 (cross-class 비교 제거)
df['Bond_TR_Score'] = np.nan
for cls in df.loc[mask, 'class'].dropna().unique():
    cls_mask = mask & (df['class'] == cls)
    tr_cls = df.loc[cls_mask, 'Bond_TR_Est_pct']
    df.loc[cls_mask, 'Bond_TR_Score'] = percentile_norm(tr_cls, ascending=True).values

n_scored = df['Bond_TR_Score'].notna().sum()
print(f"  Bond_TR_Score (class-normalized): non-null={n_scored}, "
      f"range=[{df['Bond_TR_Score'].min():.3f}, {df['Bond_TR_Score'].max():.3f}]")

# ─── 5a. Sentiment_Score 정제 — generic 뉴스 무효화 ──────────────────────────
# Yahoo가 해당 ticker를 못 찾으면 일반 시장 뉴스를 반환 →
# 동일한 News_Sentiment_Raw 값이 여러 다른 ticker에서 반복됨 → 가짜 signal
# 기준: 같은 raw 값(소수점 5자리)이 5개 초과의 서로 다른 Eqty Ticker에서 나타나면 generic으로 판단

df['_news_raw_r'] = df['News_Sentiment_Raw'].round(5)
df['_eqty_clean'] = df['Eqty Ticker'].astype(str).str.split().str[0]

valid_news = df.dropna(subset=['News_Sentiment_Raw', '_eqty_clean'])
ticker_count_per_raw = (
    valid_news.groupby('_news_raw_r')['_eqty_clean']
    .nunique()
)
GENERIC_THRESHOLD = 5  # 동일 raw 값이 5개 초과 ticker에서 나오면 generic
generic_raw_vals = set(ticker_count_per_raw[ticker_count_per_raw > GENERIC_THRESHOLD].index)

generic_mask = df['_news_raw_r'].isin(generic_raw_vals)
df['Sentiment_Score_clean'] = df['Sentiment_Score'].copy()
df.loc[generic_mask, 'Sentiment_Score_clean'] = np.nan
# 티커별 generic 여부 플래그 (Score_Sentiment 시트 표시용)
df['News_Generic_Flag'] = ''
df.loc[generic_mask & df['News_Sentiment_Raw'].notna(), 'News_Generic_Flag'] = 'GENERIC (invalidated)'
df.loc[df['News_Sentiment_Raw'].isna(), 'News_Generic_Flag'] = 'No news data'

n_generic = generic_mask.sum()
n_valid   = df['Sentiment_Score_clean'].notna().sum()
print(f"  Sentiment 정제: generic raw값 {len(generic_raw_vals)}개, "
      f"무효화 {n_generic}행 → 유효 {n_valid}행 남음")
df.drop(columns=['_news_raw_r', '_eqty_clean'], inplace=True)

# ─── 5b. AI_Macro_Score 산출 ──────────────────────────────────────────────────
print("Computing AI_Macro_Score...")
df = compute_ai_macro_score(df)

# ─── 5c. Integrated_Score (equal weight 0.20 × 5) ─────────────────────────────
# NaN → 0: 해당 컴포넌트에 유효한 데이터 없으면 0점(중립) 기여
df['Integrated_Score'] = (
    df['Bond_TR_Score'].fillna(0)           * 0.20 +
    df['Eq_Mom_Score'].fillna(0)            * 0.20 +
    df['Eq_Fund_Score'].fillna(0)           * 0.20 +
    df['Sentiment_Score_clean'].fillna(0)   * 0.20 +
    df['AI_Macro_Score'].fillna(0)          * 0.20
)
df.loc[~mask, 'Integrated_Score'] = np.nan

# Rank within class (excluding 'off')
df['Integrated_Rank_in_Class'] = np.nan
for cls in df.loc[mask, 'class'].dropna().unique():
    cls_mask = mask & (df['class'] == cls)
    ranked = df.loc[cls_mask, 'Integrated_Score'].rank(ascending=False, method='min')
    df.loc[cls_mask, 'Integrated_Rank_in_Class'] = ranked

def top_flag(rank):
    if pd.isna(rank): return ''
    r = int(rank)
    if r <= 3:  return '★★★ TOP3'
    if r <= 10: return '★★ TOP10'
    if r <= 25: return '★ TOP25'
    return ''

df['Top_Pick_Flag'] = df['Integrated_Rank_in_Class'].apply(top_flag)
print(f"  Integrated_Score computed. Top flags: {df['Top_Pick_Flag'].value_counts().to_dict()}")

# ─── 6. Sort for display ───────────────────────────────────────────────────────
df_out = (
    df[mask]
    .sort_values(['class', 'Integrated_Rank_in_Class'], na_position='last')
    .reset_index(drop=True)
)
print(f"  Sorted active rows: {len(df_out)}")

# ─── 7. Helpers ────────────────────────────────────────────────────────────────
FILL_TITLE = PatternFill('solid', fgColor='1F3864')   # dark blue
FILL_ID    = PatternFill('solid', fgColor='BDD7EE')   # light blue
FILL_COMP  = PatternFill('solid', fgColor='FFF2CC')   # yellow
FILL_SCORE = PatternFill('solid', fgColor='E2EFDA')   # light green
FILL_POS   = PatternFill('solid', fgColor='C6EFCE')   # green (positive score)
FILL_NEG   = PatternFill('solid', fgColor='FFC7CE')   # red   (negative score)

FONT_TOP3  = Font(name='Arial', bold=True, color='FF0000', size=10)
FONT_TOP10 = Font(name='Arial', bold=True, color='C55A11', size=10)
FONT_TOP25 = Font(name='Arial', bold=True, color='1F3864', size=10)
FILL_TOP3  = PatternFill('solid', fgColor='FFFF00')
FILL_TOP10 = PatternFill('solid', fgColor='FCE4D6')
FILL_TOP25 = PatternFill('solid', fgColor='DDEBF7')

def _to_py(val):
    if val is None: return None
    try:
        if pd.isna(val): return None
    except (TypeError, ValueError): pass
    if isinstance(val, (np.integer,)): return int(val)
    if isinstance(val, (np.floating,)): return float(val)
    if isinstance(val, pd.Timestamp): return val.to_pydatetime()
    return val

def _make_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)

def _title(ws, text, n_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=text)
    c.font      = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill      = FILL_TITLE
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

def _headers(ws, hdrs, id_set, comp_set, score_set):
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = Font(name='Arial', bold=True, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if h in score_set:  c.fill = FILL_SCORE
        elif h in comp_set: c.fill = FILL_COMP
        else:               c.fill = FILL_ID

def _write_rows(ws, df_in, hdrs, score_set, fmt_map, flag_col=None, color_score_cols=None):
    for ri, row in df_in.iterrows():
        er = ri + 3
        for ci, h in enumerate(hdrs, 1):
            raw = row.get(h, np.nan) if h in row.index else np.nan
            val = _to_py(raw)
            c = ws.cell(row=er, column=ci)
            c.value = val
            c.font  = Font(name='Arial', size=10)
            if h in score_set:
                c.fill = PatternFill('solid', fgColor='F4FFF4')
            if h in fmt_map:
                c.number_format = fmt_map[h]
            if flag_col and h == flag_col:
                flag = str(val) if val else ''
                c.value = flag
                if flag == '★★★ TOP3':   c.font = FONT_TOP3;  c.fill = FILL_TOP3
                elif flag == '★★ TOP10': c.font = FONT_TOP10; c.fill = FILL_TOP10
                elif flag == '★ TOP25':  c.font = FONT_TOP25; c.fill = FILL_TOP25
            if color_score_cols and h in color_score_cols and val is not None:
                if isinstance(val, (int, float)) and not math.isnan(val):
                    if val > 0.3:   c.fill = FILL_POS
                    elif val < -0.3: c.fill = FILL_NEG

def _finalize(ws, n_cols, col_widths):
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(n_cols)}2'
    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

# ─── 8. Open workbook ──────────────────────────────────────────────────────────
print("\nOpening workbook...")
wb = load_workbook(FILE)

# 불필요한 시트 삭제 — 보존: Detail_Scored, Methodology, Score_* 5개
KEEP_SHEETS = {'Detail_Scored', 'Methodology',
               'Score_BondTR', 'Score_EqMom', 'Score_EqFund',
               'Score_Sentiment', 'Score_AI', 'Score_Integrated'}
for sname in list(wb.sheetnames):
    if sname not in KEEP_SHEETS:
        del wb[sname]
        print(f"  Deleted sheet: {sname}")

# ─── Sheet 1: Score_BondTR ────────────────────────────────────────────────────
print("  Creating Score_BondTR...")
ws = _make_sheet(wb, 'Score_BondTR')
id_c    = ['class', 'Des', 'ISIN', 'Ticker', 'Cpn', 'OAS', 'OASD']
comp_c  = ['Carry_2.5M_pct', 'Spread_vs_Class_Median', 'Compression_Score_pct',
            'DPFundamentalRating', 'DPSpreadRating', 'DP_Rating_Score']
score_c = ['Bond_TR_Est_pct', 'Bond_TR_Score']
hdrs    = id_c + comp_c + score_c
fmt     = {'Cpn':'0.000','OAS':'0.0','OASD':'0.00',
           'Carry_2.5M_pct':'0.0000','Spread_vs_Class_Median':'0.0',
           'Compression_Score_pct':'0.0000','DP_Rating_Score':'0.0000',
           'Bond_TR_Est_pct':'0.0000','Bond_TR_Score':'0.0000'}
_title(ws, 'Bond TR Score  │  Carry + Compression + DP Rating (Fundamental & Spread) → Bond_TR_Score [-1, +1]', len(hdrs))
_headers(ws, hdrs, set(id_c), set(comp_c), set(score_c))
_write_rows(ws, df_out, hdrs, set(score_c), fmt)
_finalize(ws, len(hdrs), {1:12, 2:28, 3:16, 4:10, 5:7, 6:7, 7:7,
                           8:11, 9:11, 10:13, 11:16, 12:16, 13:11, 14:13, 15:13})

# ─── Sheet 2: Score_EqMom ─────────────────────────────────────────────────────
print("  Creating Score_EqMom...")
ws = _make_sheet(wb, 'Score_EqMom')
id_c    = ['class', 'Des', 'ISIN', 'Ticker', 'Eqty Ticker']
comp_c  = ['Eq_Ret_1M', 'Eq_Ret_3M', 'Eq_Vol_30D', 'Eq_vs_52w_High']
score_c = ['Eq_Mom_Score']
hdrs    = id_c + comp_c + score_c
fmt     = {'Eq_Ret_1M':'0.00%','Eq_Ret_3M':'0.00%',
           'Eq_Vol_30D':'0.00%','Eq_vs_52w_High':'0.00%',
           'Eq_Mom_Score':'0.0000'}
_title(ws, 'Equity Momentum Score  │  1M/3M Return + 30D Volatility + 52W High → Eq_Mom_Score [-1, +1]', len(hdrs))
_headers(ws, hdrs, set(id_c), set(comp_c), set(score_c))
_write_rows(ws, df_out, hdrs, set(score_c), fmt)
_finalize(ws, len(hdrs), {1:12, 2:28, 3:16, 4:10, 5:12,
                           6:11, 7:11, 8:11, 9:13, 10:13})

# ─── Sheet 3: Score_EqFund ────────────────────────────────────────────────────
print("  Creating Score_EqFund...")
ws = _make_sheet(wb, 'Score_EqFund')
id_c    = ['class', 'Des', 'ISIN', 'Ticker', 'Eqty Ticker']
comp_c  = ['Debt_to_Equity', 'Profit_Margin', 'Revenue_Growth',
            'Current_Ratio', 'EV_EBITDA', 'PE_Ratio']
score_c = ['Eq_Fund_Score']
hdrs    = id_c + comp_c + score_c
fmt     = {'Debt_to_Equity':'0.00','Profit_Margin':'0.00%','Revenue_Growth':'0.00%',
           'Current_Ratio':'0.00','EV_EBITDA':'0.00','PE_Ratio':'0.00',
           'Eq_Fund_Score':'0.0000'}
_title(ws, 'Equity Fundamental Score  │  D/E · Margin · Growth · CR · EV/EBITDA · P/E → Eq_Fund_Score [-1, +1]', len(hdrs))
_headers(ws, hdrs, set(id_c), set(comp_c), set(score_c))
_write_rows(ws, df_out, hdrs, set(score_c), fmt)
_finalize(ws, len(hdrs), {1:12, 2:28, 3:16, 4:10, 5:12,
                           6:11, 7:11, 8:11, 9:11, 10:11, 11:11, 12:13})

# ─── Sheet 4: Score_Sentiment ─────────────────────────────────────────────────
print("  Creating Score_Sentiment...")
ws = _make_sheet(wb, 'Score_Sentiment')
id_c    = ['class', 'Des', 'ISIN', 'Ticker', 'Eqty Ticker']
comp_c  = ['News_Sentiment_Raw', 'News_Article_Count', 'Google_News_Count', 'News_Generic_Flag',
           'Top_Headline', 'Top_Headline_Score',
           'Trends_Momentum', 'Trends_Factor',
           'Sentiment_Score', 'Sentiment_Score_clean']
score_c = ['Sentiment_Score_clean']
hdrs    = id_c + comp_c
fmt     = {'News_Sentiment_Raw':'0.00000','News_Article_Count':'0','Google_News_Count':'0',
           'Top_Headline_Score':'0.0000',
           'Trends_Momentum':'0.0000','Trends_Factor':'0.0000',
           'Sentiment_Score':'0.0000','Sentiment_Score_clean':'0.0000'}
_title(ws, 'Sentiment Score  │  Yahoo Finance News (VADER, recency-weighted) + Google Trends → Sentiment_Score_clean [-1,+1]  │  GENERIC = invalidated (→ 0 in Integrated)', len(hdrs))
_headers(ws, hdrs, set(id_c), set(comp_c), set(score_c))

# News_Generic_Flag 컬럼에 색상 강조 적용
FILL_GENERIC = PatternFill('solid', fgColor='FFE0E0')  # 연한 빨강 — GENERIC
FILL_NODATA  = PatternFill('solid', fgColor='F2F2F2')  # 회색 — 데이터 없음

for ri, row in df_out.iterrows():
    er = ri + 3
    for ci, h in enumerate(hdrs, 1):
        raw = row.get(h, np.nan) if h in row.index else np.nan
        val = _to_py(raw)
        c = ws.cell(row=er, column=ci)
        c.value = val
        c.font  = Font(name='Arial', size=10)
        if h in fmt:
            c.number_format = fmt[h]
        if h in score_c:
            c.fill = PatternFill('solid', fgColor='F4FFF4')
        if h == 'News_Generic_Flag':
            flag = str(val) if val else ''
            if flag == 'GENERIC (invalidated)':
                c.fill = FILL_GENERIC
                c.font = Font(name='Arial', size=10, bold=True, color='CC0000')
            elif flag == 'No news data':
                c.fill = FILL_NODATA
                c.font = Font(name='Arial', size=10, color='888888')
            if h == 'Top_Headline':
                c.alignment = Alignment(wrap_text=True, vertical='center')
            if h == 'Top_Headline_Score' and val is not None:
                try:
                    fval = float(val)
                    if fval >= 0.5:    # 강한 긍정 → 초록
                        c.fill = PatternFill('solid', fgColor='C6EFCE')
                        c.font = Font(name='Arial', size=10, bold=True, color='375623')
                    elif fval <= -0.5: # 강한 부정 → 빨강
                        c.fill = PatternFill('solid', fgColor='FFC7CE')
                        c.font = Font(name='Arial', size=10, bold=True, color='9C0006')
                except (TypeError, ValueError):
                    pass

_finalize(ws, len(hdrs), {1:12, 2:28, 3:16, 4:10, 5:12,
                           6:13, 7:10, 8:12, 9:22,
                           10:60, 11:11,
                           12:13, 13:13, 14:13, 15:13})

# ─── Sheet 5: Score_AI ────────────────────────────────────────────────────────
print("  Creating Score_AI...")
ws = _make_sheet(wb, 'Score_AI')
FILL_AI  = PatternFill('solid', fgColor='EDE7F6')   # 연보라 — AI 점수
FILL_AI_H= PatternFill('solid', fgColor='4A148C')   # 진보라 — 타이틀

id_c_ai   = ['class', 'Des', 'ISIN', 'Ticker', 'OAD', 'BCLASS3', 'Industry Sector',
             'DPFundamentalRating', 'Issuer Rtg']
comp_c_ai = ['AI_Sector_Score', 'AI_Maturity_Score', 'AI_RatingBuf_Score']
score_c_ai= ['AI_Macro_Score']
hdrs_ai   = id_c_ai + comp_c_ai + score_c_ai

# Custom title with purple
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(hdrs_ai))
tc = ws.cell(row=1, column=1,
    value='AI Macro Score  |  Sector Positioning (x0.40) + Maturity Curve (x0.35) + Rating Buffer (x0.25)  |  Based on macro reasoning, not quant rank')
tc.font      = Font(name='Arial', bold=True, size=11, color='FFFFFF')
tc.fill      = FILL_AI_H
tc.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 22

for ci, h in enumerate(hdrs_ai, 1):
    c = ws.cell(row=2, column=ci, value=h)
    c.font      = Font(name='Arial', bold=True, size=10)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if h in score_c_ai:  c.fill = FILL_AI
    elif h in comp_c_ai: c.fill = FILL_COMP
    else:                c.fill = FILL_ID

fmt_ai = {'OAD':'0.00',
          'AI_Sector_Score':'0.00','AI_Maturity_Score':'0.00','AI_RatingBuf_Score':'0.00',
          'AI_Macro_Score':'0.0000'}

for ri, row in df_out.iterrows():
    er = ri + 3
    for ci, h in enumerate(hdrs_ai, 1):
        raw = row.get(h, np.nan) if h in row.index else np.nan
        val = _to_py(raw)
        c   = ws.cell(row=er, column=ci)
        c.value = val
        c.font  = Font(name='Arial', size=10)
        if h in fmt_ai: c.number_format = fmt_ai[h]
        if h == 'AI_Macro_Score' and val is not None:
            try:
                fv = float(val)
                if fv > 0.3:    c.fill = FILL_POS
                elif fv < -0.3: c.fill = FILL_NEG
                else:           c.fill = FILL_AI
            except: c.fill = FILL_AI
        if h in comp_c_ai and val is not None:
            try:
                fv = float(val)
                if fv > 0:   c.fill = PatternFill('solid', fgColor='E8F5E9')
                elif fv < 0: c.fill = PatternFill('solid', fgColor='FFEBEE')
            except: pass

ws.freeze_panes = 'A3'
ws.auto_filter.ref = f'A2:{get_column_letter(len(hdrs_ai))}2'
col_widths_ai = {1:12, 2:28, 3:16, 4:10, 5:7, 6:18, 7:20, 8:16, 9:12,
                 10:14, 11:14, 12:14, 13:14}
for ci, w in col_widths_ai.items():
    ws.column_dimensions[get_column_letter(ci)].width = w

# ── Macro View 섹션 (Score_AI 시트 오른쪽, col 15~18) ────────────────────────
import re as _re
from ai_macro_score import SUBGROUP_SCORE_MAP as _SUBMAP

# ai_macro_score.py 에서 인라인 코멘트 추출
_rationale = {}
_ai_src = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ai_macro_score.py')
with open(_ai_src, 'r', encoding='utf-8') as _f:
    for _line in _f:
        _m = _re.match(r"\s+'([^']+)':\s*([-\d.]+),\s*#\s*(.+)", _line)
        if _m:
            _rationale[_m.group(1)] = _m.group(3).strip()

MV_COL = len(hdrs_ai) + 2   # 한 칸 띄우고 시작 (col 15)

FILL_MV_HDR  = PatternFill('solid', fgColor='4A148C')   # 진보라
FILL_MV_BG   = PatternFill('solid', fgColor='F3E5F5')   # 연보라
FILL_MV_POS  = PatternFill('solid', fgColor='C8E6C9')   # 연초록
FILL_MV_NEG  = PatternFill('solid', fgColor='FFCDD2')   # 연빨강
FILL_MV_NEUT = PatternFill('solid', fgColor='F5F5F5')   # 연회색
FILL_MV_THEME= PatternFill('solid', fgColor='311B92')   # 더 진한 보라 (테마)

# ── 매크로 배경 4가지 테마 ──────────────────────────────────────────────────
MACRO_THEMES = [
    (
        "① Fed Policy & Tariff Inflation Risk",
        "[Current] Fed is in an easing cycle (cumulative -100bp since Sep 2024), but Trump's sweeping tariffs (10% universal + 25% autos/steel/aluminum + sector-specific) are reigniting inflation fears. Core PCE remains sticky above 2.5%, and market pricing for 2025 cuts has been sharply pared back.\n"
        "[Risk] If tariffs fully pass through to CPI, the Fed may be forced to pause or reverse — a stagflation-lite scenario where growth slows but rates stay high. This is the key tail risk for credit.\n"
        "[Positioning] Prefer shorter duration, high-carry defensive names. Avoid cyclicals with thin margins and high import cost exposure. Favor regulated utilities, domestic healthcare, and infrastructure where pricing power is intact."
    ),
    (
        "② IG Spread Level — Historically Tight, Carry Dominates",
        "[Current] US IG OAS is near post-GFC tights (~85-95bp range as of Q1 2026), implying limited price appreciation upside from further compression. The risk/reward for spread tightening is asymmetric — downside (widening) is larger than upside.\n"
        "[Implication] In a tight-spread environment, total return is increasingly driven by carry (coupon income) rather than capital gains. Selection alpha comes from picking bonds with superior carry-to-risk profiles within each rating/maturity bucket.\n"
        "[Positioning] Bond_TR Score emphasizes carry (YTW-based) and DP rating buffer — both reward high-coupon, fundamentally cheap bonds. AI_Macro reinforces by avoiding sectors where spread widening risk is elevated (autos, retail, leveraged cyclicals)."
    ),
    (
        "③ Yield Curve — Steepening Pressure, Front-End Anchored",
        "[Current] The UST curve has been bear-steepening, with 10Y-2Y spread widening as fiscal deficit concerns push up term premium. The 2-year is anchored by near-term Fed expectations, while 30-year yields face upward pressure from supply (Treasury issuance) and inflation uncertainty.\n"
        "[Maturity Sweet Spot] OAD 4-7 years (approximately 5-7Y bonds): captures meaningful carry and spread duration without excessive long-end term premium risk. OAD 7-10Y remains acceptable. Bonds with OAD 13Y+ are penalized — duration extension into an unfavorable part of the curve.\n"
        "[Positioning] Maturity_Score reflects this view: peak score at OAD 4-7 (+1.0), declining sharply for 13Y+ (-0.50 to -1.00). Avoid long-dated BBB bonds where both spread widening and rate risk are elevated simultaneously."
    ),
    (
        "④ Tariff Exposure by Sector — Winners & Losers",
        "[Direct Tariff Losers] Autos & Parts: 25% import tariff on finished vehicles + parts supply chain disruption → Ford, GM suppliers, Japanese/Korean OEMs facing significant cost inflation and potential volume declines. Score: -0.70 to -0.90.\n"
        "Retail (Apparel, Dept Stores, Electronics): High China-sourced inventory → margin compression. Retailers with weak pricing power (mid-tier dept stores, specialty apparel) are most at risk. Score: -0.40 to -0.60.\n"
        "Consumer Discretionary Manufacturing (Toys, Footwear, Motorcycles): Heavily Asia-manufactured → direct tariff impact on COGS. Score: -0.40 to -0.55.\n"
        "[Relative Winners] Domestic Infrastructure & Defense: No tariff exposure, government contract revenue, benefiting from reshoring/defense spending tailwinds. Score: +0.35 to +0.40.\n"
        "Regulated Utilities (Electric, Gas, Water): Pass-through pricing to ratepayers, recession-resistant demand, largely immune to trade policy. Score: +0.75 to +0.85.\n"
        "Domestic Healthcare (Hospitals, HMOs, Labs): Primarily domestic service delivery, no import exposure, defensive demand. Score: +0.60 to +0.75.\n"
        "Steel Producers (domestic): Protected by 25% tariff on imported steel — domestic producers gain pricing power. Score: +0.20.\n"
        "[Monitoring] Semiconductors: US-China export controls tightening; supply chain restructuring ongoing. Currently neutral-to-negative pending clarity on CHIPS Act implementation and China retaliation scope."
    ),
]

# 타이틀
ws.merge_cells(start_row=1, start_column=MV_COL, end_row=1, end_column=MV_COL+3)
tc = ws.cell(row=1, column=MV_COL,
    value='MACRO VIEW  |  AI Sector Score Rationale  |  Based on April 2026 macro environment')
tc.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
tc.fill = FILL_MV_HDR
tc.alignment = Alignment(horizontal='center', vertical='center')

# 테마 블록 (row 2~5) — 라벨(col MV_COL) + 본문(col MV_COL+1 ~ MV_COL+3 merged)
for ti, (theme_label, theme_body) in enumerate(MACRO_THEMES):
    tr = ti + 2
    lc = ws.cell(row=tr, column=MV_COL, value=theme_label)
    lc.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    lc.fill = FILL_MV_THEME
    lc.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

    ws.merge_cells(start_row=tr, start_column=MV_COL+1, end_row=tr, end_column=MV_COL+3)
    bc = ws.cell(row=tr, column=MV_COL+1, value=theme_body)
    bc.font = Font(name='Arial', size=8.5)
    bc.fill = FILL_MV_BG
    bc.alignment = Alignment(vertical='top', wrap_text=True)

# 서브그룹 테이블 헤더 (row 6)
mv_subhdr_row = 6
for ci_off, (hdr_txt, hdr_fill) in enumerate([
    ('Industry Subgroup', FILL_ID),
    ('Score',             FILL_AI),
    ('Rationale',         FILL_COMP),
    ('Direction',         FILL_AI),
]):
    c = ws.cell(row=mv_subhdr_row, column=MV_COL+ci_off, value=hdr_txt)
    c.font = Font(name='Arial', bold=True, size=9)
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 서브그룹 데이터 (score 내림차순 정렬)
_sorted_subs = sorted(_SUBMAP.items(), key=lambda x: x[1], reverse=True)
for si, (subgroup, score) in enumerate(_sorted_subs):
    dr = mv_subhdr_row + 1 + si
    rationale_txt = _rationale.get(subgroup, '')
    direction = ('▲ Overweight' if score > 0.3
                 else '▼ Underweight' if score < -0.3
                 else '— Neutral')

    sc = ws.cell(row=dr, column=MV_COL, value=subgroup)
    sc.font = Font(name='Arial', size=8.5)
    sc.alignment = Alignment(vertical='center')

    vc = ws.cell(row=dr, column=MV_COL+1, value=round(score, 2))
    vc.number_format = '0.00'
    vc.font = Font(name='Arial', size=8.5, bold=True)
    vc.alignment = Alignment(horizontal='center', vertical='center')

    rc = ws.cell(row=dr, column=MV_COL+2, value=rationale_txt)
    rc.font = Font(name='Arial', size=8.5)
    rc.alignment = Alignment(vertical='center', wrap_text=True)

    dc = ws.cell(row=dr, column=MV_COL+3, value=direction)
    dc.font = Font(name='Arial', size=8.5,
                   color=('375623' if score > 0.3 else '9C0006' if score < -0.3 else '555555'))
    dc.alignment = Alignment(horizontal='center', vertical='center')

    # 행 배경색
    row_fill = (FILL_MV_POS if score > 0.3 else
                FILL_MV_NEG if score < -0.3 else FILL_MV_NEUT)
    for col_off in range(4):
        ws.cell(row=dr, column=MV_COL+col_off).fill = row_fill

# 매크로 뷰 열 너비
ws.column_dimensions[get_column_letter(MV_COL)].width   = 32   # theme label / subgroup
ws.column_dimensions[get_column_letter(MV_COL+1)].width = 7    # score
ws.column_dimensions[get_column_letter(MV_COL+2)].width = 80   # rationale (wide for theme text)
ws.column_dimensions[get_column_letter(MV_COL+3)].width = 15
ws.row_dimensions[1].height = 22
THEME_ROW_HEIGHTS = [52, 60, 60, 100]   # 테마별 행 높이 (텍스트 길이 반영)
for ti, rh in enumerate(THEME_ROW_HEIGHTS):
    ws.row_dimensions[ti+2].height = rh
ws.row_dimensions[mv_subhdr_row].height = 18
print(f"  Score_AI Macro View: {len(_sorted_subs)} subgroups written")

# ─── Sheet 6: Score_Integrated ────────────────────────────────────────────────
print("  Creating Score_Integrated...")
ws = _make_sheet(wb, 'Score_Integrated')
id_c    = ['class', 'Des', 'ISIN', 'Ticker', 'Cpn', 'OAS', 'LQA', 'Issuer Rtg', 'BCLASS3', 'Industry Sector']
comp_c  = ['Bond_TR_Score', 'Eq_Mom_Score', 'Eq_Fund_Score', 'Sentiment_Score', 'AI_Macro_Score']
score_c = ['Integrated_Score', 'Integrated_Rank_in_Class', 'Top_Pick_Flag']
hdrs    = id_c + comp_c + score_c
fmt     = {'Cpn':'0.000','OAS':'0.0','LQA':'0.0',
           'Bond_TR_Score':'0.0000','Eq_Mom_Score':'0.0000',
           'Eq_Fund_Score':'0.0000','Sentiment_Score':'0.0000',
           'AI_Macro_Score':'0.0000',
           'Integrated_Score':'0.0000','Integrated_Rank_in_Class':'0'}
_title(ws, 'Integrated Score  |  Bond_TR+EqMom+EqFund+Sentiment+AI_Macro (x0.20 each)  |  All scores [-1,+1]', len(hdrs))
_headers(ws, hdrs, set(id_c), set(comp_c), set(score_c))
_write_rows(ws, df_out, hdrs, set(score_c), fmt,
            flag_col='Top_Pick_Flag',
            color_score_cols={'Bond_TR_Score','Eq_Mom_Score','Eq_Fund_Score','Sentiment_Score','AI_Macro_Score'})
_finalize(ws, len(hdrs), {1:12, 2:28, 3:16, 4:10, 5:7, 6:7, 7:8,
                           8:10, 9:14, 10:16,
                           11:13, 12:13, 13:13, 14:13, 15:13,
                           16:13, 17:10, 18:14})

# ─── 9. Update Detail_Scored: Score/Rank/Flag columns ─────────────────────────
print("\n  Updating Detail_Scored columns...")
ws_d = wb['Detail_Scored']
col_map = {str(ws_d.cell(2, c).value).strip(): c
           for c in range(1, ws_d.max_column + 1)
           if ws_d.cell(2, c).value}

def _get_or_add_col(ws, cm, col_name, fill_color='E2EFDA'):
    """컬럼이 없으면 마지막 열에 추가하고 col_map을 갱신."""
    if col_name in cm:
        return cm[col_name]
    new_idx = ws.max_column + 1
    c = ws.cell(row=2, column=new_idx, value=col_name)
    c.font      = Font(name='Arial', bold=True, size=10)
    c.fill      = PatternFill('solid', fgColor=fill_color)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cm[col_name] = new_idx
    return new_idx

# ── 추가가 필요한 컬럼 정의 (색상별 그룹) ──────────────────────────────────────
# BondTR 관련 (연두)
for cn in ['DP_Rating_Score', 'Bond_TR_Score']:
    _get_or_add_col(ws_d, col_map, cn, 'E2EFDA')

# AI 점수 관련 (연보라)
for cn in ['AI_Sector_Score', 'AI_Maturity_Score', 'AI_RatingBuf_Score', 'AI_Macro_Score']:
    _get_or_add_col(ws_d, col_map, cn, 'EDE7F6')

# Sentiment 정제 관련 (연노랑)
for cn in ['Sentiment_Score_clean', 'News_Generic_Flag']:
    _get_or_add_col(ws_d, col_map, cn, 'FFF2CC')

print(f"    col_map size: {len(col_map)}  |  max_col: {ws_d.max_column}")

# ── ISIN → 점수 매핑 딕셔너리 구성 ────────────────────────────────────────────
score_cols_needed = [
    'DP_Rating_Score', 'Bond_TR_Est_pct', 'Bond_TR_Score',
    'AI_Sector_Score', 'AI_Maturity_Score', 'AI_RatingBuf_Score', 'AI_Macro_Score',
    'Sentiment_Score_clean', 'News_Generic_Flag',
    'Integrated_Score', 'Integrated_Rank_in_Class', 'Top_Pick_Flag',
]
isin_to_scores = {}
for _, row in df.iterrows():
    isin = str(row['ISIN']).strip() if pd.notna(row['ISIN']) else None
    if isin and isin not in isin_to_scores:
        isin_to_scores[isin] = {c: row[c] for c in score_cols_needed if c in row.index}

# ── 셀 포맷 정의 ───────────────────────────────────────────────────────────────
flag_fmt = {
    '★★★ TOP3': (Font(name='Arial', bold=True, color='FF0000', size=10), PatternFill('solid', fgColor='FFFF00')),
    '★★ TOP10': (Font(name='Arial', bold=True, color='C55A11', size=10), PatternFill('solid', fgColor='FCE4D6')),
    '★ TOP25':  (Font(name='Arial', bold=True, color='1F3864', size=10), PatternFill('solid', fgColor='DDEBF7')),
    '':         (Font(name='Arial', size=10), PatternFill(fill_type=None)),
}

FILL_GENERIC_D = PatternFill('solid', fgColor='FFE0E0')
FILL_NODATA_D  = PatternFill('solid', fgColor='F2F2F2')

score_update_cols = [
    'DP_Rating_Score', 'Bond_TR_Est_pct', 'Bond_TR_Score',
    'AI_Sector_Score', 'AI_Maturity_Score', 'AI_RatingBuf_Score', 'AI_Macro_Score',
    'Sentiment_Score_clean', 'News_Generic_Flag',
    'Integrated_Score', 'Integrated_Rank_in_Class', 'Top_Pick_Flag',
]

updated = 0
for excel_row in range(3, ws_d.max_row + 1):
    isin_col = col_map.get('ISIN')
    if not isin_col:
        break
    isin = ws_d.cell(row=excel_row, column=isin_col).value
    if not isin:
        continue
    scores = isin_to_scores.get(str(isin).strip())
    if not scores:
        continue
    for col_name in score_update_cols:
        if col_name not in col_map:
            continue
        val = _to_py(scores.get(col_name))
        c = ws_d.cell(row=excel_row, column=col_map[col_name])
        c.value = val
        c.font  = Font(name='Arial', size=10)

        if col_name == 'Top_Pick_Flag':
            flag = str(val) if val else ''
            c.value = flag
            fnt, fll = flag_fmt.get(flag, flag_fmt[''])
            c.font = fnt; c.fill = fll

        elif col_name == 'News_Generic_Flag':
            flag = str(val) if val else ''
            if flag == 'GENERIC (invalidated)':
                c.fill = FILL_GENERIC_D
                c.font = Font(name='Arial', size=10, bold=True, color='CC0000')
            elif flag == 'No news data':
                c.fill = FILL_NODATA_D
                c.font = Font(name='Arial', size=10, color='888888')

        elif col_name in ('AI_Sector_Score', 'AI_Maturity_Score',
                          'AI_RatingBuf_Score', 'AI_Macro_Score'):
            if val is not None:
                try:
                    fv = float(val)
                    if fv > 0.3:    c.fill = PatternFill('solid', fgColor='E8F5E9')
                    elif fv < -0.3: c.fill = PatternFill('solid', fgColor='FFEBEE')
                    else:           c.fill = PatternFill('solid', fgColor='EDE7F6')
                except (TypeError, ValueError):
                    pass

        elif col_name in ('DP_Rating_Score', 'Bond_TR_Score',
                          'Sentiment_Score_clean', 'Integrated_Score'):
            if val is not None:
                try:
                    fv = float(val)
                    if fv > 0.3:    c.fill = PatternFill('solid', fgColor='C6EFCE')
                    elif fv < -0.3: c.fill = PatternFill('solid', fgColor='FFC7CE')
                    else:           c.fill = PatternFill('solid', fgColor='F4FFF4')
                except (TypeError, ValueError):
                    pass

    updated += 1

print(f"    Detail_Scored: {updated} rows updated  |  cols written: {len(score_update_cols)}")

# ─── 10. Save ──────────────────────────────────────────────────────────────────
import os
print("\nSaving workbook...")
wb.save(FILE)
sz = os.path.getsize(FILE)
print(f"  Saved: {FILE}  ({sz/1024/1024:.2f} MB)")
print("\n  Sheets created:")
for s in wb.sheetnames:
    print(f"    · {s}")
print("\nDONE")
