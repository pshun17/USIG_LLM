"""기존 SCORED 파일 데이터를 읽어 가벼운 버전으로 재생성"""
import sys, io, math, re, warnings, os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
warnings.filterwarnings('ignore')

BASE = r'C:\Users\sh.park\Documents\USIG_LLM'
AS_OF = '2026-05-11'
SRC = BASE + r'\LUACSTAT_2026_05_11_SCORED.xlsx'
OUT     = SRC  # 중간 xlsx
OUT_BSB = SRC.replace('.xlsx', '.xlsb')  # 최종 xlsb

print('Reading data...')
df_out = pd.read_excel(SRC, sheet_name='Detail_Scored', header=1)
print(f'  {len(df_out):,} rows x {len(df_out.columns)} cols')

# ── Fills / Fonts
FILL_TITLE = PatternFill('solid', fgColor='1F3864')
FILL_ID    = PatternFill('solid', fgColor='BDD7EE')
FILL_COMP  = PatternFill('solid', fgColor='FFF2CC')
FILL_SCORE = PatternFill('solid', fgColor='E2EFDA')
FILL_POS   = PatternFill('solid', fgColor='C6EFCE')
FILL_NEG   = PatternFill('solid', fgColor='FFC7CE')
FILL_TOP3  = PatternFill('solid', fgColor='FFFF00')
FILL_TOP10 = PatternFill('solid', fgColor='FCE4D6')
FILL_TOP25 = PatternFill('solid', fgColor='DDEBF7')
FILL_AI_H  = PatternFill('solid', fgColor='4A148C')
FILL_AI    = PatternFill('solid', fgColor='EDE7F6')
FILL_GENERIC = PatternFill('solid', fgColor='FFE0E0')
FILL_NODATA  = PatternFill('solid', fgColor='F2F2F2')
FILL_MV_HDR  = PatternFill('solid', fgColor='4A148C')
FILL_MV_BG   = PatternFill('solid', fgColor='F3E5F5')
FILL_MV_THEME= PatternFill('solid', fgColor='311B92')
FILL_MV_POS  = PatternFill('solid', fgColor='C8E6C9')
FILL_MV_NEG  = PatternFill('solid', fgColor='FFCDD2')
FILL_MV_NEUT = PatternFill('solid', fgColor='F5F5F5')
FONT_TOP3  = Font(bold=True, color='FF0000', size=10)
FONT_TOP10 = Font(bold=True, color='C55A11', size=10)
FONT_TOP25 = Font(bold=True, color='1F3864', size=10)

def _v(val):
    if val is None: return None
    try:
        if pd.isna(val): return None
    except: pass
    if isinstance(val, (np.integer,)): return int(val)
    if isinstance(val, (np.floating,)): return float(val)
    if isinstance(val, pd.Timestamp): return val.to_pydatetime()
    return val

wb = Workbook(); wb.remove(wb.active)
def mks(n): return wb.create_sheet(n)

def wtitle(ws, text, n):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n)
    c=ws.cell(row=1,column=1,value=text)
    c.font=Font(bold=True,size=11,color='FFFFFF'); c.fill=FILL_TITLE
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=22

def whdrs(ws,hdrs,id_s,comp_s,sc_s):
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(row=2,column=ci,value=h)
        c.font=Font(bold=True,size=10)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        c.fill=FILL_SCORE if h in sc_s else(FILL_COMP if h in comp_s else FILL_ID)

def wrows(ws,hdrs,sc_s,fmt,flag_col=None,color_cols=None):
    fmt_ci={ci:fmt[h] for ci,h in enumerate(hdrs,1) if h in fmt}
    for ri,row in df_out.iterrows():
        er=ri+3
        flag_val=str(_v(row.get(flag_col,''))or'') if flag_col else ''
        for ci,h in enumerate(hdrs,1):
            val=_v(row.get(h,np.nan) if h in row.index else np.nan)
            c=ws.cell(row=er,column=ci); c.value=val
            if ci in fmt_ci: c.number_format=fmt_ci[ci]
            if flag_col and h==flag_col:
                c.value=flag_val
                if   flag_val=='★★★ TOP3':  c.font=FONT_TOP3;  c.fill=FILL_TOP3
                elif flag_val=='★★ TOP10':  c.font=FONT_TOP10; c.fill=FILL_TOP10
                elif flag_val=='★ TOP25':   c.font=FONT_TOP25; c.fill=FILL_TOP25
                continue
            if color_cols and h in color_cols and val is not None:
                try:
                    fv=float(val)
                    if not math.isnan(fv):
                        if fv>0.3:    c.fill=FILL_POS
                        elif fv<-0.3: c.fill=FILL_NEG
                except: pass

def wfin(ws,n,widths):
    ws.freeze_panes='A3'
    ws.auto_filter.ref=f'A2:{get_column_letter(n)}2'
    for ci,w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width=w

# ── Score_BondTR
print('  Score_BondTR...')
ws=mks('Score_BondTR')
id_c=['class','Des','ISIN','Ticker','Cpn','OAS','OASD']
cp_c=['Carry_2.5M_pct','Compression_Score_pct','DPFundamentalRating','DPSpreadRating','DP_Rating_Score']
sc_c=['Bond_TR_Est_pct','Bond_TR_Score']
hdrs=id_c+cp_c+sc_c
wtitle(ws,'Bond TR Score  |  Carry + Compression + DP Rating -> Bond_TR_Score [-1,+1]',len(hdrs))
whdrs(ws,hdrs,set(id_c),set(cp_c),set(sc_c))
wrows(ws,hdrs,set(sc_c),{'Cpn':'0.000','OAS':'0.0','OASD':'0.00','Carry_2.5M_pct':'0.0000',
      'Compression_Score_pct':'0.0000','DP_Rating_Score':'0.0000','Bond_TR_Est_pct':'0.0000','Bond_TR_Score':'0.0000'})
wfin(ws,len(hdrs),{1:12,2:28,3:16,4:10,5:7,6:7,7:7,8:11,9:13,10:16,11:16,12:11,13:13,14:13})

# ── Score_EqMom
print('  Score_EqMom...')
ws=mks('Score_EqMom')
id_c=['class','Des','ISIN','Ticker','Eqty Ticker']
cp_c=['Eq_Ret_1M','Eq_Ret_3M','Eq_Vol_30D','Eq_vs_52w_High']
sc_c=['Eq_Mom_Score']
hdrs=id_c+cp_c+sc_c
wtitle(ws,f'Equity Momentum Score  |  As of {AS_OF}',len(hdrs))
whdrs(ws,hdrs,set(id_c),set(cp_c),set(sc_c))
wrows(ws,hdrs,set(sc_c),{'Eq_Ret_1M':'0.00%','Eq_Ret_3M':'0.00%','Eq_Vol_30D':'0.00%','Eq_vs_52w_High':'0.00%','Eq_Mom_Score':'0.0000'})
wfin(ws,len(hdrs),{1:12,2:28,3:16,4:10,5:12,6:11,7:11,8:11,9:13,10:13})

# ── Score_EqFund
print('  Score_EqFund...')
ws=mks('Score_EqFund')
id_c=['class','Des','ISIN','Ticker','Eqty Ticker']
cp_c=['Debt_to_Equity','Profit_Margin','Revenue_Growth','Current_Ratio','EV_EBITDA','PE_Ratio']
sc_c=['Eq_Fund_Score']
hdrs=id_c+cp_c+sc_c
wtitle(ws,'Equity Fundamental Score  |  D/E Margin Growth CR EV/EBITDA PE -> Eq_Fund_Score [-1,+1]',len(hdrs))
whdrs(ws,hdrs,set(id_c),set(cp_c),set(sc_c))
wrows(ws,hdrs,set(sc_c),{'Debt_to_Equity':'0.00','Profit_Margin':'0.00%','Revenue_Growth':'0.00%',
      'Current_Ratio':'0.00','EV_EBITDA':'0.00','PE_Ratio':'0.00','Eq_Fund_Score':'0.0000'})
wfin(ws,len(hdrs),{1:12,2:28,3:16,4:10,5:12,6:11,7:11,8:11,9:11,10:11,11:11,12:13})

# ── Score_Sentiment
print('  Score_Sentiment...')
ws=mks('Score_Sentiment')
id_c=['class','Des','ISIN','Ticker','Eqty Ticker']
cp_c=['News_Sentiment_Raw','News_Article_Count','Google_News_Count','News_Generic_Flag',
      'Top_Headline','Top_Headline_Score','Trends_Momentum','Trends_Factor','Sentiment_Score','Sentiment_Score_clean']
sc_c=['Sentiment_Score_clean']
hdrs=id_c+cp_c
wtitle(ws,'Sentiment Score  |  Yahoo/Google News VADER + Google Trends',len(hdrs))
whdrs(ws,hdrs,set(id_c),set(cp_c),set(sc_c))
sent_fmt={'News_Sentiment_Raw':'0.00000','News_Article_Count':'0','Google_News_Count':'0',
          'Top_Headline_Score':'0.0000','Trends_Momentum':'0.0000','Trends_Factor':'0.0000',
          'Sentiment_Score':'0.0000','Sentiment_Score_clean':'0.0000'}
sent_fmt_ci={ci:sent_fmt[h] for ci,h in enumerate(hdrs,1) if h in sent_fmt}
flag_ci=next((ci for ci,h in enumerate(hdrs,1) if h=='News_Generic_Flag'),None)
hline_ci=next((ci for ci,h in enumerate(hdrs,1) if h=='Top_Headline'),None)
hlsco_ci=next((ci for ci,h in enumerate(hdrs,1) if h=='Top_Headline_Score'),None)
for ri,row in df_out.iterrows():
    er=ri+3
    for ci,h in enumerate(hdrs,1):
        val=_v(row.get(h,np.nan) if h in row.index else np.nan)
        c=ws.cell(row=er,column=ci); c.value=val
        if ci in sent_fmt_ci: c.number_format=sent_fmt_ci[ci]
        if flag_ci and ci==flag_ci:
            flag=str(val) if val else ''
            if   flag=='GENERIC (invalidated)': c.fill=FILL_GENERIC; c.font=Font(bold=True,color='CC0000',size=10)
            elif flag=='No news data':           c.fill=FILL_NODATA
        if hline_ci and ci==hline_ci: c.alignment=Alignment(wrap_text=True,vertical='center')
        if hlsco_ci and ci==hlsco_ci and val is not None:
            try:
                fv=float(val)
                if   fv>=0.5:  c.fill=PatternFill('solid',fgColor='C6EFCE'); c.font=Font(bold=True,color='375623',size=10)
                elif fv<=-0.5: c.fill=PatternFill('solid',fgColor='FFC7CE'); c.font=Font(bold=True,color='9C0006',size=10)
            except: pass
wfin(ws,len(hdrs),{1:12,2:28,3:16,4:10,5:12,6:13,7:10,8:12,9:22,10:60,11:11,12:13,13:13,14:13,15:13})

# ── Score_AI
print('  Score_AI...')
sys.path.insert(0, BASE)
from ai_macro_score import SUBGROUP_SCORE_MAP as _SUBMAP
_rationale={}
with open(BASE+r'\ai_macro_score.py','r',encoding='utf-8') as f:
    for line in f:
        m2=re.match(r"    '([^']+)':\s*([-\d.]+),\s*#\s*(.+)", line)
        if m2: _rationale[m2.group(1)]=m2.group(3).strip()

ws=mks('Score_AI')
id_c_ai=['class','Des','ISIN','Ticker','OAD','BCLASS3','Industry Subgroup','DPFundamentalRating','Issuer Rtg']
cp_c_ai=['AI_Sector_Score','AI_Maturity_Score','AI_RatingBuf_Score']
sc_c_ai=['AI_Macro_Score']
hdrs_ai=id_c_ai+cp_c_ai+sc_c_ai
ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(hdrs_ai))
tc=ws.cell(row=1,column=1,value=f'AI Macro Score  |  Sector x0.40 + Maturity x0.35 + RatingBuf x0.25  |  As of {AS_OF}')
tc.font=Font(bold=True,size=11,color='FFFFFF'); tc.fill=FILL_AI_H
tc.alignment=Alignment(horizontal='center',vertical='center'); ws.row_dimensions[1].height=22
for ci,h in enumerate(hdrs_ai,1):
    c=ws.cell(row=2,column=ci,value=h); c.font=Font(bold=True,size=10)
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
    c.fill=FILL_AI if h in sc_c_ai else(FILL_COMP if h in cp_c_ai else FILL_ID)
ai_fmt={'OAD':'0.00','AI_Sector_Score':'0.00','AI_Maturity_Score':'0.00','AI_RatingBuf_Score':'0.00','AI_Macro_Score':'0.0000'}
ai_fmt_ci={ci:ai_fmt[h] for ci,h in enumerate(hdrs_ai,1) if h in ai_fmt}
ai_sc_ci=next((ci for ci,h in enumerate(hdrs_ai,1) if h=='AI_Macro_Score'),None)
ai_cp_cis={ci for ci,h in enumerate(hdrs_ai,1) if h in cp_c_ai}
for ri,row in df_out.iterrows():
    er=ri+3
    for ci,h in enumerate(hdrs_ai,1):
        val=_v(row.get(h,np.nan) if h in row.index else np.nan)
        c=ws.cell(row=er,column=ci); c.value=val
        if ci in ai_fmt_ci: c.number_format=ai_fmt_ci[ci]
        if ai_sc_ci and ci==ai_sc_ci and val is not None:
            try:
                fv=float(val)
                c.fill=FILL_POS if fv>0.3 else(FILL_NEG if fv<-0.3 else FILL_AI)
            except: pass
        if ci in ai_cp_cis and val is not None:
            try:
                fv=float(val)
                if fv>0: c.fill=PatternFill('solid',fgColor='E8F5E9')
                elif fv<0: c.fill=PatternFill('solid',fgColor='FFEBEE')
            except: pass
ws.freeze_panes='A3'
ws.auto_filter.ref=f'A2:{get_column_letter(len(hdrs_ai))}2'
for ci,w in {1:12,2:28,3:16,4:10,5:7,6:18,7:26,8:16,9:12,10:14,11:14,12:14,13:14}.items():
    ws.column_dimensions[get_column_letter(ci)].width=w
MV=len(hdrs_ai)+2
MACRO_THEMES=[
    ('① Fed Policy & Tariff','[Current] Fed easing cycle (-100bp) but Trump tariffs reigniting inflation. Core PCE sticky >2.5%.\n[Risk] Stagflation-lite: growth slows, rates stay high.\n[Positioning] Shorter duration, high-carry defensives. Favor utilities, healthcare, infrastructure.'),
    ('② IG Spread Level','[Current] US IG OAS near post-GFC tights (~85-95bp). Asymmetric risk/reward — downside>>upside.\n[Implication] Total return carry-driven. Alpha from carry-to-risk selection within bucket.\n[Positioning] Bond_TR Score: carry (YTW) + DP rating buffer. AI_Macro: avoid spread-widening risk sectors.'),
    ('③ Yield Curve','[Current] Bear-steepening: fiscal deficits push up term premium. 2Y anchored by Fed.\n[Sweet Spot] OAD 4-7Y: carry + spread duration without long-end risk. OAD 13Y+ penalized.\n[Positioning] Maturity_Score: peak +1.0 at OAD 4-7, drops to -0.50/-1.00 for 13Y+.'),
    ('④ Tariff Sectors','[Losers] Autos/Parts (-0.90): 25% tariff + supply chain. Retail Apparel (-0.60): China COGS. Consumer Mfg (-0.55).\n[Winners] Utilities (+0.85): pass-through pricing. Healthcare (+0.75): domestic service. Defense (+0.40): spending tailwinds. Steel (+0.20): tariff protection.\n[Monitor] Semiconductors: US-China export controls ongoing.'),
]
ws.merge_cells(start_row=1,start_column=MV,end_row=1,end_column=MV+3)
mc=ws.cell(row=1,column=MV,value='MACRO VIEW  |  AI Sector Score Rationale  |  April 2026')
mc.font=Font(bold=True,size=10,color='FFFFFF'); mc.fill=FILL_MV_HDR
mc.alignment=Alignment(horizontal='center',vertical='center')
for ti,(label,body) in enumerate(MACRO_THEMES):
    tr=ti+2
    lc=ws.cell(row=tr,column=MV,value=label)
    lc.font=Font(bold=True,size=9,color='FFFFFF'); lc.fill=FILL_MV_THEME
    lc.alignment=Alignment(horizontal='center',vertical='top',wrap_text=True)
    ws.merge_cells(start_row=tr,start_column=MV+1,end_row=tr,end_column=MV+3)
    bc=ws.cell(row=tr,column=MV+1,value=body)
    bc.font=Font(size=8.5); bc.fill=FILL_MV_BG
    bc.alignment=Alignment(vertical='top',wrap_text=True)
mv_sh=6
for ci_off,(h,f) in enumerate([('Industry Subgroup',FILL_ID),('Score',FILL_AI),('Rationale',FILL_COMP),('Direction',FILL_AI)]):
    c=ws.cell(row=mv_sh,column=MV+ci_off,value=h)
    c.font=Font(bold=True,size=9); c.fill=f
    c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
for si,(sg,sc) in enumerate(sorted(_SUBMAP.items(),key=lambda x:x[1],reverse=True)):
    dr=mv_sh+1+si
    direction='▲ Overweight' if sc>0.3 else('▼ Underweight' if sc<-0.3 else '— Neutral')
    rf=FILL_MV_POS if sc>0.3 else(FILL_MV_NEG if sc<-0.3 else FILL_MV_NEUT)
    for co,(val,bold,nf) in enumerate([(sg,False,None),(round(sc,2),True,'0.00'),(_rationale.get(sg,''),False,None),(direction,False,None)]):
        c=ws.cell(row=dr,column=MV+co,value=val); c.fill=rf
        c.font=Font(size=8.5,bold=bold,color=('375623' if sc>0.3 else '9C0006' if sc<-0.3 else '555555') if co==3 else '000000')
        c.alignment=Alignment(horizontal='center' if co in(1,3) else 'left',vertical='center',wrap_text=(co==2))
        if nf: c.number_format=nf
ws.column_dimensions[get_column_letter(MV)].width=30
ws.column_dimensions[get_column_letter(MV+1)].width=7
ws.column_dimensions[get_column_letter(MV+2)].width=80
ws.column_dimensions[get_column_letter(MV+3)].width=15
for ti,rh in enumerate([52,60,60,100]): ws.row_dimensions[ti+2].height=rh
ws.row_dimensions[mv_sh].height=18

# ── Score_Integrated
print('  Score_Integrated...')
ws=mks('Score_Integrated')
id_c=['class','Des','ISIN','Ticker','Cpn','OAS','LQA','Issuer Rtg','BCLASS3','Industry Subgroup']
cp_c=['Bond_TR_Score','Eq_Mom_Score','Eq_Fund_Score','Sentiment_Score','AI_Macro_Score']
sc_c=['Integrated_Score','Integrated_Rank_in_Class','Top_Pick_Flag']
hdrs=id_c+cp_c+sc_c
wtitle(ws,f'Integrated Score  |  5 Components x0.20  |  As of {AS_OF}',len(hdrs))
whdrs(ws,hdrs,set(id_c),set(cp_c),set(sc_c))
wrows(ws,hdrs,set(sc_c),
      {'Cpn':'0.000','OAS':'0.0','LQA':'0.0','Bond_TR_Score':'0.0000','Eq_Mom_Score':'0.0000',
       'Eq_Fund_Score':'0.0000','Sentiment_Score':'0.0000','AI_Macro_Score':'0.0000',
       'Integrated_Score':'0.0000','Integrated_Rank_in_Class':'0'},
      flag_col='Top_Pick_Flag',
      color_cols={'Bond_TR_Score','Eq_Mom_Score','Eq_Fund_Score','Sentiment_Score','AI_Macro_Score'})
wfin(ws,len(hdrs),{1:12,2:28,3:16,4:10,5:7,6:7,7:8,8:10,9:14,10:26,11:13,12:13,13:13,14:13,15:13,16:13,17:10,18:14})

# ── Detail_Scored (서식 없이 값만)
print('  Detail_Scored...')
ws=mks('Detail_Scored')
keep_cols=[c for c in df_out.columns if not str(c).startswith('_')]
wtitle(ws,f'LUACSTAT Detail  |  As of {AS_OF}  |  {len(df_out):,} Active Bonds',len(keep_cols))
for ci,h in enumerate(keep_cols,1):
    c=ws.cell(row=2,column=ci,value=h)
    c.font=Font(bold=True,size=9); c.fill=FILL_ID
    c.alignment=Alignment(horizontal='center',wrap_text=True)
for ri,row in df_out.iterrows():
    for ci,h in enumerate(keep_cols,1):
        ws.cell(row=ri+3,column=ci,value=_v(row.get(h,np.nan)))
ws.freeze_panes='A3'
ws.auto_filter.ref=f'A2:{get_column_letter(len(keep_cols))}2'

print('Saving xlsx...')
wb.save(OUT)
print(f'  xlsx: {os.path.getsize(OUT)/1024/1024:.1f} MB  →  converting to .xlsb ...')

try:
    import win32com.client as _win32
    _xl = _win32.Dispatch('Excel.Application')
    _xl.Visible = False
    _xl.DisplayAlerts = False
    _wb = _xl.Workbooks.Open(os.path.abspath(OUT))
    _wb.SaveAs(os.path.abspath(OUT_BSB), FileFormat=50)
    _wb.Close(False)
    _xl.Quit()
    os.remove(OUT)
    OUT = OUT_BSB
    print(f'  xlsb: {os.path.getsize(OUT)/1024/1024:.1f} MB')
except Exception as e:
    print(f'  [경고] xlsb 변환 실패 ({e}) → xlsx 유지')

print(f'Saved: {OUT}')
print('DONE')
