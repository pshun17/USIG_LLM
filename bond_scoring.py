import pandas as pd
import numpy as np
import time
import os
import json
import datetime
import warnings
import requests
import urllib3
urllib3.disable_warnings()
warnings.filterwarnings('ignore')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT_FILE = 'C:/Users/sh.park/Documents/USIG_LLM/LUACSTAT_2026_03_31.xlsx'
OUTPUT_FILE = 'C:/Users/sh.park/Documents/USIG_LLM/LUACSTAT_2026_03_31_SCORED.xlsx'

SESSION = requests.Session()
SESSION.verify = False
SESSION.headers.update({
    'User-Agent': ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                   'AppleWebKit/537.36 (KHTML, like Gecko) '
                   'Chrome/122.0.0.0 Safari/537.36'),
    'Accept': 'application/json',
    'Accept-Language': 'en-US,en;q=0.9',
})

# ── STEP 1: Load ──────────────────────────────────────────────────────────────
print("Loading data...")
df = pd.read_excel(INPUT_FILE, sheet_name='Detail_1', header=7)
print(f"  Loaded {len(df)} rows x {len(df.columns)} columns")

for col in ['OAS', 'OASD', 'OAD', 'Cpn', 'Yield to Worst', '1Y Dflt']:
    df[col] = pd.to_numeric(df[col], errors='coerce')

# ── STEP 2: Bond-Level Scoring ─────────────────────────────────────────────────
print("Computing bond-level scores...")

mask = (df['class'] != 'off') & df['OAS'].notna() & df['OASD'].notna()

# A. Carry Score  (YTM-based: Yield to Worst / 12 * 2.5)
carry_mask = mask & df['Yield to Worst'].notna()
df.loc[carry_mask, 'Carry_2.5M_pct'] = (
    df.loc[carry_mask, 'Yield to Worst'] / 12 * 2.5
)

# B. Class-level OAS stats
def class_oas_stats(group):
    oas = group['OAS']
    group = group.copy()
    group['Class_OAS_Median'] = oas.median()
    group['Class_OAS_Q25'] = oas.quantile(0.25)
    group['Class_OAS_Q75'] = oas.quantile(0.75)
    group['Class_OAS_Pctile'] = oas.rank(pct=True) * 100
    group['Spread_vs_Class_Median'] = oas - oas.median()
    return group

df_active = df[mask].copy()
df_active = df_active.groupby('class', group_keys=False).apply(class_oas_stats)

for col in ['Class_OAS_Median', 'Class_OAS_Q25', 'Class_OAS_Q75',
            'Class_OAS_Pctile', 'Spread_vs_Class_Median']:
    df[col] = np.nan
    df.loc[df_active.index, col] = df_active[col].values

# C. Compression Score  (1Y Default Probability implied spread vs OAS)
# Spread_floor(bp) = 1Y_Dflt(%) × (1 - Recovery_40%) × 100 = 1Y_Dflt × 60
# Compression_gap  = max(0, OAS - Spread_floor)  → OAS wider than default-justified = undervalued
comp_mask = mask & df['1Y Dflt'].notna()
df.loc[comp_mask, 'Spread_floor_bp'] = df.loc[comp_mask, '1Y Dflt'] * 60
df.loc[comp_mask, 'Compression_gap_bp'] = (
    df.loc[comp_mask, 'OAS'] - df.loc[comp_mask, 'Spread_floor_bp']
)
df.loc[comp_mask, 'Compression_Score_pct'] = (
    df.loc[comp_mask, 'Compression_gap_bp'] * df.loc[comp_mask, 'OASD'] / 100
)

# D. DP Rating Score (Bloomberg DPFundamentalRating + DPSpreadRating, replaces Outlook_Score)
DP_RATING_MAP = {
    'AAA': 1, 'AA1': 2, 'AA2': 3, 'AA3': 4,
    'A1': 5, 'A2': 6, 'A3': 7,
    'BAA1': 8, 'BAA2': 9, 'BAA3': 10,
    'BA1': 11, 'BA2': 12, 'BA3': 13,
    'B1': 14, 'B2': 15, 'B3': 16,
    'CAA1': 17, 'CAA2': 18, 'CAA3': 19,
    'CA': 20, 'C': 21,
}

df['_dp_fund_num'] = df['DPFundamentalRating'].map(DP_RATING_MAP)
df['_dp_spd_num']  = df['DPSpreadRating'].map(DP_RATING_MAP)

# 갭 = DPSpreadRating - DPFundamentalRating
# 양수 → 시장이 펀더멘털보다 나쁘게 pricing → 저평가, 스프레드 압축 여지 → 좋은 신호
# 음수 → 시장이 펀더멘털보다 좋게 pricing → 고평가 → 나쁜 신호
df['_dp_gap'] = df['_dp_spd_num'] - df['_dp_fund_num']

def _gap_norm(series):
    """갭 클수록 저평가 → 높은 점수. scaled to [-1, +1]."""
    n = series.notna().sum()
    if n < 2:
        return pd.Series(np.nan, index=series.index)
    ranked = series.rank(method='average', na_option='keep')
    return (ranked - 1) / (n - 1) * 2 - 1

df['DP_Rating_Score'] = np.nan
df.loc[mask, 'DP_Rating_Score'] = _gap_norm(df.loc[mask, '_dp_gap']).values
df.drop(columns=['_dp_fund_num', '_dp_spd_num', '_dp_gap'], inplace=True)

# E. Bond TR Estimate (DP Rating Score replaces Outlook_Score)
df.loc[mask, 'Bond_TR_Est_pct'] = (
    df.loc[mask, 'Carry_2.5M_pct'].fillna(0) +
    df.loc[mask, 'Compression_Score_pct'].fillna(0) +
    df.loc[mask, 'DP_Rating_Score'].fillna(0) * 0.05
)

# ── STEP 3: Equity Data via Yahoo Finance API ──────────────────────────────────
print("Fetching equity tickers...")

def clean_ticker(t):
    if pd.isna(t):
        return None
    t = str(t).strip()
    if t in ('', 'N/A', 'nan'):
        return None
    parts = t.split()
    return parts[0] if parts else None

df['_eq_ticker_clean'] = df['Eqty Ticker'].map(clean_ticker)
unique_tickers = [t for t in df['_eq_ticker_clean'].dropna().unique() if t]
print(f"  {len(unique_tickers)} unique equity tickers")

today = datetime.date.today()
six_mo_ago_ts = int((datetime.datetime.combine(
    today - datetime.timedelta(days=183),
    datetime.time()) ).timestamp())
today_ts = int(datetime.datetime.combine(today, datetime.time()).timestamp())

def fetch_price_history(ticker):
    """Fetch 6-month price history from Yahoo Finance."""
    url = (f'https://query2.finance.yahoo.com/v8/finance/chart/{ticker}'
           f'?period1={six_mo_ago_ts}&period2={today_ts}&interval=1d'
           f'&events=history&includeAdjustedClose=true')
    try:
        r = SESSION.get(url, timeout=15)
        if r.status_code != 200:
            return None
        data = r.json()
        result = data.get('chart', {}).get('result')
        if not result:
            return None
        res = result[0]
        timestamps = res.get('timestamp', [])
        closes = res.get('indicators', {}).get('adjclose', [{}])[0].get('adjclose', [])
        if not timestamps or not closes:
            closes = res.get('indicators', {}).get('quote', [{}])[0].get('close', [])
        if not timestamps or not closes:
            return None
        dates = pd.to_datetime(timestamps, unit='s', utc=True).tz_convert(None)
        prices = pd.Series(closes, index=dates, dtype=float).dropna()
        return prices
    except Exception:
        return None

def fetch_fundamentals(ticker):
    """Fetch key financial ratios from Yahoo Finance quoteSummary."""
    url = (f'https://query2.finance.yahoo.com/v10/finance/quoteSummary/{ticker}'
           f'?modules=financialData,defaultKeyStatistics,summaryDetail')
    try:
        r = SESSION.get(url, timeout=15)
        if r.status_code != 200:
            return {}
        data = r.json()
        result = data.get('quoteSummary', {}).get('result')
        if not result:
            return {}
        res = result[0]
        fd = res.get('financialData', {})
        ks = res.get('defaultKeyStatistics', {})
        metrics = {}
        def g(d, key):
            v = d.get(key)
            if isinstance(v, dict):
                v = v.get('raw')
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                return float(v)
            return None
        for src, key, col in [
            (fd, 'debtToEquity', 'Debt_to_Equity'),
            (fd, 'profitMargins', 'Profit_Margin'),
            (fd, 'revenueGrowth', 'Revenue_Growth'),
            (fd, 'currentRatio', 'Current_Ratio'),
            (ks, 'enterpriseToEbitda', 'EV_EBITDA'),
            (ks, 'trailingPE', 'PE_Ratio'),
        ]:
            v = g(src, key)
            if v is not None:
                metrics[col] = v
        return metrics
    except Exception:
        return {}

EQ_COLS = ['Eq_Ret_1M', 'Eq_Ret_3M', 'Eq_Vol_30D', 'Eq_vs_52w_High',
           'Debt_to_Equity', 'Profit_Margin', 'Revenue_Growth',
           'Current_Ratio', 'EV_EBITDA', 'PE_Ratio']

one_mo_ago = pd.Timestamp(today - datetime.timedelta(days=31))
three_mo_ago = pd.Timestamp(today - datetime.timedelta(days=92))

eq_data = {}
BATCH_SIZE = 50
batches = [unique_tickers[i:i+BATCH_SIZE]
           for i in range(0, len(unique_tickers), BATCH_SIZE)]
print(f"  Fetching in {len(batches)} batches of {BATCH_SIZE}...")

for b_idx, batch in enumerate(batches):
    if b_idx % 10 == 0:
        print(f"    Batch {b_idx+1}/{len(batches)}...")
    for ticker in batch:
        metrics = {}
        try:
            prices = fetch_price_history(ticker)
            if prices is not None and len(prices) >= 2:
                cur = float(prices.iloc[-1])
                # 1M return
                p_1m = prices[prices.index <= one_mo_ago]
                if len(p_1m) > 0:
                    metrics['Eq_Ret_1M'] = float(cur / p_1m.iloc[-1] - 1)
                # 3M return
                p_3m = prices[prices.index <= three_mo_ago]
                if len(p_3m) > 0:
                    metrics['Eq_Ret_3M'] = float(cur / p_3m.iloc[-1] - 1)
                # 30D realized vol
                log_ret = np.log(prices / prices.shift(1)).dropna()
                if len(log_ret) >= 20:
                    metrics['Eq_Vol_30D'] = float(
                        log_ret.iloc[-30:].std() * np.sqrt(252))
                # vs 52w high
                hi_52w = float(prices.max())
                if hi_52w > 0:
                    metrics['Eq_vs_52w_High'] = float(cur / hi_52w - 1)
        except Exception:
            pass
        try:
            fund = fetch_fundamentals(ticker)
            metrics.update(fund)
        except Exception:
            pass
        eq_data[ticker] = metrics
    time.sleep(1)

n_with_data = sum(1 for v in eq_data.values() if v)
print(f"  Equity data fetched for {n_with_data}/{len(unique_tickers)} tickers")

# ── Build equity DataFrame ────────────────────────────────────────────────────
eq_rows = []
for t in unique_tickers:
    row = {'_eq_ticker_clean': t}
    row.update(eq_data.get(t, {}))
    eq_rows.append(row)
eq_df = pd.DataFrame(eq_rows)

for c in EQ_COLS:
    if c not in eq_df.columns:
        eq_df[c] = np.nan

# ── STEP 4: Equity Scoring ────────────────────────────────────────────────────
print("Computing equity scores...")

def rank_normalize(series, ascending=True):
    s = pd.to_numeric(series, errors='coerce')
    valid = s.notna()
    result = pd.Series(np.nan, index=s.index)
    if valid.sum() < 2:
        return result
    ranked = s[valid].rank(ascending=ascending, method='average')
    n = valid.sum()
    result[valid] = (ranked / n) * 2 - 1
    return result

def rank_normalize_01(series, ascending=True):
    s = pd.to_numeric(series, errors='coerce')
    valid = s.notna()
    result = pd.Series(np.nan, index=s.index)
    if valid.sum() < 2:
        return result
    ranked = s[valid].rank(ascending=ascending, method='average')
    n = valid.sum()
    result[valid] = ranked / n
    return result

n1 = rank_normalize(eq_df['Eq_Ret_1M'], ascending=True)
n3 = rank_normalize(eq_df['Eq_Ret_3M'], ascending=True)
nv = rank_normalize(eq_df['Eq_Vol_30D'], ascending=False)
nh = rank_normalize(eq_df['Eq_vs_52w_High'], ascending=True)
mom_components = pd.DataFrame({'n1': n1, 'n3': n3, 'nv': nv, 'nh': nh})
eq_df['Eq_Mom_Score'] = mom_components.mean(axis=1)

fd_series = rank_normalize_01(eq_df['Debt_to_Equity'], ascending=False)
fp = rank_normalize_01(eq_df['Profit_Margin'], ascending=True)
fg = rank_normalize_01(eq_df['Revenue_Growth'], ascending=True)
fc = rank_normalize_01(eq_df['Current_Ratio'], ascending=True)
ev_clean = eq_df['EV_EBITDA'].copy()
ev_clean[ev_clean <= 0] = np.nan
fe = rank_normalize_01(ev_clean, ascending=False)
fund_components = pd.DataFrame({'fd': fd_series, 'fp': fp, 'fg': fg,
                                'fc': fc, 'fe': fe})
fund_01 = fund_components.mean(axis=1)
eq_df['Eq_Fund_Score'] = fund_01 * 2 - 1

# ── Map equity scores back to bonds ──────────────────────────────────────────
eq_map = eq_df.set_index('_eq_ticker_clean')

for col in EQ_COLS + ['Eq_Mom_Score', 'Eq_Fund_Score']:
    if col in eq_map.columns:
        df[col] = df['_eq_ticker_clean'].map(eq_map[col])
    else:
        df[col] = np.nan

df.drop(columns=['_eq_ticker_clean'], inplace=True)

# ── STEP 5: Integrated Score ──────────────────────────────────────────────────
print("Computing integrated scores...")

# Normalize Bond_TR_Est_pct → [-1, +1] within each class
df['Bond_TR_Score'] = np.nan
for cls in df.loc[mask, 'class'].dropna().unique():
    cls_mask = mask & (df['class'] == cls)
    tr_cls = df.loc[cls_mask, 'Bond_TR_Est_pct']
    n_tr = tr_cls.notna().sum()
    if n_tr > 1:
        tr_ranked = tr_cls.rank(method='average', na_option='keep')
        df.loc[cls_mask, 'Bond_TR_Score'] = (tr_ranked - 1) / (n_tr - 1) * 2 - 1

# Equal 0.25 weight per component; Sentiment (0.25) will be added in sentiment_update.py
df['Integrated_Score'] = np.nan
df.loc[mask, 'Integrated_Score'] = (
    df.loc[mask, 'Bond_TR_Score'].fillna(0) * 0.25 +
    df.loc[mask, 'Eq_Mom_Score'].fillna(0) * 0.25 +
    df.loc[mask, 'Eq_Fund_Score'].fillna(0) * 0.25
)

df['Bond_TR_Rank_in_Class'] = np.nan
df['Integrated_Rank_in_Class'] = np.nan

for cls, grp in df[mask].groupby('class'):
    for score_col, rank_col in [
        ('Bond_TR_Est_pct', 'Bond_TR_Rank_in_Class'),
        ('Integrated_Score', 'Integrated_Rank_in_Class'),
    ]:
        valid = grp[score_col].notna()
        if valid.any():
            ranked = grp.loc[valid, score_col].rank(ascending=False, method='min')
            df.loc[ranked.index, rank_col] = ranked.values

def top_pick_flag(rank):
    if pd.isna(rank):
        return ''
    r = int(rank)
    if r <= 3:
        return '★★★ TOP3'
    elif r <= 10:
        return '★★ TOP10'
    elif r <= 25:
        return '★ TOP25'
    return ''

df['Top_Pick_Flag'] = df['Integrated_Rank_in_Class'].map(top_pick_flag)

# ── STEP 6: Write Excel Output ────────────────────────────────────────────────
print("Writing Excel output...")

NEW_COLS = [
    'Carry_2.5M_pct', 'Class_OAS_Median', 'Class_OAS_Q25', 'Class_OAS_Q75',
    'Class_OAS_Pctile', 'Spread_vs_Class_Median', 'Compression_Score_pct',
    'DP_Rating_Score', 'Bond_TR_Est_pct', 'Bond_TR_Score',
    'Eq_Ret_1M', 'Eq_Ret_3M', 'Eq_Vol_30D', 'Eq_vs_52w_High', 'Eq_Mom_Score',
    'Debt_to_Equity', 'Profit_Margin', 'Revenue_Growth', 'Current_Ratio',
    'EV_EBITDA', 'PE_Ratio', 'Eq_Fund_Score',
    'Integrated_Score', 'Bond_TR_Rank_in_Class', 'Integrated_Rank_in_Class',
    'Top_Pick_Flag'
]

ORIG_COLS = [c for c in df.columns if c not in NEW_COLS]
all_cols = ORIG_COLS + NEW_COLS
df_out = df[all_cols].copy()

def to_py(v):
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return float(v)
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    return v

DARK_BLUE = '1F3864'
LIGHT_BLUE_HDR = 'BDD7EE'
LIGHT_GREEN_HDR = 'E2EFDA'
BLUE_HDR2 = '4472C4'

def make_fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def center_align(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def left_align(wrap=False):
    return Alignment(horizontal='left', vertical='center', wrap_text=wrap)

wb = Workbook()

# ── Sheet 1: Detail_Scored ────────────────────────────────────────────────────
print("  Writing Detail_Scored...")
ws1 = wb.active
ws1.title = 'Detail_Scored'

total_cols = len(all_cols)

# Row 1: Title
title_text = ('Bloomberg US IG Corporate Bond — Total Return Scoring Model '
              '(2-3M Horizon) | As of 2026-03-31')
ws1.append([title_text] + [''] * (total_cols - 1))
ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
t_cell = ws1.cell(1, 1)
t_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
t_cell.fill = make_fill(DARK_BLUE)
t_cell.alignment = center_align()
ws1.row_dimensions[1].height = 22

# Row 2: Headers
ws1.append(all_cols)
for c_idx, col_name in enumerate(all_cols, 1):
    cell = ws1.cell(2, c_idx)
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.alignment = center_align(wrap=True)
    cell.fill = make_fill(LIGHT_GREEN_HDR if col_name in NEW_COLS
                          else LIGHT_BLUE_HDR)

ws1.freeze_panes = 'A3'
ws1.auto_filter.ref = f'A2:{get_column_letter(total_cols)}2'

pct_cols = {'Carry_2.5M_pct', 'Compression_Score_pct', 'Bond_TR_Est_pct',
            'Eq_Ret_1M', 'Eq_Ret_3M', 'Eq_Vol_30D', 'Eq_vs_52w_High',
            'Integrated_Score', 'Profit_Margin', 'Revenue_Growth',
            'Eq_Mom_Score', 'Eq_Fund_Score', 'Outlook_Score'}
spread_cols = {'OAS', 'OASD', 'OAD', 'Spd', 'Z-Spd',
               'Class_OAS_Median', 'Class_OAS_Q25', 'Class_OAS_Q75',
               'Spread_vs_Class_Median', 'Class_OAS_Pctile',
               'Mkt Val', 'Cpn', 'Yield to Mat', 'Yield to Worst',
               'Mod Dur to Worst', 'Mty (Yrs)', 'Debt_to_Equity',
               'Current_Ratio', 'EV_EBITDA', 'PE_Ratio'}
rank_cols = {'Bond_TR_Rank_in_Class', 'Integrated_Rank_in_Class'}

COL_WIDTHS = {
    'Des': 30, 'ISIN': 16, 'class': 12, 'Top_Pick_Flag': 14,
    'Company Name': 20, 'Ticker': 10, 'Eqty Ticker': 12,
    'Parent Ticker': 12, 'BCLASS2': 14, 'BCLASS3': 14, 'BCLASS4': 14,
    'Industry Sector': 16, 'Industry Group': 16,
}

col_idx_map = {col: i+1 for i, col in enumerate(all_cols)}
for col_name, c_idx in col_idx_map.items():
    ltr = get_column_letter(c_idx)
    ws1.column_dimensions[ltr].width = COL_WIDTHS.get(
        col_name,
        12 if col_name in (pct_cols | spread_cols | rank_cols) else 14)

FLAG_STYLES = {
    '★★★ TOP3':  {'font_color': 'FF0000', 'bg': 'FFFF00'},
    '★★ TOP10':  {'font_color': 'C55A11', 'bg': 'FCE4D6'},
    '★ TOP25':   {'font_color': '1F3864', 'bg': 'DDEBF7'},
}

print("  Writing data rows...")
flag_col_name = 'Top_Pick_Flag'
for r_idx, row in enumerate(df_out.itertuples(index=False), 3):
    row_vals = list(row)
    for c_idx, (col_name, val) in enumerate(zip(all_cols, row_vals), 1):
        cell = ws1.cell(r_idx, c_idx)
        cell.value = to_py(val)
        if col_name in pct_cols:
            cell.number_format = '0.0000%'
        elif col_name in spread_cols:
            cell.number_format = '0.00'
        elif col_name in rank_cols:
            cell.number_format = '0'
        if col_name == flag_col_name and str(val) in FLAG_STYLES:
            st = FLAG_STYLES[str(val)]
            cell.font = Font(name='Arial', size=10, bold=True,
                             color=st['font_color'])
            cell.fill = make_fill(st['bg'])
    if r_idx % 1000 == 0:
        print(f"    Written {r_idx - 2} rows...")

# ── Sheet 2: Equity_Data ──────────────────────────────────────────────────────
print("  Writing Equity_Data...")
ws2 = wb.create_sheet('Equity_Data')

ticker_to_company = {}
for ticker in unique_tickers:
    match = df['Eqty Ticker'].apply(
        lambda x: str(x).strip().split()[0] if pd.notna(x) else '') == ticker
    if match.any():
        ticker_to_company[ticker] = df.loc[match, 'Company Name'].iloc[0]

eq_summary_rows = []
for ticker in unique_tickers:
    row = {'Ticker': ticker,
           'Company Name': ticker_to_company.get(ticker, '')}
    d = eq_data.get(ticker, {})
    for c in EQ_COLS:
        row[c] = d.get(c, np.nan)
    # Scores
    t_mask = eq_df['_eq_ticker_clean'] == ticker if '_eq_ticker_clean' in eq_df.columns else pd.Series(False, index=eq_df.index)
    if t_mask.any():
        row['Eq_Mom_Score'] = to_py(eq_df.loc[t_mask, 'Eq_Mom_Score'].values[0])
        row['Eq_Fund_Score'] = to_py(eq_df.loc[t_mask, 'Eq_Fund_Score'].values[0])
    else:
        row['Eq_Mom_Score'] = None
        row['Eq_Fund_Score'] = None
    row['Data_Available'] = 'Yes' if any(not pd.isna(v) if isinstance(v, float) else (v is not None) for v in d.values()) else 'No'
    eq_summary_rows.append(row)

eq_summary = pd.DataFrame(eq_summary_rows)
eq_cols_order = ['Ticker', 'Company Name', 'Eq_Ret_1M', 'Eq_Ret_3M',
                 'Eq_Vol_30D', 'Eq_vs_52w_High', 'Debt_to_Equity',
                 'Profit_Margin', 'Revenue_Growth', 'Current_Ratio',
                 'EV_EBITDA', 'PE_Ratio', 'Eq_Mom_Score', 'Eq_Fund_Score',
                 'Data_Available']
eq_summary = eq_summary[[c for c in eq_cols_order if c in eq_summary.columns]]
eq_summary = eq_summary.sort_values('Eq_Mom_Score', ascending=False,
                                    na_position='last')

ws2.append(eq_cols_order)
for c_idx, col_name in enumerate(eq_cols_order, 1):
    cell = ws2.cell(1, c_idx)
    cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    cell.fill = make_fill(BLUE_HDR2)
    cell.alignment = center_align()
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f'A1:{get_column_letter(len(eq_cols_order))}1'

eq_pct_cols = {'Eq_Ret_1M', 'Eq_Ret_3M', 'Eq_Vol_30D', 'Eq_vs_52w_High',
               'Profit_Margin', 'Revenue_Growth', 'Eq_Mom_Score', 'Eq_Fund_Score'}
eq_dec_cols = {'Debt_to_Equity', 'Current_Ratio', 'EV_EBITDA', 'PE_Ratio'}

for r_idx, row in enumerate(eq_summary.itertuples(index=False), 2):
    for c_idx, (col_name, val) in enumerate(zip(eq_cols_order, list(row)), 1):
        cell = ws2.cell(r_idx, c_idx)
        cell.value = to_py(val)
        if col_name in eq_pct_cols:
            cell.number_format = '0.0000%'
        elif col_name in eq_dec_cols:
            cell.number_format = '0.00'

for i, col_name in enumerate(eq_cols_order, 1):
    ws2.column_dimensions[get_column_letter(i)].width = (
        18 if col_name == 'Company Name' else 14)

# ── Sheet 3: Top_Picks_by_Class ───────────────────────────────────────────────
print("  Writing Top_Picks_by_Class...")
ws3 = wb.create_sheet('Top_Picks_by_Class')

TOP_COLS = ['class', 'Des', 'ISIN', 'Ticker', 'Cpn', 'OAS', 'OASD',
            'Carry_2.5M_pct', 'Compression_Score_pct', 'Bond_TR_Est_pct',
            'Eq_Ret_1M', 'Eq_Ret_3M', 'Eq_Mom_Score', 'Eq_Fund_Score',
            'Integrated_Score', 'Integrated_Rank_in_Class', 'Top_Pick_Flag',
            'Issuer Rtg', "S&P Outlook", "Moody's Outlook",
            'BCLASS3', 'Industry Sector']

ws3.append([('Top 5 Integrated Score Candidates by Class '
             '(2-3M Horizon) | As of 2026-03-31')] +
           [''] * (len(TOP_COLS) - 1))
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(TOP_COLS))
t3 = ws3.cell(1, 1)
t3.font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
t3.fill = make_fill(DARK_BLUE)
t3.alignment = center_align()
ws3.row_dimensions[1].height = 24

ws3.append(TOP_COLS)
for c_idx, col_name in enumerate(TOP_COLS, 1):
    cell = ws3.cell(2, c_idx)
    cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    cell.fill = make_fill(BLUE_HDR2)
    cell.alignment = center_align(wrap=True)
ws3.freeze_panes = 'A3'
ws3.auto_filter.ref = f'A2:{get_column_letter(len(TOP_COLS))}2'

top_picks_frames = []
for cls in sorted(df[mask]['class'].unique()):
    grp = df[(df['class'] == cls) & mask].copy()
    grp = grp.sort_values('Integrated_Rank_in_Class')
    top_picks_frames.append(grp.head(5))

if top_picks_frames:
    top_df = pd.concat(top_picks_frames, ignore_index=True)
    for c in TOP_COLS:
        if c not in top_df.columns:
            top_df[c] = np.nan

    classes_in_order = top_df['class'].unique()
    class_shade = {cls: (i % 2 == 0) for i, cls in enumerate(classes_in_order)}
    GREY_FILL = make_fill('F2F2F2')
    WHITE_FILL = make_fill('FFFFFF')

    top_pct_cols = {'Carry_2.5M_pct', 'Compression_Score_pct', 'Bond_TR_Est_pct',
                    'Eq_Ret_1M', 'Eq_Ret_3M', 'Eq_Mom_Score', 'Eq_Fund_Score',
                    'Integrated_Score'}
    top_dec_cols = {'OAS', 'OASD', 'Cpn'}
    top_rank_cols = {'Integrated_Rank_in_Class'}

    for r_idx, row in enumerate(top_df[TOP_COLS].itertuples(index=False), 3):
        row_vals = list(row)
        cls_val = row_vals[0]
        use_grey = class_shade.get(cls_val, False)
        base_fill = GREY_FILL if use_grey else WHITE_FILL
        for c_idx, (col_name, val) in enumerate(zip(TOP_COLS, row_vals), 1):
            cell = ws3.cell(r_idx, c_idx)
            cell.value = to_py(val)
            if col_name in top_pct_cols:
                cell.number_format = '0.0000%'
            elif col_name in top_dec_cols:
                cell.number_format = '0.00'
            elif col_name in top_rank_cols:
                cell.number_format = '0'
            if col_name == 'Top_Pick_Flag' and str(val) in FLAG_STYLES:
                st = FLAG_STYLES[str(val)]
                cell.font = Font(name='Arial', size=10, bold=True,
                                 color=st['font_color'])
                cell.fill = make_fill(st['bg'])
            else:
                cell.fill = base_fill

TOP_WIDTHS = {
    'class': 14, 'Des': 30, 'ISIN': 16, 'Ticker': 10, 'Cpn': 10,
    'OAS': 10, 'OASD': 10, 'Top_Pick_Flag': 14,
    'Industry Sector': 16, 'BCLASS3': 14, 'Issuer Rtg': 12,
}
for i, col_name in enumerate(TOP_COLS, 1):
    ws3.column_dimensions[get_column_letter(i)].width = TOP_WIDTHS.get(col_name, 14)

# ── Sheet 4: Methodology ──────────────────────────────────────────────────────
print("  Writing Methodology...")
ws4 = wb.create_sheet('Methodology')

METHODOLOGY = [
    ("BLOOMBERG US IG CORPORATE BOND — SCORING METHODOLOGY", True),
    ("", False),
    ("DATA SOURCE", True),
    ("Input file: Bloomberg LUAC Index constituents as of 2026-03-31", False),
    ("Universe: 8,704 US Investment Grade Corporate Bond index members", False),
    ("Equity data: Yahoo Finance (fetched at script runtime)", False),
    ("", False),
    ("CARRY CALCULATION (Carry_2.5M_pct)", True),
    ("Formula: OAS x OASD / 12 x 2.5 / 100", False),
    ("Interpretation: Estimated carry in % of face value over a 2.5-month horizon,", False),
    ("based on option-adjusted spread and spread duration.", False),
    ("", False),
    ("COMPRESSION POTENTIAL (Compression_Score_pct)", True),
    ("Formula: max(0, Bond_OAS - Class_OAS_Median) x OASD / 100", False),
    ("Interpretation: Estimated price gain if the bond's spread compresses to the", False),
    ("median spread within its class group. Zero for bonds already tighter than median.", False),
    ("", False),
    ("OUTLOOK SCORE (Outlook_Score)", True),
    ("S&P Outlook: POS=+1, STABLE=0, NEG=-1", False),
    ("Moody's Outlook: POS=+1, STABLE=0, NEG=-1", False),
    ("Outlook_Score = average of S&P and Moody's scores. Range: [-1, +1]", False),
    ("Adjustment: adds/subtracts up to 5bps to Bond_TR_Est_pct", False),
    ("", False),
    ("BOND TOTAL RETURN ESTIMATE (Bond_TR_Est_pct)", True),
    ("Formula: Carry_2.5M_pct + Compression_Score_pct + Outlook_Score x 0.05", False),
    ("", False),
    ("EQUITY MOMENTUM SIGNALS (Eq_Mom_Score)", True),
    ("Components (each normalized to [-1,+1] via rank percentile):", False),
    ("  - 1-month equity return (higher = better)", False),
    ("  - 3-month equity return (higher = better)", False),
    ("  - 30-day realized annualized volatility (lower = better)", False),
    ("  - Price vs 52-week high (higher = better)", False),
    ("Eq_Mom_Score = simple average of 4 components. Range: [-1, +1]", False),
    ("", False),
    ("FUNDAMENTAL SIGNALS (Eq_Fund_Score)", True),
    ("Components (each normalized to [0,1] via rank percentile, then rescaled to [-1,+1]):", False),
    ("  - Debt-to-Equity ratio (lower = better)", False),
    ("  - Profit Margin (higher = better)", False),
    ("  - Revenue Growth (higher = better)", False),
    ("  - Current Ratio (higher = better)", False),
    ("  - EV/EBITDA (lower = better, negatives excluded)", False),
    ("Eq_Fund_Score = average of available components, rescaled to [-1, +1]", False),
    ("", False),
    ("INTEGRATED SCORE WEIGHTS", True),
    ("Integrated_Score = Bond_TR_Est_pct x 0.60", False),
    ("                 + Eq_Mom_Score x 0.025", False),
    ("                 + Eq_Fund_Score x 0.015", False),
    ("Bonds without equity data: equity components treated as 0", False),
    ("", False),
    ("RANKINGS & FLAGS", True),
    ("Integrated_Rank_in_Class: descending rank within each class group (1=best)", False),
    ("Bond_TR_Rank_in_Class: descending rank by Bond_TR_Est_pct within class", False),
    ("Top_Pick_Flag: ★★★ TOP3 (rank <=3), ★★ TOP10 (rank <=10), ★ TOP25 (rank <=25)", False),
    ("", False),
    ("HOW TO USE THE OUTPUT", True),
    ("1. 'Detail_Scored': Full universe with all computed metrics. Filter by class,", False),
    ("   Top_Pick_Flag, or sort by Integrated_Score to identify candidates.", False),
    ("2. 'Top_Picks_by_Class': Pre-filtered top 5 per class for quick review.", False),
    ("3. 'Equity_Data': Underlying equity fundamentals and momentum for reference.", False),
    ("4. All numeric values are hardcoded from Python — no live Excel formulas.", False),
    ("5. Horizon: 2-3 months. Scores reflect conditions as of the input data date", False),
    ("   and equity data fetch date.", False),
    ("", False),
    ("DISCLAIMERS", True),
    ("This model is for informational purposes only. Past performance does not", False),
    ("guarantee future results. Scores are quantitative estimates and should be", False),
    ("combined with fundamental analysis and portfolio construction guidelines.", False),
]

for r_idx, (text, is_bold) in enumerate(METHODOLOGY, 1):
    cell = ws4.cell(r_idx, 1)
    cell.value = text
    cell.font = Font(name='Arial', size=10, bold=is_bold)
    cell.alignment = left_align(wrap=True)
ws4.column_dimensions['A'].width = 85

# ── Save ──────────────────────────────────────────────────────────────────────
print("Saving workbook...")
wb.save(OUTPUT_FILE)

size = os.path.getsize(OUTPUT_FILE)
print(f"DONE — {OUTPUT_FILE}")
print(f"File size: {size:,} bytes ({size/1024/1024:.1f} MB)")
