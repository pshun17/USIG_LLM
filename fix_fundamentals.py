import sys
import io
import re
import time
import warnings
import os
import json

import requests
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import urllib3

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
warnings.filterwarnings('ignore')
urllib3.disable_warnings()

EXCEL_PATH = 'C:/Users/sh.park/Documents/USIG_LLM/LUACSTAT_2026_03_31_SCORED.xlsx'

# ─────────────────────────────────────────────────────────────
# STEP 1: Read file and extract tickers
# ─────────────────────────────────────────────────────────────
print("=" * 60)
print("STEP 1: Reading Excel and extracting tickers")
print("=" * 60)

df = pd.read_excel(EXCEL_PATH, sheet_name='Detail_Scored', header=1)
print(f"Detail_Scored shape: {df.shape}")

# Build ticker list per instructions
tickers_from_ticker_col = df['Ticker'].dropna().astype(str).str.strip()
tickers_from_ticker_col = tickers_from_ticker_col[~tickers_from_ticker_col.str.contains(r'\d', regex=True)]
tickers_from_ticker_col = tickers_from_ticker_col[tickers_from_ticker_col.str.len() >= 1]

tickers_from_eqty = df['Eqty Ticker'].dropna().astype(str).str.split().str[0]
tickers_from_eqty = tickers_from_eqty[~tickers_from_eqty.str.contains(r'\d', regex=True)]
tickers_from_eqty = tickers_from_eqty[tickers_from_eqty.str.len() >= 1]

all_tickers = list(
    set(tickers_from_ticker_col.tolist() + tickers_from_eqty.tolist())
    - {'', 'nan', 'N/A', 'NaN'}
)
print(f"Total tickers: {len(all_tickers)}")

# ─────────────────────────────────────────────────────────────
# STEP 2: Get Yahoo Finance crumb
# ─────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 2: Getting Yahoo Finance crumb")
print("=" * 60)

def init_session():
    session = requests.Session()
    session.verify = False
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
    })
    # Visit Yahoo Finance to set cookies
    session.get('https://finance.yahoo.com', timeout=20)
    time.sleep(2)

    crumb = None
    for url in [
        'https://query1.finance.yahoo.com/v1/test/getcrumb',
        'https://query2.finance.yahoo.com/v1/test/getcrumb',
    ]:
        try:
            r = session.get(url, timeout=15)
            if r.status_code == 200 and r.text and '"' not in r.text[:5]:
                crumb = r.text.strip()
                print(f"  Got crumb: {crumb}")
                break
        except Exception as e:
            print(f"  Crumb attempt failed at {url}: {e}")
            continue

    if not crumb:
        print("  WARNING: No crumb obtained, trying without")
    return session, crumb

session, crumb = init_session()

# ─────────────────────────────────────────────────────────────
# STEP 3: Test with 5 known tickers
# ─────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 3: Testing with 5 known tickers")
print("=" * 60)

api_headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'application/json',
    'Accept-Language': 'en-US,en;q=0.9',
}

def fetch_fundamentals(ticker, session, crumb):
    def get_raw(d, key):
        v = d.get(key, {})
        if isinstance(v, dict):
            return v.get('raw', None)
        return None

    # Build endpoint list
    if crumb:
        endpoints = [
            f'https://query1.finance.yahoo.com/v10/finance/quoteSummary/{ticker}?modules=financialData%2CdefaultKeyStatistics%2CsummaryDetail&crumb={crumb}',
            f'https://query2.finance.yahoo.com/v10/finance/quoteSummary/{ticker}?modules=financialData%2CdefaultKeyStatistics%2CsummaryDetail&crumb={crumb}',
            f'https://query1.finance.yahoo.com/v11/finance/quoteSummary/{ticker}?modules=financialData%2CdefaultKeyStatistics%2CsummaryDetail&crumb={crumb}',
        ]
    else:
        endpoints = [
            f'https://query1.finance.yahoo.com/v10/finance/quoteSummary/{ticker}?modules=financialData%2CdefaultKeyStatistics%2CsummaryDetail',
            f'https://query2.finance.yahoo.com/v10/finance/quoteSummary/{ticker}?modules=financialData%2CdefaultKeyStatistics%2CsummaryDetail',
        ]

    for url in endpoints:
        try:
            r = session.get(url, headers=api_headers, timeout=20)
            if r.status_code == 401:
                return 'auth_error'
            if r.status_code == 200:
                data = r.json()
                result = data.get('quoteSummary', {}).get('result', [])
                if result:
                    fd = result[0].get('financialData', {})
                    ks = result[0].get('defaultKeyStatistics', {})
                    sd = result[0].get('summaryDetail', {})
                    return {
                        'Debt_to_Equity': get_raw(fd, 'debtToEquity'),
                        'Profit_Margin':  get_raw(fd, 'profitMargins'),
                        'Revenue_Growth': get_raw(fd, 'revenueGrowth'),
                        'Current_Ratio':  get_raw(fd, 'currentRatio'),
                        'EV_EBITDA':      get_raw(ks, 'enterpriseToEbitda'),
                        'PE_Ratio':       get_raw(sd, 'trailingPE'),
                    }
        except Exception:
            continue
    return None

test_tickers = ['AAPL', 'MSFT', 'JPM', 'XOM', 'T']
test_ok = 0
for t in test_tickers:
    res = fetch_fundamentals(t, session, crumb)
    if res and res != 'auth_error' and any(v is not None for v in res.values()):
        test_ok += 1
        print(f"  {t}: OK  PM={res['Profit_Margin']:.4f}  PE={res['PE_Ratio']}")
    else:
        print(f"  {t}: FAIL  result={res}")
    time.sleep(0.3)

print(f"Test result: {test_ok}/5 tickers OK")
if test_ok < 3:
    print("Low test success – reinitialising session...")
    session, crumb = init_session()

CACHE_PATH = 'C:/Users/sh.park/Documents/USIG_LLM/fund_data_cache.json'

# ─────────────────────────────────────────────────────────────
# STEP 4: Mass fetch fundamentals for all tickers
# ─────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 4: Mass-fetching fundamentals for all tickers")
print("=" * 60)

fund_data = {}
success_count = 0
fail_count = 0
total = len(all_tickers)
BATCH = 30
REFRESH_EVERY = 200

print(f"Fetching {total} tickers (batch={BATCH}, crumb refresh every {REFRESH_EVERY})...")

for i, ticker in enumerate(all_tickers):
    # Refresh crumb every REFRESH_EVERY requests
    if i > 0 and i % REFRESH_EVERY == 0:
        print(f"  Refreshing session at ticker {i}/{total}...")
        try:
            session, crumb = init_session()
        except Exception as e:
            print(f"  Warning – could not refresh: {e}")

    result = fetch_fundamentals(ticker, session, crumb)

    # Re-auth on 401
    if result == 'auth_error':
        print(f"  Auth error at {ticker} (i={i}), refreshing session...")
        try:
            session, crumb = init_session()
            result = fetch_fundamentals(ticker, session, crumb)
        except Exception:
            result = None

    if result and result != 'auth_error' and any(v is not None for v in result.values()):
        fund_data[ticker] = result
        success_count += 1
    else:
        fail_count += 1

    if (i + 1) % 100 == 0:
        print(f"  Progress: {i+1}/{total}  success={success_count}  fail={fail_count}")

    # Rate limiting: sleep(0.5) every 30 tickers
    if i > 0 and i % BATCH == 0:
        time.sleep(0.5)
    else:
        time.sleep(0.07)

print(f"\nFetch complete: {success_count}/{total} tickers with data ({success_count/total*100:.1f}%)")

# Save cache so we can resume from Step 5 if needed
with open(CACHE_PATH, 'w', encoding='utf-8') as f:
    json.dump(fund_data, f)
print(f"Cache saved to {CACHE_PATH}")

# ─────────────────────────────────────────────────────────────
# STEP 5: Compute Eq_Fund_Score
# ─────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 5: Computing Eq_Fund_Score")
print("=" * 60)

fund_cols = ['Debt_to_Equity', 'Profit_Margin', 'Revenue_Growth', 'Current_Ratio', 'EV_EBITDA', 'PE_Ratio']

# Build fund DataFrame
fund_df = pd.DataFrame.from_dict(fund_data, orient='index')
fund_df.index.name = 'ticker'
fund_df = fund_df.reset_index()

print(f"Tickers with fundamental data: {len(fund_df)}")
for col in fund_cols:
    if col in fund_df.columns:
        print(f"  {col}: {fund_df[col].notna().sum()} non-null")

# Apply caps / adjustments before normalization
fund_df_norm = fund_df.copy()
if 'Current_Ratio' in fund_df_norm.columns:
    fund_df_norm['Current_Ratio'] = fund_df_norm['Current_Ratio'].clip(upper=3.0)
if 'PE_Ratio' in fund_df_norm.columns:
    fund_df_norm['PE_Ratio'] = fund_df_norm['PE_Ratio'].where(fund_df_norm['PE_Ratio'] > 0)
    fund_df_norm['PE_Ratio'] = fund_df_norm['PE_Ratio'].clip(upper=100.0)
if 'EV_EBITDA' in fund_df_norm.columns:
    fund_df_norm['EV_EBITDA'] = fund_df_norm['EV_EBITDA'].where(fund_df_norm['EV_EBITDA'] > 0)

def percentile_rank(series, higher_is_better=True):
    if higher_is_better:
        ranked = series.rank(method='average', na_option='keep')
    else:
        ranked = (-series).rank(method='average', na_option='keep')
    n = series.notna().sum()
    if n == 0:
        return pd.Series(np.nan, index=series.index)
    return (ranked - 1) / (n - 1) if n > 1 else ranked / n

norm_cols = {}
for col, higher_is_better in [
    ('Debt_to_Equity', False),
    ('Profit_Margin',  True),
    ('Revenue_Growth', True),
    ('Current_Ratio',  True),
    ('EV_EBITDA',      False),
    ('PE_Ratio',       False),
]:
    if col in fund_df_norm.columns:
        norm_cols[col] = percentile_rank(fund_df_norm[col], higher_is_better)

norm_df = pd.DataFrame(norm_cols, index=fund_df_norm.index)
norm_df['ticker'] = fund_df_norm['ticker']

score_cols = [c for c in fund_cols if c in norm_df.columns]
norm_df['available_count'] = norm_df[score_cols].notna().sum(axis=1)
norm_df['Eq_Fund_Score_raw'] = norm_df[score_cols].mean(axis=1)
norm_df.loc[norm_df['available_count'] < 2, 'Eq_Fund_Score_raw'] = np.nan
norm_df['Eq_Fund_Score'] = norm_df['Eq_Fund_Score_raw'] * 2 - 1
norm_df.loc[norm_df['available_count'] < 2, 'Eq_Fund_Score'] = np.nan

fund_df = fund_df.merge(norm_df[['ticker', 'Eq_Fund_Score']], on='ticker', how='left')

print(f"\nEq_Fund_Score:")
print(f"  Non-null: {fund_df['Eq_Fund_Score'].notna().sum()}")
print(f"  Range: [{fund_df['Eq_Fund_Score'].min():.4f}, {fund_df['Eq_Fund_Score'].max():.4f}]")

# ─────────────────────────────────────────────────────────────
# STEP 6: Map back to df and recompute Integrated_Score
# ─────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 6: Recomputing Integrated_Score and rankings")
print("=" * 60)

# fund_data keyed by tickers from 'all_tickers', which come from
# both 'Ticker' column (already clean) and first word of 'Eqty Ticker'.
# Build lookup: ticker_symbol -> fund row
fund_by_ticker = {row['ticker']: row for _, row in fund_df.iterrows()}

# Map each df row to its fundamental data
# Priority: Ticker column, then first word of Eqty Ticker
for col in fund_cols + ['Eq_Fund_Score']:
    df[col] = np.nan

matched = 0
for idx, row in df.iterrows():
    # Try Ticker column first
    t1 = str(row.get('Ticker', '')).strip()
    t2 = str(row.get('Eqty Ticker', '')).strip().split()[0] if pd.notna(row.get('Eqty Ticker')) else ''

    fund_row = fund_by_ticker.get(t1) if t1 in fund_by_ticker else fund_by_ticker.get(t2)
    if fund_row is not None:
        for col in fund_cols + ['Eq_Fund_Score']:
            if col in fund_row.index:
                df.at[idx, col] = fund_row[col]
        matched += 1

print(f"df rows matched to fundamentals: {matched}")
for col in fund_cols + ['Eq_Fund_Score']:
    print(f"  {col}: {df[col].notna().sum()} non-null")

# Recompute Integrated_Score
eq_mom  = df['Eq_Mom_Score'].fillna(0)
eq_fund = df['Eq_Fund_Score'].fillna(0)
sent    = df['Sentiment_Score'].fillna(0) if 'Sentiment_Score' in df.columns else pd.Series(0, index=df.index)
bond_tr = df['Bond_TR_Est_pct']

# Normalize Bond_TR_Est_pct → [-1, +1] within each class
bond_tr_score = pd.Series(np.nan, index=bond_tr.index)
active_mask = bond_tr.notna() & df['class'].notna() & (df['class'].astype(str).str.lower() != 'off')
for cls in df.loc[active_mask, 'class'].unique():
    cls_mask = active_mask & (df['class'] == cls)
    tr_cls = bond_tr[cls_mask]
    n_tr = tr_cls.notna().sum()
    if n_tr > 1:
        tr_ranked = tr_cls.rank(method='average', na_option='keep')
        bond_tr_score[cls_mask] = (tr_ranked - 1) / (n_tr - 1) * 2 - 1
df['Bond_TR_Score'] = bond_tr_score

df['Integrated_Score'] = (bond_tr_score.fillna(0) * 0.25
                          + eq_mom  * 0.25
                          + eq_fund * 0.25
                          + sent    * 0.25)

# Recompute Integrated_Rank_in_Class (desc, 1=best), exclude 'off'
df['Integrated_Rank_in_Class'] = np.nan
for cls, grp in df.groupby('class'):
    if str(cls).lower() == 'off':
        continue
    mask = df['class'] == cls
    ranks = df.loc[mask, 'Integrated_Score'].rank(method='min', ascending=False)
    df.loc[mask, 'Integrated_Rank_in_Class'] = ranks

# Recompute Top_Pick_Flag
def assign_flag(rank):
    if pd.isna(rank):
        return ''
    r = int(rank)
    if r <= 3:  return '★★★ TOP3'
    if r <= 10: return '★★ TOP10'
    if r <= 25: return '★ TOP25'
    return ''

df['Top_Pick_Flag'] = df['Integrated_Rank_in_Class'].apply(assign_flag)

print(f"\nTop_Pick_Flag distribution:")
print(df['Top_Pick_Flag'].value_counts().to_string())

# ─────────────────────────────────────────────────────────────
# STEP 7: Update Excel with openpyxl
# ─────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 7: Updating Excel file")
print("=" * 60)

ARIAL = Font(name='Arial')

def to_py(val):
    """Convert numpy types to Python native; NaN → None."""
    if val is None:
        return None
    if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
        return None
    if isinstance(val, (np.floating,)):
        return float(val)
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.bool_,)):
        return bool(val)
    return val

flag_styles = {
    '★★★ TOP3': {
        'font': Font(name='Arial', bold=True, color='FF0000'),
        'fill': PatternFill('solid', fgColor='FFFF00')
    },
    '★★ TOP10': {
        'font': Font(name='Arial', bold=True, color='C55A11'),
        'fill': PatternFill('solid', fgColor='FCE4D6')
    },
    '★ TOP25': {
        'font': Font(name='Arial', bold=True, color='1F3864'),
        'fill': PatternFill('solid', fgColor='DDEBF7')
    },
    '': {
        'font': Font(name='Arial', bold=False, color='000000'),
        'fill': PatternFill(fill_type=None)
    },
}

wb = load_workbook(EXCEL_PATH)

# ── Detail_Scored ──────────────────────────────────
ws_detail = wb['Detail_Scored']
HEADER_ROW = 2

col_map = {}
for cell in ws_detail[HEADER_ROW]:
    if cell.value is not None:
        col_map[str(cell.value).strip()] = cell.column

print(f"Detail_Scored header found {len(col_map)} columns")

update_cols = [
    'Debt_to_Equity', 'Profit_Margin', 'Revenue_Growth', 'Current_Ratio',
    'EV_EBITDA', 'PE_Ratio', 'Eq_Fund_Score',
    'Integrated_Score', 'Integrated_Rank_in_Class', 'Top_Pick_Flag'
]
for col in update_cols:
    if col in col_map:
        print(f"  Column '{col}' -> col {col_map[col]}")
    else:
        print(f"  WARNING: '{col}' NOT in header!")

df_reset = df.reset_index(drop=True)
for df_idx, row in df_reset.iterrows():
    excel_row = df_idx + 3   # data starts at row 3

    for col_name in update_cols:
        if col_name not in col_map:
            continue
        col_idx = col_map[col_name]
        val = to_py(row.get(col_name, np.nan))

        cell = ws_detail.cell(row=excel_row, column=col_idx)
        cell.value = val

        if col_name == 'Top_Pick_Flag':
            flag_val = str(val) if val is not None else ''
            style = flag_styles.get(flag_val, flag_styles[''])
            cell.font = style['font']
            cell.fill = style['fill']

print(f"  Detail_Scored: {len(df_reset)} rows updated")

# ── Equity_Data ────────────────────────────────────
ws_equity = wb['Equity_Data']

eq_col_map = {}
for cell in ws_equity[2]:   # header is row 2 (row 1 = title)
    if cell.value is not None:
        eq_col_map[str(cell.value).strip()] = cell.column

print(f"Equity_Data header found {len(eq_col_map)} columns: {list(eq_col_map.keys())[:10]}")

eq_fund_update_cols = [
    'Debt_to_Equity', 'Profit_Margin', 'Revenue_Growth', 'Current_Ratio',
    'EV_EBITDA', 'PE_Ratio', 'Eq_Fund_Score'
]

eq_ticker_col = eq_col_map.get('Ticker')
if eq_ticker_col is None:
    print("  WARNING: 'Ticker' column not found in Equity_Data!")
else:
    eq_updated = 0
    for eq_row in range(3, ws_equity.max_row + 1):
        ticker_val = ws_equity.cell(row=eq_row, column=eq_ticker_col).value
        if not ticker_val:
            continue
        clean_t = str(ticker_val).strip()
        if not clean_t or clean_t == 'nan':
            continue

        fund_row = fund_by_ticker.get(clean_t)
        if fund_row is None:
            continue

        for col_name in eq_fund_update_cols:
            if col_name not in eq_col_map:
                continue
            val = to_py(fund_row.get(col_name, np.nan))
            ws_equity.cell(row=eq_row, column=eq_col_map[col_name]).value = val
        eq_updated += 1

    print(f"  Equity_Data: {eq_updated} rows updated")

# ── Top_Picks_by_Class ─────────────────────────────
if 'Top_Picks_by_Class' in wb.sheetnames:
    del wb['Top_Picks_by_Class']

ws_top = wb.create_sheet('Top_Picks_by_Class')

top_cols = [
    'class', 'Des', 'ISIN', 'Ticker', 'Cpn', 'OAS', 'OASD',
    'Carry_2.5M_pct', 'Compression_Score_pct', 'Bond_TR_Est_pct',
    'Eq_Ret_1M', 'Eq_Ret_3M', 'Eq_Mom_Score', 'Eq_Fund_Score',
    'Integrated_Score', 'Integrated_Rank_in_Class', 'Top_Pick_Flag',
    'Issuer Rtg', 'S&P Outlook', "Moody's Outlook", 'BCLASS3', 'Industry Sector'
]

# Title row
title_cell = ws_top.cell(row=1, column=1)
title_cell.value = 'Top 5 Integrated Score Candidates by Class (2-3M Horizon) | As of 2026-03-31'
title_cell.font = Font(name='Arial', bold=True)

# Header row
header_font = Font(name='Arial', bold=True)
header_fill = PatternFill('solid', fgColor='D9E1F2')
for c_idx, col_name in enumerate(top_cols, start=1):
    cell = ws_top.cell(row=2, column=c_idx)
    cell.value = col_name
    cell.font = header_font
    cell.fill = header_fill

# Top 5 per class
top_df = (
    df[df['class'].notna() & (df['class'].astype(str).str.lower() != 'off')]
    .sort_values(['class', 'Integrated_Rank_in_Class'])
    .groupby('class')
    .head(5)
    .reset_index(drop=True)
)

data_row = 3
for _, row in top_df.iterrows():
    for c_idx, col_name in enumerate(top_cols, start=1):
        val = to_py(row.get(col_name, None))
        cell = ws_top.cell(row=data_row, column=c_idx)
        cell.value = val
        cell.font = Font(name='Arial')

        if col_name == 'Top_Pick_Flag':
            flag_val = str(val) if val is not None else ''
            style = flag_styles.get(flag_val, flag_styles[''])
            cell.font = style['font']
            cell.fill = style['fill']
    data_row += 1

print(f"  Top_Picks_by_Class: {len(top_df)} rows written")

# Save
wb.save(EXCEL_PATH)
print(f"\nFile saved: {EXCEL_PATH}")

# ─────────────────────────────────────────────────────────────
# STEP 8: Summary
# ─────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("STEP 8: Summary")
print("=" * 60)

file_size = os.path.getsize(EXCEL_PATH)
print(f"File size: {file_size / 1024 / 1024:.2f} MB")

df_verify = pd.read_excel(EXCEL_PATH, sheet_name='Detail_Scored', header=1)
print(f"\nFundamentals fetched: {success_count} / {total} tickers")
print(f"Bonds with Eq_Fund_Score (non-null): {df_verify['Eq_Fund_Score'].notna().sum()}")
print("\nNon-null counts per fundamental column:")
for col in fund_cols + ['Eq_Fund_Score']:
    nn = df_verify[col].notna().sum() if col in df_verify.columns else 0
    print(f"  {col}: {nn}")

print(f"\nSample rows with non-null Eq_Fund_Score:")
sample = df_verify[df_verify['Eq_Fund_Score'].notna()][
    ['Ticker', 'Eqty Ticker', 'Debt_to_Equity', 'Profit_Margin',
     'Revenue_Growth', 'Current_Ratio', 'EV_EBITDA', 'PE_Ratio',
     'Eq_Fund_Score', 'Integrated_Score', 'Top_Pick_Flag']
].head(5)
print(sample.to_string())

print("\nFUNDAMENTALS COMPLETE")
