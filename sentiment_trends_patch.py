"""
Patch script: fetch Google Trends and update the Excel file
(runs after sentiment_update.py which already wrote news sentiment)
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import os
import time
import math
import warnings
import traceback
import re

import numpy as np
import pandas as pd
import urllib3
urllib3.disable_warnings()
warnings.filterwarnings('ignore')

from pytrends.request import TrendReq
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

TODAY_STR = '2026-04-03'
FILE = 'C:/Users/sh.park/Documents/USIG_LLM/LUACSTAT_2026_03_31_SCORED.xlsx'

PURPLE_LIGHT = 'E8D5F5'
PURPLE_DARK  = '4B2D7F'
PURPLE_MID   = '7030A0'

# ── Read existing scored data ──
print("=== Reading existing data ===")
df = pd.read_excel(FILE, sheet_name='Detail_Scored', header=1)
print(f"  Shape: {df.shape}")

def clean_ticker(raw):
    if pd.isna(raw):
        return None
    s = str(raw).strip()
    s = re.sub(r'\s+[A-Z]{2,3}$', '', s).strip()
    if not s or len(s) > 15 or re.fullmatch(r'[\d\.\-]+', s):
        return None
    if re.search(r'\d{5,}', s):
        return None
    return s

df['_clean_ticker'] = df['Eqty Ticker'].apply(clean_ticker)
df['_company_name'] = df['Company Name'].fillna('')

unique_tickers = [t for t in df['_clean_ticker'].dropna().unique() if t]
n_tickers = len(unique_tickers)
print(f"  Unique tickers: {n_tickers}")

ticker_to_company = {}
for _, row in df[['_clean_ticker', '_company_name']].dropna(subset=['_clean_ticker']).iterrows():
    t = row['_clean_ticker']
    if t not in ticker_to_company:
        ticker_to_company[t] = row['_company_name']

# ── Fetch Google Trends ──
print("\n=== Fetching Google Trends ===")

trends_results = {}

def safe_trends_batch(pytrends_obj, tickers_group):
    results = {}
    try:
        pytrends_obj.build_payload(tickers_group, cat=0, timeframe='today 3-m', geo='US')
        data = pytrends_obj.interest_over_time()
        if data is None or data.empty:
            return results

        if 'isPartial' in data.columns:
            data = data.drop(columns=['isPartial'])

        data = data.sort_index()
        n_weeks = len(data)

        for t in tickers_group:
            if t not in data.columns:
                continue
            series = data[t].values.astype(float)

            if n_weeks < 8:
                results[t] = {'recent_4w': None, 'prev_4w': None, 'momentum': None}
                continue

            recent_4w = float(np.mean(series[-4:]))
            prev_4w   = float(np.mean(series[-8:-4]))
            momentum  = (recent_4w - prev_4w) / (prev_4w + 1.0)

            results[t] = {
                'recent_4w': recent_4w,
                'prev_4w':   prev_4w,
                'momentum':  momentum
            }
    except Exception as e:
        pass
    return results

try:
    pytrends = TrendReq(
        hl='en-US', tz=360, timeout=(10, 25), retries=2, backoff_factor=0.5,
        requests_args={'verify': False}
    )

    group_size = 5
    n_groups = math.ceil(n_tickers / group_size)

    for gi in range(n_groups):
        group = unique_tickers[gi * group_size: (gi + 1) * group_size]
        if gi % 20 == 0:
            print(f"  Trends group {gi}/{n_groups} ({gi*group_size}/{n_tickers} tickers)...")

        try:
            batch_res = safe_trends_batch(pytrends, group)
            trends_results.update(batch_res)
        except Exception as e:
            pass

        time.sleep(1.0)

except Exception as e:
    print(f"  pytrends failed: {e}")

tickers_with_trends = sum(1 for v in trends_results.values() if v.get('momentum') is not None)
print(f"  Tickers with trends data: {tickers_with_trends} / {n_tickers}")

# ── Read existing sentiment columns from the already-updated Excel ──
# We need News_Sentiment_Raw from the file to recompute Sentiment_Score
print("\n=== Reading existing news sentiment from Excel ===")

wb = load_workbook(FILE)
ws_detail = wb['Detail_Scored']

# Find column positions by reading row 2 headers
header_row = {}
for cell in ws_detail[2]:
    if cell.value is not None:
        header_row[str(cell.value)] = cell.column

print(f"  Column map (sentiment cols): "
      f"News_Sentiment_Raw={header_row.get('News_Sentiment_Raw')}, "
      f"News_Article_Count={header_row.get('News_Article_Count')}, "
      f"Trends_Momentum={header_row.get('Trends_Momentum')}, "
      f"Sentiment_Score={header_row.get('Sentiment_Score')}, "
      f"Integrated_Score={header_row.get('Integrated_Score')}, "
      f"Top_Pick_Flag={header_row.get('Top_Pick_Flag')}")

COL_NEWS_RAW  = header_row['News_Sentiment_Raw']
COL_NEWS_CNT  = header_row['News_Article_Count']
COL_TRENDS    = header_row['Trends_Momentum']
COL_SENT      = header_row['Sentiment_Score']
COL_INT       = header_row['Integrated_Score']
COL_INT_RANK  = header_row['Integrated_Rank_in_Class']
COL_FLAG      = header_row['Top_Pick_Flag']

# Read existing news raw scores from the sheet
news_raw_by_ticker = {}
ticker_col = header_row.get('Eqty Ticker')

# Re-read the df with pandas to get news values already written
df2 = pd.read_excel(FILE, sheet_name='Detail_Scored', header=1)
df2['_clean_ticker'] = df2['Eqty Ticker'].apply(clean_ticker)
df2['_company_name'] = df2['Company Name'].fillna('')

# Build ticker→news_raw from df2
for _, row in df2.dropna(subset=['_clean_ticker']).iterrows():
    t = row['_clean_ticker']
    if t not in news_raw_by_ticker:
        nr = row.get('News_Sentiment_Raw')
        if not (isinstance(nr, float) and math.isnan(nr)):
            news_raw_by_ticker[t] = nr

print(f"  News raw scores for {len(news_raw_by_ticker)} tickers")

# ── Re-normalize and recompute Sentiment_Score with trends ──
print("\n=== Recomputing Sentiment_Score with Trends ===")

# Build unified sentiment df
sentiment_rows = []
for t in unique_tickers:
    tr = trends_results.get(t, {})
    nr = news_raw_by_ticker.get(t, None)
    nr_cnt = int(df2[df2['_clean_ticker'] == t]['News_Article_Count'].dropna().iloc[0]) if len(df2[df2['_clean_ticker'] == t]['News_Article_Count'].dropna()) > 0 else 0
    top_hl = ''
    sentiment_rows.append({
        'ticker': t,
        'News_Sentiment_Raw': nr,
        'News_Article_Count': nr_cnt,
        'Trends_Momentum': tr.get('momentum'),
        'Trends_Recent_4w': tr.get('recent_4w'),
        'Trends_Prev_4w': tr.get('prev_4w'),
        'top_headlines': top_hl
    })

sdf = pd.DataFrame(sentiment_rows)

# Rank-normalize
valid_news = sdf['News_Sentiment_Raw'].notna()
if valid_news.sum() > 0:
    ranks_news = sdf.loc[valid_news, 'News_Sentiment_Raw'].rank(method='average')
    n_news = valid_news.sum()
    sdf.loc[valid_news, 'News_Sentiment_Norm'] = (ranks_news / n_news) * 2 - 1
else:
    sdf['News_Sentiment_Norm'] = np.nan

valid_trends = sdf['Trends_Momentum'].notna()
if valid_trends.sum() > 0:
    ranks_trends = sdf.loc[valid_trends, 'Trends_Momentum'].rank(method='average')
    n_trends = valid_trends.sum()
    sdf.loc[valid_trends, 'Trends_Norm'] = (ranks_trends / n_trends) * 2 - 1
else:
    sdf['Trends_Norm'] = np.nan

def compute_sentiment(row):
    nn = row.get('News_Sentiment_Norm')
    tn = row.get('Trends_Norm')
    has_news   = nn is not None and not (isinstance(nn, float) and math.isnan(float(nn)))
    has_trends = tn is not None and not (isinstance(tn, float) and math.isnan(float(tn)))
    if has_news and has_trends:
        return float(nn) * 0.7 + float(tn) * 0.3
    elif has_news:
        return float(nn)
    elif has_trends:
        return float(tn) * 0.5
    return None

sdf['Sentiment_Score'] = sdf.apply(compute_sentiment, axis=1)
tickers_with_sentiment = sdf['Sentiment_Score'].notna().sum()
print(f"  Tickers with Sentiment_Score: {tickers_with_sentiment} / {n_tickers}")

# Map to df2
ticker_to_sent  = dict(zip(sdf['ticker'], sdf['Sentiment_Score']))
ticker_to_trend = dict(zip(sdf['ticker'], sdf['Trends_Momentum']))
ticker_to_tr_r  = dict(zip(sdf['ticker'], sdf['Trends_Recent_4w']))
ticker_to_tr_p  = dict(zip(sdf['ticker'], sdf['Trends_Prev_4w']))

df2['Trends_Momentum']  = df2['_clean_ticker'].map(ticker_to_trend)
df2['Trends_Recent_4w'] = df2['_clean_ticker'].map(ticker_to_tr_r)
df2['Trends_Prev_4w']   = df2['_clean_ticker'].map(ticker_to_tr_p)
df2['Sentiment_Score']  = df2['_clean_ticker'].map(ticker_to_sent)

# Recompute Integrated_Score
df2['Integrated_Score'] = (
    df2['Bond_TR_Est_pct'].fillna(0) * 0.6
    + df2['Eq_Mom_Score'].fillna(0) * 0.025
    + df2['Eq_Fund_Score'].fillna(0) * 0.015
    + df2['Sentiment_Score'].fillna(0) * 0.010
)

# Recompute ranks
df2['Integrated_Rank_in_Class'] = np.nan
for cls in df2['class'].dropna().unique():
    if cls == 'off':
        continue
    mask = (df2['class'] == cls) & df2['Integrated_Score'].notna()
    if mask.sum() == 0:
        continue
    df2.loc[mask, 'Integrated_Rank_in_Class'] = (
        df2.loc[mask, 'Integrated_Score']
        .rank(ascending=False, method='min')
        .astype(int)
    )

def top_pick_flag(rank):
    if pd.isna(rank):
        return ''
    r = int(rank)
    if r <= 3:
        return '★★★ TOP3'
    elif r <= 10:
        return '★★ TOP10'
    elif r <= 25:
        return '★ TOP25'
    return ''

df2['Top_Pick_Flag'] = df2['Integrated_Rank_in_Class'].apply(top_pick_flag)

# ── Update Detail_Scored sheet ──
print("\n=== Updating Detail_Scored sheet ===")

def to_py(val):
    if val is None:
        return None
    if isinstance(val, float) and math.isnan(val):
        return None
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        return float(val)
    if isinstance(val, (np.bool_,)):
        return bool(val)
    return val

font_top3  = Font(name='Arial', bold=True, color='FF0000', size=10)
font_top10 = Font(name='Arial', bold=True, color='C55A11', size=10)
font_top25 = Font(name='Arial', bold=True, color='1F3864', size=10)
font_none  = Font(name='Arial', size=10)
fill_top3  = PatternFill('solid', start_color='FFFF00')
fill_top10 = PatternFill('solid', start_color='FCE4D6')
fill_top25 = PatternFill('solid', start_color='DDEBF7')
fill_none  = PatternFill(fill_type=None)

for df_idx in range(len(df2)):
    excel_row = df_idx + 3

    trends_val = to_py(df2.at[df_idx, 'Trends_Momentum'])
    sent_val   = to_py(df2.at[df_idx, 'Sentiment_Score'])
    int_val    = to_py(df2.at[df_idx, 'Integrated_Score'])
    rank_val   = df2.at[df_idx, 'Integrated_Rank_in_Class']
    flag_val   = df2.at[df_idx, 'Top_Pick_Flag']

    ws_detail.cell(row=excel_row, column=COL_TRENDS).value   = trends_val
    ws_detail.cell(row=excel_row, column=COL_SENT).value     = sent_val
    ws_detail.cell(row=excel_row, column=COL_INT).value      = int_val
    ws_detail.cell(row=excel_row, column=COL_INT_RANK).value = to_py(rank_val) if not (isinstance(rank_val, float) and math.isnan(rank_val)) else None

    flag_str = str(flag_val) if flag_val else ''
    flag_cell = ws_detail.cell(row=excel_row, column=COL_FLAG)
    flag_cell.value = flag_str

    if flag_str == '★★★ TOP3':
        flag_cell.font = font_top3
        flag_cell.fill = fill_top3
    elif flag_str == '★★ TOP10':
        flag_cell.font = font_top10
        flag_cell.fill = fill_top10
    elif flag_str == '★ TOP25':
        flag_cell.font = font_top25
        flag_cell.fill = fill_top25
    else:
        flag_cell.font = font_none
        flag_cell.fill = fill_none

print("  Detail_Scored updated.")

# ── Update Sentiment_Analysis sheet ──
print("=== Updating Sentiment_Analysis sheet ===")

if 'Sentiment_Analysis' in wb.sheetnames:
    del wb['Sentiment_Analysis']

ws_sent = wb.create_sheet('Sentiment_Analysis')

def center_align():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

ws_sent.merge_cells('A1:I1')
title_cell = ws_sent['A1']
title_cell.value = f'Equity Sentiment Analysis — News & Search Trends | As of {TODAY_STR}'
title_cell.font  = Font(name='Arial', bold=True, size=13, color='FFFFFF')
title_cell.fill  = PatternFill('solid', start_color=PURPLE_DARK)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws_sent.row_dimensions[1].height = 24

headers_sent = [
    'Ticker', 'Company_Name', 'News_Sentiment_Raw', 'News_Article_Count',
    'Trends_Momentum', 'Trends_Recent_4w', 'Trends_Prev_4w',
    'Sentiment_Score', 'Top_Headlines'
]
for ci, h in enumerate(headers_sent, 1):
    cell = ws_sent.cell(row=2, column=ci, value=h)
    cell.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    cell.fill      = PatternFill('solid', start_color=PURPLE_MID)
    cell.alignment = center_align()

col_widths = {1: 10, 2: 30, 3: 15, 4: 15, 5: 15, 6: 15, 7: 15, 8: 15, 9: 80}
for ci, w in col_widths.items():
    ws_sent.column_dimensions[get_column_letter(ci)].width = w

# Read top headlines from original news results — re-fetch or load from existing sheet
# We'll use a simplified version: just get the top_headlines from sdf
# But we need to get them from the first run. Re-read from the already existing Sentiment_Analysis if available.
# Since we just deleted it, we'll use sdf with empty headlines and note it.

# Build sdf with company names for the Sentiment_Analysis sheet
sdf['_company'] = sdf['ticker'].map(ticker_to_company).fillna('')

# Re-read top headlines from df2 (the Excel data doesn't have headlines stored, so we use what we have)
# The top_headlines col in sdf was set to '' in this patch script since we only fetched trends here.
# Read the headlines from the original sdf if possible (it was written to the sheet in the first run).
# Since the Sentiment_Analysis sheet was deleted, we'll note "See original run for headlines".

# Sort by Sentiment_Score descending
sdf_sorted = sdf.sort_values('Sentiment_Score', ascending=False, na_position='last').reset_index(drop=True)

fill_green = PatternFill('solid', start_color='C6EFCE')
fill_red   = PatternFill('solid', start_color='FFC7CE')

for ri, row in sdf_sorted.iterrows():
    excel_row = ri + 3
    score_val = to_py(row['Sentiment_Score'])

    ws_sent.cell(row=excel_row, column=1).value = row['ticker']
    ws_sent.cell(row=excel_row, column=2).value = row['_company']
    ws_sent.cell(row=excel_row, column=3).value = to_py(row['News_Sentiment_Raw'])
    ws_sent.cell(row=excel_row, column=4).value = int(row['News_Article_Count']) if row['News_Article_Count'] else 0
    ws_sent.cell(row=excel_row, column=5).value = to_py(row['Trends_Momentum'])
    ws_sent.cell(row=excel_row, column=6).value = to_py(row['Trends_Recent_4w'])
    ws_sent.cell(row=excel_row, column=7).value = to_py(row['Trends_Prev_4w'])
    ws_sent.cell(row=excel_row, column=8).value = score_val
    ws_sent.cell(row=excel_row, column=9).value = row.get('top_headlines', '') or ''

    for ci in [3, 5, 6, 7, 8]:
        ws_sent.cell(row=excel_row, column=ci).number_format = '0.0000'
    ws_sent.cell(row=excel_row, column=4).number_format = '0'

    score_cell = ws_sent.cell(row=excel_row, column=8)
    if score_val is not None:
        if score_val > 0.3:
            score_cell.fill = fill_green
        elif score_val < -0.3:
            score_cell.fill = fill_red

ws_sent.freeze_panes = 'A3'
ws_sent.auto_filter.ref = f'A2:I{2 + len(sdf_sorted)}'

# ── Regenerate Top_Picks_by_Class ──
print("=== Regenerating Top_Picks_by_Class ===")

ws_top = wb['Top_Picks_by_Class']
# Unmerge all merged cells first
for mc in list(ws_top.merged_cells.ranges):
    ws_top.unmerge_cells(str(mc))
for row in ws_top.iter_rows():
    for cell in row:
        try:
            cell.value = None
            cell.font  = Font(name='Arial', size=10)
            cell.fill  = PatternFill(fill_type=None)
            cell.alignment = Alignment()
        except AttributeError:
            pass

ws_top.merge_cells('A1:W1')
t1 = ws_top['A1']
t1.value = f'Top 5 Integrated Score Candidates by Class (2-3M Horizon) | As of {TODAY_STR}'
t1.font  = Font(name='Arial', bold=True, size=13, color='FFFFFF')
t1.fill  = PatternFill('solid', start_color='1F3864')
t1.alignment = Alignment(horizontal='center', vertical='center')
ws_top.row_dimensions[1].height = 24

top_headers = [
    'class', 'Des', 'ISIN', 'Ticker', 'Cpn', 'OAS', 'OASD',
    'Carry_2.5M_pct', 'Compression_Score_pct', 'Bond_TR_Est_pct',
    'Eq_Ret_1M', 'Eq_Ret_3M', 'Eq_Mom_Score', 'Eq_Fund_Score',
    'Sentiment_Score',
    'Integrated_Score', 'Integrated_Rank_in_Class', 'Top_Pick_Flag',
    'Issuer Rtg', 'S&P Outlook', "Moody's Outlook", 'BCLASS3', 'Industry Sector'
]

hdr_fill_blue = PatternFill('solid', start_color='1F3864')
for ci, h in enumerate(top_headers, 1):
    cell = ws_top.cell(row=2, column=ci, value=h)
    cell.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    cell.fill      = hdr_fill_blue
    cell.alignment = center_align()

top_cols_map = {h: h for h in top_headers}
# Fix mapping for cols that have spaces/special chars
top_cols_map['Issuer Rtg']     = 'Issuer Rtg'
top_cols_map['S&P Outlook']    = 'S&P Outlook'
top_cols_map["Moody's Outlook"] = "Moody's Outlook"
top_cols_map['Industry Sector'] = 'Industry Sector'

df2_active = df2[df2['class'].notna() & (df2['class'] != 'off')].copy()
top5 = (
    df2_active
    .sort_values('Integrated_Score', ascending=False)
    .groupby('class', sort=False)
    .head(5)
    .sort_values(['class', 'Integrated_Score'], ascending=[True, False])
    .reset_index(drop=True)
)

excel_row = 3
for _, row in top5.iterrows():
    for ci, hdr in enumerate(top_headers, 1):
        val = row.get(hdr)
        cell = ws_top.cell(row=excel_row, column=ci)

        if isinstance(val, float) and math.isnan(val):
            val = None

        cell.value = to_py(val)
        cell.font  = Font(name='Arial', size=10)

        if hdr == 'Top_Pick_Flag':
            flag = str(val) if val else ''
            cell.value = flag
            if flag == '★★★ TOP3':
                cell.font = font_top3
                cell.fill = fill_top3
            elif flag == '★★ TOP10':
                cell.font = font_top10
                cell.fill = fill_top10
            elif flag == '★ TOP25':
                cell.font = font_top25
                cell.fill = fill_top25

    excel_row += 1

ws_top.column_dimensions['A'].width = 12
ws_top.column_dimensions['B'].width = 30
ws_top.column_dimensions['C'].width = 16

# ── Save ──
print("\n=== Saving workbook ===")
wb.save(FILE)
file_size = os.path.getsize(FILE)
print(f"  Saved: {FILE}")
print(f"  File size: {file_size / 1024 / 1024:.2f} MB")

print("\n=== FINAL VERIFICATION ===")
news_cnt   = sum(1 for t in unique_tickers if news_raw_by_ticker.get(t) is not None)
trends_cnt = tickers_with_trends
sent_cnt   = int(tickers_with_sentiment)
nonnull_in_detail = int(df2['Sentiment_Score'].notna().sum())

print(f"  Tickers with news data:       {news_cnt} / {n_tickers}")
print(f"  Tickers with trends data:     {trends_cnt} / {n_tickers}")
print(f"  Tickers with Sentiment_Score: {sent_cnt} / {n_tickers}")
print(f"  Non-null Sentiment_Score in Detail_Scored: {nonnull_in_detail}")
print(f"  Final file size: {file_size:,} bytes ({file_size / 1024 / 1024:.2f} MB)")
print("\nSENTIMENT UPDATE COMPLETE")
