"""
Sentiment Update Script for LUACSTAT_2026_03_31_SCORED.xlsx
- Fetches Yahoo Finance news + VADER NLP sentiment
- Fetches Google Trends momentum via pytrends
- Computes composite Sentiment_Score
- Updates Integrated_Score with new weighting
- Regenerates rankings and updates all sheets
"""

import sys
import io
import os
import time
import math
import warnings
import traceback
import re
from datetime import datetime, timezone

import numpy as np
import pandas as pd
import requests

from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

warnings.filterwarnings('ignore')

TODAY = datetime(2026, 4, 3, tzinfo=timezone.utc)
FILE = 'C:/Users/sh.park/Documents/USIG_LLM/LUACSTAT_2026_03_31_SCORED.xlsx'

PURPLE_LIGHT  = 'E8D5F5'
PURPLE_DARK   = '4B2D7F'
PURPLE_MID    = '7030A0'

# ──────────────────────────────────────────────────────────────
# STEP 1 — Read data
# ──────────────────────────────────────────────────────────────
print("=== STEP 1: Reading existing file ===")
df = pd.read_excel(FILE, sheet_name='Detail_Scored', header=1)
print(f"  Loaded {len(df)} rows x {len(df.columns)} cols")

def clean_ticker(raw):
    if pd.isna(raw):
        return None
    s = str(raw).strip()
    # Remove exchange suffix like " US", " LN", " JP", " CN" etc.
    s = re.sub(r'\s+[A-Z]{2,3}$', '', s).strip()
    # Filter out numeric-only, too long, or clearly invalid
    if not s or len(s) > 15 or re.fullmatch(r'[\d\.\-]+', s):
        return None
    # Filter rows with numeric-looking tickers (Bloomberg IDs)
    if re.search(r'\d{5,}', s):
        return None
    return s

# Build per-row clean ticker and company name
df['_clean_ticker'] = df['Eqty Ticker'].apply(clean_ticker)
df['_company_name'] = df['Company Name'].fillna('')

# Unique tickers for API calls
unique_tickers = [t for t in df['_clean_ticker'].dropna().unique() if t]
print(f"  Unique clean tickers: {len(unique_tickers)}")
print(f"  Sample: {unique_tickers[:10]}")

# Map ticker → company name (first occurrence)
ticker_to_company = {}
for _, row in df[['_clean_ticker', '_company_name']].dropna(subset=['_clean_ticker']).iterrows():
    t = row['_clean_ticker']
    if t not in ticker_to_company:
        ticker_to_company[t] = row['_company_name']

# ──────────────────────────────────────────────────────────────
# STEP 2 — Yahoo Finance News + VADER
# ──────────────────────────────────────────────────────────────
print("\n=== STEP 2: Fetching Yahoo Finance news ===")

analyzer = SentimentIntensityAnalyzer()

session = requests.Session()
session.headers.update({
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'application/json',
    'Accept-Language': 'en-US,en;q=0.9',
})

def get_yahoo_crumb():
    """Get Yahoo Finance crumb for authenticated API calls."""
    try:
        r = session.get('https://fc.yahoo.com', timeout=10, verify=False)
        cookie = r.cookies.get('A3')
        r2 = session.get('https://query1.finance.yahoo.com/v1/test/getcrumb', timeout=10, verify=False)
        if r2.status_code == 200 and r2.text:
            return r2.text.strip()
    except:
        pass
    return None

crumb = get_yahoo_crumb()
print(f"  Yahoo crumb: {'obtained' if crumb else 'not available (will try without)'}")

def fetch_yahoo_news(ticker, crumb=None):
    """Fetch up to 20 news headlines from Yahoo Finance."""
    headlines = []
    urls = [
        f'https://query1.finance.yahoo.com/v1/finance/search?q={ticker}&newsCount=20&quotesCount=0',
        f'https://query2.finance.yahoo.com/v1/finance/search?q={ticker}&newsCount=20&quotesCount=0',
    ]
    if crumb:
        urls = [u + f'&crumb={crumb}' for u in urls]

    for url in urls:
        try:
            r = session.get(url, timeout=12, verify=False)
            if r.status_code == 200:
                data = r.json()
                news_items = data.get('news', [])
                for item in news_items[:20]:
                    title = item.get('title', '')
                    pub_time = item.get('providerPublishTime', None)
                    if title:
                        headlines.append({'title': title, 'pub_time': pub_time})
                if headlines:
                    return headlines
        except Exception:
            pass

    # RSS fallback
    try:
        rss_url = f'https://feeds.finance.yahoo.com/rss/2.0/headline?s={ticker}&region=US&lang=en-US'
        r = session.get(rss_url, timeout=12, verify=False)
        if r.status_code == 200:
            import xml.etree.ElementTree as ET
            root = ET.fromstring(r.text)
            for item in root.findall('.//item')[:20]:
                title_el = item.find('title')
                pub_el = item.find('pubDate')
                if title_el is not None and title_el.text:
                    pub_time = None
                    if pub_el is not None and pub_el.text:
                        try:
                            from email.utils import parsedate_to_datetime
                            dt = parsedate_to_datetime(pub_el.text)
                            pub_time = int(dt.timestamp())
                        except:
                            pass
                    headlines.append({'title': title_el.text, 'pub_time': pub_time})
    except Exception:
        pass

    return headlines

def compute_vader_score(headlines):
    """Compute recency-weighted VADER sentiment from list of headline dicts.
    Returns (weighted_mean, count, most_impactful_headline, most_impactful_score)
    """
    if not headlines:
        return None, 0, '', None

    today_ts = TODAY.timestamp()
    weights = []
    scores = []
    titles = []

    for h in headlines:
        compound = analyzer.polarity_scores(h['title'])['compound']

        # Recency weighting
        if h['pub_time'] is not None:
            days_old = max(0, (today_ts - h['pub_time']) / 86400)
        else:
            days_old = 7  # assume 1 week old if unknown

        weight = math.exp(-days_old / 14.0)
        weights.append(weight)
        scores.append(compound)
        titles.append(h['title'])

    total_weight = sum(weights)
    if total_weight == 0:
        return None, len(headlines), '', None

    weighted_mean = sum(s * w for s, w in zip(scores, weights)) / total_weight

    # 가장 임팩트 있는 헤드라인: 가중치 적용 score (|compound × weight|) 기준 최대
    weighted_scores = [abs(s * w) for s, w in zip(scores, weights)]
    max_idx = weighted_scores.index(max(weighted_scores))
    top_headline  = titles[max_idx]
    top_score     = round(scores[max_idx], 4)

    return weighted_mean, len(headlines), top_headline, top_score

def fetch_google_news(ticker, company_name=''):
    """Fetch up to 15 headlines from Google News RSS for direction signal."""
    import xml.etree.ElementTree as ET
    from email.utils import parsedate_to_datetime
    headlines = []
    queries = [ticker]
    if company_name and company_name != ticker:
        queries.append(company_name)
    for query in queries:
        try:
            q = requests.utils.quote(f'{query} stock')
            url = f'https://news.google.com/rss/search?q={q}&hl=en-US&gl=US&ceid=US:en'
            r = session.get(url, timeout=12, verify=False)
            if r.status_code == 200 and '<item>' in r.text:
                root = ET.fromstring(r.text)
                for item in root.findall('.//item')[:15]:
                    title_el = item.find('title')
                    pub_el   = item.find('pubDate')
                    if title_el is not None and title_el.text:
                        pub_time = None
                        if pub_el is not None and pub_el.text:
                            try:
                                pub_time = int(parsedate_to_datetime(pub_el.text).timestamp())
                            except Exception:
                                pass
                        headlines.append({'title': title_el.text.strip(), 'pub_time': pub_time})
                if headlines:
                    return headlines
        except Exception:
            pass
    return headlines

# Fetch news for all tickers
news_results = {}  # ticker → {'raw': float, 'count': int, 'google_count': int, 'headlines': [str,...]}

batch_size = 30
n_tickers = len(unique_tickers)

for i, ticker in enumerate(unique_tickers):
    if i % 100 == 0:
        print(f"  Progress: {i}/{n_tickers} tickers processed...")

    try:
        # Yahoo Finance 뉴스 수집
        yahoo_headlines = fetch_yahoo_news(ticker, crumb)
        # Google News RSS 수집 (별도 소스 — 방향성 판단 보강)
        google_headlines = fetch_google_news(ticker, ticker_to_company.get(ticker, ''))
        google_count = len(google_headlines)
        # 중복 제거 후 병합 (title 기준)
        existing_titles = {h['title'] for h in yahoo_headlines}
        merged = yahoo_headlines[:]
        for gh in google_headlines:
            if gh['title'] not in existing_titles:
                merged.append(gh)
                existing_titles.add(gh['title'])
        raw_score, count, top_headline, top_score = compute_vader_score(merged)
        top_headlines = [h['title'] for h in merged[:5]]
        news_results[ticker] = {
            'raw': raw_score,
            'count': count,
            'google_count': google_count,
            'headlines': top_headlines,
            'top_headline': top_headline,   # 가장 임팩트 있는 헤드라인 (|compound×weight| 최대)
            'top_score': top_score,         # 해당 헤드라인의 VADER compound score
        }
    except Exception as e:
        news_results[ticker] = {'raw': None, 'count': 0, 'google_count': 0, 'headlines': [],
                                'top_headline': '', 'top_score': None}

    if (i + 1) % batch_size == 0:
        time.sleep(0.3)

tickers_with_news = sum(1 for v in news_results.values() if v['raw'] is not None)
print(f"  Tickers with news data: {tickers_with_news} / {n_tickers}")

# Generic 뉴스 무효화: 같은 raw 값이 5개 초과 ticker에서 나타나면 Yahoo가 공통 뉴스를 반환한 것
raw_series = {t: round(v['raw'], 5) for t, v in news_results.items() if v['raw'] is not None}
from collections import Counter
raw_val_counts = Counter(raw_series.values())
GENERIC_THRESHOLD = 5
generic_raw_vals = {val for val, cnt in raw_val_counts.items() if cnt > GENERIC_THRESHOLD}
n_invalidated = 0
for t in news_results:
    if news_results[t]['raw'] is not None:
        if round(news_results[t]['raw'], 5) in generic_raw_vals:
            news_results[t]['raw'] = None
            news_results[t]['count'] = 0
            n_invalidated += 1
print(f"  Generic raw값 {len(generic_raw_vals)}개 감지 → {n_invalidated}개 ticker 무효화")

# ──────────────────────────────────────────────────────────────
# STEP 3 — Google Trends (pytrends)
# ──────────────────────────────────────────────────────────────
print("\n=== STEP 3: Fetching Google Trends data ===")

from pytrends.request import TrendReq

trends_results = {}  # ticker → {'recent_4w': float, 'prev_4w': float, 'momentum': float}

def safe_trends_batch(pytrends_obj, tickers_group):
    """Fetch trends for a group of up to 5 tickers, return dict of results."""
    results = {}
    try:
        pytrends_obj.build_payload(
            tickers_group, cat=0, timeframe='today 3-m', geo='US'
        )
        data = pytrends_obj.interest_over_time()
        if data is None or data.empty:
            return results

        # data has weekly rows; drop 'isPartial' col
        if 'isPartial' in data.columns:
            data = data.drop(columns=['isPartial'])

        # Sort by date ascending
        data = data.sort_index()
        n_weeks = len(data)

        for t in tickers_group:
            if t not in data.columns:
                continue
            series = data[t].values.astype(float)

            if n_weeks < 8:
                # Not enough data
                results[t] = {'recent_4w': None, 'prev_4w': None, 'momentum': None}
                continue

            recent_4w = float(np.mean(series[-4:]))
            prev_4w   = float(np.mean(series[-8:-4]))
            momentum  = (recent_4w - prev_4w) / (prev_4w + 1.0)

            results[t] = {
                'recent_4w': recent_4w,
                'prev_4w': prev_4w,
                'momentum': momentum
            }
    except Exception as e:
        pass
    return results

try:
    pytrends = TrendReq(hl='en-US', tz=360, timeout=(10, 25), retries=2, backoff_factor=0.5,
                        requests_args={'verify': False})

    group_size = 5
    n_groups = math.ceil(n_tickers / group_size)

    for gi in range(n_groups):
        group = unique_tickers[gi * group_size : (gi + 1) * group_size]
        if gi % 20 == 0:
            print(f"  Trends group {gi}/{n_groups}...")

        try:
            batch_res = safe_trends_batch(pytrends, group)
            trends_results.update(batch_res)
        except Exception:
            pass

        time.sleep(1.0)

except Exception as e:
    print(f"  pytrends initialization failed: {e}. Skipping trends.")

tickers_with_trends = sum(
    1 for v in trends_results.values()
    if v.get('momentum') is not None
)
print(f"  Tickers with trends data: {tickers_with_trends} / {n_tickers}")

# ──────────────────────────────────────────────────────────────
# STEP 4 — Compute Composite Sentiment Score
# ──────────────────────────────────────────────────────────────
print("\n=== STEP 4: Computing Sentiment_Score ===")

# Build sentiment dataframe
sentiment_rows = []
for t in unique_tickers:
    nr = news_results.get(t, {})
    tr = trends_results.get(t, {})
    sentiment_rows.append({
        'ticker': t,
        'News_Sentiment_Raw': nr.get('raw'),
        'News_Article_Count': nr.get('count', 0),
        'Google_News_Count': nr.get('google_count', 0),
        'Top_Headline': nr.get('top_headline', ''),
        'Top_Headline_Score': nr.get('top_score'),
        'Trends_Momentum': tr.get('momentum'),
        'Trends_Recent_4w': tr.get('recent_4w'),
        'Trends_Prev_4w': tr.get('prev_4w'),
        'top_headlines': ' | '.join(nr.get('headlines', []))
    })

sdf = pd.DataFrame(sentiment_rows)

# Rank-normalize News_Sentiment_Raw → [-1, +1]
valid_news = sdf['News_Sentiment_Raw'].notna()
if valid_news.sum() > 0:
    ranks_news = sdf.loc[valid_news, 'News_Sentiment_Raw'].rank(method='average')
    n_news = valid_news.sum()
    sdf.loc[valid_news, 'News_Sentiment_Norm'] = (ranks_news / n_news) * 2 - 1
else:
    sdf['News_Sentiment_Norm'] = np.nan

# ── Option A: 뉴스 중심, 트렌드로 세기 보정 ─────────────────────────────────
# 뉴스가 방향(부호) 결정, 트렌드는 세기(amplitude) 조절만 담당
# Trends_Norm → Trends_Factor [0.7, 1.3]: 트렌드 상승 = 증폭, 하락 = 감쇠 (부호 유지)

# Rank-normalize Trends_Momentum → [-1, +1]
valid_trends = sdf['Trends_Momentum'].notna()
if valid_trends.sum() > 0:
    ranks_trends = sdf.loc[valid_trends, 'Trends_Momentum'].rank(method='average')
    n_trends = valid_trends.sum()
    sdf.loc[valid_trends, 'Trends_Norm'] = (ranks_trends / n_trends) * 2 - 1
else:
    sdf['Trends_Norm'] = np.nan

# Trends_Factor: 뉴스 점수에 곱할 증폭 계수
#   트렌드 상승(+1) → 1.3배 증폭  |  중립(0) → 1.0배  |  하락(-1) → 0.7배
sdf['Trends_Factor'] = sdf['Trends_Norm'].apply(
    lambda x: round(1.0 + float(x) * 0.3, 4) if pd.notna(x) else None
)

# Composite Sentiment_Score (뉴스 중심, 트렌드 보정)
def compute_sentiment(row):
    nn = row.get('News_Sentiment_Norm')
    tf = row.get('Trends_Factor')
    tn = row.get('Trends_Norm')
    has_news   = nn is not None and not (isinstance(nn, float) and math.isnan(nn))
    has_trends = tf is not None and not (isinstance(tf, float) and math.isnan(tf))

    if has_news and has_trends:
        # 뉴스 점수 × 트렌드 증폭 계수 (부호는 뉴스가 결정, 세기만 조절)
        score = float(nn) * float(tf)
        return max(-1.0, min(1.0, score))  # [-1, +1] 유지
    elif has_news:
        return float(nn)
    elif has_trends:
        # 뉴스 없음: 트렌드만으로는 방향 불명 → 50% 가중
        return float(tn) * 0.5
    else:
        return None

sdf['Sentiment_Score'] = sdf.apply(compute_sentiment, axis=1)

tickers_with_sentiment = sdf['Sentiment_Score'].notna().sum()
print(f"  Tickers with Sentiment_Score: {tickers_with_sentiment} / {n_tickers}")

# Map back to main df
ticker_to_sentiment     = dict(zip(sdf['ticker'], sdf['Sentiment_Score']))
ticker_to_news_raw      = dict(zip(sdf['ticker'], sdf['News_Sentiment_Raw']))
ticker_to_news_cnt      = dict(zip(sdf['ticker'], sdf['News_Article_Count']))
ticker_to_google_cnt    = dict(zip(sdf['ticker'], sdf['Google_News_Count']))
ticker_to_top_headline  = dict(zip(sdf['ticker'], sdf['Top_Headline']))
ticker_to_top_hl_score  = dict(zip(sdf['ticker'], sdf['Top_Headline_Score']))
ticker_to_trend_mom     = dict(zip(sdf['ticker'], sdf['Trends_Momentum']))
ticker_to_trend_factor  = dict(zip(sdf['ticker'], sdf['Trends_Factor']))

df['News_Sentiment_Raw']  = df['_clean_ticker'].map(ticker_to_news_raw)
df['News_Article_Count']  = df['_clean_ticker'].map(ticker_to_news_cnt).fillna(0).astype(int)
df['Google_News_Count']   = df['_clean_ticker'].map(ticker_to_google_cnt).fillna(0).astype(int)
df['Top_Headline']        = df['_clean_ticker'].map(ticker_to_top_headline)
df['Top_Headline_Score']  = df['_clean_ticker'].map(ticker_to_top_hl_score)
df['Trends_Momentum']     = df['_clean_ticker'].map(ticker_to_trend_mom)
df['Trends_Factor']       = df['_clean_ticker'].map(ticker_to_trend_factor)
df['Sentiment_Score']     = df['_clean_ticker'].map(ticker_to_sentiment)

# ──────────────────────────────────────────────────────────────
# STEP 5 — Recompute Integrated_Score
# ──────────────────────────────────────────────────────────────
print("\n=== STEP 5: Recomputing Integrated_Score ===")

# Normalize Bond_TR_Est_pct → [-1, +1] within each class
df['Bond_TR_Score'] = np.nan
active_mask = df['Bond_TR_Est_pct'].notna() & df['class'].notna() & (df['class'] != 'off')
for cls in df.loc[active_mask, 'class'].unique():
    cls_mask = active_mask & (df['class'] == cls)
    tr_cls = df.loc[cls_mask, 'Bond_TR_Est_pct']
    n_tr = tr_cls.notna().sum()
    if n_tr > 1:
        tr_ranked = tr_cls.rank(method='average', na_option='keep')
        df.loc[cls_mask, 'Bond_TR_Score'] = (tr_ranked - 1) / (n_tr - 1) * 2 - 1

# Equal 0.25 weight per component, all in [-1, +1]
df['Integrated_Score'] = (
    df['Bond_TR_Score'].fillna(0) * 0.25
    + df['Eq_Mom_Score'].fillna(0) * 0.25
    + df['Eq_Fund_Score'].fillna(0) * 0.25
    + df['Sentiment_Score'].fillna(0) * 0.25
)

# Recompute ranks within class (excluding 'off')
df['Integrated_Rank_in_Class'] = np.nan

for cls in df['class'].dropna().unique():
    if cls == 'off':
        continue
    mask = (df['class'] == cls) & df['Integrated_Score'].notna()
    if mask.sum() == 0:
        continue
    df.loc[mask, 'Integrated_Rank_in_Class'] = (
        df.loc[mask, 'Integrated_Score']
        .rank(ascending=False, method='min')
        .astype(int)
    )

def top_pick_flag(rank):
    if pd.isna(rank):
        return ''
    r = int(rank)
    if r <= 3:
        return '★★★ TOP3'
    elif r <= 10:
        return '★★ TOP10'
    elif r <= 25:
        return '★ TOP25'
    return ''

df['Top_Pick_Flag'] = df['Integrated_Rank_in_Class'].apply(top_pick_flag)
print(f"  Integrated_Score recomputed for {df['Integrated_Score'].notna().sum()} rows")

# ──────────────────────────────────────────────────────────────
# STEP 6 — Update Detail_Scored (column-name-based, safe for re-runs)
# ──────────────────────────────────────────────────────────────
print("\n=== STEP 6: Updating Excel workbook ===")

wb = load_workbook(FILE)
ws_detail = wb['Detail_Scored']

# ── 6a. Build column-name → index map from header row (row 2) ──────────────
def build_col_map(ws):
    return {str(ws.cell(2, c).value).strip(): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(2, c).value}

def get_or_add_col(ws, col_map, col_name):
    """Return existing column index, or append new column at the end."""
    if col_name in col_map:
        return col_map[col_name]
    new_idx = ws.max_column + 1
    c = ws.cell(row=2, column=new_idx, value=col_name)
    c.font      = Font(name="Arial", bold=True, size=10)
    c.fill      = PatternFill("solid", fgColor=PURPLE_LIGHT)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    col_map[col_name] = new_idx
    return new_idx

col_map = build_col_map(ws_detail)
print(f"  Detail_Scored: {ws_detail.max_column} cols before update")

# Columns to write (will append if missing)
WRITE_COLS = [
    "Bond_TR_Score",
    "News_Sentiment_Raw",
    "News_Article_Count",
    "Google_News_Count",
    "Top_Headline",
    "Top_Headline_Score",
    "Trends_Momentum",
    "Trends_Factor",
    "Sentiment_Score",
    "Integrated_Score",
    "Integrated_Rank_in_Class",
    "Top_Pick_Flag",
    "Carry_2.5M_pct",
    "Compression_Score_pct",
    "Bond_TR_Est_pct",
]

col_idx = {name: get_or_add_col(ws_detail, col_map, name) for name in WRITE_COLS}
print(f"  Columns mapped: {col_idx}")

# Helper
def to_py(val):
    if val is None: return None
    if isinstance(val, float) and math.isnan(val): return None
    if isinstance(val, (np.integer,)): return int(val)
    if isinstance(val, (np.floating,)): return float(val)
    if isinstance(val, (np.bool_,)): return bool(val)
    return val

fill_top3  = PatternFill("solid", fgColor="FFFF00")
fill_top10 = PatternFill("solid", fgColor="FCE4D6")
fill_top25 = PatternFill("solid", fgColor="DDEBF7")
fill_none  = PatternFill(fill_type=None)
font_top3  = Font(name="Arial", bold=True, color="FF0000", size=10)
font_top10 = Font(name="Arial", bold=True, color="C55A11", size=10)
font_top25 = Font(name="Arial", bold=True, color="1F3864", size=10)
font_none  = Font(name="Arial", size=10)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── 6b. Write data rows ────────────────────────────────────────────────────
print("  Writing data rows...")

def _write(df_col, excel_col, fmt=None, int_cast=False):
    for df_idx in range(len(df)):
        er = df_idx + 3
        raw = df.at[df_idx, df_col] if df_col in df.columns else None
        val = to_py(int(raw) if int_cast and raw is not None and raw == raw else raw)
        c = ws_detail.cell(row=er, column=excel_col)
        c.value = val
        if fmt:
            c.number_format = fmt

_write("Bond_TR_Score",       col_idx["Bond_TR_Score"],       "0.0000")
_write("News_Sentiment_Raw",  col_idx["News_Sentiment_Raw"],  "0.00000")
_write("News_Article_Count",  col_idx["News_Article_Count"],  "0", int_cast=True)
_write("Google_News_Count",   col_idx["Google_News_Count"],   "0", int_cast=True)
_write("Top_Headline",        col_idx["Top_Headline"])
_write("Top_Headline_Score",  col_idx["Top_Headline_Score"],  "0.0000")
_write("Trends_Momentum",     col_idx["Trends_Momentum"],     "0.0000")
_write("Trends_Factor",       col_idx["Trends_Factor"],       "0.0000")
_write("Sentiment_Score",     col_idx["Sentiment_Score"],     "0.0000")
_write("Integrated_Score",    col_idx["Integrated_Score"],    "0.0000")
_write("Integrated_Rank_in_Class", col_idx["Integrated_Rank_in_Class"], "0")
_write("Carry_2.5M_pct",      col_idx["Carry_2.5M_pct"],     "0.0000")
_write("Compression_Score_pct",col_idx["Compression_Score_pct"],"0.0000")
_write("Bond_TR_Est_pct",     col_idx["Bond_TR_Est_pct"],     "0.0000")

# Top_Pick_Flag with color
flag_col = col_idx["Top_Pick_Flag"]
for df_idx in range(len(df)):
    er   = df_idx + 3
    flag = df.at[df_idx, "Top_Pick_Flag"]
    c    = ws_detail.cell(row=er, column=flag_col)
    c.value = str(flag) if flag else ""
    if flag == "★★★ TOP3":   c.font = font_top3;  c.fill = fill_top3
    elif flag == "★★ TOP10": c.font = font_top10; c.fill = fill_top10
    elif flag == "★ TOP25":  c.font = font_top25; c.fill = fill_top25
    else:                       c.font = font_none;  c.fill = fill_none

print(f"  Written {len(df)} rows to Detail_Scored")

# ── 6c. Delete obsolete sheets, keep canonical set ─────────────────────────
KEEP = {"Detail_Scored", "Methodology",
        "Score_BondTR", "Score_EqMom", "Score_EqFund",
        "Score_Sentiment", "Score_Integrated"}
for sname in list(wb.sheetnames):
    if sname not in KEEP:
        del wb[sname]
        print(f"  Deleted sheet: {sname}")

# ── 6d. Save ───────────────────────────────────────────────────────────────
print("\nSaving workbook...")
wb.save(FILE)
file_size = os.path.getsize(FILE)
print(f"  Saved: {FILE}  ({file_size/1024/1024:.2f} MB)")

# ── 6e. Rebuild Score_* sheets via build_score_sheets ─────────────────────
print("\nRebuilding Score sheets...")
import subprocess
build_script = os.path.join(os.path.dirname(os.path.abspath(FILE)), "build_score_sheets.py")
result = subprocess.run(
    [sys.executable, build_script],
    capture_output=True, text=True, encoding="utf-8"
)
print(result.stdout)
if result.returncode != 0:
    print("build_score_sheets.py ERROR:", result.stderr[:500])

# ── Summary ────────────────────────────────────────────────────────────────
print("\n=== DONE ===")
print(f"  Tickers processed:            {n_tickers}")
print(f"  Tickers with news data:       {tickers_with_news} / {n_tickers}")
print(f"  Tickers with trends data:     {tickers_with_trends} / {n_tickers}")
print(f"  Tickers with Sentiment_Score: {tickers_with_sentiment} / {n_tickers}")
non_null = df['Sentiment_Score'].notna().sum()
print(f"  Non-null Sentiment in df:     {non_null}")
file_size = os.path.getsize(FILE)
print(f"  Final file size: {file_size:,} bytes ({file_size/1024/1024:.2f} MB)")
