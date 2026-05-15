"""
score_monthly.py  ─  Full 5-Component Monthly Scoring Script
═══════════════════════════════════════════════════════════════
월별 LUACSTAT 원본 파일 → 전체 스코어 산출 → *_SCORED.xlsx 저장

Components:
  ① Bond_TR_Score   : Carry + Compression + DP_Rating  (Bloomberg data)
  ② AI_Macro_Score  : Sector × 0.40 + Maturity × 0.35 + RatingBuf × 0.25
  ③ Eq_Mom_Score    : 1M/3M return, 30D vol, 52W high  (yfinance)
  ④ Eq_Fund_Score   : D/E, Margin, Growth, CR, EV/EBITDA  (Yahoo Finance API)
  ⑤ Sentiment_Score : Yahoo News + Google News VADER + Google Trends

Usage:
  python score_monthly.py LUACSTAT_2026_05_11.xlsx
  (인수 없으면 폴더 내 최신 LUACSTAT_*.xlsx 자동 선택)
"""

import sys, os, re, io, math, time, warnings, glob
from datetime import date, datetime, timedelta, timezone
from collections import Counter
import xml.etree.ElementTree as ET
from email.utils import parsedate_to_datetime

import numpy as np
import pandas as pd
import requests
import yfinance as yf
import urllib3

from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.filterwarnings('ignore')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

BASE_DIR = r'C:\Users\sh.park\Documents\USIG_LLM'
sys.path.insert(0, BASE_DIR)
from ai_macro_score import compute_ai_macro_score, SUBGROUP_SCORE_MAP as _SUBMAP

# ══════════════════════════════════════════════════════════════
# 0. 파일 선택
# ══════════════════════════════════════════════════════════════
if len(sys.argv) > 1:
    IN_FILE = sys.argv[1]
    if not os.path.isabs(IN_FILE):
        IN_FILE = os.path.join(BASE_DIR, IN_FILE)
else:
    candidates = sorted([f for f in glob.glob(os.path.join(BASE_DIR, 'LUACSTAT_*.xlsx'))
                         if 'SCORED' not in f])
    if not candidates:
        raise FileNotFoundError("LUACSTAT_*.xlsx 파일을 찾을 수 없습니다.")
    IN_FILE = candidates[-1]

m = re.search(r'(\d{4}_\d{2}_\d{2})', os.path.basename(IN_FILE))
DATE_TAG = m.group(1) if m else 'UNKNOWN'
AS_OF    = date(int(DATE_TAG[:4]), int(DATE_TAG[5:7]), int(DATE_TAG[8:10])) if m else date.today()
AS_OF_DT = datetime(AS_OF.year, AS_OF.month, AS_OF.day, tzinfo=timezone.utc)
OUT_FILE = IN_FILE.replace('.xlsx', '_SCORED.xlsx')

print(f"\n{'='*65}")
print(f"  USIG Monthly Scoring  |  {os.path.basename(IN_FILE)}")
print(f"  As-of: {AS_OF}   Output: {os.path.basename(OUT_FILE)}")
print(f"{'='*65}\n")

# ══════════════════════════════════════════════════════════════
# 1. 파일 로드 (헤더 행 자동 감지)
# ══════════════════════════════════════════════════════════════
print("[1/7] Loading file...")
xl   = pd.ExcelFile(IN_FILE)
SHEET = xl.sheet_names[0]
raw_top = pd.read_excel(IN_FILE, sheet_name=SHEET, header=None, nrows=15)
hdr_row = next((i for i, row in raw_top.iterrows()
                if 'Des' in list(row.values) and 'ISIN' in list(row.values)), None)
if hdr_row is None:
    raise ValueError("헤더 행을 찾을 수 없습니다.")

df = pd.read_excel(IN_FILE, sheet_name=SHEET, header=hdr_row)
df = df[df['ISIN'].notna() & df['ISIN'].astype(str).str.match(r'^[A-Z]{2}\w+')].reset_index(drop=True)
for col in ['OAS','OASD','Yield to Worst','1Y Dflt','OAD','Cpn','LQA']:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')

mask = (df['class'].notna() & df['class'].astype(str).str.lower().ne('off')
        & df['OAS'].notna() & df['OASD'].notna())
print(f"  Loaded {len(df):,} bonds  |  Active: {mask.sum():,}")

# ══════════════════════════════════════════════════════════════
# 2. Bond_TR_Score
# ══════════════════════════════════════════════════════════════
print("\n[2/7] Bond_TR_Score...")

DP_RATING_MAP = {'AAA':1,'AA1':2,'AA2':3,'AA3':4,'A1':5,'A2':6,'A3':7,
                 'BAA1':8,'BAA2':9,'BAA3':10,'BA1':11,'BA2':12,'BA3':13,
                 'B1':14,'B2':15,'B3':16,'CAA1':17,'CAA2':18,'CAA3':19,'CA':20,'C':21}

def percentile_norm(series, ascending=True):
    n = series.notna().sum()
    if n < 2: return pd.Series(np.nan, index=series.index)
    ranked = series.rank(method='average', na_option='keep') if ascending \
             else (-series).rank(method='average', na_option='keep')
    return (ranked - 1) / (n - 1) * 2 - 1

cm = mask & df['Yield to Worst'].notna()
df.loc[cm, 'Carry_2.5M_pct'] = df.loc[cm, 'Yield to Worst'] / 12 * 2.5

cm2 = mask & df['1Y Dflt'].notna()
df.loc[cm2, 'Spread_floor_bp']       = df.loc[cm2, '1Y Dflt'] * 60
df.loc[cm2, 'Compression_gap_bp']    = df.loc[cm2, 'OAS'] - df.loc[cm2, 'Spread_floor_bp']
df.loc[cm2, 'Compression_Score_pct'] = df.loc[cm2, 'Compression_gap_bp'] * df.loc[cm2, 'OASD'] / 100

df['_dp_gap'] = df['DPSpreadRating'].map(DP_RATING_MAP) - df['DPFundamentalRating'].map(DP_RATING_MAP)
df['DP_Rating_Score'] = np.nan
df.loc[mask, 'DP_Rating_Score'] = percentile_norm(df.loc[mask, '_dp_gap']).values
df.drop(columns=['_dp_gap'], inplace=True)

df.loc[mask, 'Bond_TR_Est_pct'] = (
    df.loc[mask, 'Carry_2.5M_pct'].fillna(0) +
    df.loc[mask, 'Compression_Score_pct'].fillna(0) +
    df.loc[mask, 'DP_Rating_Score'].fillna(0) * 0.05
)
df['Bond_TR_Score'] = np.nan
for cls in df.loc[mask, 'class'].dropna().unique():
    cm3 = mask & (df['class'] == cls)
    df.loc[cm3, 'Bond_TR_Score'] = percentile_norm(df.loc[cm3, 'Bond_TR_Est_pct']).values

print(f"  Bond_TR_Score: {df['Bond_TR_Score'].notna().sum():,} non-null  "
      f"[{df['Bond_TR_Score'].min():.3f}, {df['Bond_TR_Score'].max():.3f}]")

# ══════════════════════════════════════════════════════════════
# 3. AI_Macro_Score
# ══════════════════════════════════════════════════════════════
print("\n[3/7] AI_Macro_Score...")
df = compute_ai_macro_score(df)

# ══════════════════════════════════════════════════════════════
# 4. Eq_Mom_Score  (yfinance)
# ══════════════════════════════════════════════════════════════
print("\n[4/7] Eq_Mom_Score (yfinance)...")

VALID_TK = re.compile(r'^[A-Z][A-Z0-9./\-]{0,9}$')

def pick_ticker(row):
    for col in ['Ticker', 'Parent Ticker', 'Eqty Ticker']:
        v = str(row.get(col, '')).strip().split()[0] if pd.notna(row.get(col)) else ''
        if v and VALID_TK.match(v):
            return v
    return None

df['_eq_ticker'] = df.apply(pick_ticker, axis=1)
tickers = [t for t in df['_eq_ticker'].dropna().unique() if t]
print(f"  Valid tickers: {len(tickers):,}")

p_start = (AS_OF - timedelta(days=400)).isoformat()
p_end   = AS_OF.isoformat()
price_dict = {}
CHUNK = 200
for i in range(0, len(tickers), CHUNK):
    chunk = tickers[i:i+CHUNK]
    try:
        raw = yf.download(chunk, start=p_start, end=p_end,
                          auto_adjust=True, progress=False, threads=True)
        close = raw['Close'] if 'Close' in raw.columns else raw.xs('Close', axis=1, level=0)
        for tk in chunk:
            if tk in close.columns:
                s = close[tk].dropna()
                if len(s) > 20:
                    price_dict[tk] = s
    except Exception as e:
        print(f"  chunk {i//CHUNK+1} warn: {e}")
print(f"  Price data fetched: {len(price_dict):,} tickers")

def eq_mom_metrics(ticker):
    if ticker not in price_dict: return np.nan, np.nan, np.nan, np.nan
    px  = price_dict[ticker]
    idx = px.index[px.index <= pd.Timestamp(AS_OF)]
    if not len(idx): return np.nan, np.nan, np.nan, np.nan
    last = float(px[idx[-1]])
    def ret(days):
        pi = px.index[px.index <= pd.Timestamp(AS_OF) - timedelta(days=days)]
        return last / float(px[pi[-1]]) - 1 if len(pi) else np.nan
    r1m  = ret(30);  r3m = ret(90)
    rec  = px[px.index >= pd.Timestamp(AS_OF) - timedelta(days=45)]
    vol30 = float(rec.pct_change().std() * np.sqrt(252)) if len(rec) > 5 else np.nan
    yr   = px[px.index >= pd.Timestamp(AS_OF) - timedelta(days=365)]
    vs52 = last / float(yr.max()) if len(yr) and yr.max() > 0 else np.nan
    return r1m, r3m, vol30, vs52

rows_eq = df['_eq_ticker'].apply(eq_mom_metrics)
df['Eq_Ret_1M']     = [r[0] for r in rows_eq]
df['Eq_Ret_3M']     = [r[1] for r in rows_eq]
df['Eq_Vol_30D']    = [r[2] for r in rows_eq]
df['Eq_vs_52w_High']= [r[3] for r in rows_eq]

def rank_norm(series): return percentile_norm(series, ascending=True)

eq_mask = mask & df['Eq_Ret_1M'].notna()
if eq_mask.sum() > 10:
    df.loc[eq_mask, 'Eq_Mom_Score'] = (
        rank_norm(df.loc[eq_mask, 'Eq_Ret_1M'])    * 0.35 +
        rank_norm(df.loc[eq_mask, 'Eq_Ret_3M'])    * 0.35 +
        rank_norm(-df.loc[eq_mask, 'Eq_Vol_30D'])  * 0.15 +
        rank_norm(df.loc[eq_mask, 'Eq_vs_52w_High'])* 0.15
    )
    print(f"  Eq_Mom_Score: {df['Eq_Mom_Score'].notna().sum():,} non-null  "
          f"[{df['Eq_Mom_Score'].min():.3f}, {df['Eq_Mom_Score'].max():.3f}]  "
          f"coverage={eq_mask.sum()/mask.sum()*100:.1f}%")
else:
    df['Eq_Mom_Score'] = np.nan
    print("  Eq_Mom_Score: 데이터 부족")

# ══════════════════════════════════════════════════════════════
# 5. Eq_Fund_Score  (Yahoo Finance API)
# ══════════════════════════════════════════════════════════════
print("\n[5/7] Eq_Fund_Score (Yahoo Finance API)...")

API_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'application/json',
    'Accept-Language': 'en-US,en;q=0.9',
}

def init_yf_session():
    s = requests.Session(); s.verify = False
    s.headers.update({'User-Agent': API_HEADERS['User-Agent']})
    s.get('https://fc.yahoo.com', timeout=10)
    r = s.get('https://query1.finance.yahoo.com/v1/test/getcrumb', timeout=10)
    crumb = r.text.strip() if r.status_code == 200 and r.text else ''
    return s, crumb

def fetch_fundamentals(session, crumb, ticker):
    modules = 'defaultKeyStatistics,financialData,summaryDetail'
    for base in ['https://query1.finance.yahoo.com', 'https://query2.finance.yahoo.com']:
        url = f'{base}/v10/finance/quoteSummary/{ticker}?modules={modules}&crumb={crumb}'
        try:
            r = session.get(url, headers=API_HEADERS, timeout=10)
            if r.status_code == 401: return None
            res = r.json().get('quoteSummary', {}).get('result', [])
            if res:
                fd = res[0].get('financialData', {})
                ks = res[0].get('defaultKeyStatistics', {})
                sd = res[0].get('summaryDetail', {})
                def g(d, k):
                    v = d.get(k, {}); return v.get('raw') if isinstance(v, dict) else None
                result = {
                    'Debt_to_Equity': g(fd,'debtToEquity'),
                    'Profit_Margin':  g(fd,'profitMargins'),
                    'Revenue_Growth': g(fd,'revenueGrowth'),
                    'Current_Ratio':  g(fd,'currentRatio'),
                    'EV_EBITDA':      g(ks,'enterpriseToEbitda'),
                    'PE_Ratio':       g(sd,'trailingPE'),
                }
                if any(v is not None for v in result.values()):
                    return result
        except Exception:
            continue
    return None

# Unique tickers for fundamentals
VALID_FUND_TK = re.compile(r'^[A-Z][A-Z0-9\.\-]{0,9}$')
def clean_for_yahoo(raw):
    if pd.isna(raw): return None
    s = str(raw).strip()
    s = re.sub(r'\s+[A-Z]{2,3}$', '', s).strip()
    return s if VALID_FUND_TK.match(s) else None

df['_fund_ticker'] = df['Eqty Ticker'].apply(clean_for_yahoo)
# Fallback to Ticker / Parent Ticker
for col in ['Ticker', 'Parent Ticker']:
    need = df['_fund_ticker'].isna()
    df.loc[need, '_fund_ticker'] = df.loc[need, col].apply(clean_for_yahoo)

fund_tickers = [t for t in df['_fund_ticker'].dropna().unique() if t]
print(f"  Fund tickers to query: {len(fund_tickers):,}")

sess_fund, crumb_fund = init_yf_session()
print(f"  Yahoo crumb: {'OK' if crumb_fund else 'failed (will retry)'}")

fundamentals = {}
ok = fail = 0
for i, tk in enumerate(fund_tickers):
    if i > 0 and i % 200 == 0:
        print(f"  Fund progress: {i}/{len(fund_tickers)}  ok={ok}")
        try: sess_fund, crumb_fund = init_yf_session()
        except: pass
    res = fetch_fundamentals(sess_fund, crumb_fund, tk)
    if res is None:
        try: sess_fund, crumb_fund = init_yf_session()
        except: pass
        res = fetch_fundamentals(sess_fund, crumb_fund, tk)
    if res:
        fundamentals[tk] = res; ok += 1
    else:
        fail += 1
    time.sleep(0.05)
    if i % 30 == 29: time.sleep(0.5)

print(f"  Fundamentals fetched: {ok}/{len(fund_tickers)} ({ok/max(len(fund_tickers),1)*100:.1f}%)")

FUND_COLS = ['Debt_to_Equity','Profit_Margin','Revenue_Growth','Current_Ratio','EV_EBITDA','PE_Ratio']
for col in FUND_COLS + ['Eq_Fund_Score']:
    df[col] = np.nan

if fundamentals:
    fund_df = pd.DataFrame.from_dict(fundamentals, orient='index').reset_index()
    fund_df.columns = ['ticker'] + [c for c in fund_df.columns[1:]]
    # Apply caps
    if 'Current_Ratio' in fund_df: fund_df['Current_Ratio'] = fund_df['Current_Ratio'].clip(upper=3.0)
    if 'PE_Ratio'      in fund_df:
        fund_df['PE_Ratio'] = fund_df['PE_Ratio'].where(fund_df['PE_Ratio'] > 0).clip(upper=100.0)
    if 'EV_EBITDA'     in fund_df:
        fund_df['EV_EBITDA'] = fund_df['EV_EBITDA'].where(fund_df['EV_EBITDA'] > 0)
    # Rank-normalize
    def pct_rank(s, higher_is_better=True):
        n = s.notna().sum()
        if n < 2: return pd.Series(np.nan, index=s.index)
        r = s.rank(method='average', na_option='keep') if higher_is_better \
            else (-s).rank(method='average', na_option='keep')
        return (r - 1) / (n - 1)
    norm_cols = {col: pct_rank(fund_df[col], hib) for col, hib in [
        ('Debt_to_Equity',False),('Profit_Margin',True),('Revenue_Growth',True),
        ('Current_Ratio',True),('EV_EBITDA',False),('PE_Ratio',False)
    ] if col in fund_df.columns}
    norm_df = pd.DataFrame(norm_cols, index=fund_df.index)
    norm_df['ticker'] = fund_df['ticker']
    avail = norm_df[[c for c in FUND_COLS if c in norm_df]].notna().sum(axis=1)
    raw_s  = norm_df[[c for c in FUND_COLS if c in norm_df]].mean(axis=1)
    norm_df['Eq_Fund_Score'] = (raw_s * 2 - 1).where(avail >= 2)
    fund_lookup = dict(zip(norm_df['ticker'], norm_df['Eq_Fund_Score']))
    for fc in FUND_COLS:
        if fc in fund_df.columns:
            fl = dict(zip(fund_df['ticker'], fund_df[fc]))
            df[fc] = df['_fund_ticker'].map(fl)
    df['Eq_Fund_Score'] = df['_fund_ticker'].map(fund_lookup)

n_fund = df['Eq_Fund_Score'].notna().sum()
print(f"  Eq_Fund_Score: {n_fund:,} non-null  "
      f"[{df['Eq_Fund_Score'].min():.3f}, {df['Eq_Fund_Score'].max():.3f}]" if n_fund else "  Eq_Fund_Score: 0")

# ══════════════════════════════════════════════════════════════
# 6. Sentiment_Score  (Yahoo News + Google News + VADER + Trends)
# ══════════════════════════════════════════════════════════════
print("\n[6/7] Sentiment_Score (Yahoo/Google News VADER + Trends)...")

analyzer = SentimentIntensityAnalyzer()
sess_news = requests.Session()
sess_news.verify = False
sess_news.headers.update({'User-Agent': API_HEADERS['User-Agent'],
                          'Accept': 'application/json'})
try:
    sess_news.get('https://fc.yahoo.com', timeout=10)
    r2 = sess_news.get('https://query1.finance.yahoo.com/v1/test/getcrumb', timeout=10)
    news_crumb = r2.text.strip() if r2.status_code == 200 else None
except:
    news_crumb = None
print(f"  News crumb: {'OK' if news_crumb else 'N/A'}")

def fetch_yahoo_news(ticker, crumb=None):
    urls = [
        f'https://query1.finance.yahoo.com/v1/finance/search?q={ticker}&newsCount=20&quotesCount=0',
        f'https://query2.finance.yahoo.com/v1/finance/search?q={ticker}&newsCount=20&quotesCount=0',
    ]
    if crumb: urls = [u + f'&crumb={crumb}' for u in urls]
    for url in urls:
        try:
            r = sess_news.get(url, timeout=12)
            if r.status_code == 200:
                items = r.json().get('news', [])
                hs = [{'title': it['title'], 'pub_time': it.get('providerPublishTime')}
                      for it in items[:20] if it.get('title')]
                if hs: return hs
        except: pass
    # RSS fallback
    try:
        r = sess_news.get(f'https://feeds.finance.yahoo.com/rss/2.0/headline?s={ticker}&region=US&lang=en-US', timeout=12)
        if r.status_code == 200:
            root = ET.fromstring(r.text)
            hs = []
            for it in root.findall('.//item')[:20]:
                te = it.find('title'); pe = it.find('pubDate')
                if te is not None and te.text:
                    pt = None
                    if pe is not None and pe.text:
                        try: pt = int(parsedate_to_datetime(pe.text).timestamp())
                        except: pass
                    hs.append({'title': te.text, 'pub_time': pt})
            if hs: return hs
    except: pass
    return []

def fetch_google_news(ticker, company=''):
    hs = []
    for q in ([ticker] + ([company] if company and company != ticker else [])):
        try:
            qe = requests.utils.quote(f'{q} stock')
            r  = sess_news.get(f'https://news.google.com/rss/search?q={qe}&hl=en-US&gl=US&ceid=US:en', timeout=12)
            if r.status_code == 200 and '<item>' in r.text:
                root = ET.fromstring(r.text)
                for it in root.findall('.//item')[:15]:
                    te = it.find('title'); pe = it.find('pubDate')
                    if te is not None and te.text:
                        pt = None
                        if pe is not None and pe.text:
                            try: pt = int(parsedate_to_datetime(pe.text).timestamp())
                            except: pass
                        hs.append({'title': te.text.strip(), 'pub_time': pt})
                if hs: return hs
        except: pass
    return hs

def vader_score(headlines):
    if not headlines: return None, 0, '', None
    ts = AS_OF_DT.timestamp()
    scores, weights, titles = [], [], []
    for h in headlines:
        c = analyzer.polarity_scores(h['title'])['compound']
        days = max(0, (ts - h['pub_time']) / 86400) if h.get('pub_time') else 7
        w = math.exp(-days / 14.0)
        scores.append(c); weights.append(w); titles.append(h['title'])
    tw = sum(weights)
    if tw == 0: return None, len(headlines), '', None
    wm = sum(s*w for s, w in zip(scores, weights)) / tw
    mi = max(range(len(scores)), key=lambda i: abs(scores[i]*weights[i]))
    return wm, len(headlines), titles[mi], round(scores[mi], 4)

# Build ticker→company map
df['_clean_news_ticker'] = df['Eqty Ticker'].apply(clean_for_yahoo)
ticker_to_company = {}
for _, row in df[['_clean_news_ticker','Company Name']].dropna(subset=['_clean_news_ticker']).iterrows():
    t = row['_clean_news_ticker']
    if t and t not in ticker_to_company:
        ticker_to_company[t] = str(row.get('Company Name', ''))

news_tickers = [t for t in df['_clean_news_ticker'].dropna().unique() if t]
print(f"  News tickers: {len(news_tickers):,}")

news_results = {}
for i, tk in enumerate(news_tickers):
    if i % 100 == 0:
        print(f"  News progress: {i}/{len(news_tickers)}...")
    try:
        yh = fetch_yahoo_news(tk, news_crumb)
        gn = fetch_google_news(tk, ticker_to_company.get(tk, ''))
        existing = {h['title'] for h in yh}
        merged = yh + [h for h in gn if h['title'] not in existing]
        raw, cnt, top_h, top_s = vader_score(merged)
        news_results[tk] = {'raw': raw, 'count': cnt, 'google_count': len(gn),
                            'top_headline': top_h, 'top_score': top_s}
    except:
        news_results[tk] = {'raw': None, 'count': 0, 'google_count': 0,
                            'top_headline': '', 'top_score': None}
    if (i+1) % 30 == 0: time.sleep(0.3)

# Generic 무효화
raw_cnt = Counter(round(v['raw'],5) for v in news_results.values() if v['raw'] is not None)
generic = {val for val, cnt in raw_cnt.items() if cnt > 5}
n_inv = 0
for t in news_results:
    if news_results[t]['raw'] is not None and round(news_results[t]['raw'],5) in generic:
        news_results[t]['raw'] = None; n_inv += 1
print(f"  News collected. Generic 무효화: {n_inv} tickers")

# Google Trends
trends_results = {}
try:
    from pytrends.request import TrendReq
    pt = TrendReq(hl='en-US', tz=360, timeout=(10,25), retries=2, backoff_factor=0.5,
                  requests_args={'verify': False})
    grps = [news_tickers[i:i+5] for i in range(0, len(news_tickers), 5)]
    for gi, grp in enumerate(grps):
        if gi % 20 == 0: print(f"  Trends: {gi}/{len(grps)} groups...")
        try:
            pt.build_payload(grp, cat=0, timeframe='today 3-m', geo='US')
            data = pt.interest_over_time()
            if data is not None and not data.empty:
                if 'isPartial' in data.columns: data = data.drop(columns=['isPartial'])
                data = data.sort_index()
                for t in grp:
                    if t in data.columns and len(data) >= 8:
                        s = data[t].values.astype(float)
                        r4 = float(np.mean(s[-4:])); p4 = float(np.mean(s[-8:-4]))
                        trends_results[t] = {'momentum': (r4 - p4) / (p4 + 1.0)}
        except: pass
        time.sleep(1.0)
    print(f"  Trends: {len(trends_results)} tickers")
except ImportError:
    print("  pytrends not installed — skipping Trends")

# Map sentiment to df
df['News_Sentiment_Raw']   = np.nan
df['News_Article_Count']   = np.nan
df['Google_News_Count']    = np.nan
df['Top_Headline']         = ''
df['Top_Headline_Score']   = np.nan
df['Trends_Momentum']      = np.nan
df['Trends_Factor']        = np.nan
df['News_Generic_Flag']    = ''

TRENDS_SCALE = 0.3
for idx, row in df.iterrows():
    t = row.get('_clean_news_ticker')
    if not t or t not in news_results: continue
    nr = news_results[t]
    df.at[idx, 'News_Sentiment_Raw'] = nr['raw']
    df.at[idx, 'News_Article_Count'] = nr['count']
    df.at[idx, 'Google_News_Count']  = nr['google_count']
    df.at[idx, 'Top_Headline']       = nr['top_headline'] or ''
    df.at[idx, 'Top_Headline_Score'] = nr['top_score']
    if t in trends_results and trends_results[t].get('momentum') is not None:
        tm = trends_results[t]['momentum']
        df.at[idx, 'Trends_Momentum'] = tm
        df.at[idx, 'Trends_Factor']   = np.clip(tm * TRENDS_SCALE, -1, 1)
    if nr['raw'] is None and nr['count'] > 0:
        df.at[idx, 'News_Generic_Flag'] = 'GENERIC (invalidated)'
    elif nr['raw'] is None:
        df.at[idx, 'News_Generic_Flag'] = 'No news data'

# Sentiment_Score = VADER + Trends 합산
df['Sentiment_Score'] = np.nan
has_news = df['News_Sentiment_Raw'].notna()
df.loc[has_news, 'Sentiment_Score'] = (
    df.loc[has_news, 'News_Sentiment_Raw'] * 0.70 +
    df.loc[has_news, 'Trends_Factor'].fillna(0) * 0.30
).clip(-1, 1)

# Cross-sectional rank normalize
sent_mask = mask & df['Sentiment_Score'].notna()
if sent_mask.sum() > 10:
    df.loc[sent_mask, 'Sentiment_Score_clean'] = percentile_norm(df.loc[sent_mask, 'Sentiment_Score']).values
else:
    df['Sentiment_Score_clean'] = np.nan

n_sent = df['Sentiment_Score_clean'].notna().sum()
print(f"  Sentiment_Score_clean: {n_sent:,} non-null")

# ══════════════════════════════════════════════════════════════
# 7. Integrated_Score
# ══════════════════════════════════════════════════════════════
print("\n[7/7] Integrated_Score + writing output...")

df['Integrated_Score'] = (
    df['Bond_TR_Score'].fillna(0)          * 0.20 +
    df['Eq_Mom_Score'].fillna(0)           * 0.20 +
    df['Eq_Fund_Score'].fillna(0)          * 0.20 +
    df.get('Sentiment_Score_clean', pd.Series(0, index=df.index)).fillna(0) * 0.20 +
    df['AI_Macro_Score'].fillna(0)         * 0.20
)
df.loc[~mask, 'Integrated_Score'] = np.nan

df['Integrated_Rank_in_Class'] = np.nan
for cls in df.loc[mask, 'class'].dropna().unique():
    cm4 = mask & (df['class'] == cls)
    df.loc[cm4, 'Integrated_Rank_in_Class'] = df.loc[cm4, 'Integrated_Score'].rank(ascending=False, method='min')

def top_flag(r):
    if pd.isna(r): return ''
    r = int(r)
    return '★★★ TOP3' if r <= 3 else ('★★ TOP10' if r <= 10 else ('★ TOP25' if r <= 25 else ''))
df['Top_Pick_Flag'] = df['Integrated_Rank_in_Class'].apply(top_flag)

df_out = (df[mask].sort_values(['class','Integrated_Rank_in_Class'], na_position='last')
                  .reset_index(drop=True))
print(f"  Active bonds: {len(df_out):,}  |  TOP3: {(df['Top_Pick_Flag']=='★★★ TOP3').sum()}")

# ══════════════════════════════════════════════════════════════
# 8. Write SCORED.xlsx
# ══════════════════════════════════════════════════════════════
FILL_TITLE = PatternFill('solid', fgColor='1F3864')
FILL_ID    = PatternFill('solid', fgColor='BDD7EE')
FILL_COMP  = PatternFill('solid', fgColor='FFF2CC')
FILL_SCORE = PatternFill('solid', fgColor='E2EFDA')
FILL_POS   = PatternFill('solid', fgColor='C6EFCE')
FILL_NEG   = PatternFill('solid', fgColor='FFC7CE')
FILL_TOP3  = PatternFill('solid', fgColor='FFFF00')
FILL_TOP10 = PatternFill('solid', fgColor='FCE4D6')
FILL_TOP25 = PatternFill('solid', fgColor='DDEBF7')
FILL_AI_H  = PatternFill('solid', fgColor='4A148C')
FILL_AI    = PatternFill('solid', fgColor='EDE7F6')
FILL_MV_HDR   = PatternFill('solid', fgColor='4A148C')
FILL_MV_BG    = PatternFill('solid', fgColor='F3E5F5')
FILL_MV_THEME = PatternFill('solid', fgColor='311B92')
FILL_MV_POS   = PatternFill('solid', fgColor='C8E6C9')
FILL_MV_NEG   = PatternFill('solid', fgColor='FFCDD2')
FILL_MV_NEUT  = PatternFill('solid', fgColor='F5F5F5')
FILL_GENERIC  = PatternFill('solid', fgColor='FFE0E0')
FILL_NODATA   = PatternFill('solid', fgColor='F2F2F2')
FONT_TOP3     = Font(name='Arial', bold=True, color='FF0000', size=10)
FONT_TOP10    = Font(name='Arial', bold=True, color='C55A11', size=10)
FONT_TOP25    = Font(name='Arial', bold=True, color='1F3864', size=10)

def _v(val):
    if val is None: return None
    try:
        if pd.isna(val): return None
    except: pass
    if isinstance(val, (np.integer,)): return int(val)
    if isinstance(val, (np.floating,)): return float(val)
    if isinstance(val, pd.Timestamp): return val.to_pydatetime()
    return val

wb = Workbook(); wb.remove(wb.active)

def mks(name): return wb.create_sheet(name)

def wtitle(ws, text, n):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill = FILL_TITLE
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

def whdrs(ws, hdrs, id_s, comp_s, sc_s):
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.fill = FILL_SCORE if h in sc_s else (FILL_COMP if h in comp_s else FILL_ID)

def wrows(ws, hdrs, sc_s, fmt, flag_col=None, color_cols=None):
    # 헤더 행에 number_format 미리 설정 (열 전체 적용 → 셀마다 반복 불필요)
    fmt_ci = {ci: fmt[h] for ci, h in enumerate(hdrs, 1) if h in fmt}
    for ci, nf in fmt_ci.items():
        ws.column_dimensions[get_column_letter(ci)].number_format = nf

    for ri, row in df_out.iterrows():
        er = ri + 3
        flag_val = str(_v(row.get(flag_col, '')) or '') if flag_col else ''
        is_top3  = flag_val == '★★★ TOP3'
        is_top10 = flag_val == '★★ TOP10'
        is_top25 = flag_val == '★ TOP25'

        for ci, h in enumerate(hdrs, 1):
            val = _v(row.get(h, np.nan) if h in row.index else np.nan)
            c = ws.cell(row=er, column=ci)
            c.value = val
            # number format (필요한 셀만)
            if ci in fmt_ci:
                c.number_format = fmt_ci[ci]
            # TOP 플래그 열 서식
            if flag_col and h == flag_col:
                c.value = flag_val
                if   is_top3:  c.font = FONT_TOP3;  c.fill = FILL_TOP3
                elif is_top10: c.font = FONT_TOP10; c.fill = FILL_TOP10
                elif is_top25: c.font = FONT_TOP25; c.fill = FILL_TOP25
                continue
            # score 열 색상 (TOP 행만 추가 색상, 나머지는 양수/음수만)
            if color_cols and h in color_cols and val is not None:
                try:
                    fv = float(val)
                    if not math.isnan(fv):
                        if   fv > 0.3:  c.fill = FILL_POS
                        elif fv < -0.3: c.fill = FILL_NEG
                except: pass

def wfin(ws, n, widths):
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(n)}2'
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

# ── Sheet 1: Score_BondTR ─────────────────────────────────────
ws = mks('Score_BondTR')
id_c = ['class','Des','ISIN','Ticker','Cpn','OAS','OASD']
cp_c = ['Carry_2.5M_pct','Compression_Score_pct','DPFundamentalRating','DPSpreadRating','DP_Rating_Score']
sc_c = ['Bond_TR_Est_pct','Bond_TR_Score']
hdrs = id_c + cp_c + sc_c
wtitle(ws,'Bond TR Score  │  Carry + Compression + DP Rating → Bond_TR_Score [-1,+1]',len(hdrs))
whdrs(ws, hdrs, set(id_c), set(cp_c), set(sc_c))
wrows(ws, hdrs, set(sc_c),
      {'Cpn':'0.000','OAS':'0.0','OASD':'0.00','Carry_2.5M_pct':'0.0000',
       'Compression_Score_pct':'0.0000','DP_Rating_Score':'0.0000',
       'Bond_TR_Est_pct':'0.0000','Bond_TR_Score':'0.0000'})
wfin(ws,len(hdrs),{1:12,2:28,3:16,4:10,5:7,6:7,7:7,8:11,9:13,10:16,11:16,12:11,13:13,14:13})

# ── Sheet 2: Score_EqMom ──────────────────────────────────────
ws = mks('Score_EqMom')
id_c = ['class','Des','ISIN','Ticker','Eqty Ticker']
cp_c = ['Eq_Ret_1M','Eq_Ret_3M','Eq_Vol_30D','Eq_vs_52w_High']
sc_c = ['Eq_Mom_Score']
hdrs = id_c + cp_c + sc_c
wtitle(ws,f'Equity Momentum Score  │  1M/3M Return + Vol + 52W High  |  As of {AS_OF}',len(hdrs))
whdrs(ws, hdrs, set(id_c), set(cp_c), set(sc_c))
wrows(ws, hdrs, set(sc_c),
      {'Eq_Ret_1M':'0.00%','Eq_Ret_3M':'0.00%','Eq_Vol_30D':'0.00%',
       'Eq_vs_52w_High':'0.00%','Eq_Mom_Score':'0.0000'})
wfin(ws,len(hdrs),{1:12,2:28,3:16,4:10,5:12,6:11,7:11,8:11,9:13,10:13})

# ── Sheet 3: Score_EqFund ─────────────────────────────────────
ws = mks('Score_EqFund')
id_c = ['class','Des','ISIN','Ticker','Eqty Ticker']
cp_c = ['Debt_to_Equity','Profit_Margin','Revenue_Growth','Current_Ratio','EV_EBITDA','PE_Ratio']
sc_c = ['Eq_Fund_Score']
hdrs = id_c + cp_c + sc_c
wtitle(ws,'Equity Fundamental Score  │  D/E · Margin · Growth · CR · EV/EBITDA · P/E → Eq_Fund_Score [-1,+1]',len(hdrs))
whdrs(ws, hdrs, set(id_c), set(cp_c), set(sc_c))
wrows(ws, hdrs, set(sc_c),
      {'Debt_to_Equity':'0.00','Profit_Margin':'0.00%','Revenue_Growth':'0.00%',
       'Current_Ratio':'0.00','EV_EBITDA':'0.00','PE_Ratio':'0.00','Eq_Fund_Score':'0.0000'})
wfin(ws,len(hdrs),{1:12,2:28,3:16,4:10,5:12,6:11,7:11,8:11,9:11,10:11,11:11,12:13})

# ── Sheet 4: Score_Sentiment ──────────────────────────────────
ws = mks('Score_Sentiment')
id_c = ['class','Des','ISIN','Ticker','Eqty Ticker']
cp_c = ['News_Sentiment_Raw','News_Article_Count','Google_News_Count','News_Generic_Flag',
        'Top_Headline','Top_Headline_Score','Trends_Momentum','Trends_Factor',
        'Sentiment_Score','Sentiment_Score_clean']
sc_c = ['Sentiment_Score_clean']
hdrs = id_c + cp_c
wtitle(ws,'Sentiment Score  │  Yahoo/Google News (VADER, recency-weighted) + Google Trends',len(hdrs))
whdrs(ws, hdrs, set(id_c), set(cp_c), set(sc_c))
sent_fmt = {'News_Sentiment_Raw':'0.00000','News_Article_Count':'0','Google_News_Count':'0',
            'Top_Headline_Score':'0.0000','Trends_Momentum':'0.0000','Trends_Factor':'0.0000',
            'Sentiment_Score':'0.0000','Sentiment_Score_clean':'0.0000'}
sent_fmt_ci = {ci: sent_fmt[h] for ci, h in enumerate(hdrs, 1) if h in sent_fmt}
flag_ci   = next((ci for ci, h in enumerate(hdrs,1) if h == 'News_Generic_Flag'), None)
hline_ci  = next((ci for ci, h in enumerate(hdrs,1) if h == 'Top_Headline'), None)
hlsco_ci  = next((ci for ci, h in enumerate(hdrs,1) if h == 'Top_Headline_Score'), None)
for ri, row in df_out.iterrows():
    er = ri + 3
    for ci, h in enumerate(hdrs, 1):
        val = _v(row.get(h, np.nan) if h in row.index else np.nan)
        c = ws.cell(row=er, column=ci); c.value = val
        if ci in sent_fmt_ci: c.number_format = sent_fmt_ci[ci]
        if flag_ci and ci == flag_ci:
            flag = str(val) if val else ''
            if   flag == 'GENERIC (invalidated)': c.fill = FILL_GENERIC; c.font = Font(bold=True, color='CC0000', size=10)
            elif flag == 'No news data':           c.fill = FILL_NODATA
        if hline_ci and ci == hline_ci:
            c.alignment = Alignment(wrap_text=True, vertical='center')
        if hlsco_ci and ci == hlsco_ci and val is not None:
            try:
                fv = float(val)
                if   fv >=  0.5: c.fill = PatternFill('solid',fgColor='C6EFCE'); c.font=Font(bold=True,color='375623',size=10)
                elif fv <= -0.5: c.fill = PatternFill('solid',fgColor='FFC7CE'); c.font=Font(bold=True,color='9C0006',size=10)
            except: pass
wfin(ws,len(hdrs),{1:12,2:28,3:16,4:10,5:12,6:13,7:10,8:12,9:22,10:60,11:11,12:13,13:13,14:13,15:13})

# ── Sheet 5: Score_AI ─────────────────────────────────────────
ws = mks('Score_AI')
id_c_ai  = ['class','Des','ISIN','Ticker','OAD','BCLASS3','Industry Subgroup','DPFundamentalRating','Issuer Rtg']
cp_c_ai  = ['AI_Sector_Score','AI_Maturity_Score','AI_RatingBuf_Score']
sc_c_ai  = ['AI_Macro_Score']
hdrs_ai  = id_c_ai + cp_c_ai + sc_c_ai
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(hdrs_ai))
tc = ws.cell(row=1,column=1, value=f'AI Macro Score  |  Sector×0.40 + Maturity×0.35 + RatingBuf×0.25  |  As of {AS_OF}')
tc.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
tc.fill = FILL_AI_H; tc.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 22
for ci, h in enumerate(hdrs_ai, 1):
    c = ws.cell(row=2, column=ci, value=h)
    c.font = Font(name='Arial', bold=True, size=10)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.fill = FILL_AI if h in sc_c_ai else (FILL_COMP if h in cp_c_ai else FILL_ID)
fmt_ai = {'OAD':'0.00','AI_Sector_Score':'0.00','AI_Maturity_Score':'0.00',
          'AI_RatingBuf_Score':'0.00','AI_Macro_Score':'0.0000'}
ai_fmt_ci = {ci: fmt_ai[h] for ci, h in enumerate(hdrs_ai,1) if h in fmt_ai}
ai_sc_ci  = next((ci for ci,h in enumerate(hdrs_ai,1) if h=='AI_Macro_Score'), None)
ai_cp_cis = {ci for ci,h in enumerate(hdrs_ai,1) if h in cp_c_ai}
for ri, row in df_out.iterrows():
    er = ri + 3
    for ci, h in enumerate(hdrs_ai, 1):
        val = _v(row.get(h, np.nan) if h in row.index else np.nan)
        c = ws.cell(row=er, column=ci); c.value = val
        if ci in ai_fmt_ci: c.number_format = ai_fmt_ci[ci]
        if ai_sc_ci and ci == ai_sc_ci and val is not None:
            try:
                fv = float(val)
                c.fill = FILL_POS if fv > 0.3 else (FILL_NEG if fv < -0.3 else FILL_AI)
            except: pass
        if ci in ai_cp_cis and val is not None:
            try:
                fv = float(val)
                if   fv > 0: c.fill = PatternFill('solid', fgColor='E8F5E9')
                elif fv < 0: c.fill = PatternFill('solid', fgColor='FFEBEE')
            except: pass
ws.freeze_panes = 'A3'
ws.auto_filter.ref = f'A2:{get_column_letter(len(hdrs_ai))}2'
for ci, w in {1:12,2:28,3:16,4:10,5:7,6:18,7:26,8:16,9:12,10:14,11:14,12:14,13:14}.items():
    ws.column_dimensions[get_column_letter(ci)].width = w

# Macro View 섹션 (Score_AI 오른쪽)
_rationale = {}
with open(os.path.join(BASE_DIR,'ai_macro_score.py'), 'r', encoding='utf-8') as _f:
    for _line in _f:
        _m2 = re.match(r"\s+'([^']+)':\s*([-\d.]+),\s*#\s*(.+)", _line)
        if _m2: _rationale[_m2.group(1)] = _m2.group(3).strip()

MV = len(hdrs_ai) + 2
MACRO_THEMES = [
    ("① Fed Policy & Tariff Inflation Risk",
     "[Current] Fed in easing cycle (-100bp since Sep 2024) but Trump tariffs (10% universal, 25% autos/steel/aluminum) reigniting inflation. Core PCE sticky above 2.5%.\n"
     "[Risk] Tariff pass-through → Fed forced to pause. Stagflation-lite: growth slows, rates stay high. Key tail risk.\n"
     "[Positioning] Shorter duration, high-carry defensives. Avoid cyclicals with thin margins/high import exposure. Favor regulated utilities, domestic healthcare, infrastructure."),
    ("② IG Spread Level — Historically Tight, Carry Dominates",
     "[Current] US IG OAS near post-GFC tights (~85-95bp Q1 2026). Asymmetric risk/reward — downside (widening) >> upside.\n"
     "[Implication] Total return increasingly carry-driven. Selection alpha from superior carry-to-risk profiles within each rating/maturity bucket.\n"
     "[Positioning] Bond_TR Score: carry (YTW-based) + DP rating buffer. AI_Macro: avoid sectors with elevated spread widening risk."),
    ("③ Yield Curve — Steepening Pressure, Front-End Anchored",
     "[Current] Bear-steepening: 10Y-2Y widening as fiscal deficits push up term premium. 2Y anchored by Fed; 30Y faces pressure from Treasury supply.\n"
     "[Sweet Spot] OAD 4-7Y: carry + spread duration without long-end term premium risk. OAD 7-10Y acceptable. OAD 13Y+ penalized.\n"
     "[Positioning] Maturity_Score: peak +1.0 at OAD 4-7, drops to -0.50/-1.00 for 13Y+. Avoid long-dated BBB."),
    ("④ Tariff Exposure by Sector — Winners & Losers",
     "[Losers] Autos/Parts (-0.70 to -0.90): 25% import tariff + parts supply chain disruption. Retail Apparel/Dept (-0.40 to -0.60): China-sourced inventory → margin compression. Consumer Mfg (-0.40 to -0.55): direct COGS impact.\n"
     "[Winners] Regulated Utilities (+0.75 to +0.85): pass-through pricing, immune to trade policy. Domestic Healthcare (+0.60 to +0.75): domestic delivery, no import exposure. Defense (+0.35 to +0.40): spending tailwinds. Domestic Steel (+0.20): protected by 25% tariff.\n"
     "[Monitor] Semiconductors: US-China export controls tightening; supply chain restructuring ongoing."),
]
ws.merge_cells(start_row=1, start_column=MV, end_row=1, end_column=MV+3)
mc = ws.cell(row=1, column=MV, value='MACRO VIEW  |  AI Sector Score Rationale  |  April 2026 Macro Environment')
mc.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
mc.fill = FILL_MV_HDR; mc.alignment = Alignment(horizontal='center', vertical='center')
for ti, (label, body) in enumerate(MACRO_THEMES):
    tr = ti + 2
    lc = ws.cell(row=tr, column=MV, value=label)
    lc.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
    lc.fill = FILL_MV_THEME; lc.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    ws.merge_cells(start_row=tr, start_column=MV+1, end_row=tr, end_column=MV+3)
    bc = ws.cell(row=tr, column=MV+1, value=body)
    bc.font = Font(name='Arial', size=8.5); bc.fill = FILL_MV_BG
    bc.alignment = Alignment(vertical='top', wrap_text=True)
mv_subhdr = 6
for ci_off,(h,f) in enumerate([('Industry Subgroup',FILL_ID),('Score',FILL_AI),('Rationale',FILL_COMP),('Direction',FILL_AI)]):
    c = ws.cell(row=mv_subhdr, column=MV+ci_off, value=h)
    c.font = Font(name='Arial', bold=True, size=9); c.fill = f
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for si,(sg,sc) in enumerate(sorted(_SUBMAP.items(), key=lambda x: x[1], reverse=True)):
    dr = mv_subhdr + 1 + si
    direction = '▲ Overweight' if sc>0.3 else ('▼ Underweight' if sc<-0.3 else '— Neutral')
    rf = FILL_MV_POS if sc>0.3 else (FILL_MV_NEG if sc<-0.3 else FILL_MV_NEUT)
    for co,(val,bold,num) in enumerate([(sg,False,None),(round(sc,2),True,'0.00'),(_rationale.get(sg,''),False,None),(direction,False,None)]):
        c = ws.cell(row=dr, column=MV+co, value=val); c.fill = rf
        c.font = Font(name='Arial', size=8.5, bold=bold,
                      color=('375623' if sc>0.3 else '9C0006' if sc<-0.3 else '555555') if co==3 else '000000')
        c.alignment = Alignment(horizontal='center' if co in (1,3) else 'left', vertical='center', wrap_text=(co==2))
        if num: c.number_format = num
ws.column_dimensions[get_column_letter(MV)].width   = 30
ws.column_dimensions[get_column_letter(MV+1)].width = 7
ws.column_dimensions[get_column_letter(MV+2)].width = 80
ws.column_dimensions[get_column_letter(MV+3)].width = 15
for ti, rh in enumerate([52,60,60,100]):
    ws.row_dimensions[ti+2].height = rh
ws.row_dimensions[mv_subhdr].height = 18

# ── Sheet 6: Score_Integrated ─────────────────────────────────
ws = mks('Score_Integrated')
id_c = ['class','Des','ISIN','Ticker','Cpn','OAS','LQA','Issuer Rtg','BCLASS3','Industry Subgroup']
cp_c = ['Bond_TR_Score','Eq_Mom_Score','Eq_Fund_Score','Sentiment_Score','AI_Macro_Score']
sc_c = ['Integrated_Score','Integrated_Rank_in_Class','Top_Pick_Flag']
hdrs = id_c + cp_c + sc_c
wtitle(ws,f'Integrated Score  |  Bond_TR+EqMom+EqFund+Sentiment+AI_Macro (×0.20 each)  |  As of {AS_OF}',len(hdrs))
whdrs(ws, hdrs, set(id_c), set(cp_c), set(sc_c))
wrows(ws, hdrs, set(sc_c),
      {'Cpn':'0.000','OAS':'0.0','LQA':'0.0',
       'Bond_TR_Score':'0.0000','Eq_Mom_Score':'0.0000','Eq_Fund_Score':'0.0000',
       'Sentiment_Score':'0.0000','AI_Macro_Score':'0.0000',
       'Integrated_Score':'0.0000','Integrated_Rank_in_Class':'0'},
      flag_col='Top_Pick_Flag',
      color_cols={'Bond_TR_Score','Eq_Mom_Score','Eq_Fund_Score','Sentiment_Score','AI_Macro_Score'})
wfin(ws,len(hdrs),{1:12,2:28,3:16,4:10,5:7,6:7,7:8,8:10,9:14,10:26,
                    11:13,12:13,13:13,14:13,15:13,16:13,17:10,18:14})

# ── Sheet 7: Detail_Scored ────────────────────────────────────
ws = mks('Detail_Scored')
keep_cols = [c for c in df_out.columns if not c.startswith('_')]
wtitle(ws, f'LUACSTAT Detail  |  As of {AS_OF}  |  {len(df_out):,} Active Bonds', len(keep_cols))
for ci, h in enumerate(keep_cols, 1):
    c = ws.cell(row=2, column=ci, value=h)
    c.font = Font(name='Arial', bold=True, size=9)
    c.fill = FILL_ID; c.alignment = Alignment(horizontal='center', wrap_text=True)
# 서식 없이 값만 기록 (속도/용량 최적화)
for ri, row in df_out.iterrows():
    for ci, h in enumerate(keep_cols, 1):
        ws.cell(row=ri+3, column=ci, value=_v(row.get(h, np.nan)))
ws.freeze_panes = 'A3'
ws.auto_filter.ref = f'A2:{get_column_letter(len(keep_cols))}2'

# ── Save ──────────────────────────────────────────────────────
wb.save(OUT_FILE)

print(f"\n{'='*65}")
print(f"  Saved: {OUT_FILE}")
print(f"  Sheets: {wb.sheetnames}")
print(f"\n  Score Summary:")
print(f"    Bond_TR_Score    : {df_out['Bond_TR_Score'].notna().sum():>5,} bonds")
print(f"    Eq_Mom_Score     : {df_out['Eq_Mom_Score'].notna().sum():>5,} bonds")
print(f"    Eq_Fund_Score    : {df_out['Eq_Fund_Score'].notna().sum():>5,} bonds")
print(f"    Sentiment_Score  : {df_out.get('Sentiment_Score_clean', pd.Series()).notna().sum():>5,} bonds")
print(f"    AI_Macro_Score   : {df_out['AI_Macro_Score'].notna().sum():>5,} bonds")
print(f"    Integrated_Score : {df_out['Integrated_Score'].notna().sum():>5,} bonds")
print(f"    TOP3 picks       : {(df_out['Top_Pick_Flag']=='★★★ TOP3').sum():>5,}")
print(f"{'='*65}")
print("DONE")
