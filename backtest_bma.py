# -*- coding: utf-8 -*-
"""
backtest_bma.py
Bond_TR x 0.33 + Eq_Mom x 0.33 + AI_Macro x 0.34
- Ticker 컬럼 우선 사용 (Bloomberg 전용코드 필터링)
- price cache 재수집 (eq_price_cache_v2.pkl)
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import os, re, ssl, time, warnings
import numpy as np
import pandas as pd
from datetime import date
from itertools import groupby
from collections import defaultdict

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import matplotlib.gridspec as gridspec

warnings.filterwarnings('ignore')
ssl._create_default_https_context = ssl._create_unverified_context
os.environ['CURL_CA_BUNDLE'] = ''

# ── 설정 ──────────────────────────────────────────────────────────────────────
PATH_OLD   = r"Z:\USIGAIQ\USIG Market data\LUACSTAT"
PATH_NEW   = r"Z:\USIGAIQ\USIG Market data\LUACSTAT_PREP"
EQ_CACHE   = r"C:\Users\sh.park\Documents\USIG_LLM\eq_price_cache_v2.pkl"
OUT_PNG    = r"C:\Users\sh.park\Documents\USIG_LLM\backtest_bma.png"
OUT_XLSX   = r"C:\Users\sh.park\Documents\USIG_LLM\backtest_bma.xlsx"

BACKTEST_START = date(2021, 1, 1)
TOP_N          = 3
MIN_CLASS_SIZE = 5

W_BOND = 0.33
W_MOM  = 0.33
W_AI   = 0.34

# ── 매핑 테이블 ───────────────────────────────────────────────────────────────
RTG_TEXT_TO_NUM = {
    'AAA':1,'AA+':2,'AA':3,'AA-':4,'A+':5,'A':6,'A-':7,
    'BBB+':8,'BBB':9,'BBB-':10,'BB+':11,'BB':12,'BB-':13,
}
RATING_CUT = 7

DP_RATING_MAP = {
    'AAA':1,'AA1':2,'AA2':3,'AA3':4,'A1':5,'A2':6,'A3':7,
    'BAA1':8,'BAA2':9,'BAA3':10,'BA1':11,'BA2':12,'BA3':13,
    'B1':14,'B2':15,'B3':16,'CAA1':17,'CAA2':18,'CAA3':19,'CA':20,'C':21,
}

SUBGROUP_SCORE_MAP = {
    'Electric-Integrated':0.85,'Electric-Distribution':0.85,'Electric-Transmission':0.80,
    'Electric-Generation':0.75,'Gas-Distribution':0.80,'Water':0.85,
    'Non-hazardous Waste Disp':0.70,'Medical-Drugs':0.70,'Medical-Hospitals':0.75,
    'Medical-HMO':0.65,'Medical-Biomedical/Gene':0.60,'Medical Products':0.60,
    'Medical Instruments':0.55,'Medical Labs&Testing Srv':0.65,'Pharmacy Services':0.60,
    'Medical-Whsle Drug Dist':0.55,'Medical Imaging Systems':0.55,'Diagnostic Equipment':0.55,
    'Drug Delivery Systems':0.55,'Medical-Generic Drugs':0.50,'Phys Practice Mgmnt':0.60,
    'Medical-Outptnt/Home Med':0.60,'Diversified Banking Inst':0.55,
    'Super-Regional Banks-US':0.60,'Commer Banks-Eastern US':0.55,
    'Commer Banks-Southern US':0.55,'Commer Banks-Central US':0.50,
    'Commer Banks-Western US':0.50,'Commer Banks Non-US':0.30,'Money Center Banks':0.45,
    'Fiduciary Banks':0.65,'Life/Health Insurance':0.50,'Property/Casualty Ins':0.55,
    'Reinsurance':0.50,'Multi-line Insurance':0.45,'Insurance Brokers':0.55,
    'Financial Guarantee Ins':0.20,'Finance-Invest Bnkr/Brkr':0.35,
    'Invest Mgmnt/Advis Serv':0.40,'Private Equity':0.20,'Finance-Credit Card':0.30,
    'Finance-Leasing Compan':0.25,'Finance-Auto Loans':0.10,'Finance-Other Services':0.30,
    'Finance-Mtge Loan/Banker':0.20,'Finance-Consumer Loans':0.10,'Venture Capital':0.00,
    'Investment Companies':0.15,'Pipelines':0.30,'Oil Comp-Integrated':0.10,
    'Oil Comp-Explor&Prodtn':-0.10,'Oil Refining&Marketing':-0.05,'Oil-Field Services':-0.20,
    'Oil&Gas Drilling':-0.25,'Agricultural Chemicals':0.20,
    'Auto-Cars/Light Trucks':-0.90,'Auto-Med&Heavy Duty Trks':-0.70,
    'Auto/Trk Prts&Equip-Orig':-0.80,'Retail-Automobile':-0.60,
    'Retail-Discount':-0.20,'Retail-Building Products':-0.30,'Retail-Major Dept Store':-0.50,
    'Retail-Apparel/Shoe':-0.60,'Retail-Auto Parts':-0.40,'Retail-Sporting Goods':-0.40,
    'Retail-Consumer Electron':-0.50,'Retail-Restaurants':-0.10,'Retail-Gardening Prod':-0.30,
    'Hotels&Motels':-0.15,'Casino Hotels':-0.20,'Cruise Lines':-0.25,
    'E-Commerce/Products':-0.20,'E-Commerce/Services':-0.10,'Internet Content-Entmnt':0.00,
    'Apparel Manufacturers':-0.55,'Athletic Footwear':-0.55,'Recreational Vehicles':-0.40,
    'Beverages-Non-alcoholic':0.55,'Beverages-Wine/Spirits':0.45,'Brewery':0.45,
    'Food-Misc/Diversified':0.50,'Food-Retail':0.40,'Food-Confectionery':0.50,
    'Food-Meat Products':0.45,'Food-Baking':0.45,'Food-Wholesale/Distrib':0.40,
    'Poultry':0.40,'Coffee':0.45,'Cosmetics&Toiletries':0.55,'Soap&Cleaning Prepar':0.50,
    'Consumer Products-Misc':0.40,'Tobacco':0.60,'Agricultural Operations':0.35,
    'Electronic Compo-Semicon':-0.30,'Semicon Compo-Intg Circu':-0.25,
    'Semiconductor Equipment':-0.35,'Enterprise Software/Serv':0.10,
    'Applications Software':0.05,'Computer Services':0.15,'Data Processing/Mgmt':0.20,
    'Computer Aided Design':0.00,'Software Tools':0.00,'E-Marketing/Info':0.00,
    'Decision Support Softwar':0.05,'Computers':-0.15,'Computers-Memory Devices':-0.20,
    'Computers-Other':-0.10,'Electronic Connectors':-0.20,'Electronic Compo-Misc':-0.15,
    'Electronic Parts Distrib':-0.10,'Electronic Measur Instr':-0.10,
    'Electronic Secur Devices':-0.05,'Electronic Forms':0.00,'Networking Products':-0.10,
    'Wireless Equipment':-0.15,'Telecom Eq Fiber Optics':-0.10,
    'Telecommunication Equip':-0.10,'Industrial Automat/Robot':-0.10,
    'Instruments-Controls':-0.05,'Telephone-Integrated':0.10,'Cellular Telecom':0.05,
    'Cable/Satellite TV':0.00,'Multimedia':0.00,'Broadcast Serv/Program':0.00,
    'Advertising Agencies':-0.15,'Advertising Services':-0.15,'Telecom Services':0.10,
    'Web Portals/ISP':0.00,'Aerospace/Defense':0.40,'Aerospace/Defense-Equip':0.35,
    'Machinery-Farm':0.20,'Machinery-General Indust':-0.10,'Machinery-Constr&Mining':-0.10,
    'Machinery-Pumps':-0.05,'Machinery-Electric Util':0.10,'Tools-Hand Held':-0.10,
    'Diversified Manufact Op':-0.05,'Industrial Gases':0.30,'Chemicals-Diversified':-0.10,
    'Chemicals-Specialty':-0.05,'Coatings/Paint':-0.10,'Containers-Paper/Plastic':-0.20,
    'Paper&Related Products':-0.15,'Commercial Serv-Finance':0.10,'Commercial Services':0.05,
    'Consulting Services':0.10,'Non-Profit Charity':0.00,'Distribution/Wholesale':-0.10,
    'Office Automation&Equip':-0.15,'Office Supplies&Forms':-0.15,
    'Bldg Prod-Cement/Aggreg':-0.10,'Bldg Prod-Air&Heating':-0.10,'Bldg Prod-Wood':-0.15,
    'Bldg&Construct Prod-Misc':-0.10,'Bldg-Residential/Commer':-0.20,
    'Building-Maint&Service':-0.05,'Shipbuilding':0.20,'Power Conv/Supply Equip':0.10,
    'Transport-Rail':0.30,'Transport-Services':0.00,'Transport-Equip&Leasng':-0.10,
    'Transport-Truck':-0.10,'Transport-Marine':-0.10,'Airlines':-0.40,
    'Steel-Producers':0.20,'Steel Pipe&Tube':0.15,'Metal-Iron':0.10,'Metal-Aluminum':0.10,
    'Metal-Copper':0.00,'Metal-Diversified':0.00,'Diversified Minerals':0.00,
    'Gold Mining':0.30,'Metal Processors&Fabrica':0.05,'Schools':0.40,'Toys':-0.40,
    'Entertainment Software':0.00,'Engineering/R&D Services':0.15,'Vitamins&Nutrition Prod':0.30,
    'Rental Auto/Equipment':-0.10,'Mach Tools&Rel Products':-0.10,
    'Motorcycle/Motor Scooter':-0.50,'Chemicals-Plastics':-0.15,
    'Miscellaneous Manufactur':-0.10,'REITS-Industrial':0.10,'REITS-Warehouse/Industr':0.20,
    'REITS-Diversified':-0.10,'REITS-Apartments':-0.05,'REITS-Shopping Centers':-0.30,
    'REITS-Regional Malls':-0.40,'REITS-Office Property':-0.35,'REITS-Health Care':0.30,
    'REITS-Storage':0.10,'REITS-Single Tenant':0.00,'REITS-Hotels':-0.20,
    'REITS-Manufactured Homes':0.00,'REITS-Mortgage':-0.20,'Real Estate Mgmnt/Servic':-0.10,
}
BCLASS3_SCORE_MAP = {
    'Utility':0.75,'Healthcare':0.60,'Banking':0.50,'Financial Other':0.30,
    'Insurance':0.50,'Energy':0.05,'Consumer Cyclical':0.20,
    'Consumer Non-Cyclical':0.45,'Technology':0.00,'Communications':0.05,
    'Industrial':0.00,'Transportation':0.00,'Basic Industry':-0.05,
    'Real Estate':-0.10,'Government Related':0.30,'Sovereign':0.20,
    'Quasi-Government':0.25,'Supranational':0.40,
}
RATING_BUFFER_MAP = {
    'AAA':0.50,'AA1':0.55,'AA2':0.55,'AA3':0.50,
    'A1':0.70,'A2':0.70,'A3':0.65,
    'BAA1':0.25,'BAA2':0.05,'BAA3':-0.65,
    'BA1':-0.85,'BA2':-1.00,'BA3':-1.00,
    'B1':-1.00,'B2':-1.00,'B3':-1.00,
}

# ── 유효 Yahoo 티커 판별 ──────────────────────────────────────────────────────
VALID_TICKER_RE = re.compile(r'^[A-Z][A-Z0-9./\-]{0,9}$')

def clean_ticker(val):
    """Bloomberg Ticker 컬럼에서 Yahoo Finance 형식 티커 추출."""
    if pd.isna(val): return None
    t = str(val).strip()
    if t in ('', 'N/A', 'nan', 'None'): return None
    # "STT US" -> "STT"
    t = t.split()[0]
    # 숫자로 시작하거나 Bloomberg 전용코드 제외
    if not VALID_TICKER_RE.match(t): return None
    return t

def maturity_bucket(yrs):
    if pd.isna(yrs): return None
    if yrs <= 3:    return '~3Y'
    elif yrs <= 7:  return '3~7Y'
    elif yrs <= 15: return '7~15Y'
    else:           return '15Y+'

def maturity_score(oad):
    if pd.isna(oad): return 0.0
    oad = float(oad)
    if oad < 2:    return  0.00
    elif oad < 4:  return  0.30
    elif oad < 7:  return  1.00
    elif oad < 10: return  0.70
    elif oad < 13: return  0.00
    elif oad < 16: return -0.50
    else:          return -1.00

def percentile_norm(series):
    n = series.notna().sum()
    if n < 2: return pd.Series(np.nan, index=series.index)
    ranked = series.rank(method='average', na_option='keep')
    return (ranked - 1) / (n - 1) * 2 - 1

# ═══════════════════════════════════════════════════════════════════════════════
# 1. 월별 파일 목록
# ═══════════════════════════════════════════════════════════════════════════════
pattern_f = re.compile(r'^LUACSTAT_(\d{4})_(\d{2})_(\d{2})\.(xls|xlsx)$')
dated_dict = {}
for p in [PATH_OLD, PATH_NEW]:
    for f in os.listdir(p):
        m = pattern_f.match(f)
        if m:
            y,mo,d,ext = int(m.group(1)),int(m.group(2)),int(m.group(3)),m.group(4)
            try:
                dt = date(y,mo,d)
                if dt not in dated_dict or ext=='xlsx':
                    dated_dict[dt] = (os.path.join(p,f), ext)
            except: pass

dated_sorted = sorted(dated_dict.items())
monthly = []
for (yr,mo), grp in groupby(dated_sorted, key=lambda x:(x[0].year,x[0].month)):
    monthly.append(list(grp)[-1])
monthly_bt = [(d,fp) for d,(fp,ext) in monthly if d.year >= BACKTEST_START.year-1]

print(f"Files: {len(monthly_bt)}, Range: {monthly_bt[0][0]} ~ {monthly_bt[-1][0]}")

# ═══════════════════════════════════════════════════════════════════════════════
# 2. 파일 로딩
# ═══════════════════════════════════════════════════════════════════════════════
def load_file(filepath):
    df_raw = pd.read_excel(filepath, header=None, nrows=15)
    header_row = None
    for i, row in df_raw.iterrows():
        if 'ISIN' in [str(v) for v in row.values] and \
           any(x in [str(v) for v in row.values] for x in ['OAS','Yield to Worst']):
            header_row = i; break
    if header_row is None: return None
    df = pd.read_excel(filepath, header=header_row)
    if 'Composite Rating Num' not in df.columns and 'BB Comp' in df.columns:
        df['Composite Rating Num'] = df['BB Comp'].map(RTG_TEXT_TO_NUM)
    elif 'Composite Rating Num' not in df.columns and 'Index Rtg' in df.columns:
        df['Composite Rating Num'] = df['Index Rtg'].map(RTG_TEXT_TO_NUM)
    if 'Mty (Yrs)' not in df.columns and 'Eff Mty (Yrs)' in df.columns:
        df.rename(columns={'Eff Mty (Yrs)': 'Mty (Yrs)'}, inplace=True)
    if 'OASD' not in df.columns and 'Spd Dur' in df.columns:
        df.rename(columns={'Spd Dur': 'OASD'}, inplace=True)
    isin_mask = df['ISIN'].astype(str).str.match(r'^[A-Z]{2}[A-Z0-9]{10}$', na=False) if 'ISIN' in df.columns else pd.Series(False, index=df.index)
    df_bonds = df[isin_mask].copy()
    for col in ['OAS','OASD','OAD','Yield to Worst','Mty (Yrs)',
                'Composite Rating Num','1Y Dflt','Total Return - 1mo']:
        if col in df_bonds.columns:
            df_bonds[col] = pd.to_numeric(df_bonds[col], errors='coerce')
    df_bonds['ISIN'] = df_bonds['ISIN'].astype(str).str.strip()
    df_bonds = df_bonds[df_bonds['ISIN'].str.match(r'^[A-Z]{2}[A-Z0-9]{10}$', na=False)]
    return df_bonds

# ═══════════════════════════════════════════════════════════════════════════════
# 3. 티커 수집 (Ticker 컬럼 우선)
# ═══════════════════════════════════════════════════════════════════════════════
print("\n[Phase 1] Scanning for clean equity tickers...")

TICKER_COLS_PRIORITY = ['Ticker', 'Parent Ticker', 'Eqty Ticker']

all_tickers = set()
for dt_scan, fp_scan in monthly_bt:
    try:
        df_s = load_file(fp_scan)
        if df_s is None: continue
        for col in TICKER_COLS_PRIORITY:
            if col in df_s.columns:
                ts = df_s[col].apply(clean_ticker).dropna().unique()
                all_tickers.update(ts)
                break
    except: pass

all_tickers = sorted(all_tickers)
print(f"  Valid tickers: {len(all_tickers)}")
print(f"  Sample: {all_tickers[:20]}")

# ═══════════════════════════════════════════════════════════════════════════════
# 4. 주가 히스토리 수집 (v2 캐시)
# ═══════════════════════════════════════════════════════════════════════════════
PRICE_START = '2019-01-01'
PRICE_END   = date.today().strftime('%Y-%m-%d')

if os.path.exists(EQ_CACHE):
    print(f"\n  Loading v2 price cache...")
    price_cache = pd.read_pickle(EQ_CACHE)
    print(f"  Cache loaded: {len(price_cache)} tickers")
else:
    print(f"\n[Phase 2] Fetching price history for {len(all_tickers)} tickers...")
    price_cache = {}
    BATCH = 50

    try:
        import yfinance as yf
        try:
            import curl_cffi.requests as cffi_req
            yf.utils.requests = cffi_req
        except ImportError:
            pass

        for i in range(0, len(all_tickers), BATCH):
            batch = all_tickers[i:i+BATCH]
            pct = (i+len(batch))/len(all_tickers)*100
            print(f"  Batch {i//BATCH+1}/{(len(all_tickers)-1)//BATCH+1} ({pct:.0f}%)...", end='', flush=True)
            try:
                raw = yf.download(batch, start=PRICE_START, end=PRICE_END,
                                  progress=False, auto_adjust=True, threads=True)
                if isinstance(raw.columns, pd.MultiIndex):
                    close = raw['Close'] if 'Close' in raw.columns.get_level_values(0) else raw.iloc[:,0:1]
                else:
                    close = raw[['Close']] if 'Close' in raw.columns else raw
                ok = 0
                for t in batch:
                    if t in close.columns:
                        s = close[t].dropna()
                        if len(s) > 20:
                            price_cache[t] = s
                            ok += 1
                print(f" {ok}/{len(batch)} valid")
            except Exception as e:
                print(f" WARN batch: {e}")
                for t in batch:
                    if t in price_cache: continue
                    try:
                        s = yf.download(t, start=PRICE_START, end=PRICE_END,
                                        progress=False, auto_adjust=True)
                        if 'Close' in s.columns and len(s) > 20:
                            price_cache[t] = s['Close'].dropna()
                    except: pass
            time.sleep(0.3)

    except ImportError:
        print("  yfinance not available")

    pd.to_pickle(price_cache, EQ_CACHE)
    n_valid = sum(1 for v in price_cache.values() if len(v) > 0)
    print(f"  V2 price cache saved: {n_valid} tickers with data")

n_valid = sum(1 for v in price_cache.values() if len(v) > 0)
print(f"  Price cache: {n_valid}/{len(price_cache)} tickers with data")

# ═══════════════════════════════════════════════════════════════════════════════
# 5. 월별 Eq_Mom 계산 함수
# ═══════════════════════════════════════════════════════════════════════════════
def compute_eq_mom_scores(tickers, as_of_date):
    as_of_ts = pd.Timestamp(as_of_date)
    one_mo   = as_of_ts - pd.DateOffset(months=1)
    three_mo = as_of_ts - pd.DateOffset(months=3)
    yr52     = as_of_ts - pd.DateOffset(weeks=52)
    rows = []
    for t in tickers:
        prices = price_cache.get(t)
        if prices is None or len(prices) < 10:
            rows.append({'t': t}); continue
        p = prices[prices.index <= as_of_ts]
        if len(p) < 5:
            rows.append({'t': t}); continue
        cur = float(p.iloc[-1])
        m = {'t': t}
        p1 = p[p.index <= one_mo]
        if len(p1) > 0: m['r1'] = cur / float(p1.iloc[-1]) - 1
        p3 = p[p.index <= three_mo]
        if len(p3) > 0: m['r3'] = cur / float(p3.iloc[-1]) - 1
        lr = np.log(p / p.shift(1)).dropna()
        if len(lr) >= 20: m['vol'] = float(lr.iloc[-30:].std() * np.sqrt(252))
        p52 = p[p.index >= yr52]
        if len(p52) > 0:
            hi = float(p52.max())
            if hi > 0: m['h52'] = cur / hi - 1
        rows.append(m)
    mdf = pd.DataFrame(rows).set_index('t')
    def rn(s, asc=True):
        s = pd.to_numeric(s, errors='coerce'); v = s.notna()
        res = pd.Series(np.nan, index=s.index)
        if v.sum() < 2: return res
        ranked = s[v].rank(ascending=asc, method='average')
        res[v] = (ranked - 1) / (v.sum() - 1) * 2 - 1
        return res
    c = pd.DataFrame({
        'n1': rn(mdf.get('r1',  pd.Series(dtype=float)), True),
        'n3': rn(mdf.get('r3',  pd.Series(dtype=float)), True),
        'nv': rn(mdf.get('vol', pd.Series(dtype=float)), False),
        'nh': rn(mdf.get('h52', pd.Series(dtype=float)), True),
    })
    mdf['score'] = c.mean(axis=1)
    return {t: float(mdf.loc[t,'score']) for t in tickers
            if t in mdf.index and not pd.isna(mdf.loc[t,'score'])}

# ═══════════════════════════════════════════════════════════════════════════════
# 6. 백테스팅
# ═══════════════════════════════════════════════════════════════════════════════
print(f"\n[Phase 3] Running backtest: Bond_TR x{W_BOND} + Eq_Mom x{W_MOM} + AI_Macro x{W_AI}")

port_records  = []
all_picks_log = []

for i in range(len(monthly_bt) - 1):
    dt_t,  fp_t  = monthly_bt[i]
    dt_t1, fp_t1 = monthly_bt[i + 1]
    if dt_t < BACKTEST_START: continue

    df_t = load_file(fp_t)
    if df_t is None or len(df_t) < 100:
        print(f"  SKIP {dt_t}: load failed")
        continue

    df_t1 = load_file(fp_t1)
    if df_t1 is None: continue

    df_t1['_tr'] = pd.to_numeric(df_t1['Total Return - 1mo'], errors='coerce')
    isin_to_tr = dict(zip(df_t1['ISIN'], df_t1['_tr']))

    # 벤치마크
    if 'Mkt Val' in df_t1.columns:
        df_t1['_mv'] = pd.to_numeric(df_t1['Mkt Val'], errors='coerce')
        vb = df_t1.dropna(subset=['_mv','_tr'])
        vb = vb[vb['_mv'] > 0]
        bench_tr = float(np.average(vb['_tr'], weights=vb['_mv'])) if len(vb) > 100 else float(df_t1['_tr'].mean())
    else:
        bench_tr = float(df_t1['_tr'].dropna().mean())

    # 필터
    mask = (
        df_t['OAS'].notna() & df_t['OAS'].between(-50,600) &
        df_t['OASD'].notna() & df_t['OASD'].gt(0) &
        df_t['Yield to Worst'].notna() & df_t['Yield to Worst'].gt(0) &
        df_t['Mty (Yrs)'].notna() &
        df_t['BCLASS3'].notna() &
        df_t['Composite Rating Num'].notna()
    )
    df = df_t[mask].copy()
    df['_mty_bkt'] = df['Mty (Yrs)'].apply(maturity_bucket)
    df['_rtg_bkt'] = df['Composite Rating Num'].apply(
        lambda x: 'A-이상' if x <= RATING_CUT else 'BBB+이하')
    df['_bclass3'] = df['BCLASS3'].astype(str).str.strip()
    df['_class'] = df['_bclass3'] + '|' + df['_mty_bkt'] + '|' + df['_rtg_bkt']

    # Bond_TR 요소
    df['_carry']    = df['Yield to Worst'] / 12 * 2.5
    df['_comp_raw'] = (df['OAS'] - df['1Y Dflt'].fillna(0) * 60) * df['OASD']
    if 'DPFundamentalRating' in df.columns and 'DPSpreadRating' in df.columns:
        df['_dp_gap']   = df['DPSpreadRating'].map(DP_RATING_MAP) - df['DPFundamentalRating'].map(DP_RATING_MAP)
        df['_dp_score'] = percentile_norm(df['_dp_gap'])
    else:
        df['_dp_score'] = 0.0
    df['_bond_tr_est'] = df['_carry'].fillna(0) + df['_comp_raw'].fillna(0) + df['_dp_score'].fillna(0)*0.05

    # AI_Macro 요소
    sg_col = next((c for c in ['Industry Subgroup','BCLASS4','BCLASS3'] if c in df.columns), None)
    if sg_col:
        df['_ai_sector'] = df[sg_col].astype(str).map(SUBGROUP_SCORE_MAP)
        fb = df['_ai_sector'].isna()
        if fb.any(): df.loc[fb,'_ai_sector'] = df.loc[fb,'_bclass3'].map(BCLASS3_SCORE_MAP)
        df['_ai_sector'] = df['_ai_sector'].fillna(0.0)
    else:
        df['_ai_sector'] = 0.0
    df['_ai_mty'] = df['OAD'].apply(maturity_score) if 'OAD' in df.columns else \
                    df['Mty (Yrs)'].apply(lambda y: maturity_score(y*0.9) if not pd.isna(y) else 0.0)
    rtg_col = next((c for c in ['DPFundamentalRating','Issuer Rtg'] if c in df.columns), None)
    df['_ai_rtg'] = df[rtg_col].map(RATING_BUFFER_MAP).fillna(0.0) if rtg_col else 0.0
    df['_ai_macro'] = (df['_ai_sector']*0.40 + df['_ai_mty']*0.35 + df['_ai_rtg']*0.25).clip(-1,1)

    # Eq_Mom: Ticker 컬럼 우선 사용
    eq_col = next((c for c in TICKER_COLS_PRIORITY if c in df.columns), None)
    if eq_col:
        df['_eq_tk'] = df[eq_col].apply(clean_ticker)
        uniq_t = [t for t in df['_eq_tk'].dropna().unique() if t]
        mom_sc = compute_eq_mom_scores(uniq_t, dt_t)
        df['_eq_mom'] = df['_eq_tk'].map(mom_sc)
    else:
        df['_eq_mom'] = np.nan

    n_mom_valid = df['_eq_mom'].notna().sum()

    # 클래스별 스코어링 & Top-N 선택
    picks = {}
    for cls, grp in df.groupby('_class'):
        if len(grp) < MIN_CLASS_SIZE: continue
        score_b = percentile_norm(grp['_bond_tr_est'])
        score_m = percentile_norm(grp['_eq_mom']) if grp['_eq_mom'].notna().sum() >= 3 \
                  else pd.Series(0.0, index=grp.index)
        score_a = grp['_ai_macro']
        integrated = (
            score_b.fillna(0) * W_BOND +
            score_m.fillna(0) * W_MOM  +
            score_a.fillna(0) * W_AI
        )
        top_idx = integrated.nlargest(TOP_N).index
        picks[cls] = grp.loc[top_idx, 'ISIN'].tolist()

    n_classes  = len(picks)
    n_selected = sum(len(v) for v in picks.values())

    # 포트폴리오 수익률
    all_isins = [isin for isins in picks.values() for isin in isins]
    tr_valid  = [float(isin_to_tr[i]) for i in all_isins
                 if i in isin_to_tr and not pd.isna(isin_to_tr.get(i, np.nan))]
    if len(tr_valid) < 5: continue

    port_tr = np.mean(tr_valid)
    port_records.append({
        'date':       dt_t1,
        'port_tr':    port_tr / 100,
        'bench_tr':   bench_tr / 100,
        'n_classes':  n_classes,
        'n_selected': n_selected,
        'n_matched':  len(tr_valid),
        'n_mom_valid':n_mom_valid,
    })
    all_picks_log.append({'score_date': dt_t, 'return_date': dt_t1, 'picks': picks})

    print(f"  {dt_t} -> {dt_t1}: cls={n_classes}, sel={n_selected}, "
          f"mom_valid={n_mom_valid}, port={port_tr:.3f}%, bench={bench_tr:.3f}%")

if not port_records:
    print("ERROR: No records!")
    exit(1)

res = pd.DataFrame(port_records).set_index('date')
res.index = pd.to_datetime(res.index)
print(f"\n  Total months: {len(res)}")

# ═══════════════════════════════════════════════════════════════════════════════
# 7. 성과 통계
# ═══════════════════════════════════════════════════════════════════════════════
def cum_wealth(ret, start=100):
    return start * (1 + ret).cumprod()

bench   = res['bench_tr'].dropna()
port    = res['port_tr']
aligned = port.loc[bench.index]
excess  = aligned - bench

b_ann = (1+bench.mean())**12 - 1
m_ann = (1+aligned.mean())**12 - 1
b_vol = bench.std()*np.sqrt(12)
m_vol = aligned.std()*np.sqrt(12)
ir    = excess.mean()/excess.std()*np.sqrt(12) if excess.std()>0 else 0
total = (1+aligned).prod()-1
alpha_ann = excess.mean()*12

print("\n" + "="*65)
print(f"Bond_TR x{W_BOND} + Eq_Mom x{W_MOM} + AI_Macro x{W_AI}")
print(f"{'':20} {'LUACTRUU':>10} {'Model':>10}")
print(f"{'Total Return':20} {(1+bench).prod()-1:>10.1%} {total:>10.1%}")
print(f"{'Ann. Return':20} {b_ann:>10.2%} {m_ann:>10.2%}")
print(f"{'Volatility':20} {b_vol:>10.2%} {m_vol:>10.2%}")
print(f"{'Sharpe':20} {b_ann/b_vol:>10.2f} {m_ann/m_vol:>10.2f}")
print(f"{'Max Drawdown':20} {(cum_wealth(bench)/cum_wealth(bench).cummax()-1).min():>10.2%} "
      f"{(cum_wealth(aligned)/cum_wealth(aligned).cummax()-1).min():>10.2%}")
print(f"{'Win Rate':20} {(bench>0).mean():>10.0%} {(aligned>0).mean():>10.0%}")
print(f"{'Info Ratio':20} {'':>10} {ir:>10.2f}")
print(f"{'Ann. Alpha':20} {'':>10} {alpha_ann:>10.2%}")
print("="*65)

print("\nAnnual Performance:")
print(f"  {'Year':<5} {'Bench':>8} {'Model':>8} {'Alpha':>8}")
bench_v = res['bench_tr'].dropna()
for yr in sorted(res.index.year.unique()):
    b = (1+bench_v[bench_v.index.year==yr]).prod()-1 if (bench_v.index.year==yr).any() else np.nan
    p = (1+port[port.index.year==yr]).prod()-1
    bs = f'{b:>7.2%}' if not np.isnan(b) else '    N/A'
    as_ = f'{p-b:>+7.2%}' if not np.isnan(b) else '    N/A'
    print(f"  {yr}  {bs}  {p:>7.2%}  {as_}")

# 평균 Eq_Mom 커버리지
avg_mom = res['n_mom_valid'].mean()
avg_sel = res['n_selected'].mean()
print(f"\n  Avg Eq_Mom coverage: {avg_mom:.0f}/{avg_sel:.0f} bonds ({avg_mom/avg_sel*100:.1f}%)")

# ═══════════════════════════════════════════════════════════════════════════════
# 8. 차트
# ═══════════════════════════════════════════════════════════════════════════════
EVENTS = [
    ('2022-01', 'Fed 긴축\n시작'), ('2022-06', '금리\n피크'),
    ('2023-03', 'SVB\n사태'), ('2023-10', '금리\n재급등'), ('2024-09', 'Fed\n피벗'),
]
COLORS = {'bench': '#1F3864', 'model': '#C00000', 'excess': '#70AD47'}

fig = plt.figure(figsize=(22, 14))
gs  = gridspec.GridSpec(3, 3, figure=fig, hspace=0.45, wspace=0.32, top=0.91, bottom=0.06)

# Plot 1: 누적수익
ax1 = fig.add_subplot(gs[0, :2])
cw_bench = cum_wealth(bench)
cw_model = cum_wealth(aligned)
ax1.fill_between(cw_model.index, cw_model.values, cw_bench.values,
                 where=(cw_model.values >= cw_bench.values), alpha=0.15, color=COLORS['model'])
ax1.fill_between(cw_model.index, cw_model.values, cw_bench.values,
                 where=(cw_model.values < cw_bench.values), alpha=0.15, color='gray')
ax1.plot(cw_bench.index, cw_bench.values, color=COLORS['bench'], lw=2.5,
         label=f"LUACTRUU: {(1+bench).prod()-1:.1%}", zorder=5)
ax1.plot(cw_model.index, cw_model.values, color=COLORS['model'], lw=2.3,
         label=f"Model B{int(W_BOND*100)}+M{int(W_MOM*100)}+A{int(W_AI*100)}: {total:.1%}", zorder=7)
for ev_dt, ev_lbl in EVENTS:
    xv = pd.Timestamp(ev_dt)
    if res.index[0] <= xv <= res.index[-1]:
        ax1.axvline(xv, color='#AAAAAA', lw=0.9, ls='--', alpha=0.7)
ax1.axhline(100, color='gray', lw=0.5, ls=':', alpha=0.6)
ax1.set_title(
    f'누적 수익률 ($100) -- Bond_TR x{W_BOND} + Eq_Mom x{W_MOM} + AI_Macro x{W_AI} / Top{TOP_N}/Class\n'
    f'Eq_Mom: clean Ticker col 사용 (Bloomberg 코드 필터링)',
    fontsize=11, fontweight='bold', pad=6)
ax1.set_ylabel('포트폴리오 가치 ($)', fontsize=10)
ax1.yaxis.set_major_formatter(mtick.FormatStrFormatter('$%.0f'))
ax1.legend(fontsize=10, loc='upper left', framealpha=0.9)
ax1.grid(True, alpha=0.2)
ax1.set_xlim(res.index[0], res.index[-1])

# Plot 2: 성과 테이블
ax2 = fig.add_subplot(gs[0, 2])
ax2.axis('off')
rows = [
    ['', 'LUACTRUU', f'B{int(W_BOND*100)}+M{int(W_MOM*100)}+A{int(W_AI*100)}'],
    ['총수익', f"{(1+bench).prod()-1:.1%}", f"{total:.1%}"],
    ['연환산수익', f"{b_ann:.2%}", f"{m_ann:.2%}"],
    ['변동성', f"{b_vol:.2%}", f"{m_vol:.2%}"],
    ['Sharpe', f"{b_ann/b_vol:.2f}", f"{m_ann/m_vol:.2f}"],
    ['MaxDD',
     f"{(cum_wealth(bench)/cum_wealth(bench).cummax()-1).min():.2%}",
     f"{(cum_wealth(aligned)/cum_wealth(aligned).cummax()-1).min():.2%}"],
    ['월별승률', f"{(bench>0).mean():.0%}", f"{(aligned>0).mean():.0%}"],
    ['누적 초과수익', '', f"{(1+aligned).prod()-(1+bench).prod():.1%}"],
    ['연환산 Alpha', '', f"{alpha_ann:.2%}"],
    ['Info Ratio', '', f"{ir:.2f}"],
    ['Eq_Mom 커버리지', '', f"{avg_mom/avg_sel*100:.0f}%"],
]
tbl = ax2.table(cellText=rows[1:], colLabels=rows[0],
                cellLoc='center', loc='center', bbox=[0,0,1,1])
tbl.auto_set_font_size(False); tbl.set_fontsize(9)
for (r,c), cell in tbl.get_celld().items():
    cell.set_edgecolor('#DDDDDD'); cell.set_height(0.085)
    if r == 0:
        cell.set_facecolor('#1F3864'); cell.set_text_props(color='white', fontweight='bold')
    elif c == 2 and r > 0:
        cell.set_facecolor('#FFF0F0'); cell.set_text_props(fontweight='bold', color='#C00000')
    elif r % 2 == 0: cell.set_facecolor('#F5F5F5')
ax2.set_title('성과 요약', fontsize=10, fontweight='bold', pad=4)

# Plot 3: 연도별 바
ax3 = fig.add_subplot(gs[1, :2])
years_list = sorted(res.index.year.unique())
bench_yr = [(1+bench[bench.index.year==y]).prod()-1 if (bench.index.year==y).any() else np.nan for y in years_list]
model_yr = [(1+aligned[aligned.index.year==y]).prod()-1 for y in years_list]
x = np.arange(len(years_list)); w_ = 0.35
ax3.bar(x-w_/2, [v*100 if not (v is None or np.isnan(v)) else 0 for v in bench_yr],
        w_, label='LUACTRUU', color=COLORS['bench'], alpha=0.85)
ax3.bar(x+w_/2, [v*100 for v in model_yr],
        w_, label=f'Model B{int(W_BOND*100)}+M{int(W_MOM*100)}+A{int(W_AI*100)}',
        color=COLORS['model'], alpha=0.85)
ax3.axhline(0, color='black', lw=0.8)
for i,(bv,mv) in enumerate(zip(bench_yr, model_yr)):
    av = mv-bv if bv is not None and not np.isnan(bv) else None
    if av is not None:
        ax3.text(x[i], max(mv*100,(bv or 0)*100)+0.4, f'a{av*100:+.1f}%',
                 ha='center', va='bottom', fontsize=8, color=COLORS['excess'], fontweight='bold')
ax3.set_xticks(x)
ax3.set_xticklabels([str(y) for y in years_list], fontsize=10)
ax3.set_title('연도별 수익률 비교 (%)', fontsize=12, fontweight='bold', pad=6)
ax3.yaxis.set_major_formatter(mtick.PercentFormatter())
ax3.legend(fontsize=10); ax3.grid(True, alpha=0.2, axis='y')

# Plot 4: 누적 초과수익
ax4 = fig.add_subplot(gs[1, 2])
exc_cum = (1 + excess).cumprod() - 1
ax4.plot(exc_cum.index, exc_cum.values*100, color=COLORS['excess'], lw=2)
ax4.fill_between(exc_cum.index, exc_cum.values*100, 0,
                 where=(exc_cum.values >= 0), alpha=0.3, color=COLORS['excess'])
ax4.fill_between(exc_cum.index, exc_cum.values*100, 0,
                 where=(exc_cum.values < 0), alpha=0.3, color='#FF6B6B')
ax4.axhline(0, color='black', lw=0.8)
ax4.set_title('누적 초과수익 (vs 벤치마크)', fontsize=10, fontweight='bold', pad=4)
ax4.yaxis.set_major_formatter(mtick.PercentFormatter())
ax4.grid(True, alpha=0.2)

# Plot 5: 월별 초과수익 분포
ax5 = fig.add_subplot(gs[2, 0])
exc_pct = excess * 100
ax5.hist(exc_pct.values, bins=25, color=COLORS['model'], alpha=0.7, edgecolor='white')
ax5.axvline(0, color='black', lw=1)
ax5.axvline(exc_pct.mean(), color='orange', lw=1.5, ls='--', label=f'Mean: {exc_pct.mean():.2f}%')
ax5.set_title('월별 초과수익 분포 (%)', fontsize=10, fontweight='bold', pad=4)
ax5.legend(fontsize=9); ax5.grid(True, alpha=0.2)

# Plot 6: Eq_Mom 커버리지 추이
ax6 = fig.add_subplot(gs[2, 1:])
cov_pct = res['n_mom_valid'] / res['n_selected'] * 100
ax6.plot(cov_pct.index, cov_pct.values, color='#4472C4', lw=2)
ax6.fill_between(cov_pct.index, cov_pct.values, alpha=0.2, color='#4472C4')
ax6.axhline(cov_pct.mean(), color='orange', lw=1.5, ls='--',
            label=f'평균 커버리지: {cov_pct.mean():.1f}%')
ax6.set_title('Eq_Mom 커버리지 (선택 종목 중 주식데이터 있는 비율)', fontsize=10, fontweight='bold', pad=4)
ax6.set_ylabel('커버리지 (%)', fontsize=9)
ax6.yaxis.set_major_formatter(mtick.PercentFormatter())
ax6.legend(fontsize=9); ax6.grid(True, alpha=0.2)
ax6.set_ylim(0, 110)

fig.suptitle(
    f'Bloomberg US IG Corp -- B+M+A (33/33/34) Backtest  '
    f'({res.index[0].strftime("%Y-%m")} ~ {res.index[-1].strftime("%Y-%m")})\n'
    f'Bond_TR x{W_BOND} + Eq_Mom x{W_MOM} + AI_Macro x{W_AI}  | Top{TOP_N}/Class | Ticker col 개선',
    fontsize=13, fontweight='bold', y=0.97
)

plt.savefig(OUT_PNG, dpi=150, bbox_inches='tight', facecolor='white')
print(f"\n[DONE] Chart: {OUT_PNG}")

# ═══════════════════════════════════════════════════════════════════════════════
# 9. Excel
# ═══════════════════════════════════════════════════════════════════════════════
try:
    from openpyxl import Workbook as OWB
    from openpyxl.styles import Font as OF, PatternFill as OFill, Alignment as OA

    wb = OWB()
    ws = wb.active; ws.title = 'Monthly'
    hdrs = ['Date','Port(%)','Bench(%)','Excess(%)','Classes','Selected','Matched','MomCov(%)']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = OF(bold=True, color='FFFFFF')
        c.fill = OFill('solid', fgColor='1F3864')
        c.alignment = OA(horizontal='center')
    for ri, (idx, row) in enumerate(res.iterrows(), 2):
        bv = row['bench_tr']*100
        pv = row['port_tr']*100
        vals = [idx.strftime('%Y-%m'), round(pv,3), round(bv,3),
                round(pv-bv,3), row['n_classes'], row['n_selected'], row['n_matched'],
                round(row['n_mom_valid']/row['n_selected']*100,1) if row['n_selected']>0 else 0]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=ri, column=ci, value=v)
    wb.save(OUT_XLSX)
    print(f"[DONE] Excel: {OUT_XLSX}")
except Exception as e:
    print(f"[WARN] Excel failed: {e}")
